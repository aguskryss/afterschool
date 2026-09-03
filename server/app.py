from flask import Flask, request, jsonify, send_from_directory, Response, redirect
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, create_refresh_token, jwt_required, get_jwt_identity, get_jwt
import base64
import bcrypt
import binascii
from collections import namedtuple
# Aliased rather than `import html`: `html` is a local variable holding a
# rendered body all over this file, and the shadowing would be a landmine.
from html import escape as _escape
import psycopg2
import os
import secrets
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
import csv
import io
import json
import threading
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from pywebpush import webpush, WebPushException
    PUSH_AVAILABLE = True
except ImportError:
    PUSH_AVAILABLE = False

try:
    import pyotp
    import scheduler
    TOTP_AVAILABLE = True
except ImportError:
    TOTP_AVAILABLE = False

from flask_jwt_extended import decode_token, verify_jwt_in_request
try:
    from server.database import get_db, get_db_transaction, init_db
    from server import tenancy
    from server import database as database_module
    from server import bulk_invites
    from server import photo_storage
    from server import roster_staging
    from server import scheduler
    from server import daily_routing
    from server import pickup_import
    from server.roster_import import parse_grade
except ImportError:
    from database import get_db, get_db_transaction, init_db
    import tenancy
    import database as database_module
    import bulk_invites
    import photo_storage
    import roster_staging
    import scheduler
    import daily_routing
    import pickup_import
    from roster_import import parse_grade

try:
    from server import pickup_events
except ImportError:
    import pickup_events

# ─── RATE LIMITING (persisted in login_attempts table) ────────────────────
_MAX_LOGIN_ATTEMPTS = 5
_ACCOUNT_LOCKOUT_MINUTES = 15

def _is_rate_limited(ip):
    """Check if an IP is rate-limited for login attempts in the last 15 minutes."""
    if not ip:
        return False
    try:
        db = get_db()
        row = db.execute(
            """SELECT COUNT(*) AS n FROM login_attempts
               WHERE ip_address = %s AND attempted_at > NOW() - INTERVAL '15 minutes'""",
            (ip,)
        ).fetchone()
        db.close()
        return (row['n'] if row else 0) >= _MAX_LOGIN_ATTEMPTS
    except Exception as e:
        print(f"[ERROR] _is_rate_limited: {e}")
        return False

def _record_login_attempt(ip, email=''):
    """Record a failed login attempt in the DB."""
    if not ip:
        return
    try:
        db = get_db()
        db.execute(
            "INSERT INTO login_attempts (email, ip_address) VALUES (%s, %s)",
            (email or '', ip)
        )
        db.commit()
        db.close()
    except Exception as e:
        print(f"[ERROR] _record_login_attempt: {e}")

def _clear_login_attempts(ip):
    """Clear login attempts for an IP on successful login."""
    if not ip:
        return
    try:
        db = get_db()
        db.execute("DELETE FROM login_attempts WHERE ip_address = %s", (ip,))
        db.commit()
        db.close()
    except Exception as e:
        print(f"[ERROR] _clear_login_attempts: {e}")

def parse_date(date_str):
    """Validate that date_str is YYYY-MM-DD. Returns the string or raises ValueError."""
    datetime.strptime(date_str or '', '%Y-%m-%d')
    return date_str

# CTE that resolves "child IS effectively absent on :date" combining one-off
# absences, active recurring rules for the date's weekday, minus per-day
# exceptions. Parameter order when used: (date, dow, date, date, date).
ABSENT_CTE = """
WITH absent AS (
    SELECT child_id FROM absences WHERE absence_date = %s
    UNION
    SELECT child_id FROM recurring_absences
     WHERE day_of_week = %s
       AND start_date <= %s
       AND (end_date IS NULL OR end_date >= %s)
)
SELECT child_id FROM absent
WHERE child_id NOT IN (
    SELECT child_id FROM absence_exceptions WHERE exception_date = %s
)
"""

def absent_child_ids_for_date(db, date_str):
    """Return set of child IDs effectively absent on date_str.

    Combines one-off absences and active recurring rules for the date's
    weekday, minus any absence_exceptions overrides.
    """
    dow = datetime.strptime(date_str, '%Y-%m-%d').strftime('%A')
    rows = db.execute(ABSENT_CTE, (date_str, dow, date_str, date_str, date_str)).fetchall()
    return {r['child_id'] for r in rows}


# ─── WHERE A CHILD IS ───────────────────────────────────────────────────────
# One row per child per day carries `status`; see sql/31_add_child_status.sql
# for why these six and not the spec's seven.
#
# `picked_up` is the state the headcount calls "in the building". Collecting a
# child from school and walking them in is one tap by one counselor — there is
# no second tap anywhere in the program, and a state nobody transitions into
# would be permanently, silently wrong. `arrived` slots in between later if
# that ever changes.

STATUS_WAITING = 'waiting'
STATUS_PICKED_UP = 'picked_up'
STATUS_CHECKED_OUT = 'checked_out'
STATUS_ABSENT = 'absent'
STATUS_NOT_FOUND = 'not_found'
STATUS_ISSUE = 'issue'

CHILD_STATUSES = (
    STATUS_WAITING, STATUS_PICKED_UP, STATUS_CHECKED_OUT,
    STATUS_ABSENT, STATUS_NOT_FOUND, STATUS_ISSUE,
)

# The single place that keeps `status` and `on_bus` from disagreeing. Every
# attendance query, export and trend chart in the app still reads on_bus, at
# every JCC, so it is written alongside the status rather than replaced.
_ON_BUS_FOR_STATUS = {
    STATUS_WAITING: 0,
    STATUS_PICKED_UP: 1,
    STATUS_CHECKED_OUT: 1,   # they were here; checked_out_at is what says they left
    STATUS_ABSENT: 0,
    STATUS_NOT_FOUND: 0,
    STATUS_ISSUE: 0,
}


def derive_status(status, on_bus, checked_out_at):
    """The status of an attendance row, for rows that predate the column.

    Every row written before sql/31 has status NULL, and so does every row at a
    JCC whose counselors only ever tap present/absent. Reading those the way
    the app read them before the column existed is what makes applying the
    migration change no number on any screen.
    """
    if status:
        return status
    if checked_out_at:
        return STATUS_CHECKED_OUT
    if on_bus:
        return STATUS_PICKED_UP
    return None


# An assignment applies on a date when the date falls inside its window.
# A NULL bound is an open one, so a row with both NULL is permanent — which is
# what every row written before assignments had dates is.
_ASSIGNMENT_IN_EFFECT = """(
    (cs.effective_from IS NULL OR cs.effective_from <= %s::date)
    AND (cs.effective_to IS NULL OR cs.effective_to >= %s::date)
)"""

# A child belongs to a user either as the primary account (children.parent_id,
# the roster's Contact #1 — required, ON DELETE CASCADE) or as a linked
# second contact (child_contacts.user_id, Contact #2 — optional, ON DELETE
# SET NULL). See the comment on child_contacts in server/database.py for why
# the second seam exists. `c` must be the alias of a `children` row in scope;
# every use takes the same %s twice (user_id, user_id).
_CHILD_BELONGS_TO_USER = """(
    c.parent_id = %s
    OR EXISTS (SELECT 1 FROM child_contacts cc2
                WHERE cc2.child_id = c.id AND cc2.user_id = %s)
)"""


def parent_owns_child(db, user_id, child_id):
    """Whether this user may act on this child — as Contact #1 or a linked
    Contact #2. RLS already keeps the query inside one organization; this is
    the ownership check within it."""
    return db.execute(
        f"SELECT 1 FROM children c WHERE c.id = %s AND {_CHILD_BELONGS_TO_USER}",
        (child_id, user_id, user_id),
    ).fetchone() is not None


def child_owner_ids(db, child_ids):
    """Every user who may act as a parent for any of these children — the
    account children.parent_id points at, plus a linked Contact #2, if any.
    Deduped. Used everywhere a single child's `parent_id` used to be handed
    straight to notify_parent() or a pickup_events.publish() payload — now
    both accounts hear about it, not just Contact #1.
    """
    if not child_ids:
        return []
    rows = db.execute("""
        SELECT DISTINCT parent_id AS user_id FROM children WHERE id = ANY(%s)
        UNION
        SELECT DISTINCT user_id FROM child_contacts
         WHERE child_id = ANY(%s) AND user_id IS NOT NULL
    """, (child_ids, child_ids)).fetchall()
    return [r['user_id'] for r in rows]


def child_owner_map(db, child_ids):
    """child_owner_ids, for many children at once, in one round trip.

    A loop calling child_owner_ids() once per child — and notify_parent()
    once per owner inside that, each opening its own connection — is fine
    for a single release or a single photo (at most two owners) but was
    exactly the bug that broke the daily attendance check: with maxconn=8
    (server/database.py) and real concurrent traffic on top of it, a
    hundred-plus-child organization ran the pool dry mid-loop. Same shape
    of fix _run_broadcast_pushes already uses for the same reason.
    """
    if not child_ids:
        return {}
    rows = db.execute("""
        SELECT id AS child_id, parent_id AS user_id FROM children
         WHERE id = ANY(%s)
        UNION
        SELECT cc.child_id, cc.user_id
          FROM child_contacts cc
         WHERE cc.child_id = ANY(%s) AND cc.user_id IS NOT NULL
    """, (child_ids, child_ids)).fetchall()
    out: dict[int, list[int]] = {}
    for r in rows:
        out.setdefault(r['child_id'], []).append(r['user_id'])
    return out


def counselor_school_ids(db, user_id, on_date=None):
    """The schools a counselor covers on a date, or None for "all of them".

    Three states, and telling the second from the third is the whole point:

      * No assignment rows at all      -> None, meaning every school.
      * Rows, some in effect on `on_date` -> just those.
      * Rows, none in effect on `on_date` -> [], meaning no schools.

    The first is fail-open on purpose. A JCC that has not divided its staff up
    yet needs them working on day one, and a counselor staring at an empty
    roster cannot mark anyone present or release anyone. That is the state a
    new JCC lands in by default, so it has to be the usable one.

    The third exists only because assignments gained dates, and it must NOT
    collapse into the first. A counselor covering one school for one afternoon
    has rows; when that window closes, treating them as unassigned would hand
    them every school in the organization at midnight — the exact opposite of
    what the admin asked for, arriving silently.

    Either way this is bounded by the tenant: RLS keeps every one of these
    queries inside the caller's organization, so a counselor can never reach
    another JCC's children whatever this returns.
    """
    on_date = on_date or today_for_org(db)
    rows = db.execute(f"""
        SELECT cs.school_id,
               {_ASSIGNMENT_IN_EFFECT} AS in_effect
          FROM counselor_schools cs
         WHERE cs.counselor_id = %s
    """, (on_date, on_date, user_id)).fetchall()
    if not rows:
        return None
    return [r['school_id'] for r in rows if r['in_effect']]


def counselor_covers_child(db, user_id, child_id, on_date=None):
    """Whether a counselor may act on this child on a date.

    True when the child's school is one of theirs, and also when they have no
    assignment at all — see counselor_school_ids for why that reads as "all"
    rather than "none", and why an expired assignment does not.
    """
    school_ids = counselor_school_ids(db, user_id, on_date)
    if school_ids is None:
        row = db.execute("SELECT 1 FROM children WHERE id = %s", (child_id,)).fetchone()
    elif not school_ids:
        return False
    else:
        row = db.execute(
            "SELECT 1 FROM children WHERE id = %s AND school_id = ANY(%s)",
            (child_id, school_ids),
        ).fetchone()
    return row is not None

# Load or generate VAPID keys — persist in DB so they survive Render deploys
VAPID_PRIVATE_KEY = None
VAPID_PUBLIC_KEY = None

def _generate_vapid_keys_raw():
    """Generate a new VAPID key pair as base64url-encoded strings (the format
    pywebpush expects). Private key is the raw 32-byte P-256 scalar; public
    key is the 65-byte uncompressed X9.62 point."""
    from cryptography.hazmat.primitives.asymmetric import ec
    from cryptography.hazmat.primitives import serialization
    import base64
    private_key = ec.generate_private_key(ec.SECP256R1())
    raw_private = private_key.private_numbers().private_value.to_bytes(32, 'big')
    private_b64url = base64.urlsafe_b64encode(raw_private).rstrip(b'=').decode()
    raw_public = private_key.public_key().public_bytes(
        encoding=serialization.Encoding.X962,
        format=serialization.PublicFormat.UncompressedPoint
    )
    public_b64url = base64.urlsafe_b64encode(raw_public).rstrip(b'=').decode()
    return {'private_key': private_b64url, 'public_key': public_b64url}

def _load_or_create_vapid_keys():
    """Load VAPID keys from env vars or DB; generate + store if missing."""
    global VAPID_PRIVATE_KEY, VAPID_PUBLIC_KEY

    # 1) Prefer environment variables (most reliable for production)
    env_priv = os.environ.get('VAPID_PRIVATE_KEY', '')
    env_pub = os.environ.get('VAPID_PUBLIC_KEY', '')
    if env_priv and env_pub:
        VAPID_PRIVATE_KEY = env_priv
        VAPID_PUBLIC_KEY = env_pub
        print("VAPID keys loaded from environment variables")
        return

    # 2) Try loading from database (persists across Render deploys)
    try:
        import psycopg2, psycopg2.extras
        # The owner role: this both creates app_settings and writes to it, and
        # the low-privilege application role deliberately cannot create tables.
        DATABASE_URL = (os.environ.get('ADMIN_DATABASE_URL')
                        or os.environ.get('DATABASE_URL', ''))
        if not DATABASE_URL:
            # Stay quiet: _check_required_config below reports the missing
            # database properly. Announcing it here as a VAPID problem buries
            # the real cause under an unrelated symptom.
            return
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS app_settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        """)
        # app_settings is infrastructure, not tenant data, so the application
        # role reads it without any organization pinned.
        cur.execute("GRANT SELECT, INSERT, UPDATE ON app_settings TO PUBLIC")
        conn.commit()
        cur.execute("SELECT value FROM app_settings WHERE key = 'vapid_private_key'")
        row_priv = cur.fetchone()
        cur.execute("SELECT value FROM app_settings WHERE key = 'vapid_public_key'")
        row_pub = cur.fetchone()

        # Legacy PEM-format keys cannot be consumed by pywebpush, which
        # expects a base64url raw 32-byte scalar. Detect and force regenerate.
        legacy_pem = row_priv and '-----BEGIN' in row_priv['value']

        if row_priv and row_pub and not legacy_pem:
            VAPID_PRIVATE_KEY = row_priv['value']
            VAPID_PUBLIC_KEY = row_pub['value']
            conn.close()
            print("VAPID keys loaded from database")
            return

        # 3) Generate new keys and store in DB
        keys = _generate_vapid_keys_raw()
        cur.execute(
            "INSERT INTO app_settings (key, value) VALUES ('vapid_private_key', %s) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
            (keys['private_key'],)
        )
        cur.execute(
            "INSERT INTO app_settings (key, value) VALUES ('vapid_public_key', %s) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
            (keys['public_key'],)
        )
        conn.commit()
        conn.close()
        VAPID_PRIVATE_KEY = keys['private_key']
        VAPID_PUBLIC_KEY = keys['public_key']
        # Also wipe stale push subscriptions since keys changed (or legacy keys
        # never produced deliverable pushes in the first place).
        try:
            conn2 = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
            conn2.cursor().execute("DELETE FROM push_subscriptions")
            conn2.commit()
            conn2.close()
        except:
            pass
        reason = "legacy PEM format detected, migrated" if legacy_pem else "no keys found"
        print(f"VAPID keys generated and saved to database ({reason}); old push subscriptions cleared")
    except Exception as e:
        print(f"Could not load/generate VAPID keys: {e}")

_load_or_create_vapid_keys()

app = Flask(__name__, static_folder='../public', static_url_path='')

# ─── STARTUP CONFIGURATION CHECK ──────────────────────────────────────────
# Report everything that's missing at once. Failing on the first one means a
# deploy-fix-deploy loop where each round surfaces exactly one more variable,
# and the messages that follow are worse than the cause: a missing DATABASE_URL
# announces itself as "Could not load/generate VAPID keys".
def _check_required_config():
    # Computed here rather than read from _is_production, which is defined
    # below: this check has to run before anything else can fail confusingly.
    in_production = (
        os.environ.get('FLASK_ENV', '').lower() == 'production'
        or os.environ.get('RENDER', '') == 'true'
    )
    problems = []
    if not os.environ.get('DATABASE_URL'):
        problems.append(
            'DATABASE_URL — the application database connection. Must be a '
            'session-mode or direct connection (port 5432); the transaction '
            'pooler accepts NOTIFY but never delivers it, so live pickup would '
            'silently reach nobody.'
        )
    if in_production and not os.environ.get('ALLOWED_ORIGINS'):
        problems.append(
            "ALLOWED_ORIGINS — the site's own URL, e.g. "
            "https://afterschool.kikarlabs.com. Required in "
            "production so CORS isn't left open to any origin."
        )
    if not os.environ.get('ADMIN_EMAIL'):
        problems.append(
            'ADMIN_EMAIL — who the first organization\'s administrator will be. '
            'Needed to seed that account on the very first boot.'
        )
    if not problems:
        return
    lines = ['', '=' * 72, 'Kikar Afterschool cannot start. Missing configuration:', '']
    for item in problems:
        lines.append(f'  * {item}')
    lines += [
        '',
        'Set these in the Render dashboard under Environment, then redeploy.',
        'See docs/deploy.md for where each value comes from.',
        '=' * 72,
        '',
    ]
    raise RuntimeError('\n'.join(lines))


_check_required_config()


# ─── CORS: restrict to known origins ──────────────────────────────────────
_is_production = (
    os.environ.get('FLASK_ENV', '').lower() == 'production'
    or os.environ.get('RENDER', '') == 'true'
)
_allowed_origins = os.environ.get('ALLOWED_ORIGINS', '').split(',')
_allowed_origins = [o.strip() for o in _allowed_origins if o.strip()]
if _allowed_origins:
    CORS(app, origins=_allowed_origins, supports_credentials=True)
else:
    if _is_production:
        raise RuntimeError(
            "ALLOWED_ORIGINS must be set in production. Refusing to start with open CORS."
        )
    print("[SECURITY WARNING] ALLOWED_ORIGINS not set — falling back to localhost only for dev. Set ALLOWED_ORIGINS for any non-dev environment.")
    CORS(app, origins=['http://localhost:5001', 'http://127.0.0.1:5001', 'http://localhost:5000', 'http://127.0.0.1:5000'], supports_credentials=True)

# ─── JWT: require strong secret, long-lived sessions ─────────────────────
# Tokens are long-lived on purpose: admins, counselors, and parents should
# stay logged in until they tap Logout manually. The frontend already
# auto-refreshes on page load and on 401 via /api/auth/refresh, so a client
# that's been idle for weeks still comes back signed in.
_jwt_secret = os.environ.get('JWT_SECRET_KEY', '')
if not _jwt_secret:
    if _is_production:
        raise RuntimeError(
            "JWT_SECRET_KEY must be set in production. Refusing to start with an ephemeral key."
        )
    _jwt_secret = secrets.token_urlsafe(64)
    print("[SECURITY WARNING] JWT_SECRET_KEY not set. Generated ephemeral key — set it in environment for production.")
app.config['JWT_SECRET_KEY'] = _jwt_secret
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(days=365)
app.config['JWT_REFRESH_TOKEN_EXPIRES'] = timedelta(days=365)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB max upload

jwt = JWTManager(app)

def current_org_id():
    """Organization of the request in flight, for code that needs it explicitly
    (publishing events, building channel names) rather than implicitly through
    the pinned connection."""
    from flask import g
    return getattr(g, 'organization_id', None)


# ─── TENANT CONTEXT ───────────────────────────────────────────────────────
# Resolved once per request, before any handler runs, so every get_db() in
# that request pins the same organization. Unauthenticated requests leave it
# unset, which makes reads return nothing and writes fail — the safe default.
@app.before_request
def _bind_tenant():
    from flask import g
    g.organization_id = None
    g.is_superadmin = False
    try:
        verify_jwt_in_request(optional=True)
        claims = get_jwt() or {}
    except Exception:
        # A malformed or expired token is not this hook's problem; the route's
        # own @jwt_required will reject it with the right status.
        return
    if claims.get('role') == 'superadmin':
        g.is_superadmin = True
    g.organization_id = claims.get('org')


# ─── MODULE ACCESS ────────────────────────────────────────────────────────
# Modules are what a JCC buys, so refusing one has to happen here. The client
# hides screens for modules that are off (`hasModule`, web/src/lib/auth.ts),
# but that is presentation: anyone holding a valid token can call the endpoint
# directly, and until this hook existed they got the data.
#
# The route-to-module map lives in tenancy.MODULE_ROUTES, next to the module
# catalogue itself.

def _request_modules(db=None):
    """Module settings for the organization in flight, read once per request.

    Pass `db` when the caller already holds a connection. The pool is capped at
    8 (see server/database.py) against 64 gunicorn threads, so a handler that
    checks a module mid-query while holding its own connection would take two
    slots for one request. On module-gated routes the before_request hook has
    already filled this cache and no connection is opened at all; it's the core
    routes that ask about a module partway through — the roster and its
    check_in_out times — where the second checkout would happen.
    """
    from flask import g
    cached = getattr(g, '_org_modules', None)
    if cached is not None:
        return cached
    org_id = getattr(g, 'organization_id', None)
    if org_id is None:
        g._org_modules = {}
        return g._org_modules
    sql = "SELECT modules FROM organizations WHERE id = %s"
    if db is not None:
        row = db.execute(sql, (org_id,)).fetchone()
    else:
        own = get_db()
        try:
            row = own.execute(sql, (org_id,)).fetchone()
        finally:
            own.close()
    g._org_modules = (row['modules'] or {}) if row else {}
    return g._org_modules


def module_on(key, db=None):
    """Whether the request's organization has `key` enabled.

    For handlers that stay reachable either way but change what they return —
    a core endpoint enriching its response with a paid module's data, say.
    Routes that belong wholly to a module are refused by the hook below
    instead, and never reach their handler.

    Pass the connection you already hold, if you hold one; see _request_modules.
    """
    from flask import g
    if getattr(g, 'is_superadmin', False):
        return True
    return tenancy.module_enabled(_request_modules(db), key)


# ─── WHAT DAY IS IT ────────────────────────────────────────────────────────
# Every "today" in this file is a calendar date at a JCC — which children are
# expected, who is marked absent, which day's attendance a counselor is
# filling in. None of them is a timestamp.
#
# datetime.today() answers that with the container's clock, and Render runs
# UTC. From roughly 7pm Eastern onward that clock is already tomorrow, so the
# dashboard would roll over mid-pickup: the weekday changes, the expected
# roster changes, and every attendance row written that afternoon stops
# matching. On a Friday evening it rolls to Saturday, no weekday has
# registrations, and the whole board reads "nothing scheduled".
#
# organizations.timezone has been there the whole time, and until now only
# the spreadsheet exports read it.

DEFAULT_TIMEZONE = 'America/New_York'


def org_timezone(db=None):
    """The IANA zone this organization runs on, read once per request.

    Cached on `g` like the module settings, and for the same reason: the pool
    is capped at 8 against 64 threads, so asking the database what day it is
    on every call would be a connection per date. Pass `db` when the caller
    already holds one.
    """
    try:
        from flask import g
        cached = getattr(g, '_org_timezone', None)
    except RuntimeError:
        # No request context — a background thread. It has no organization to
        # ask about, so the default is the only answer available.
        return DEFAULT_TIMEZONE
    if cached is not None:
        return cached

    org_id = getattr(g, 'organization_id', None)
    tzname = None
    if org_id is not None:
        sql = "SELECT timezone FROM organizations WHERE id = %s"
        try:
            if db is not None:
                row = db.execute(sql, (org_id,)).fetchone()
            else:
                own = get_db()
                try:
                    row = own.execute(sql, (org_id,)).fetchone()
                finally:
                    own.close()
            tzname = row['timezone'] if row else None
        except Exception as e:
            # Never fail a request over this. A wrong-by-hours date is bad;
            # a 500 on the dashboard is worse.
            print(f"[WARN] org_timezone: {e}")
    g._org_timezone = tzname or DEFAULT_TIMEZONE
    return g._org_timezone


def now_for_org(db=None):
    """The current local time at the JCC, as an aware datetime.

    Use this for anything that needs the parts of a date — the year, the
    month, the weekday, or arithmetic like "thirty days ago".
    """
    try:
        tz = ZoneInfo(org_timezone(db))
    except Exception:
        # An unknown zone in the row shouldn't take the request down with it.
        tz = ZoneInfo(DEFAULT_TIMEZONE)
    return datetime.now(tz)


def today_for_org(db=None):
    """Today's date at the JCC as YYYY-MM-DD — the default for every endpoint
    that takes an optional ?date=."""
    return now_for_org(db).strftime('%Y-%m-%d')


@app.before_request
def _enforce_module_access():
    from flask import g
    # Preflights carry no credentials; refusing them breaks CORS rather than
    # protecting anything. The real request that follows is still checked.
    if request.method == 'OPTIONS':
        return None
    key = tenancy.module_for_path(request.path)
    if key is None:
        return None
    # Superadmins administer the catalogue itself, so no module can lock them
    # out. Requests with no organization yet are left to @jwt_required, which
    # rejects them with 401 rather than a misleading 403.
    if getattr(g, 'is_superadmin', False) or getattr(g, 'organization_id', None) is None:
        return None
    if not tenancy.module_enabled(_request_modules(), key):
        return jsonify({'error': 'This feature is not enabled for your organization'}), 403
    return None


# ─── SECURITY HEADERS ─────────────────────────────────────────────────────
@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    if request.is_secure or os.environ.get('FORCE_HTTPS'):
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    # CSP: allow inline styles/scripts for the existing SPA frontend.
    # Multiple script CDNs listed so Chart.js can fall back if one is blocked.
    #
    # img-src ALLOWS ANY https HOST, DELIBERATELY.
    #   Every image this product shows comes from somewhere else. Photographs
    #   are signed URLs on Supabase; an organization's logo is either an
    #   uploaded object on Supabase or a link to the JCC's own website, which
    #   can be any host at all. `'self' data:` blocked all of it — the logo
    #   rendered as a broken icon with nothing saying why, and the photo
    #   gallery had the same fault, unnoticed because the `photos` module is
    #   off by default.
    #
    #   Naming the Supabase origin instead would fix the logo we upload and
    #   leave the pasted-URL fallback broken, which is the same silent failure
    #   one step later.
    #
    #   What this costs is small and worth being clear about: an injected
    #   <img> could signal a load, or smuggle data into a URL. What it does not
    #   cost is code execution — images do not execute, and script-src is
    #   unchanged. Note that script-src already allows 'unsafe-inline', so an
    #   attacker who can inject markup here has a far better tool than an image
    #   tag; tightening this directive while that one stands would be theatre.
    #   The order to fix them in is script-src first.
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://unpkg.com https://cdnjs.cloudflare.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdnjs.cloudflare.com; "
        "font-src 'self' https://fonts.gstatic.com https://cdnjs.cloudflare.com; "
        "img-src 'self' data: https:; "
        "connect-src 'self'; "
        "frame-ancestors 'none';"
    )
    return response

try:
    from server.superadmin import bp as superadmin_bp
except ImportError:
    from superadmin import bp as superadmin_bp
app.register_blueprint(superadmin_bp)

# Initialize database on import (needed for gunicorn)
init_db()

# Row level security is the backstop under the per-request scoping. A role with
# BYPASSRLS ignores every policy, which would leave the isolation configured
# but inert — so say so loudly at boot rather than discovering it after a leak.
try:
    _rls_db = get_db()
    if not tenancy.rls_is_enforced(_rls_db.execute("SELECT 1")):
        print("[SECURITY WARNING] The application database role bypasses row "
              "level security. Tenant isolation is relying on application "
              "scoping alone. Point DATABASE_URL at a role without BYPASSRLS.")
    else:
        print("[tenancy] Row level security is enforced for this connection.")
    _rls_db.close()
except Exception as _e:
    print(f"[tenancy] Could not verify row level security: {_e}")

# Optionally seed permanent test accounts (parent + counselor + school + children).
# Enable by setting SEED_TEST_ACCOUNTS=1 in the environment. Idempotent.
if os.environ.get('SEED_TEST_ACCOUNTS', '').lower() in ('1', 'true', 'yes'):
    try:
        try:
            from server.seed_test_data import seed_test_accounts
        except ImportError:
            from seed_test_data import seed_test_accounts
        seed_test_accounts(verbose=True)
    except Exception as e:
        print(f"[seed_test_accounts] skipped due to error: {e}")

import re
def validate_password(pw):
    """Returns error message if password is weak, or None if strong enough."""
    if len(pw) < 8:
        return 'Password must be at least 8 characters.'
    if not re.search(r'[A-Z]', pw):
        return 'Password must contain at least one uppercase letter.'
    if not re.search(r'[!@#$%^&*()_+\-=\[\]{}|;:\'",.<>?/\\`~]', pw):
        return 'Password must contain at least one special character (!@#$%^&* etc.).'
    return None

# Email config.
#
# Render blocks outbound SMTP (port 587/465 are not routable from the host —
# any connect to smtp.gmail.com hangs ~135s and dies with "Errno 101 Network
# is unreachable"). To work around this we route email through Resend's HTTPS
# API by default. Set EMAIL_PROVIDER=smtp explicitly if you want the legacy
# SMTP path (e.g. for local dev).
EMAIL_CONFIG = {
    'provider': (os.environ.get('EMAIL_PROVIDER') or 'resend').lower(),
    'smtp_host': 'smtp.gmail.com',
    'smtp_port': 587,
    'smtp_user': os.environ.get('EMAIL_USER', ''),
    'smtp_pass': os.environ.get('EMAIL_PASS', ''),
    'resend_api_key': os.environ.get('RESEND_API_KEY', ''),
    'from_name': os.environ.get('EMAIL_FROM_NAME', 'Kikar Afterschool'),
    # El dominio del remitente tiene que estar verificado en Resend o la API
    # responde 403 y no sale nada. El default es kikarlabs.com; setear
    # EMAIL_FROM en Render para no depender de él.
    'from_email': os.environ.get(
        'EMAIL_FROM',
        os.environ.get('EMAIL_USER', 'noreply@kikarlabs.com'),
    ),
    # rstrip: a trailing slash here doubles up with the leading '/' every
    # caller appends (e.g. BASE_URL="https://x.com/" -> "https://x.com//app/home").
    'base_url': os.environ.get('BASE_URL', 'http://localhost:5000').rstrip('/'),
}

# The domain every organization sends from lives in tenancy.EMAIL_SENDER_DOMAIN,
# next to the schema and the CHECK constraint that make it unforgeable. See the
# comment there before changing how a sender address is built.

# Distinguishes "no URL was supplied" from "this photo has no URL", which is a
# real state: an object Supabase would not sign renders as a gap in the grid.
_UNSET = object()

# The teal the invitation templates used before any of them could be branded.
# Still the fallback for an organization that has not set brand_primary.
DEFAULT_BRAND_PRIMARY = '#005F6B'

# Everything about an outgoing email that depends on *which* JCC it is from.
#
# Resolved once, in request context, and passed down — never looked up inside
# the send functions. See email_identity() for why that distinction matters.
EmailIdentity = namedtuple(
    'EmailIdentity',
    'from_email from_name reply_to org_name logo_url brand_primary',
)


def platform_identity():
    """Kikar itself, for mail that belongs to no organization.

    Also the fallback when an organization's row cannot be read. A send that
    goes out with platform branding is worse than one with the JCC's — but it
    is much better than an account creation that fails because a SELECT did.
    """
    return EmailIdentity(
        from_email=EMAIL_CONFIG['from_email'],
        from_name=EMAIL_CONFIG['from_name'],
        reply_to=None,
        org_name=EMAIL_CONFIG['from_name'],
        logo_url=None,
        brand_primary=DEFAULT_BRAND_PRIMARY,
    )


def _ambient_org_id():
    """The organization of whatever context we are in, or None.

    Three different places know this and they do not overlap:

      * a normal request       flask.g, set by _bind_tenant
      * the bulk invite job    the thread pin, set by _run_all_invites_job
      * a bare worker thread   neither — and `g` does not merely return None
                               there, it raises RuntimeError

    That last case is the whole reason this function exists rather than a call
    to current_org_id(): reading `g` outside an app context is an exception, so
    a send from a plain threading.Thread would blow up instead of degrading.
    """
    try:
        from flask import g, has_app_context
        if has_app_context():
            org_id = getattr(g, 'organization_id', None)
            if org_id:
                return org_id
    except Exception:
        pass
    return getattr(database_module._thread_local, 'organization_id', None)


def email_identity(db=None):
    """Resolve who this organization's mail comes from.

    CALL THIS IN REQUEST CONTEXT AND PASS THE RESULT DOWN. Every send function
    below takes an identity as a required argument rather than resolving one
    itself, and that is deliberate: send_email_async() hands the work to a bare
    threading.Thread, where there is no `g` and no thread pin, so a lookup done
    at send time would quietly produce platform branding for a JCC's parents.
    Requiring the argument turns that into a TypeError at import time instead of
    a wrong From: header nobody notices.

    Passing `db` reuses a connection the caller already holds. Worth doing: the
    pool is capped at 8 (server/database.py) and opening a second connection
    while holding one narrows it for everybody.

    Never raises. An organization that cannot be read falls back to Kikar.
    """
    org_id = _ambient_org_id()
    if not org_id:
        return platform_identity()

    owned = db is None
    try:
        if owned:
            db = get_db()
        row = db.execute(
            "SELECT slug, name, logo_url, logo_path, brand_primary, "
            "       email_from_local, email_from_name, email_reply_to "
            "  FROM organizations WHERE id = %s",
            (org_id,),
        ).fetchone()
    except Exception as e:
        print(f"email identity lookup failed for org {org_id}: {e.__class__.__name__}")
        return platform_identity()
    finally:
        if owned and db is not None:
            try:
                db.close()
            except Exception:
                pass

    if not row:
        return platform_identity()

    # The slug is the fallback local part, so `jccns` sends as
    # jccns@kikarlabs.com with nothing configured. email_from_local overrides it
    # when a JCC wants a different word in front of the @. A slug that is not a
    # legal local part (none are today — _SLUG is stricter) falls back to the
    # platform address rather than building a malformed one.
    from_email = (
        tenancy.org_from_email(row['email_from_local'], row['slug'])
        or EMAIL_CONFIG['from_email']
    )
    reply_to = (row['email_reply_to'] or '').strip() or None
    return EmailIdentity(
        from_email=from_email,
        from_name=(row['email_from_name'] or row['name'] or EMAIL_CONFIG['from_name']),
        reply_to=reply_to,
        org_name=row['name'] or EMAIL_CONFIG['from_name'],
        # The uploaded logo is why this bucket is public: an invitation opened
        # three weeks later still has to render it, and a signed URL would be
        # long dead by then.
        logo_url=tenancy.org_logo_url(row),
        brand_primary=(row['brand_primary'] or '').strip() or DEFAULT_BRAND_PRIMARY,
    )

# Gate for parent invitation emails. Set PARENT_INVITES_ENABLED=1 to allow sending
# invitation emails to parents. While disabled, parent accounts are still created
# with a setup token, but no email is dispatched.
PARENT_INVITES_ENABLED = os.environ.get('PARENT_INVITES_ENABLED', '') == '1'


def _send_via_resend(to_email, subject, html_body, identity):
    """Send a single email through Resend's HTTPS API (port 443, works on Render).
    Returns (ok, reason). reason is None on success, otherwise a short string
    describing why so the admin UI can show "rate_limit" vs "bounced" vs "network".

    `identity` is required — see email_identity() for why it is not resolved
    here. Its from_email is always on EMAIL_SENDER_DOMAIN, which is the domain
    verified in Resend; sending from an unverified one comes back 403."""
    api_key = EMAIL_CONFIG['resend_api_key']
    if not api_key:
        print(f"[EMAIL SKIPPED - no Resend API key] To: {to_email} | Subject: {subject}")
        return False, 'no Resend API key configured'
    try:
        import requests as _requests
        payload = {
            'from': f"{identity.from_name} <{identity.from_email}>",
            'to': [to_email],
            'subject': subject,
            'html': html_body,
        }
        # Only when the JCC has one. An empty reply_to is not the same as an
        # absent one: Resend rejects the empty string rather than ignoring it.
        if identity.reply_to:
            payload['reply_to'] = identity.reply_to
        resp = _requests.post(
            'https://api.resend.com/emails',
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json',
            },
            json=payload,
            timeout=30,
        )
        if resp.status_code >= 400:
            body_preview = resp.text[:300]
            print(f"Resend error ({resp.status_code}) for {to_email}: {body_preview}")
            if resp.status_code == 429:
                short = 'rate_limit (429)'
            elif resp.status_code == 422:
                short = 'invalid recipient (422)'
            elif resp.status_code in (401, 403):
                short = f'auth/domain ({resp.status_code})'
            else:
                short = f'HTTP {resp.status_code}'
            return False, short
        try:
            resend_id = resp.json().get('id', '<no-id>')
        except Exception:
            resend_id = '<unparseable>'
        print(f"Resend OK for {to_email}: id={resend_id}")
        return True, None
    except Exception as e:
        print(f"Resend send error for {to_email}: {e}")
        return False, f'network: {e.__class__.__name__}'


def _send_via_smtp(to_email, subject, html_body, identity):
    """Legacy SMTP path. Note: outbound SMTP is blocked on Render — keep this
    only for local dev."""
    if not EMAIL_CONFIG['smtp_user'] or not EMAIL_CONFIG['smtp_pass']:
        print(f"[EMAIL SKIPPED - no SMTP config] To: {to_email} | Subject: {subject}")
        return False
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = f"{identity.from_name} <{identity.from_email}>"
        msg['To'] = to_email
        if identity.reply_to:
            msg['Reply-To'] = identity.reply_to
        msg.attach(MIMEText(html_body, 'html'))
        with smtplib.SMTP(EMAIL_CONFIG['smtp_host'], EMAIL_CONFIG['smtp_port']) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['smtp_user'], EMAIL_CONFIG['smtp_pass'])
            # The envelope sender stays the authenticated SMTP account: Gmail
            # rejects a MAIL FROM it does not own, so only the header carries
            # the organization here. On the Resend path — the one that actually
            # runs in production — both are the organization's address.
            server.sendmail(EMAIL_CONFIG['smtp_user'], to_email, msg.as_string())
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False


def send_email(to_email, subject, html_body, identity):
    """`identity` is required on purpose. See email_identity()."""
    if EMAIL_CONFIG['provider'] == 'resend':
        ok, _reason = _send_via_resend(to_email, subject, html_body, identity)
        return ok
    return _send_via_smtp(to_email, subject, html_body, identity)

def _issue_setup_token(db, user_id, hours_valid=168):
    """Create a one-time password-setup token in password_reset_tokens and return its URL."""
    token = secrets.token_urlsafe(32)
    expires = (datetime.now() + timedelta(hours=hours_valid)).strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        "INSERT INTO password_reset_tokens (user_id, token, expires_at) VALUES (%s, %s, %s)",
        (user_id, token, expires)
    )
    return f"{EMAIL_CONFIG['base_url']}/reset-password?token={token}"

def _email_shell(identity, heading, inner_html):
    """The frame every outgoing email shares, painted in the JCC's colours.

    organizations already carried logo_url, brand_primary and brand_accent
    before any of this — the console has been collecting them for a while and
    email was the one surface that never read them. The header colour and the
    logo come from those columns now, so a parent at one JCC does not receive a
    message wearing another one's branding.
    """
    logo = ''
    if identity.logo_url:
        logo = (
            f'<img src="{_escape(identity.logo_url, quote=True)}" alt="" '
            f'style="max-height:44px;margin:0 0 12px;display:block;border:0;">'
        )
    return f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <div style="background: {_escape(identity.brand_primary, quote=True)}; color: white; padding: 24px; border-radius: 8px 8px 0 0;">
            {logo}<h1 style="margin: 0; font-size: 22px;">{_escape(heading)}</h1>
        </div>
        <div style="background: #f8fafc; padding: 24px; border-radius: 0 0 8px 8px;">
            {inner_html}
        </div>
    </div>
    """


def _cta(identity, url, label):
    """The one button in each email, in the organization's primary colour."""
    return (
        f'<a href="{_escape(url, quote=True)}" style="display:inline-block;'
        f'background:{_escape(identity.brand_primary, quote=True)};color:white;padding:12px 24px;'
        f'border-radius:6px;text-decoration:none;margin-top:12px;font-weight:600;">'
        f'{_escape(label)}</a>'
    )


def _build_invitation_html(user_name, setup_url, identity):
    brand = _escape(identity.brand_primary, quote=True)
    org = _escape(identity.org_name)
    # Reply-To changes what we can honestly tell a parent to do. Without one,
    # replies reach nobody, so the footer has to say so rather than invite them.
    #
    # This used to carry a hardcoded phone number. It was one JCC's support line
    # printed in every organization's mail, which is wrong the moment there are
    # two of them — and it is the reason to set email_reply_to per organization
    # instead: a JCC's own contact route belongs in its own row.
    if identity.reply_to:
        footer = ('This link expires in 7 days. Questions? Just reply to this '
                  'email and it will reach us.')
    else:
        footer = (f'This link expires in 7 days. Please do not reply to this '
                  f'email — for any questions, contact {org} directly.')
    return _email_shell(identity, f"Welcome to {identity.org_name}.", f"""
            <p>Hi {_escape(user_name)},</p>
            <p>You've been added as a parent in the new {org} app — where you'll mark absences, get attendance updates, and tap "I'm Here for Pickup" when you arrive at the playground.</p>
            <p>Tap the button below to create your password and get started.</p>
            {_cta(identity, setup_url, 'Create My Password')}
            <div style="margin-top:28px;padding-top:20px;border-top:1px solid #e2e8f0;">
                <h2 style="margin:0 0 12px;color:{brand};font-family:Arial,sans-serif;font-size:16px;font-weight:700;">Quick Setup</h2>
                <ol style="margin:0;padding-left:20px;color:#334155;font-size:14px;line-height:1.6;">
                    <li style="margin-bottom:10px;"><strong style="color:{brand};">Set your password</strong> — tap the button above.</li>
                    <li style="margin-bottom:10px;">
                        <strong style="color:{brand};">Install the app on your phone</strong> — after logging in:
                        <ul style="margin:6px 0 0;padding-left:18px;color:#334155;">
                            <li style="margin-bottom:4px;"><strong>iPhone (Safari):</strong> tap Share → Add to Home Screen → Add.</li>
                            <li><strong>Android (Chrome):</strong> tap the three-dot menu → Install app → Install.</li>
                        </ul>
                    </li>
                    <li><strong style="color:{brand};">Allow notifications</strong> when prompted — so you never miss an update.</li>
                </ol>
            </div>
            <p style="margin-top:24px;color:#64748b;font-size:14px;">{footer}</p>
    """)


def parent_invitation_subject(identity):
    return f"You're in! Set up your {identity.org_name} parent account."


def send_invitation(user_email, user_name, setup_url, identity):
    return send_email(
        user_email,
        parent_invitation_subject(identity),
        _build_invitation_html(user_name, setup_url, identity),
        identity,
    )

def send_counselor_invitation(user_email, user_name, setup_url, identity):
    org = _escape(identity.org_name)
    html = _email_shell(identity, f"Welcome to {identity.org_name}!", f"""
            <p>Hello {_escape(user_name)},</p>
            <p>You have been added as a <strong>counselor</strong> in the {org} attendance system.</p>
            <p>Click below to set your password and access your account.</p>
            {_cta(identity, setup_url, 'Set My Password')}
            <p style="margin-top:24px;color:#64748b;font-size:14px;">This link expires in 7 days. If you have any questions, contact the {org} administrator.</p>
    """)
    return send_email(
        user_email,
        f"Welcome to {identity.org_name} — Set Up Your Account",
        html,
        identity,
    )

def send_admin_invitation(user_email, user_name, setup_url, identity):
    org = _escape(identity.org_name)
    html = _email_shell(identity, f"Welcome to {identity.org_name} Admin", f"""
            <p>Hello {_escape(user_name)},</p>
            <p>You have been added as an <strong>administrator</strong> of the {org} program. As an admin you can manage parents, counselors, schools, attendance, calendar events, and other admins.</p>
            <p>Click below to set your password and access your admin account.</p>
            {_cta(identity, setup_url, 'Set My Password')}
            <p style="margin-top:24px;color:#64748b;font-size:14px;">This link expires in 7 days. For security, do not share it with anyone. If you did not expect this invitation, please ignore this email.</p>
    """)
    return send_email(
        user_email,
        f"You're now a {identity.org_name} admin - Set Up Your Account",
        html,
        identity,
    )


def _notification_email_html(identity, title, body, url):
    """The email twin of a parent push — same title and body, wrapped in the
    organization's own branding.

    `body` is escaped: unlike an invitation's fixed copy, this carries text a
    staff member typed (an office reply, a broadcast) straight into an HTML
    email, and unescaped that is a way to inject markup into every parent's
    inbox.
    """
    return _email_shell(identity, title, f"""
            <p style="color:#334155;font-size:15px;line-height:1.6;">{_escape(body)}</p>
            {_cta(identity, f"{EMAIL_CONFIG['base_url']}{url}", 'Open in the app')}
    """)

# ─── AUTH ROUTES ────────────────────────────────────────────────────────────

@app.route('/api/health', methods=['GET'])
def health():
    """Unauthenticated liveness probe for the platform.

    Touches the database so a healthy answer means the app can actually serve
    a request, not merely that the process is up. Deliberately says nothing
    about which organizations exist.
    """
    try:
        db = get_db()
        try:
            db.execute('SELECT 1')
        finally:
            db.close()
    except Exception:
        return jsonify({'status': 'degraded'}), 503
    return jsonify({'status': 'ok'})


@app.route('/api/public/branding', methods=['GET'])
def public_branding():
    """Name, logo and colours for the sign-in screen, before anyone has a token.

    Single-organization deployment: there is exactly one row to show, and
    showing it here is safe because none of these four columns is sensitive —
    a JCC's name and logo are already on their own public website. Bypasses
    RLS the same way superadmin routes do (g.is_superadmin = True) because an
    anonymous visitor has no organization_id for the policy to match against.
    """
    from flask import g
    g.is_superadmin = True
    db = get_db()
    try:
        row = db.execute(
            'SELECT name, logo_path, logo_url, brand_primary, brand_accent '
            'FROM organizations ORDER BY id LIMIT 1'
        ).fetchone()
    finally:
        db.close()
    if not row:
        return jsonify({'name': None, 'logo_url': None, 'brand_primary': None, 'brand_accent': None})
    # Same resolution as everywhere else that shows a logo: an uploaded file
    # wins over a pasted link. Getting this wrong here specifically would
    # mean the sign-in screen shows a stale URL nobody can see (see
    # tenancy.org_logo_url's docstring) while every signed-in screen shows
    # the right one — a mismatch that looks like the upload failed.
    return jsonify({
        'name': row['name'],
        'logo_url': tenancy.org_logo_url(row),
        'brand_primary': row['brand_primary'],
        'brand_accent': row['brand_accent'],
    })


@app.route('/api/auth/login', methods=['POST'])
def login():
    client_ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()

    # Rate limiting check
    if _is_rate_limited(client_ip):
        return jsonify({'error': f'Too many login attempts. Please try again in {_ACCOUNT_LOCKOUT_MINUTES} minutes.'}), 429

    data = request.json or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    totp_code = data.get('totp_code', '')

    if not email or not password:
        return jsonify({'error': 'Email and password are required'}), 400

    db = get_db()
    # auth_lookup() is SECURITY DEFINER: at this point no organization is
    # pinned, so a plain SELECT on users would be hidden by RLS. See
    # tenancy._ensure_auth_lookup for why this is a function and not a
    # loophole in the policy.
    user = db.execute("SELECT * FROM auth_lookup(%s)", (email,)).fetchone()

    if not user or not bcrypt.checkpw(password.encode(), user['password_hash'].encode()):
        db.close()
        _record_login_attempt(client_ip)
        return jsonify({'error': 'Invalid email or password'}), 401

    # Backfill password_set_at for users created before this column existed.
    # If they can log in, they've set their own password (the random hash from
    # account creation is never given to them).
    if user.get('password_set_at') is None:
        db.execute("UPDATE users SET password_set_at = NOW() WHERE id = %s AND password_set_at IS NULL", (user['id'],))
        db.commit()

    # Check if 2FA is enabled for this user
    totp_row = db.execute("SELECT * FROM user_totp WHERE user_id = %s AND enabled = 1", (user['id'],)).fetchone()
    db.close()

    if totp_row:
        if not TOTP_AVAILABLE:
            return jsonify({'error': 'Server 2FA module not available. Contact administrator.'}), 500
        if not totp_code:
            return jsonify({'requires_2fa': True, 'message': 'Please enter your 2FA code'}), 200
        totp = pyotp.TOTP(totp_row['secret'])
        if not totp.verify(totp_code, valid_window=1):
            _record_login_attempt(client_ip)
            return jsonify({'error': 'Invalid 2FA code'}), 401

    _clear_login_attempts(client_ip)
    # The password checked out, so binding this request to the user's
    # organization is now legitimate — and necessary, since reading their
    # organization row below goes through RLS like everything else.
    from flask import g
    g.organization_id = user['organization_id']
    g.is_superadmin = (user['role'] == 'superadmin')

    # Stamped only here, not up by the password check above: a request that
    # stops at requires_2fa hasn't actually signed in yet, and counting it
    # would show an admin "last login: today" for someone who never got past
    # their 2FA prompt.
    db = get_db()
    db.execute("UPDATE users SET last_login_at = NOW() WHERE id = %s", (user['id'],))
    db.close()

    claims = {'role': user['role'], 'name': user['name'],
              'org': user['organization_id']}
    token = create_access_token(identity=str(user['id']), additional_claims=claims)
    refresh = create_refresh_token(identity=str(user['id']), additional_claims=claims)
    return jsonify({
        'token': token,
        'refresh_token': refresh,
        'role': user['role'],
        'name': user['name'],
        # Branding and modules ride along with sign-in so the app can paint
        # itself in the right colours on first render, instead of flashing
        # Kikar's palette and then repainting.
        'organization': _organization_context(user['organization_id']),
    })


def _organization_context(organization_id):
    """Public-facing organization data: what the client needs to brand itself
    and to know which modules it may show. Never includes anything a member of
    the organization shouldn't see."""
    if organization_id is None:
        return None
    db = get_db()
    try:
        row = db.execute(
            "SELECT id, slug, name, logo_url, logo_path, brand_primary, brand_accent, "
            "       timezone, status, modules "
            "  FROM organizations WHERE id = %s", (organization_id,)
        ).fetchone()
    finally:
        db.close()
    if not row:
        return None
    modules = dict(tenancy.default_modules())
    modules.update(row['modules'] or {})
    return {
        'id': row['id'],
        'slug': row['slug'],
        'name': row['name'],
        # One field for the client, two sources behind it. The uploaded file
        # wins; a pasted URL is the fallback. Nothing in the UI needs to know
        # which one it got — see org_logo_url().
        'logo_url': tenancy.org_logo_url(row),
        'brand_primary': row['brand_primary'],
        'brand_accent': row['brand_accent'],
        'timezone': row['timezone'],
        'status': row['status'],
        'modules': modules,
    }


@app.route('/api/auth/context', methods=['GET'])
@jwt_required()
def auth_context():
    """Re-read the organization without signing in again.

    Modules and branding change under a running session when a superadmin
    edits them; the client refreshes from here rather than making people log
    out and back in."""
    claims = get_jwt() or {}
    return jsonify({
        'role': claims.get('role'),
        'name': claims.get('name'),
        'organization': _organization_context(claims.get('org')),
    })

@app.route('/api/auth/change-password', methods=['POST'])
@jwt_required()
def change_password():
    user_id = get_jwt_identity()
    data = request.json
    current = data.get('current_password', '')
    new_pw = data.get('new_password', '')

    pw_error = validate_password(new_pw)
    if pw_error:
        return jsonify({'error': pw_error}), 400

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = %s", (user_id,)).fetchone()
    if not user or not bcrypt.checkpw(current.encode(), user['password_hash'].encode()):
        db.close()
        return jsonify({'error': 'Current password is incorrect'}), 401

    new_hash = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()
    db.execute("UPDATE users SET password_hash = %s WHERE id = %s", (new_hash, user_id))
    db.commit()
    db.close()
    return jsonify({'message': 'Password updated'})

# ─── TOKEN REFRESH ─────────────────────────────────────────────────────────

@app.route('/api/auth/refresh', methods=['POST'])
@jwt_required(refresh=True)
def refresh_token():
    identity = get_jwt_identity()
    claims = get_jwt()
    new_token = create_access_token(
        identity=identity,
        additional_claims={'role': claims.get('role'), 'name': claims.get('name'),
                           'org': claims.get('org')}
    )
    return jsonify({'token': new_token})

@app.route('/api/auth/check', methods=['GET'])
@jwt_required()
def auth_check():
    claims = get_jwt()
    return jsonify({'ok': True, 'role': claims.get('role')})

# ─── TWO-FACTOR AUTHENTICATION (2FA) ───────────────────────────────────────

@app.route('/api/auth/2fa/setup', methods=['POST'])
@jwt_required()
def setup_2fa():
    if not TOTP_AVAILABLE:
        return jsonify({'error': '2FA is not available on this server'}), 500

    user_id = get_jwt_identity()
    db = get_db()
    user = db.execute("SELECT email, name FROM users WHERE id = %s", (user_id,)).fetchone()

    # Check if 2FA is already enabled
    existing = db.execute("SELECT * FROM user_totp WHERE user_id = %s", (user_id,)).fetchone()
    if existing and existing['enabled']:
        db.close()
        return jsonify({'error': '2FA is already enabled. Disable it first to reconfigure.'}), 400

    # Generate a new TOTP secret
    secret = pyotp.random_base32()
    totp = pyotp.TOTP(secret)
    provisioning_uri = totp.provisioning_uri(name=user['email'], issuer_name='Kikar Afterschool')

    # Save secret (not yet enabled)
    if existing:
        db.execute("UPDATE user_totp SET secret = %s, enabled = 0 WHERE user_id = %s", (secret, user_id))
    else:
        db.execute("INSERT INTO user_totp (user_id, secret, enabled) VALUES (%s, %s, 0)", (user_id, secret))
    db.commit()
    db.close()

    return jsonify({
        'secret': secret,
        'provisioning_uri': provisioning_uri,
        'message': 'Scan the QR code with your authenticator app (Google Authenticator, Authy, etc.), then verify with a code.'
    })


@app.route('/api/auth/2fa/verify', methods=['POST'])
@jwt_required()
def verify_2fa():
    if not TOTP_AVAILABLE:
        return jsonify({'error': '2FA is not available on this server'}), 500

    user_id = get_jwt_identity()
    code = (request.json or {}).get('code', '')

    if not code:
        return jsonify({'error': 'Verification code is required'}), 400

    db = get_db()
    totp_row = db.execute("SELECT * FROM user_totp WHERE user_id = %s", (user_id,)).fetchone()

    if not totp_row:
        db.close()
        return jsonify({'error': 'Run 2FA setup first'}), 400

    totp = pyotp.TOTP(totp_row['secret'])
    if not totp.verify(code, valid_window=1):
        db.close()
        return jsonify({'error': 'Invalid code. Try again.'}), 400

    db.execute("UPDATE user_totp SET enabled = 1 WHERE user_id = %s", (user_id,))
    db.commit()
    db.close()

    return jsonify({'message': '2FA has been enabled successfully!'})


@app.route('/api/auth/2fa/disable', methods=['POST'])
@jwt_required()
def disable_2fa():
    user_id = get_jwt_identity()
    password = (request.json or {}).get('password', '')

    if not password:
        return jsonify({'error': 'Current password is required to disable 2FA'}), 400

    db = get_db()
    user = db.execute("SELECT password_hash FROM users WHERE id = %s", (user_id,)).fetchone()
    if not user or not bcrypt.checkpw(password.encode(), user['password_hash'].encode()):
        db.close()
        return jsonify({'error': 'Invalid password'}), 401

    db.execute("DELETE FROM user_totp WHERE user_id = %s", (user_id,))
    db.commit()
    db.close()
    return jsonify({'message': '2FA has been disabled'})


@app.route('/api/auth/2fa/status', methods=['GET'])
@jwt_required()
def get_2fa_status():
    user_id = get_jwt_identity()
    db = get_db()
    totp_row = db.execute("SELECT enabled FROM user_totp WHERE user_id = %s AND enabled = 1", (user_id,)).fetchone()
    db.close()
    return jsonify({'enabled': bool(totp_row), 'available': TOTP_AVAILABLE})


# ─── PARENT ROUTES ──────────────────────────────────────────────────────────

@app.route('/api/parent/email-notifications', methods=['GET'])
@jwt_required()
def parent_get_email_notifications():
    """A parent's own opt-in to email as a second notification channel.

    Same footing as changing a password — a preference about the account,
    not a module a JCC buys — so it stays core (tests/test_module_access.py).
    """
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    db = get_db()
    try:
        row = db.execute(
            "SELECT email, email_notifications FROM users WHERE id = %s",
            (user_id,)
        ).fetchone()
    finally:
        db.close()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    return jsonify({'enabled': bool(row['email_notifications']), 'email': row['email']})


@app.route('/api/parent/email-notifications', methods=['PATCH'])
@jwt_required()
def parent_set_email_notifications():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    enabled = bool((request.json or {}).get('enabled'))
    db = get_db()
    try:
        db.execute(
            "UPDATE users SET email_notifications = %s WHERE id = %s",
            (enabled, user_id)
        )
        db.commit()
    finally:
        db.close()
    return jsonify({'enabled': enabled})


@app.route('/api/parent/children', methods=['GET'])
@jwt_required()
def parent_get_children():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    db = get_db()

    today = now_for_org(db).date()
    today_str = today.strftime('%Y-%m-%d')
    horizon_days = 60
    horizon = today + timedelta(days=horizon_days)
    horizon_str = horizon.strftime('%Y-%m-%d')
    DOW_NAMES = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']

    # Where each child is right now, for organizations that bought daily_ops.
    # Reuses the same _plan_for_day the counselor's board and My day read, so a
    # parent can never see a room the routing engine itself would disagree
    # with. Keyed by child_id and left empty (so `.get()` below is always
    # safe) for every organization without the module, or outside the 3-6pm
    # window the program actually runs in.
    locations: dict[int, dict] = {}
    today_day = DOW_NAMES[today.weekday()]
    if module_on('daily_ops', db) and today_day in daily_routing.WEEKDAYS:
        now_time = now_for_org(db).time()
        current_block = daily_routing.block_of(now_time)
        plan, _by_id, _rooms, _absent = _plan_for_day(db, today_day, today)
        for session in plan['classes']:
            start, end = session.get('start_time'), session.get('end_time')
            if start and end and start <= now_time < end:
                for entry in session['children']:
                    locations[entry['child_id']] = {
                        'kind': 'class', 'name': session['name'],
                    }
        if current_block:
            for group in plan['care']:
                if group['time_block'] != current_block or not group['room_name']:
                    continue
                for entry in group['children']:
                    locations.setdefault(entry['child_id'], {
                        'kind': 'care', 'name': group['room_name'],
                    })

    rows = db.execute(f"""
        SELECT c.id, c.name, s.name as school, c.service_type,
               COALESCE(
                 (SELECT json_agg(r.day_of_week ORDER BY CASE r.day_of_week
                    WHEN 'Monday' THEN 1 WHEN 'Tuesday' THEN 2 WHEN 'Wednesday' THEN 3
                    WHEN 'Thursday' THEN 4 WHEN 'Friday' THEN 5 END)
                  FROM registrations r WHERE r.child_id = c.id), '[]'
               ) AS registered_days
        FROM children c
        JOIN schools s ON c.school_id = s.id
        WHERE {_CHILD_BELONGS_TO_USER} AND c.active = 1
        ORDER BY c.name
    """, (user_id, user_id)).fetchall()

    result = []
    for row in rows:
        child_id = row['id']
        days = row['registered_days'] if isinstance(row['registered_days'], list) else json.loads(row['registered_days']) if isinstance(row['registered_days'], str) else []

        rules = db.execute("""
            SELECT id, day_of_week, start_date, end_date
            FROM recurring_absences WHERE child_id = %s
            ORDER BY CASE day_of_week
                WHEN 'Monday' THEN 1 WHEN 'Tuesday' THEN 2 WHEN 'Wednesday' THEN 3
                WHEN 'Thursday' THEN 4 WHEN 'Friday' THEN 5 END
        """, (child_id,)).fetchall()
        rules_out = [{
            'id': r['id'],
            'day_of_week': r['day_of_week'],
            'start_date': str(r['start_date']),
            'end_date': str(r['end_date']) if r['end_date'] else None,
        } for r in rules]

        one_off_rows = db.execute("""
            SELECT absence_date FROM absences
            WHERE child_id = %s AND absence_date >= %s AND absence_date <= %s
        """, (child_id, today_str, horizon_str)).fetchall()
        one_off_dates = {str(r['absence_date']) for r in one_off_rows}

        exception_rows = db.execute("""
            SELECT exception_date FROM absence_exceptions
            WHERE child_id = %s AND exception_date >= %s AND exception_date <= %s
        """, (child_id, today_str, horizon_str)).fetchall()
        exception_dates = {str(r['exception_date']) for r in exception_rows}

        rule_by_dow = {}
        for r in rules:
            rule_by_dow[r['day_of_week']] = r

        upcoming = []
        for i in range(horizon_days + 1):
            d = today + timedelta(days=i)
            ds = d.strftime('%Y-%m-%d')
            if ds in exception_dates:
                continue
            dow_name = DOW_NAMES[d.weekday()]
            rule = rule_by_dow.get(dow_name)
            rule_active = False
            if rule:
                start_ok = str(rule['start_date']) <= ds
                end_ok = (rule['end_date'] is None) or (ds <= str(rule['end_date']))
                rule_active = start_ok and end_ok
            if ds in one_off_dates:
                upcoming.append({'date': ds, 'kind': 'one_off'})
            elif rule_active:
                upcoming.append({'date': ds, 'kind': 'recurring', 'rule_id': rule['id']})

        result.append({
            'id': child_id,
            'name': row['name'],
            'school': row['school'],
            'service_type': row['service_type'],
            'registered_days': days,
            'recurring_absences': rules_out,
            'absence_exceptions': sorted(exception_dates),
            'upcoming_absences': upcoming,
            'location': locations.get(child_id),
        })

    db.close()
    return jsonify(result)

@app.route('/api/parent/absences', methods=['POST'])
@jwt_required()
def parent_mark_absence():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    data = request.json
    child_id = data.get('child_id')
    absence_date = data.get('date')

    if not child_id or not absence_date:
        return jsonify({'error': 'Missing child_id or date'}), 400
    try:
        absence_date = parse_date(absence_date)
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    db = get_db()
    # Verify child belongs to this parent
    if not parent_owns_child(db, user_id, child_id):
        db.close()
        return jsonify({'error': 'Child not found'}), 404

    try:
        # Marking a one-off absence overrides any prior "attending this day"
        # exception for the same date, so the child shows as absent regardless
        # of how the day was previously cleared.
        db.execute(
            "DELETE FROM absence_exceptions WHERE child_id = %s AND exception_date = %s",
            (child_id, absence_date)
        )
        db.execute(
            "INSERT INTO absences (child_id, absence_date, marked_by) VALUES (%s, %s, %s) ON CONFLICT DO NOTHING",
            (child_id, absence_date, user_id)
        )
        db.commit()
    except Exception as e:
        db.close()
        print(f"[ERROR] parent_mark_absence: {e}")
        return jsonify({'error': 'Failed to mark absence'}), 400

    db.close()
    return jsonify({'message': 'Absence marked'})

@app.route('/api/parent/absences', methods=['DELETE'])
@jwt_required()
def parent_remove_absence():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    data = request.json
    child_id = data.get('child_id')
    absence_date = data.get('date')

    if not child_id or not absence_date:
        return jsonify({'error': 'Missing child_id or date'}), 400
    try:
        absence_date = parse_date(absence_date)
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    db = get_db()
    if not parent_owns_child(db, user_id, child_id):
        db.close()
        return jsonify({'error': 'Child not found'}), 404

    # "Remove" means "the child WILL attend this date". Delete any one-off and
    # add an exception so an active recurring rule for this weekday is also
    # overridden — single button works for both kinds of absences.
    db.execute("DELETE FROM absences WHERE child_id = %s AND absence_date = %s", (child_id, absence_date))
    db.execute(
        "INSERT INTO absence_exceptions (child_id, exception_date, marked_by) "
        "VALUES (%s, %s, %s) ON CONFLICT DO NOTHING",
        (child_id, absence_date, user_id)
    )
    db.commit()
    db.close()
    return jsonify({'message': 'Absence removed'})

@app.route('/api/parent/recurring-absences', methods=['POST'])
@jwt_required()
def parent_add_recurring_absence():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    data = request.json or {}
    child_id = data.get('child_id')
    day_of_week = data.get('day_of_week')
    end_date = data.get('end_date')  # nullable

    if not child_id or not day_of_week:
        return jsonify({'error': 'Missing child_id or day_of_week'}), 400
    if day_of_week not in ('Monday','Tuesday','Wednesday','Thursday','Friday'):
        return jsonify({'error': 'Invalid day_of_week'}), 400

    today_str = today_for_org()
    if end_date:
        try:
            end_date = parse_date(end_date)
        except ValueError:
            return jsonify({'error': 'Invalid end_date format. Use YYYY-MM-DD.'}), 400
        if end_date < today_str:
            return jsonify({'error': 'end_date must be today or later'}), 400
    else:
        end_date = None

    db = get_db()
    if not parent_owns_child(db, user_id, child_id):
        db.close()
        return jsonify({'error': 'Child not found'}), 404

    try:
        db.execute(
            """INSERT INTO recurring_absences
                 (child_id, day_of_week, start_date, end_date, marked_by)
               VALUES (%s, %s, %s, %s, %s)
               ON CONFLICT (child_id, day_of_week) DO UPDATE SET
                 start_date = EXCLUDED.start_date,
                 end_date   = EXCLUDED.end_date,
                 marked_by  = EXCLUDED.marked_by""",
            (child_id, day_of_week, today_str, end_date, user_id)
        )
        db.commit()
    except Exception as e:
        db.close()
        print(f"[ERROR] parent_add_recurring_absence: {e}")
        return jsonify({'error': 'Failed to save recurring absence'}), 400

    db.close()
    return jsonify({'message': 'Recurring absence saved'})

@app.route('/api/parent/absence-exceptions', methods=['DELETE'])
@jwt_required()
def parent_delete_absence_exception():
    """Remove a per-day "attending this day" override so the recurring rule
    (or any one-off) takes effect again. Used by the Manage Recurring view to
    undo an Attending toggle without inserting a redundant one-off."""
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    data = request.json or {}
    child_id = data.get('child_id')
    exception_date = data.get('date')
    if not child_id or not exception_date:
        return jsonify({'error': 'Missing child_id or date'}), 400
    try:
        exception_date = parse_date(exception_date)
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    db = get_db()
    if not parent_owns_child(db, user_id, child_id):
        db.close()
        return jsonify({'error': 'Child not found'}), 404

    db.execute(
        "DELETE FROM absence_exceptions WHERE child_id = %s AND exception_date = %s",
        (child_id, exception_date)
    )
    db.commit()
    db.close()
    return jsonify({'message': 'Exception removed'})

@app.route('/api/parent/recurring-absences/<int:rule_id>', methods=['DELETE'])
@jwt_required()
def parent_delete_recurring_absence(rule_id):
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    db = get_db()
    rule = db.execute(f"""
        SELECT ra.id FROM recurring_absences ra
        JOIN children c ON c.id = ra.child_id
        WHERE ra.id = %s AND {_CHILD_BELONGS_TO_USER}
    """, (rule_id, user_id, user_id)).fetchone()
    if not rule:
        db.close()
        return jsonify({'error': 'Rule not found'}), 404

    db.execute("DELETE FROM recurring_absences WHERE id = %s", (rule_id,))
    db.commit()
    db.close()
    return jsonify({'message': 'Recurring absence removed'})


# ─── PARENT-SUBMITTED MAKE-UP CLASSES ──────────────────────────────────────
#
# Parents submit one-off make-up classes from the Parent Portal side menu.
# Storage: rows in activity_roster with source='parent_makeup' and
# specific_date set. Counselor auto-assignment matches the recurring slot
# (activity_id + day_of_week + action + action_time) and copies its
# assigned_counselor_id. The cascade in admin_assign_counselor /
# admin_bulk_assign_counselor keeps make-ups in sync when admin changes the
# recurring counselor.

def _resolve_makeup_counselor(db, activity_id, day_of_week, action, action_time):
    """Return the assigned_counselor_id (or None) of the recurring roster
    slot that matches (activity, weekday, action, time). Used at parent
    submit time to auto-assign and at admin reassign time to cascade.

    Recurring slots are activity_roster rows where specific_date IS NULL.
    Exact action_time match wins; otherwise we fall back to the most recent
    recurring row for the same activity/day/action. Returns None when no
    recurring slot has a counselor — admin will see "needs assignment"."""
    row = db.execute("""
        SELECT assigned_counselor_id FROM activity_roster
         WHERE activity_id = %s
           AND day_of_week = %s
           AND action = %s
           AND specific_date IS NULL
           AND assigned_counselor_id IS NOT NULL
         ORDER BY (action_time = %s) DESC, created_at DESC
         LIMIT 1
    """, (activity_id, day_of_week, action, action_time or '')).fetchone()
    return row['assigned_counselor_id'] if row else None


def _notify_admins_of_makeup(parent_name, child_name, activity_name, makeup_date):
    try:
        db = get_db()
        admin_ids = [r['id'] for r in db.execute(
            "SELECT id FROM users WHERE role='admin'"
        ).fetchall()]
        db.close()
    except Exception as e:
        print(f"[ERROR] _notify_admins_of_makeup: {e}")
        return
    title = '🔁 New make-up class request'
    body = f'{parent_name} requested {activity_name} for {child_name} on {makeup_date}.'
    for aid in admin_ids:
        send_push_to_user_async(aid, title, body, url='/app/makeup',
                                 tag='makeup-request')


@app.route('/api/parent/makeup-classes/options', methods=['GET'])
@jwt_required()
def parent_makeup_options():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    child_id = request.args.get('child_id', type=int)
    if not child_id:
        return jsonify({'error': 'child_id required'}), 400

    db = get_db()
    child = db.execute(
        f"SELECT c.id, c.name FROM children c "
        f"WHERE c.id=%s AND {_CHILD_BELONGS_TO_USER} AND c.active=1",
        (child_id, user_id, user_id)
    ).fetchone()
    if not child:
        db.close()
        return jsonify({'error': 'Child not found'}), 404

    rows = db.execute("""
        SELECT a.id AS activity_id, a.name AS activity_name, a.category,
               ar.day_of_week, ar.action, ar.action_time, ar.school
          FROM activity_roster ar
          JOIN activities a ON a.id = ar.activity_id
         WHERE ar.child_name = %s
           AND ar.specific_date IS NULL
         ORDER BY a.name, ar.day_of_week, ar.action_time, ar.action
    """, (child['name'],)).fetchall()
    db.close()

    by_activity = {}
    for r in rows:
        aid = r['activity_id']
        if aid not in by_activity:
            by_activity[aid] = {
                'activity_id': aid,
                'activity_name': r['activity_name'],
                'category': r['category'],
                'school': r['school'],
                'slots': [],
            }
        by_activity[aid]['slots'].append({
            'day_of_week': r['day_of_week'],
            'action': r['action'],
            'action_time': r['action_time'],
        })
    return jsonify(list(by_activity.values()))


@app.route('/api/parent/makeup-classes', methods=['GET'])
@jwt_required()
def parent_list_makeup_classes():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()

    db = get_db()
    rows = db.execute("""
        SELECT ar.id, ar.child_name, ar.action, ar.action_time, ar.day_of_week,
               to_char(ar.specific_date, 'YYYY-MM-DD') AS specific_date,
               ar.parent_will_pickup, ar.assigned_counselor_id, ar.admin_seen_at,
               a.id AS activity_id, a.name AS activity_name,
               u.name AS counselor_name
          FROM activity_roster ar
          JOIN activities a ON a.id = ar.activity_id
          LEFT JOIN users u ON u.id = ar.assigned_counselor_id
         WHERE ar.requested_by_parent_id = %s
           AND ar.specific_date >= CURRENT_DATE
         ORDER BY ar.specific_date, ar.child_name, ar.action_time
    """, (user_id,)).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/parent/makeup-classes', methods=['POST'])
@jwt_required()
def parent_create_makeup_class():
    """Create a parent-submitted make-up class.

    Expected body:
      child_id        (int, required)        — must belong to caller
      activity_id     (int, required)        — child must be in its roster
      makeup_date     (YYYY-MM-DD, required) — must be today or later
      dropoff_time    (HH:MM, optional)      — required if needs_dropoff
      needs_dropoff   (bool)                 — at least one of dropoff/pickup
      pickup_time     (HH:MM, optional)      — required if needs_pickup
      needs_pickup    (bool)                 — false if parent_will_pickup
      parent_will_pickup (bool)              — convenience flag stored on row

    Inserts 1-2 rows (Drop Off, Pick Up) into activity_roster with
    source='parent_makeup'. Counselor is auto-resolved per row from the
    recurring slot for the same activity + weekday + action + time."""
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    d = request.json or {}

    child_id = d.get('child_id')
    activity_id = d.get('activity_id')
    makeup_date = (d.get('makeup_date') or '').strip()
    needs_dropoff = bool(d.get('needs_dropoff'))
    needs_pickup = bool(d.get('needs_pickup'))
    parent_will_pickup = bool(d.get('parent_will_pickup'))
    dropoff_time = (d.get('dropoff_time') or '').strip() or None
    pickup_time = (d.get('pickup_time') or '').strip() or None

    if not child_id or not activity_id or not makeup_date:
        return jsonify({'error': 'child_id, activity_id and makeup_date are required'}), 400
    try:
        parsed = datetime.strptime(makeup_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'makeup_date must be YYYY-MM-DD'}), 400
    if parsed < now_for_org().date():
        return jsonify({'error': 'makeup_date cannot be in the past'}), 400
    day_of_week = parsed.strftime('%A')
    if day_of_week in ('Saturday', 'Sunday'):
        return jsonify({'error': 'Make-up classes are only valid Monday through Friday'}), 400

    if not needs_dropoff and not needs_pickup:
        return jsonify({'error': 'At least one of drop-off or pick-up must be selected'}), 400
    if needs_dropoff and not dropoff_time:
        return jsonify({'error': 'dropoff_time required when needs_dropoff is true'}), 400
    if needs_pickup and not pickup_time:
        return jsonify({'error': 'pickup_time required when needs_pickup is true'}), 400

    db = get_db()
    child = db.execute(f"""
        SELECT c.id, c.name, s.name AS school_name
          FROM children c
          JOIN schools s ON s.id = c.school_id
         WHERE c.id=%s AND {_CHILD_BELONGS_TO_USER} AND c.active=1
    """, (child_id, user_id, user_id)).fetchone()
    if not child:
        db.close()
        return jsonify({'error': 'Child not found'}), 404

    activity = db.execute(
        "SELECT id, name FROM activities WHERE id=%s", (activity_id,)
    ).fetchone()
    if not activity:
        db.close()
        return jsonify({'error': 'Activity not found'}), 404

    enrolled = db.execute("""
        SELECT 1 FROM activity_roster
         WHERE activity_id=%s AND child_name=%s AND specific_date IS NULL
         LIMIT 1
    """, (activity_id, child['name'])).fetchone()
    if not enrolled:
        db.close()
        return jsonify({'error': 'This child is not registered in that activity'}), 400

    closure = db.execute(
        "SELECT 1 FROM calendar_events WHERE event_date=%s AND event_type='closed' LIMIT 1",
        (makeup_date,)
    ).fetchone()
    if closure:
        db.close()
        return jsonify({'error': 'School is closed on that date'}), 400

    inserted_ids = []
    rows_to_insert = []
    if needs_dropoff:
        rows_to_insert.append(('Drop Off', dropoff_time))
    if needs_pickup:
        rows_to_insert.append(('Pick Up', pickup_time))

    for action, action_time in rows_to_insert:
        counselor_id = _resolve_makeup_counselor(
            db, activity_id, day_of_week, action, action_time
        )
        cur = db.execute("""
            INSERT INTO activity_roster
              (activity_id, child_name, school, option_type, action,
               day_of_week, action_time, assigned_counselor_id, source,
               specific_date, requested_by_parent_id, parent_will_pickup)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'parent_makeup',%s,%s,%s)
            RETURNING id
        """, (activity_id, child['name'], child['school_name'], 'Full Service',
              action, day_of_week, action_time, counselor_id, makeup_date,
              user_id, parent_will_pickup))
        inserted_ids.append(cur.fetchone()['id'])
    db.commit()

    parent = db.execute("SELECT name FROM users WHERE id=%s", (user_id,)).fetchone()
    db.close()

    _notify_admins_of_makeup(
        parent['name'] if parent else 'A parent',
        child['name'], activity['name'], makeup_date
    )

    return jsonify({'ok': True, 'ids': inserted_ids}), 201


@app.route('/api/parent/makeup-classes/<int:entry_id>', methods=['DELETE'])
@jwt_required()
def parent_delete_makeup_class(entry_id):
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()

    db = get_db()
    row = db.execute("""
        SELECT id FROM activity_roster
         WHERE id=%s AND requested_by_parent_id=%s
           AND specific_date >= CURRENT_DATE
    """, (entry_id, user_id)).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Not found'}), 404
    db.execute("DELETE FROM activity_roster WHERE id=%s", (entry_id,))
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ─── COUNSELOR ROUTES ───────────────────────────────────────────────────────

@app.route('/api/counselor/roster', methods=['GET'])
@jwt_required()
def counselor_get_roster():
    claims = get_jwt()
    # Also admin, who reads it through the same Pickup screen (see
    # PickupRelease in web/) to release a child at the office rather than at a
    # school gate. counselor_school_ids(db, admin_user_id, ...) finds no
    # counselor_schools rows for an admin and returns None — "every school" —
    # which is exactly right without any admin-specific branch here.
    if claims.get('role') not in ('counselor', 'admin'):
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    date_str = request.args.get('date', today_for_org())

    # Determine day of week from date
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        day_name = date_obj.strftime('%A')
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    db = get_db()

    # The counselor's schools *for the date being viewed* — or all of them
    # when nobody has divided the counselors up yet. Scoping to the date
    # rather than to today is what makes a temporary assignment show up on
    # the day it covers: someone stepping in for Thursday sees Thursday's
    # roster when they look at Thursday, and their own on every other day.
    school_ids = counselor_school_ids(db, user_id, date_str)
    if school_ids is None:
        schools = db.execute(
            "SELECT id, name, division_type FROM schools ORDER BY name"
        ).fetchall()
    else:
        schools = db.execute("""
            SELECT id, name, division_type FROM schools
             WHERE id = ANY(%s)
             ORDER BY name
        """, (school_ids,)).fetchall()

    result = []
    is_wednesday = (day_name == 'Wednesday')

    # Effective absences for this date (one-offs + active recurring - exceptions)
    absent_ids = absent_child_ids_for_date(db, date_str)
    absent_id_list = list(absent_ids) if absent_ids else None

    # One query shape for both halves of the roster. It used to be four copies
    # of nearly the same SELECT — two for "some children are absent" and two
    # for "none are" — and the absent halves had already drifted, carrying
    # fewer columns than the attending ones.
    #
    # Everything a counselor needs at a school gate is here: the grade and the
    # dismissal hour so they know who they are looking for and when the child
    # leaves, and the allergies, which is the one field that changes what they
    # do in the moment.
    ROSTER_COLUMNS = """
        SELECT c.id, c.name, c.service_type, c.release_group,
               c.grade_label, c.grade_num, c.allergies, c.notes,
               r.dismissal_time,
               u.name AS parent_name, u.email AS parent_email
          FROM children c
          JOIN users u ON c.parent_id = u.id
          JOIN registrations r ON r.child_id = c.id
         WHERE c.school_id = %s
           AND c.active = 1
           AND r.day_of_week = %s
    """

    def roster_rows(school_id, absent_only):
        clause = ("c.id = ANY(%s::int[])" if absent_only
                  else "(%s::int[] IS NULL OR c.id <> ALL(%s::int[]))")
        params = ([school_id, day_name, absent_id_list] if absent_only
                  else [school_id, day_name, absent_id_list, absent_id_list])
        return db.execute(
            f"{ROSTER_COLUMNS} AND {clause} ORDER BY c.name", tuple(params)
        ).fetchall()

    def child_dict(ch):
        return {
            'id': ch['id'],
            'name': ch['name'],
            'service_type': ch['service_type'],
            'release_group': ch['release_group'],
            'grade': ch['grade_label'],
            'grade_num': ch['grade_num'],
            # NULL is not 6pm. For every JCC that does not run dismissal
            # times a registrations row means "enrolled this weekday" and
            # nothing more, and the row must not imply an hour nobody set.
            'dismissal_time': ch['dismissal_time'],
            'allergies': ch['allergies'],
            'notes': ch['notes'],
            'parent': ch['parent_name'],
            'parent_email': ch['parent_email'],
        }

    for school in schools:
        children = roster_rows(school['id'], absent_only=False)
        absent = roster_rows(school['id'], absent_only=True) if absent_id_list else []

        # Roster split is driven by the school's declared division_type, NOT
        # inferred from per-child release_group values. One dirty row (e.g. a
        # Ben Gamla child stamped 'K-1st') must not flip the whole school's
        # UI into grade-split mode.
        division_type = school['division_type']
        needs_split = division_type in ('grade', 'torah')
        split_type = division_type if needs_split else None
        # Grade-split schools release together on Wednesdays.
        if split_type == 'grade' and is_wednesday:
            needs_split = False
        wednesday_together = is_wednesday and division_type == 'grade'

        result.append({
            'school': school['name'],
            'school_id': school['id'],
            'date': date_str,
            'day': day_name,
            'needs_split': needs_split,
            'split_type': split_type,
            'wednesday_together': wednesday_together,
            'attending': [child_dict(ch) for ch in children],
            'absent': [child_dict(ch) for ch in absent],
        })

    # Arrival and departure times ride along for organizations that bought
    # check_in_out. Added as a separate map rather than folded into each child
    # so the roster payload keeps the shape every existing client expects.
    if module_on('check_in_out', db) and result:
        times = db.execute("""
            SELECT child_id, checked_in_at, checked_out_at
              FROM attendance_records
             WHERE attendance_date = %s
               AND (checked_in_at IS NOT NULL OR checked_out_at IS NOT NULL)
        """, (date_str,)).fetchall()
        by_child = {
            str(t['child_id']): {
                'checked_in_at': t['checked_in_at'],
                'checked_out_at': t['checked_out_at'],
            }
            for t in times
        }
        # Narrowed per school rather than attaching the organization-wide map
        # to every entry — a counselor covering four schools would otherwise
        # receive every other school's times four times over.
        for entry in result:
            ids = {str(ch['id']) for ch in entry['attending']}
            entry['times'] = {k: v for k, v in by_child.items() if k in ids}

    db.close()
    return jsonify(result)

# ─── ADMIN ROUTES ───────────────────────────────────────────────────────────

def require_admin():
    claims = get_jwt()
    return claims.get('role') == 'admin'

def _current_user_id():
    """The signed-in user's id as an int, for columns that reference users(id).

    The JWT identity is a string — create_access_token requires it — and these
    are foreign keys, so the conversion has to happen somewhere.
    """
    try:
        return int(get_jwt_identity())
    except (TypeError, ValueError):
        return None

# ─── LIVE DEMO ──────────────────────────────────────────────────────────────
# Lets an admin open the parent/counselor portals already logged in, for live
# demos. Gated behind DEMO_MODE and restricted to the seeded demo accounts only
# (see server/seed_test_data.py) — it can NEVER mint a token for a real user.

def _demo_mode_enabled():
    return os.environ.get('DEMO_MODE', '').lower() in ('1', 'true', 'yes')

# view_as value -> email of the seeded demo account it may impersonate.
_DEMO_ACCOUNTS = {
    'parent': 'testparent@jclub.com',
    'counselor': 'testcounselor@jclub.com',
}

@app.route('/api/admin/demo-status', methods=['GET'])
@jwt_required()
def admin_demo_status():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    return jsonify({'enabled': _demo_mode_enabled()})

@app.route('/api/admin/demo-token', methods=['POST'])
@jwt_required()
def admin_demo_token():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    if not _demo_mode_enabled():
        return jsonify({'error': 'Demo mode is disabled'}), 403

    view_as = (request.json or {}).get('view_as', '')
    email = _DEMO_ACCOUNTS.get(view_as)
    if not email:
        return jsonify({'error': "view_as must be 'parent' or 'counselor'"}), 400

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE email = %s", (email,)).fetchone()
    db.close()
    if not user or user['role'] != view_as:
        return jsonify({'error': 'Demo accounts not seeded. Set SEED_TEST_ACCOUNTS=1 and restart.'}), 409

    # 'org' matters as much here as at login. Without it the demo session binds
    # to no organization, every tenant-scoped read comes back empty through RLS,
    # and the demo shows a working app with nothing in it — which is a worse
    # demo than none. This predates the multi-tenant rewrite and was missed.
    claims = {'role': user['role'], 'name': user['name'],
              'org': user['organization_id']}
    token = create_access_token(identity=str(user['id']), additional_claims=claims)
    refresh = create_refresh_token(identity=str(user['id']), additional_claims=claims)
    return jsonify({'token': token, 'refresh_token': refresh, 'role': user['role'], 'name': user['name']})

@app.route('/api/admin/stats', methods=['GET'])
@jwt_required()
def admin_stats():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        db = get_db()
        stats = {
            'parents': db.execute("SELECT COUNT(*) as n FROM users WHERE role='parent'").fetchone()['n'],
            'counselors': db.execute("SELECT COUNT(*) as n FROM users WHERE role='counselor'").fetchone()['n'],
            'children': db.execute("SELECT COUNT(*) as n FROM children WHERE active = 1").fetchone()['n'],
            # Every active child, and the subset enrolled on at least one
            # weekday. The dashboard shows both because they are far apart —
            # a roster arrives with children whose class signups have not
            # happened yet, and they are counted in the first number and in
            # none of the day-by-day ones. Showing only the total invites the
            # question "why does it say 156 children and 74 expected?", which
            # has a good answer that the screen was not giving.
            'children_enrolled': db.execute("""
                SELECT COUNT(DISTINCT c.id) AS n
                  FROM children c
                  JOIN registrations r ON r.child_id = c.id
                 WHERE c.active = 1
            """).fetchone()['n'],
            'schools': db.execute("SELECT COUNT(*) as n FROM schools").fetchone()['n'],
        }
        db.close()
        return jsonify(stats)
    except Exception as e:
        print(f"[ERROR] admin_stats: {e}")
        try: db.close()
        except Exception: pass
        return jsonify({'error': 'An internal error occurred'}), 500

@app.route('/api/admin/schools', methods=['GET'])
@jwt_required()
def admin_get_schools():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    schools = db.execute("SELECT * FROM schools ORDER BY name").fetchall()
    db.close()
    return jsonify([dict(s) for s in schools])

@app.route('/api/admin/schools/<int:school_id>/roster', methods=['GET'])
@jwt_required()
def admin_school_roster(school_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    date_str = request.args.get('date', today_for_org())
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        day_name = date_obj.strftime('%A')
    except ValueError:
        return jsonify({'error': 'Invalid date'}), 400

    db = get_db()
    school = db.execute("SELECT * FROM schools WHERE id = %s", (school_id,)).fetchone()
    if not school:
        db.close()
        return jsonify({'error': 'School not found'}), 404

    # Full Service first, then others — both groups alphabetical
    children = db.execute("""
        SELECT c.id, c.name, c.service_type, u.name as parent_name, u.email as parent_email,
               CASE c.service_type WHEN 'Full Service' THEN 1 ELSE 2 END as sort_order
        FROM children c
        JOIN users u ON c.parent_id = u.id
        JOIN registrations r ON r.child_id = c.id
        WHERE c.school_id = %s
          AND c.active = 1
          AND r.day_of_week = %s
        ORDER BY sort_order, c.name
    """, (school_id, day_name)).fetchall()

    absent_ids = absent_child_ids_for_date(db, date_str)

    attending = []
    absent = []
    for ch in children:
        entry = {'id': ch['id'], 'name': ch['name'], 'service_type': ch['service_type'],
                 'parent': ch['parent_name'], 'parent_email': ch['parent_email']}
        if ch['id'] in absent_ids:
            absent.append(entry)
        else:
            attending.append(entry)

    db.close()
    return jsonify({
        'school': school['name'],
        'date': date_str,
        'day': day_name,
        'attending': attending,
        'absent': absent,
    })

@app.route('/api/admin/schools', methods=['POST'])
@jwt_required()
def admin_add_school():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    name = request.json.get('name', '').strip()
    if not name:
        return jsonify({'error': 'School name required'}), 400
    db = get_db()
    try:
        db.execute("INSERT INTO schools (name) VALUES (%s)", (name,))
        db.commit()
    except psycopg2.IntegrityError:
        db.close()
        return jsonify({'error': 'School already exists'}), 409
    db.close()
    return jsonify({'message': 'School added'})

@app.route('/api/admin/schools/<int:school_id>', methods=['PUT'])
@jwt_required()
def admin_update_school(school_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    updates = {}
    if 'name' in data:
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'School name required'}), 400
        updates['name'] = name
    if 'division_type' in data:
        division_type = (data.get('division_type') or '').strip()
        if division_type not in ('none', 'grade', 'torah'):
            return jsonify({'error': "division_type must be 'none', 'grade' or 'torah'"}), 400
        updates['division_type'] = division_type
    if not updates:
        return jsonify({'error': 'Nothing to update'}), 400

    sets = ', '.join(f"{k} = %s" for k in updates)
    db = get_db()
    try:
        cur = db.execute(
            f"UPDATE schools SET {sets} WHERE id = %s RETURNING id",
            (*updates.values(), school_id),
        )
        if cur.rowcount == 0:
            db.close()
            return jsonify({'error': 'School not found'}), 404
        db.commit()
    except psycopg2.IntegrityError:
        db.close()
        return jsonify({'error': 'A school with that name already exists'}), 409
    db.close()
    return jsonify({'message': 'School updated'})

@app.route('/api/admin/schools/<int:school_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_school(school_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        # No ON DELETE CASCADE from children.school_id on purpose: a school
        # still holding kids is a data problem to notice, not to silently
        # solve by orphaning them or wiping their enrollment.
        in_use = db.execute(
            "SELECT COUNT(*) AS n FROM children WHERE school_id = %s", (school_id,)
        ).fetchone()['n']
        if in_use:
            return jsonify({
                'error': f'{in_use} child{"ren" if in_use != 1 else ""} still '
                         f'enrolled at this school — move or remove them first.'
            }), 409
        # staff_assignments.school_id is ON DELETE CASCADE (sql/57), so a school
        # with a bus route still on the weekly template or the calendar would
        # otherwise lose that staffing silently, the same trap sql/37 already
        # guards against for rooms and classes.
        staffed = db.execute(
            "SELECT COUNT(*) AS n FROM staff_assignments WHERE school_id = %s",
            (school_id,)
        ).fetchone()['n']
        if staffed:
            return jsonify({
                'error': f'{staffed} bus assignment{"s" if staffed != 1 else ""} '
                         f'still point at this school — remove them first.'
            }), 409
        cur = db.execute("DELETE FROM schools WHERE id = %s", (school_id,))
        if cur.rowcount == 0:
            return jsonify({'error': 'School not found'}), 404
        db.commit()
    finally:
        db.close()
    return jsonify({'message': 'School deleted'})


# ─── GRADES ─────────────────────────────────────────────────────────────────
# The menu a child's grade is picked from (sql/49_add_grades.sql), not where
# it is stored — that stays children.grade_label/grade_num, filled in by
# parse_grade() either way. Deleting a grade here never touches an existing
# child's row; it only stops offering that option going forward.

def _grade_sort_order(db, grade_num):
    """K first, then 1 through 12 in order; anything unrecognized goes last.

    Mirrors what a human would expect without asking them to also pick a
    number — parse_grade() already knows K and 0 are the same rank, so a
    label that already means something sorts itself.
    """
    if grade_num is not None:
        return grade_num
    row = db.execute("SELECT COALESCE(MAX(sort_order), -1) AS n FROM grades").fetchone()
    return row['n'] + 1


def _grade_num_or_400(name):
    """Reject a grade name parse_grade() cannot turn into a number.

    Without this, this list could offer a name — 'Pre-K', a typo — that looks
    like a validated option but always resolves to grade_num NULL for any
    child assigned it: invisible to care rules, the daily board, every screen
    that groups by grade. A picker that can silently produce that is worse
    than the free text it replaced, because it looks safe.
    """
    (_label, grade_num), _code = parse_grade(name)
    if grade_num is None:
        return None, (jsonify({
            'error': f'"{name}" is not a grade this app recognises. Try K, '
                     'a number like 1-12, or 1st/2nd/3rd/etc.'
        }), 400)
    return grade_num, None


@app.route('/api/admin/grades', methods=['GET'])
@jwt_required()
def admin_get_grades():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    grades = db.execute("SELECT * FROM grades ORDER BY sort_order, name").fetchall()
    db.close()
    return jsonify([dict(g) for g in grades])

@app.route('/api/admin/grades', methods=['POST'])
@jwt_required()
def admin_add_grade():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    name = (request.json or {}).get('name', '').strip()
    if not name:
        return jsonify({'error': 'Grade name required'}), 400
    grade_num, err = _grade_num_or_400(name)
    if err:
        return err
    db = get_db()
    try:
        sort_order = _grade_sort_order(db, grade_num)
        db.execute(
            "INSERT INTO grades (name, sort_order) VALUES (%s, %s)",
            (name, sort_order),
        )
        db.commit()
    except psycopg2.IntegrityError:
        db.close()
        return jsonify({'error': 'Grade already exists'}), 409
    db.close()
    return jsonify({'message': 'Grade added'})

@app.route('/api/admin/grades/<int:grade_id>', methods=['PUT'])
@jwt_required()
def admin_update_grade(grade_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    name = (request.json or {}).get('name', '').strip()
    if not name:
        return jsonify({'error': 'Grade name required'}), 400
    grade_num, err = _grade_num_or_400(name)
    if err:
        return err
    db = get_db()
    try:
        sort_order = _grade_sort_order(db, grade_num)
        cur = db.execute(
            "UPDATE grades SET name = %s, sort_order = %s WHERE id = %s RETURNING id",
            (name, sort_order, grade_id),
        )
        if cur.rowcount == 0:
            db.close()
            return jsonify({'error': 'Grade not found'}), 404
        db.commit()
    except psycopg2.IntegrityError:
        db.close()
        return jsonify({'error': 'A grade with that name already exists'}), 409
    db.close()
    return jsonify({'message': 'Grade updated'})

@app.route('/api/admin/grades/<int:grade_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_grade(grade_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    cur = db.execute("DELETE FROM grades WHERE id = %s", (grade_id,))
    if cur.rowcount == 0:
        db.close()
        return jsonify({'error': 'Grade not found'}), 404
    db.commit()
    db.close()
    return jsonify({'message': 'Grade deleted'})


@app.route('/api/admin/counselors', methods=['GET'])
@jwt_required()
def admin_get_counselors():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    today = today_for_org(db)
    counselors = db.execute(f"""
        SELECT u.id, u.name, u.email,
               -- Invitation state, so the list can tell an active counselor
               -- from one who was invited and never finished setting up.
               u.password_set_at, u.invited_at AS last_invite_at,
               u.last_login_at,
               COALESCE(
                 (SELECT json_agg(json_build_object(
                            'school_id', s.id,
                            'school', s.name,
                            'permanent', cs.effective_from IS NULL
                                         AND cs.effective_to IS NULL,
                            'effective_from', cs.effective_from,
                            'effective_to', cs.effective_to,
                            'in_effect', {_ASSIGNMENT_IN_EFFECT})
                          ORDER BY s.name)
                  FROM schools s JOIN counselor_schools cs ON cs.school_id = s.id
                  WHERE cs.counselor_id = u.id), '[]'
               ) AS schools
        FROM users u
        WHERE u.role = 'counselor'
        ORDER BY u.name
    """, (today, today)).fetchall()
    result = []
    for c in counselors:
        schools = c['schools'] if isinstance(c['schools'], list) else json.loads(c['schools']) if isinstance(c['schools'], str) else []
        result.append({
            'id': c['id'], 'name': c['name'], 'email': c['email'],
            'schools': schools,
            # An empty list means "every school", not "no schools" — see
            # counselor_school_ids. The screen has to say which one it is,
            # because they are opposites and looked identical before.
            'covers_all_schools': len(schools) == 0,
            'password_set_at': iso_utc(c['password_set_at']),
            'last_invite_at': iso_utc(c['last_invite_at']),
            'last_login_at': iso_utc(c['last_login_at']),
        })
    db.close()
    return jsonify(result)

@app.route('/api/admin/counselors', methods=['POST'])
@jwt_required()
def admin_add_counselor():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json
    name = data.get('name', '').strip()
    email = data.get('email', '').strip().lower()
    school_ids = data.get('school_ids', [])

    if not name or not email:
        return jsonify({'error': 'Name and email required'}), 400

    random_pw = secrets.token_urlsafe(32)
    pw_hash = bcrypt.hashpw(random_pw.encode(), bcrypt.gensalt()).decode()

    db = get_db()
    try:
        cur = db.execute("INSERT INTO users (email, password_hash, role, name) VALUES (%s, %s, 'counselor', %s) RETURNING id", (email, pw_hash, name))
        counselor_id = cur.fetchone()['id']
        # Assignments made at creation are logged like any other, so the
        # history of a counselor's schools starts at their first day rather
        # than at the first time somebody edited them.
        actor = _current_user_id()
        for sid in school_ids:
            db.execute(
                "INSERT INTO counselor_schools (counselor_id, school_id, assigned_by, assigned_at) "
                "VALUES (%s, %s, %s, CURRENT_TIMESTAMP) ON CONFLICT DO NOTHING",
                (counselor_id, sid, actor))
            _log_assignment(db, 'added', counselor_id, sid, actor)
        setup_url = _issue_setup_token(db, counselor_id)
        db.commit()
    except psycopg2.IntegrityError:
        db.close()
        return jsonify({'error': 'Email already exists'}), 409

    send_counselor_invitation(email, name, setup_url, email_identity(db))
    db.close()
    return jsonify({'message': 'Counselor added and invitation sent', 'email': email})

@app.route('/api/admin/counselors/<int:counselor_id>/resend-invite', methods=['POST'])
@jwt_required()
def admin_resend_counselor_invite(counselor_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = %s AND role='counselor'", (counselor_id,)).fetchone()
    if not user:
        db.close()
        return jsonify({'error': 'Counselor not found'}), 404

    random_pw = secrets.token_urlsafe(32)
    pw_hash = bcrypt.hashpw(random_pw.encode(), bcrypt.gensalt()).decode()
    db.execute("UPDATE users SET password_hash = %s WHERE id = %s", (pw_hash, counselor_id))
    db.execute(
        "UPDATE password_reset_tokens SET used=1 WHERE user_id=%s AND used=0",
        (counselor_id,)
    )
    setup_url = _issue_setup_token(db, counselor_id)
    db.commit()

    send_counselor_invitation(user['email'], user['name'], setup_url, email_identity(db))
    db.close()
    return jsonify({'message': 'Invitation resent'})


@app.route('/api/admin/counselors/<int:counselor_id>/setup-link', methods=['POST'])
@jwt_required()
def admin_counselor_setup_link(counselor_id):
    """A one-time link that lets this counselor set their own password.

    The same link the invitation email carries, handed back instead of sent.
    Without a mail provider configured an invitation is generated, thrown away
    and reported as sent, which leaves a JCC unable to onboard a single member
    of staff and with no sign of why. The superadmin console already resolves
    this exact problem the same way when it creates a JCC's first admin
    (server/superadmin.py) — this is that, one level down.

    Two things it deliberately does NOT do, and both are why it is its own
    route rather than a field added to resend-invite:

      * It does not send email. That is the entire point.
      * It does not touch password_hash. resend-invite rotates it before
        sending, so with mail broken it locks out a counselor who already had
        a working password and tells nobody. Asking for a link must never be
        able to do that.

    Any admin of this organization can mint one, which is a real power: it
    sets someone else's password. It is the same power they already have by
    creating the account in the first place, and RLS keeps it inside the
    organization — a counselor at another JCC is a 404 here.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    db = get_db()
    try:
        if not _counselor_exists(db, counselor_id):
            return jsonify({'error': 'Counselor not found'}), 404
        # Only one setup link is live at a time. Handing out a second while the
        # first still works would leave two ways into one account with no way
        # to tell which one is loose.
        db.execute(
            "UPDATE password_reset_tokens SET used = 1 "
            " WHERE user_id = %s AND used = 0",
            (counselor_id,),
        )
        setup_url = _issue_setup_token(db, counselor_id)
        db.commit()
    except Exception as e:
        print(f"[ERROR] admin_counselor_setup_link: {e}")
        return jsonify({'error': 'Could not create a setup link'}), 500
    finally:
        db.close()

    return jsonify({'setup_url': setup_url, 'expires_in_days': 7})


# ─── COUNSELOR SCHOOL ASSIGNMENTS ───────────────────────────────────────────
# Who covers which school, and when. The program's real workflow is not "these
# are Angelie's schools" — it is "Angelie has Brown, someone called out, give
# Fischer Brown for Thursday without taking anything away from either of them".
# That needs a window and a record of who asked for it.

def _log_assignment(db, action, counselor_id, school_id, actor_id,
                    effective_from=None, effective_to=None):
    """Write one line of the assignment history.

    The counselor's and school's names are copied in rather than joined later:
    a log that stops rendering because a school was deleted is not a log. The
    same reason school_id carries no foreign key.
    """
    names = db.execute("""
        SELECT (SELECT name FROM users   WHERE id = %s) AS counselor_name,
               (SELECT name FROM schools WHERE id = %s) AS school_name,
               (SELECT name FROM users   WHERE id = %s) AS actor_name
    """, (counselor_id, school_id, actor_id)).fetchone()
    db.execute("""
        INSERT INTO counselor_school_changes
            (counselor_id, counselor_name, school_id, school_name, action,
             effective_from, effective_to, changed_by, changed_by_name)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (counselor_id, names['counselor_name'] or 'Unknown counselor',
          school_id, names['school_name'] or 'Deleted school', action,
          effective_from, effective_to,
          actor_id, names['actor_name'] or 'Unknown admin'))


def _parse_window(data):
    """Read effective_from / effective_to off a request body.

    Returns (from, to, error). Both absent is a permanent assignment, which is
    the common case and must stay the easy one to express.
    """
    window = []
    for key in ('effective_from', 'effective_to'):
        raw = data.get(key)
        if raw in (None, ''):
            window.append(None)
            continue
        try:
            window.append(parse_date(raw))
        except ValueError:
            return None, None, f'{key} must be YYYY-MM-DD'
    start, end = window
    if start and end and start > end:
        return None, None, 'effective_from must be on or before effective_to'
    return start, end, None


def _counselor_exists(db, counselor_id):
    return db.execute(
        "SELECT 1 FROM users WHERE id = %s AND role = 'counselor'",
        (counselor_id,),
    ).fetchone() is not None


@app.route('/api/admin/counselors/<int:counselor_id>/schools', methods=['GET'])
@jwt_required()
def admin_list_counselor_schools(counselor_id):
    """One counselor's assignments, each flagged with whether it applies today."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        if not _counselor_exists(db, counselor_id):
            return jsonify({'error': 'Counselor not found'}), 404
        on_date = request.args.get('date', today_for_org(db))
        try:
            parse_date(on_date)
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400
        rows = db.execute(f"""
            SELECT cs.school_id, s.name AS school,
                   cs.effective_from, cs.effective_to, cs.assigned_at,
                   u.name AS assigned_by_name,
                   {_ASSIGNMENT_IN_EFFECT} AS in_effect
              FROM counselor_schools cs
              JOIN schools s ON s.id = cs.school_id
              LEFT JOIN users u ON u.id = cs.assigned_by
             WHERE cs.counselor_id = %s
             ORDER BY s.name
        """, (on_date, on_date, counselor_id)).fetchall()
    finally:
        db.close()

    return jsonify({
        'date': on_date,
        # The distinction counselor_school_ids turns on, made explicit so the
        # screen can say it out loud rather than showing an empty list that
        # means two opposite things.
        'covers_all_schools': len(rows) == 0,
        'schools': [{
            'school_id': r['school_id'],
            'school': r['school'],
            'effective_from': r['effective_from'].isoformat() if r['effective_from'] else None,
            'effective_to': r['effective_to'].isoformat() if r['effective_to'] else None,
            'permanent': r['effective_from'] is None and r['effective_to'] is None,
            'in_effect': r['in_effect'],
            'assigned_by_name': r['assigned_by_name'],
            'assigned_at': iso_utc(r['assigned_at']),
        } for r in rows],
    })


@app.route('/api/admin/counselors/<int:counselor_id>/schools/<int:school_id>',
           methods=['POST'])
@jwt_required()
def admin_add_counselor_school(counselor_id, school_id):
    """Give a counselor a school, permanently or for a window.

    Additive: it does not touch their other schools. That is the difference
    between covering for someone and being reassigned, and the old endpoint
    could only express the second.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    start, end, error = _parse_window(data)
    if error:
        return jsonify({'error': error}), 400

    db = get_db_transaction()
    try:
        if not _counselor_exists(db, counselor_id):
            db.rollback()
            return jsonify({'error': 'Counselor not found'}), 404
        if not db.execute("SELECT 1 FROM schools WHERE id = %s",
                          (school_id,)).fetchone():
            db.rollback()
            return jsonify({'error': 'School not found'}), 404

        existing = db.execute(
            "SELECT 1 FROM counselor_schools WHERE counselor_id = %s AND school_id = %s",
            (counselor_id, school_id),
        ).fetchone()
        db.execute("""
            INSERT INTO counselor_schools
                (counselor_id, school_id, effective_from, effective_to,
                 assigned_by, assigned_at)
            VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (counselor_id, school_id) DO UPDATE
               SET effective_from = EXCLUDED.effective_from,
                   effective_to   = EXCLUDED.effective_to,
                   assigned_by    = EXCLUDED.assigned_by,
                   assigned_at    = CURRENT_TIMESTAMP
        """, (counselor_id, school_id, start, end, _current_user_id()))
        _log_assignment(db, 'updated' if existing else 'added',
                        counselor_id, school_id, _current_user_id(), start, end)
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"[ERROR] add_counselor_school: {e}")
        return jsonify({'error': 'Failed to assign the school'}), 500
    finally:
        db.close()
    return jsonify({'message': 'School assigned'}), 201


@app.route('/api/admin/counselors/<int:counselor_id>/schools/<int:school_id>',
           methods=['DELETE'])
@jwt_required()
def admin_remove_counselor_school(counselor_id, school_id):
    """Take one school off a counselor, leaving the rest alone."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db_transaction()
    try:
        row = db.execute(
            "SELECT effective_from, effective_to FROM counselor_schools "
            " WHERE counselor_id = %s AND school_id = %s",
            (counselor_id, school_id),
        ).fetchone()
        if not row:
            db.rollback()
            return jsonify({'error': 'Assignment not found'}), 404
        # Logged before the delete, so the window it had is on the record.
        _log_assignment(db, 'removed', counselor_id, school_id,
                        _current_user_id(),
                        row['effective_from'], row['effective_to'])
        db.execute(
            "DELETE FROM counselor_schools WHERE counselor_id = %s AND school_id = %s",
            (counselor_id, school_id),
        )
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"[ERROR] remove_counselor_school: {e}")
        return jsonify({'error': 'Failed to remove the school'}), 500
    finally:
        db.close()
    return jsonify({'message': 'School removed'})


@app.route('/api/admin/counselors/<int:counselor_id>/schools', methods=['PUT'])
@jwt_required()
def admin_assign_counselor_schools(counselor_id):
    """Replace a counselor's whole set of schools with the list given.

    Kept because it is the shape the add-counselor form posts, but it no
    longer deletes and re-inserts the lot. Rows that are staying put are left
    alone — they carry the window, who granted it and when, and rewriting them
    would quietly turn a Thursday-only cover into a permanent assignment and
    stamp the wrong admin's name on it.

    Every add and every removal it does still lands in the log, so a bulk set
    reads the same as the individual calls in the history.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    school_ids = (request.json or {}).get('school_ids', [])
    if not isinstance(school_ids, list):
        return jsonify({'error': 'school_ids must be a list'}), 400
    school_ids = [int(s) for s in school_ids]

    db = get_db_transaction()
    try:
        if not _counselor_exists(db, counselor_id):
            db.rollback()
            return jsonify({'error': 'Counselor not found'}), 404
        current = {
            r['school_id'] for r in db.execute(
                "SELECT school_id FROM counselor_schools WHERE counselor_id = %s",
                (counselor_id,)).fetchall()
        }
        actor = _current_user_id()
        for sid in current - set(school_ids):
            _log_assignment(db, 'removed', counselor_id, sid, actor)
            db.execute(
                "DELETE FROM counselor_schools WHERE counselor_id = %s AND school_id = %s",
                (counselor_id, sid))
        for sid in set(school_ids) - current:
            db.execute("""
                INSERT INTO counselor_schools
                    (counselor_id, school_id, assigned_by, assigned_at)
                VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
                ON CONFLICT (counselor_id, school_id) DO NOTHING
            """, (counselor_id, sid, actor))
            _log_assignment(db, 'added', counselor_id, sid, actor)
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"[ERROR] assign_counselor_schools: {e}")
        return jsonify({'error': 'Failed to assign schools'}), 500
    finally:
        db.close()
    return jsonify({'message': 'Schools assigned'})


@app.route('/api/admin/counselor-assignments/log', methods=['GET'])
@jwt_required()
def admin_counselor_assignment_log():
    """Who changed which counselor's schools, when, and for how long."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    counselor_id = request.args.get('counselor_id', type=int)
    limit = min(request.args.get('limit', default=100, type=int), 500)

    db = get_db()
    try:
        sql = """
            SELECT id, counselor_id, counselor_name, school_id, school_name,
                   action, effective_from, effective_to,
                   changed_by, changed_by_name, changed_at
              FROM counselor_school_changes
        """
        params = []
        if counselor_id:
            sql += " WHERE counselor_id = %s"
            params.append(counselor_id)
        sql += " ORDER BY changed_at DESC, id DESC LIMIT %s"
        params.append(limit)
        rows = db.execute(sql, tuple(params)).fetchall()
    finally:
        db.close()

    return jsonify([{
        'id': r['id'],
        'counselor_id': r['counselor_id'],
        'counselor_name': r['counselor_name'],
        'school_id': r['school_id'],
        'school_name': r['school_name'],
        'action': r['action'],
        'effective_from': r['effective_from'].isoformat() if r['effective_from'] else None,
        'effective_to': r['effective_to'].isoformat() if r['effective_to'] else None,
        'permanent': r['effective_from'] is None and r['effective_to'] is None,
        'changed_by_name': r['changed_by_name'],
        'changed_at': iso_utc(r['changed_at']),
    } for r in rows])

@app.route('/api/admin/parents', methods=['GET'])
@jwt_required()
def admin_get_parents():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    parents = db.execute("""
        SELECT u.id, u.name, u.email, u.phone, u.is_new, u.password_set_at,
               u.invited_at AS last_invite_at, u.last_login_at,
               COALESCE(
                 (SELECT json_agg(json_build_object(
                    'id', c.id, 'name', c.name, 'school', s.name,
                    'grade_label', c.grade_label,
                    'service_type', c.service_type, 'active', c.active,
                    'is_new', c.is_new)
                  ORDER BY c.name)
                  FROM children c JOIN schools s ON c.school_id = s.id
                  -- Reached as Contact #1 (parent_id) or a linked Contact #2
                  -- (child_contacts.user_id) — see _CHILD_BELONGS_TO_USER.
                  -- The union is over plain ids, not json, since json (unlike
                  -- jsonb) has no equality operator for UNION to dedupe with.
                  WHERE c.id IN (
                      SELECT id FROM children WHERE parent_id = u.id
                      UNION
                      SELECT child_id FROM child_contacts WHERE user_id = u.id
                  )), '[]'
               ) AS children
        FROM users u
        WHERE u.role = 'parent'
        ORDER BY u.name
    """).fetchall()
    result = []
    for p in parents:
        children = p['children'] if isinstance(p['children'], list) else json.loads(p['children']) if isinstance(p['children'], str) else []
        result.append({
            'id': p['id'],
            'name': p['name'],
            'email': p['email'],
            'phone': p['phone'] or '',
            'is_new': bool(p['is_new']),
            'password_set': p['password_set_at'] is not None,
            'last_invite_at': p['last_invite_at'].isoformat() if p['last_invite_at'] else None,
            'last_login_at': iso_utc(p['last_login_at']),
            'children': children,
        })
    db.close()
    return jsonify(result)

@app.route('/api/admin/parents', methods=['POST'])
@jwt_required()
def admin_add_parent():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json
    name = data.get('name', '').strip()
    email = data.get('email', '').strip().lower()

    if not name or not email:
        return jsonify({'error': 'Name and email required'}), 400

    random_pw = secrets.token_urlsafe(32)
    pw_hash = bcrypt.hashpw(random_pw.encode(), bcrypt.gensalt()).decode()

    db = get_db()
    try:
        cur = db.execute("INSERT INTO users (email, password_hash, role, name) VALUES (%s, %s, 'parent', %s) RETURNING id", (email, pw_hash, name))
        user_id = cur.fetchone()['id']
        setup_url = _issue_setup_token(db, user_id)
        db.commit()
    except psycopg2.IntegrityError:
        db.close()
        return jsonify({'error': 'Email already exists'}), 409

    # Invitation emails are never sent automatically when creating a parent.
    # Admins must trigger them explicitly via "Send Invites to All Parents",
    # the per-parent "Resend Invite" button, or by handing over the link below.
    #
    # The link is returned rather than logged. It used to be printed here in
    # full, which put a live password-setup credential for a named parent into
    # the application log of every deploy — readable by anyone with access to
    # Render's dashboard, and retained long after the token expires. The admin
    # who just created the account is the one person who should have it, and
    # they can mint a fresh one any time from the setup-link route.
    print(f"[PARENT ADDED] {email} — no auto-email; setup link returned to the admin.")
    msg = 'Parent added. Email the invitation, or copy their setup link.'
    db.close()
    return jsonify({
        'message': msg,
        'email': email,
        'setup_url': setup_url,
        'expires_in_days': 7,
    })

# ─── CHILDREN ───────────────────────────────────────────────────────────────
# The children table could be written, updated and deleted through the API but
# never listed, so the only way an admin could see a child was to find the
# parent who owns them. These two routes are the read side.

WEEKDAY_ORDER = ('Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday')

# Weekday ordering for the aggregated day list. Alphabetical would read
# Friday, Monday, Thursday, Tuesday, Wednesday, which is unusable on a roster.
_DAY_SORT_SQL = """CASE r.day_of_week
        WHEN 'Monday' THEN 1 WHEN 'Tuesday' THEN 2 WHEN 'Wednesday' THEN 3
        WHEN 'Thursday' THEN 4 WHEN 'Friday' THEN 5 ELSE 6 END"""


def _as_list(value):
    """json_agg comes back as a list or as text depending on the driver path."""
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        return json.loads(value)
    return []


def _child_status(child, day_name, absent_ids):
    """One word for where this child stands today.

    Two sources, in order. What a counselor recorded wins: it is an
    observation, and the attendance row carries it (or, for rows written
    before the status column existed, derive_status reads it back out of
    on_bus and checked_out_at the way this function used to).

    Everything below that is the roster talking — enrolled today, reported
    absent by a parent, not in today, no longer in the program. Those are
    predictions and defaults, and none of them should overwrite somebody who
    is standing in the building.

    The vocabulary is deliberately the same as the counselor's. It used to say
    `in_building` where the counselor said `picked_up`, and two words for one
    state is how the meanings drift apart in the first place.
    """
    if not child['active']:
        return 'inactive'
    recorded = derive_status(child.get('status'), child['on_bus'],
                             child['checked_out_at'])
    if recorded:
        return recorded
    if child['id'] in absent_ids:
        return STATUS_ABSENT
    scheduled = any(d['day'] == day_name for d in child['days'])
    return 'expected' if scheduled else 'not_scheduled'


def _parse_days_payload(data):
    """`days` from a create/update body, normalised to `[{day, class_session_ids}]`.

    Accepts the older shape too — a bare list of weekday names, still sent by
    the two "add a child" forms that have no class picker — and treats each
    one as "attending, no class" so neither caller has to change. Returns
    (days, all_class_ids, error_response); error_response is None on success.

    `dismissal_time` is carried through only when the entry includes the key
    at all — omitting it (the classes-only path most callers still use) must
    leave the stored hour untouched, not null it out. See `_sync_child_schedule`.
    """
    days = data.get('days') or []
    if not isinstance(days, list):
        return None, None, (jsonify({'error': 'days must be a list'}), 400)

    parsed = []
    all_class_ids: set[int] = set()
    for entry in days:
        if isinstance(entry, str):
            entry = {'day': entry, 'class_session_ids': []}
        if not isinstance(entry, dict) or 'day' not in entry:
            return None, None, (jsonify({
                'error': 'Each day must be a weekday name or {day, class_session_ids}'
            }), 400)
        day = entry['day']
        if day not in WEEKDAY_ORDER:
            return None, None, (jsonify({'error': f'Invalid weekday: {day}'}), 400)
        ids = entry.get('class_session_ids') or []
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return None, None, (jsonify({
                'error': f'class_session_ids for {day} must be a list of ids'
            }), 400)
        all_class_ids.update(ids)
        parsed_entry = {'day': day, 'class_session_ids': ids}
        if 'dismissal_time' in entry:
            dismissal_time = entry['dismissal_time']
            if dismissal_time is not None and (
                not isinstance(dismissal_time, int) or dismissal_time not in (3, 4, 5, 6)
            ):
                return None, None, (jsonify({
                    'error': f'dismissal_time for {day} must be 3, 4, 5, 6 or null'
                }), 400)
            parsed_entry['dismissal_time'] = dismissal_time
        parsed.append(parsed_entry)
    return parsed, all_class_ids, None


def _sync_child_schedule(db, child_id, days, all_class_ids):
    """Write this child's weekday attendance and class enrollments together.

    `days` is `[{'day': 'Monday', 'class_session_ids': [12, 15]}, ...]` — only
    the days a child should attend after this call. A day left out is a day
    they no longer attend, so neither a registration for it nor a class on it
    should survive: the two are kept as a pair here the same way the roster
    importer keeps them (`roster_staging._apply_registrations`), so a manual
    edit can never leave a stale class enrollment behind on a day the child
    was just taken off of.

    `dismissal_time` normally arrives from the roster import and is left
    alone here — a registration already on a day that stays is otherwise
    untouched. An admin correcting the hour by hand is the one exception:
    when an entry carries the `dismissal_time` key (see `_parse_days_payload`),
    it is written to that day's row, new or existing alike.

    `class_session_ids` are validated against the day they are claimed for
    before this runs — see the caller. That check is also what keeps this
    from ever enrolling a child in another organization's class: RLS makes a
    class_sessions row from another JCC invisible rather than a permission
    error, so the id simply would not have validated.
    """
    kept = [d['day'] for d in days]
    db.execute(
        "DELETE FROM registrations WHERE child_id = %s AND day_of_week <> ALL(%s)",
        (child_id, kept),
    )
    db.execute("""
        DELETE FROM class_enrollments e
              USING class_sessions cs
         WHERE e.class_session_id = cs.id AND e.child_id = %s
           AND cs.day_of_week <> ALL(%s)
    """, (child_id, kept))

    for entry in days:
        day, wanted = entry['day'], entry['class_session_ids']
        db.execute(
            "INSERT INTO registrations (child_id, day_of_week) VALUES (%s, %s) "
            "ON CONFLICT (child_id, day_of_week) DO NOTHING",
            (child_id, day),
        )
        if 'dismissal_time' in entry:
            db.execute(
                "UPDATE registrations SET dismissal_time = %s "
                "WHERE child_id = %s AND day_of_week = %s",
                (entry['dismissal_time'], child_id, day),
            )
        # Everything this child has on THIS day that is not in the new set is
        # coming off it — the rest of `wanted` gets inserted below.
        db.execute("""
            DELETE FROM class_enrollments e
                  USING class_sessions cs
             WHERE e.class_session_id = cs.id AND e.child_id = %s
               AND cs.day_of_week = %s AND cs.id <> ALL(%s)
        """, (child_id, day, wanted))
        for class_id in wanted:
            db.execute(
                "INSERT INTO class_enrollments (child_id, class_session_id) "
                "VALUES (%s, %s) ON CONFLICT DO NOTHING",
                (child_id, class_id),
            )


def _validate_class_days(db, days, all_class_ids):
    """None on success, else the (response, status) for a class on the wrong day.

    A class picked for Monday has to actually run on Monday — the picker only
    offers same-day classes, but the check belongs here too, since the id is
    the only thing that crosses the wire and nothing stops a stale or crafted
    one arriving instead.
    """
    if not all_class_ids:
        return None
    valid_day = {r['id']: r['day_of_week'] for r in db.execute(
        "SELECT id, day_of_week FROM class_sessions WHERE id = ANY(%s) AND active",
        (list(all_class_ids),),
    ).fetchall()}
    for entry in days:
        bad = [i for i in entry['class_session_ids'] if valid_day.get(i) != entry['day']]
        if bad:
            return jsonify({
                'error': f"Class {bad[0]} does not run on {entry['day']}"
            }), 400
    return None


@app.route('/api/admin/children', methods=['GET'])
@jwt_required()
def admin_get_children():
    """Every child in the organization, with schedule, guardians and status.

    Unlike the counselor roster this is not filtered to one day or one school by
    default — an admin looking for a child does not know which weekday they are
    enrolled on, which is the whole reason the screen exists.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    # Same default as every other dated endpoint, so this screen and the
    # dashboard always agree about what "today" is.
    date_str = request.args.get('date', today_for_org())
    try:
        day_name = datetime.strptime(date_str, '%Y-%m-%d').strftime('%A')
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    day_filter = request.args.get('day') or None
    if day_filter and day_filter not in WEEKDAY_ORDER:
        return jsonify({'error': 'Invalid day'}), 400

    school_filter = request.args.get('school_id') or None
    if school_filter is not None:
        try:
            school_filter = int(school_filter)
        except ValueError:
            return jsonify({'error': 'Invalid school_id'}), 400

    # Absent by default: a roster that hides withdrawn children by default is
    # what an admin wants nine times out of ten, and the tenth is one click.
    active_arg = (request.args.get('active') or 'all').lower()
    if active_arg not in ('1', '0', 'all'):
        return jsonify({'error': 'Invalid active filter'}), 400

    db = get_db()
    try:
        clauses = []
        params = [date_str]
        if school_filter is not None:
            clauses.append("c.school_id = %s")
            params.append(school_filter)
        if active_arg != 'all':
            clauses.append("c.active = %s")
            params.append(int(active_arg))
        if day_filter:
            clauses.append(
                "EXISTS (SELECT 1 FROM registrations r2"
                "         WHERE r2.child_id = c.id AND r2.day_of_week = %s)"
            )
            params.append(day_filter)
        where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''

        rows = db.execute(f"""
            SELECT c.id, c.name, c.first_name, c.last_name,
                   c.grade_label, c.grade_num, c.active, c.service_type,
                   c.arrival_mode, c.bus_rider, c.allergies,
                   c.withdrawn_at, c.withdrawn_reason,
                   s.id AS school_id, s.name AS school,
                   u.id AS parent_id, u.name AS parent_name, u.email AS parent_email,
                   ar.on_bus, ar.checked_in_at, ar.checked_out_at,
                   ar.status, ar.status_note,
                   COALESCE((
                     SELECT json_agg(json_build_object(
                              'day', r.day_of_week,
                              'dismissal_time', r.dismissal_time)
                            ORDER BY {_DAY_SORT_SQL})
                       FROM registrations r WHERE r.child_id = c.id
                   ), '[]') AS days,
                   COALESCE((
                     SELECT json_agg(json_build_object(
                              'priority', cc.priority, 'name', cc.name,
                              'phone', cc.phone, 'email', cc.email)
                            ORDER BY cc.priority)
                       FROM child_contacts cc WHERE cc.child_id = c.id
                   ), '[]') AS contacts
              FROM children c
              JOIN schools s ON s.id = c.school_id
              JOIN users u ON u.id = c.parent_id
              LEFT JOIN attendance_records ar
                     ON ar.child_id = c.id AND ar.attendance_date = %s
            {where}
             ORDER BY c.name
        """, tuple(params)).fetchall()

        absent_ids = absent_child_ids_for_date(db, date_str)
    finally:
        db.close()

    result = []
    for row in rows:
        child = dict(row)
        child['days'] = _as_list(child['days'])
        child['contacts'] = _as_list(child['contacts'])
        child['active'] = bool(child['active'])
        child['on_bus'] = bool(child['on_bus'])
        child['status'] = _child_status(child, day_name, absent_ids)
        for field in ('withdrawn_at', 'checked_in_at', 'checked_out_at'):
            child[field] = iso_utc(child[field])
        result.append(child)

    return jsonify({'date': date_str, 'day': day_name, 'children': result})


@app.route('/api/admin/children/<int:child_id>', methods=['GET'])
@jwt_required()
def admin_get_child(child_id):
    """One child, with the fields too personal or too long for a table row."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    date_str = request.args.get('date', today_for_org())
    try:
        day_name = datetime.strptime(date_str, '%Y-%m-%d').strftime('%A')
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    db = get_db()
    try:
        row = db.execute(f"""
            SELECT c.id, c.name, c.first_name, c.last_name,
                   c.grade_label, c.grade_num, c.active, c.service_type,
                   c.release_group, c.arrival_mode, c.bus_rider, c.allergies,
                   c.notes, c.dob, c.sex, c.roster_flag,
                   c.withdrawn_at, c.withdrawn_reason, c.created_at,
                   s.id AS school_id, s.name AS school,
                   u.id AS parent_id, u.name AS parent_name,
                   u.email AS parent_email, u.phone AS parent_phone,
                   ar.on_bus, ar.checked_in_at, ar.checked_out_at,
                   ar.status, ar.status_note,
                   COALESCE((
                     SELECT json_agg(json_build_object(
                              'day', r.day_of_week,
                              'dismissal_time', r.dismissal_time)
                            ORDER BY {_DAY_SORT_SQL})
                       FROM registrations r WHERE r.child_id = c.id
                   ), '[]') AS days,
                   COALESCE((
                     SELECT json_agg(json_build_object(
                              'priority', cc.priority, 'name', cc.name,
                              'phone', cc.phone, 'email', cc.email,
                              -- NULL for contact 1 (children.parent_id is
                              -- their account, not this column) and for an
                              -- unlinked contact 2. The Guardians screen uses
                              -- this to offer "Invite as parent"; the linked
                              -- account's own status (password set, last
                              -- login) is fetched separately below, in Python,
                              -- so its timestamps go through iso_utc() rather
                              -- than Postgres' own json timestamp format.
                              'user_id', cc.user_id)
                            ORDER BY cc.priority)
                       FROM child_contacts cc
                      WHERE cc.child_id = c.id
                   ), '[]') AS contacts,
                   COALESCE((
                     SELECT json_agg(json_build_object(
                              'item', cp.item, 'status', cp.status,
                              'recorded_on', cp.recorded_on,
                              'raw_value', cp.raw_value)
                            ORDER BY cp.item)
                       FROM child_compliance cp WHERE cp.child_id = c.id
                   ), '[]') AS compliance
              FROM children c
              JOIN schools s ON s.id = c.school_id
              JOIN users u ON u.id = c.parent_id
              LEFT JOIN attendance_records ar
                     ON ar.child_id = c.id AND ar.attendance_date = %s
             WHERE c.id = %s
        """, (date_str, child_id)).fetchone()

        # RLS already scopes this to the caller's organization, so a child in
        # another JCC is indistinguishable from one that does not exist. That
        # is the intended answer, not an accident.
        if not row:
            return jsonify({'error': 'Child not found'}), 404

        absent_ids = absent_child_ids_for_date(db, date_str)

        recent = db.execute("""
            SELECT ar.attendance_date, ar.on_bus,
                   ar.checked_in_at, ar.checked_out_at,
                   u.name AS submitted_by_name
              FROM attendance_records ar
              LEFT JOIN users u ON u.id = ar.submitted_by
             WHERE ar.child_id = %s
             ORDER BY ar.attendance_date DESC
             LIMIT 14
        """, (child_id,)).fetchall()

        # Where this child actually goes, day by day. Classes are read
        # straight off class_enrollments; the care room is never one of
        # those rows — nothing stores it per child — so it is worked out the
        # same way daily_routing works out anyone else's afternoon, from
        # grade against the rules in Care Rooms.
        class_rows = db.execute("""
            SELECT cs.id, cs.name, cs.day_of_week, cs.start_time, cs.end_time
              FROM class_enrollments e
              JOIN class_sessions cs ON cs.id = e.class_session_id
             WHERE e.child_id = %s AND cs.active
        """, (child_id,)).fetchall()
        care_rules = _care_rules(db)
        room_names = {r['id']: r['name']
                      for r in db.execute("SELECT id, name FROM rooms").fetchall()}

        # A linked Contact #2's own account status — kept separate from the
        # contacts json_agg above so its timestamps go through iso_utc()
        # instead of Postgres' own (Z-less) json timestamp format.
        second_guardian_accounts = {
            r['user_id']: r for r in db.execute("""
                SELECT cc.user_id, u.password_set_at IS NOT NULL AS password_set,
                       u.invited_at, u.last_login_at
                  FROM child_contacts cc
                  JOIN users u ON u.id = cc.user_id
                 WHERE cc.child_id = %s AND cc.priority = 2
            """, (child_id,)).fetchall()
        }
    finally:
        db.close()

    child = dict(row)
    child['days'] = _as_list(child['days'])
    child['contacts'] = _as_list(child['contacts'])
    for contact in child['contacts']:
        account = second_guardian_accounts.get(contact.get('user_id'))
        contact['password_set'] = bool(account['password_set']) if account else None
        contact['invited_at'] = iso_utc(account['invited_at']) if account else None
        contact['last_login_at'] = iso_utc(account['last_login_at']) if account else None
    child['compliance'] = _as_list(child['compliance'])
    child['active'] = bool(child['active'])
    child['on_bus'] = bool(child['on_bus'])
    child['status'] = _child_status(child, day_name, absent_ids)
    for field in ('withdrawn_at', 'checked_in_at', 'checked_out_at',
                  'created_at', 'dob'):
        child[field] = iso_utc(child[field])
    child['recent_attendance'] = [{
        'attendance_date': r['attendance_date'],
        'on_bus': bool(r['on_bus']),
        'checked_in_at': iso_utc(r['checked_in_at']),
        'checked_out_at': iso_utc(r['checked_out_at']),
        'submitted_by_name': r['submitted_by_name'],
    } for r in recent]

    def as_text(value):
        return value.strftime('%H:%M') if value is not None else None

    classes_by_day: dict[str, list] = {}
    for r in class_rows:
        classes_by_day.setdefault(r['day_of_week'], []).append(dict(r))
    for day_entry in child['days']:
        day = day_entry['day']
        day_classes = sorted(
            classes_by_day.get(day, []),
            key=lambda c: (c['start_time'] is None, c['start_time']),
        )
        day_entry['classes'] = [{
            'id': c['id'], 'name': c['name'],
            'start_time': as_text(c['start_time']),
            'end_time': as_text(c['end_time']),
        } for c in day_classes]
        day_entry['care'] = []
        for segment in daily_routing.care_segments(day_classes, day_entry['dismissal_time']):
            winner = (
                daily_routing.resolve_care_room(
                    care_rules, day, segment['time_block'], child['grade_num'])
                if child['grade_num'] is not None else None
            )
            day_entry['care'].append({
                'time_block': segment['time_block'],
                'room_id': winner['room_id'] if winner else None,
                'room_name': room_names.get(winner['room_id']) if winner else None,
            })

    return jsonify(child)


# ─── SECOND GUARDIAN ────────────────────────────────────────────────────────
# Giving Contact #2 (child_contacts, priority=2) a portal login of their own —
# see the comment on child_contacts in server/database.py. Contact #1 is
# untouched by all three of these: they are children.parent_id, not a row
# here, and always has had a login since the child was created.

@app.route('/api/admin/children/<int:child_id>/second-guardian', methods=['PUT'])
@jwt_required()
def admin_set_second_guardian(child_id):
    """Add or edit Contact #2's name/phone/email for this child.

    For a child added by hand, which has no child_contacts rows at all yet —
    the roster importer already writes this row for an imported child (with
    no user_id), so this is the same upsert _apply_contacts() does in
    roster_staging.py, reachable from the admin screen for the other path.

    Does not touch user_id: editing the email here is a correction to the
    contact card, not a re-link of an already-invited account — same
    separation admin_update_parent already draws for Contact #1's login.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    name = (data.get('name') or '').strip()
    phone = (data.get('phone') or '').strip() or None
    email = (data.get('email') or '').strip().lower() or None
    if not name:
        return jsonify({'error': 'A name is required'}), 400

    db = get_db()
    try:
        child = db.execute(
            "SELECT id FROM children WHERE id = %s", (child_id,)
        ).fetchone()
        if not child:
            return jsonify({'error': 'Child not found'}), 404
        row = db.execute("""
            INSERT INTO child_contacts (child_id, priority, name, phone, email)
            VALUES (%s, 2, %s, %s, %s)
            ON CONFLICT (child_id, priority) DO UPDATE
               SET name = EXCLUDED.name, phone = EXCLUDED.phone,
                   email = EXCLUDED.email
            RETURNING id, priority, name, phone, email, user_id
        """, (child_id, name, phone, email)).fetchone()
        db.commit()
    finally:
        db.close()
    return jsonify(dict(row))


@app.route('/api/admin/children/<int:child_id>/second-guardian/invite',
           methods=['POST'])
@jwt_required()
def admin_invite_second_guardian(child_id):
    """Give Contact #2 a portal login, and invite them.

    Same lookup-or-create-by-email as the roster importer's _parent_for()
    (roster_staging.py) — an email that already belongs to a users row (a
    parent at another one of their kids, say) is linked to rather than
    duplicated, and if that account already has a password set, nothing is
    emailed: it is already usable, so this only links it.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        contact = db.execute("""
            SELECT cc.id, cc.name, cc.email, cc.user_id
              FROM child_contacts cc
              JOIN children c ON c.id = cc.child_id
             WHERE cc.child_id = %s AND cc.priority = 2
        """, (child_id,)).fetchone()
        if not contact:
            return jsonify({
                'error': 'Add a name and email for the second guardian first.'
            }), 400
        if not contact['email']:
            return jsonify({'error': 'That contact has no email on file.'}), 400
        email = contact['email'].strip().lower()

        existing_user = db.execute(
            "SELECT id, name, password_set_at FROM users WHERE lower(email) = %s",
            (email,)
        ).fetchone()
        if existing_user:
            user_id = existing_user['id']
            already_usable = existing_user['password_set_at'] is not None
        else:
            pw_hash = bcrypt.hashpw(secrets.token_urlsafe(32).encode(), bcrypt.gensalt()).decode()
            user_id = db.execute(
                "INSERT INTO users (email, password_hash, role, name) "
                "VALUES (%s, %s, 'parent', %s) RETURNING id",
                (email, pw_hash, contact['name'] or email),
            ).fetchone()['id']
            already_usable = False

        db.execute(
            "UPDATE child_contacts SET user_id = %s WHERE id = %s",
            (user_id, contact['id'])
        )

        if already_usable:
            db.commit()
            return jsonify({
                'message': 'Linked to their existing account.',
                'user_id': user_id,
                'invited': False,
            })

        db.execute(
            "UPDATE password_reset_tokens SET used = 1 WHERE user_id = %s AND used = 0",
            (user_id,)
        )
        setup_url = _issue_setup_token(db, user_id)
        db.commit()
        identity = email_identity(db)
    finally:
        db.close()

    name = existing_user['name'] if existing_user else contact['name']
    ok = send_invitation(email, name or email, setup_url, identity)
    if not ok:
        return jsonify({
            'error': 'Linked, but the invitation email could not be sent. '
                     'Use "Resend invite" once mail is working, or hand over '
                     'a setup link.',
            'user_id': user_id,
            'invited': False,
        }), 502
    _mark_user_invited(user_id)
    return jsonify({'message': 'Invited', 'user_id': user_id, 'invited': True})


@app.route('/api/admin/children/<int:child_id>/second-guardian/access',
           methods=['DELETE'])
@jwt_required()
def admin_remove_second_guardian_access(child_id):
    """Take Contact #2's portal access away without deleting them as a
    contact, or the account itself (they may be Contact #1 or #2 for a
    sibling — see child_contacts.user_id ON DELETE SET NULL for why deleting
    the account outright is always safe too, just a different action:
    DELETE /api/admin/parents/<id>).
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        row = db.execute("""
            UPDATE child_contacts SET user_id = NULL
             WHERE child_id = %s AND priority = 2
            RETURNING id
        """, (child_id,)).fetchone()
        db.commit()
    finally:
        db.close()
    if not row:
        return jsonify({'error': 'No second guardian on file for this child'}), 404
    return jsonify({'message': 'Access removed'})


@app.route('/api/admin/children', methods=['POST'])
@jwt_required()
def admin_add_child():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json
    name = data.get('name', '').strip()
    parent_id = data.get('parent_id')
    school_id = data.get('school_id')
    service_type = data.get('service_type', 'Full Service').strip()

    if not name or not parent_id or not school_id:
        return jsonify({'error': 'Missing required fields'}), 400

    days, all_class_ids, err = _parse_days_payload(data)
    if err:
        return err

    # Same fields the roster importer fills in from a spreadsheet, for a child
    # entered by hand instead. grade_num is derived from grade_label the same
    # way the importer derives it, rather than trusted from the client, so a
    # care rule keyed on grade_num works the same regardless of which path
    # created the row.
    grade_label = (data.get('grade_label') or '').strip() or None
    grade_num = None
    if grade_label:
        (grade_label, grade_num), _code = parse_grade(grade_label)
    dob = data.get('dob') or None
    sex = data.get('sex') or None
    if sex not in (None, 'M', 'F'):
        return jsonify({'error': "sex must be 'M' or 'F'"}), 400
    bus_rider = data.get('bus_rider')
    arrival_mode = data.get('arrival_mode') or None
    if arrival_mode not in (None, 'bus', 'dropoff'):
        return jsonify({'error': "arrival_mode must be 'bus' or 'dropoff'"}), 400
    allergies = (data.get('allergies') or '').strip() or None
    notes = (data.get('notes') or '').strip() or None

    db = get_db()
    try:
        class_err = _validate_class_days(db, days, all_class_ids)
        if class_err:
            return class_err
        cur = db.execute(
            "INSERT INTO children (name, parent_id, school_id, service_type, "
            "grade_label, grade_num, dob, sex, bus_rider, arrival_mode, "
            "allergies, notes) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id",
            (name, parent_id, school_id, service_type, grade_label, grade_num,
             dob, sex, bus_rider, arrival_mode, allergies, notes),
        )
        child_id = cur.fetchone()['id']
        _sync_child_schedule(db, child_id, days, all_class_ids)
        db.commit()
    except psycopg2.errors.ForeignKeyViolation:
        return jsonify({'error': 'That parent or school does not exist'}), 400
    finally:
        db.close()
    return jsonify({'message': 'Child added', 'child_id': child_id})

@app.route('/api/admin/children/<int:child_id>', methods=['PUT'])
@jwt_required()
def admin_update_child(child_id):
    """Set which weekdays a child attends and which classes they're in, and/or
    the fields on their profile.

    Two things `days` is careful about, both of which it used to get wrong:

      * A body with no `days` key means "change nothing", not "attend
        nothing". It used to default to [] and then DELETE unconditionally,
        so any caller that omitted the field unenrolled the child from every
        weekday — and a child with no registrations disappears from the
        counselor's roster, the headcount and the sign-out sheet at once,
        with nothing on screen to say why.
      * Weekdays that are staying put are left alone rather than deleted and
        re-inserted, so an entry that omits `dismissal_time` never drops the
        hour the roster import set. An entry that includes the key — the
        admin's manual correction — overwrites it instead. See
        `_sync_child_schedule`.

    `days` items carry `class_session_ids` too now — see
    `_sync_child_schedule` for how that stays in step with the day itself.
    The care room a child ends up in is never part of this: it is always
    computed from grade + the rules in Care Rooms, never stored per child.

    The profile fields are the same "present means change it" rule, field by
    field — a form that only shows some of a child's data must not be able to
    null out the rest of it by omission.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}

    db = get_db()
    try:
        if 'days' in data:
            days, all_class_ids, err = _parse_days_payload(data)
            if err:
                return err
            class_err = _validate_class_days(db, days, all_class_ids)
            if class_err:
                return class_err
            _sync_child_schedule(db, child_id, days, all_class_ids)

        updates = {}
        if 'name' in data:
            name = (data.get('name') or '').strip()
            if not name:
                return jsonify({'error': 'Name cannot be empty'}), 400
            updates['name'] = name
        if 'school_id' in data:
            school_id = data.get('school_id')
            if not school_id:
                return jsonify({'error': 'School cannot be empty'}), 400
            updates['school_id'] = school_id
        if 'service_type' in data:
            updates['service_type'] = (data.get('service_type') or '').strip()
        if 'allergies' in data:
            updates['allergies'] = (data.get('allergies') or '').strip() or None
        if 'notes' in data:
            updates['notes'] = (data.get('notes') or '').strip() or None
        if 'grade_label' in data:
            grade_label = (data.get('grade_label') or '').strip() or None
            grade_num = None
            if grade_label:
                (grade_label, grade_num), _code = parse_grade(grade_label)
            updates['grade_label'] = grade_label
            updates['grade_num'] = grade_num
        if 'dob' in data:
            updates['dob'] = data.get('dob') or None
        if 'sex' in data:
            sex = data.get('sex') or None
            if sex not in (None, 'M', 'F'):
                return jsonify({'error': "sex must be 'M' or 'F'"}), 400
            updates['sex'] = sex
        if 'bus_rider' in data:
            updates['bus_rider'] = bool(data.get('bus_rider'))
        if 'arrival_mode' in data:
            arrival_mode = data.get('arrival_mode') or None
            if arrival_mode not in (None, 'bus', 'dropoff'):
                return jsonify({'error': "arrival_mode must be 'bus' or 'dropoff'"}), 400
            updates['arrival_mode'] = arrival_mode
        if 'active' in data:
            active = bool(data.get('active'))
            updates['active'] = 1 if active else 0
            # withdrawn_at/_reason exist to say why a name disappeared from the
            # roster — set them going inactive, clear them coming back so a
            # reactivated child doesn't still read as withdrawn.
            updates['withdrawn_at'] = None if active else datetime.utcnow()
            updates['withdrawn_reason'] = (
                None if active else (data.get('withdrawn_reason') or '').strip() or None
            )

        if updates:
            sets = ', '.join(f"{k} = %s" for k in updates)
            cur = db.execute(
                f"UPDATE children SET {sets} WHERE id = %s RETURNING id",
                (*updates.values(), child_id),
            )
            if cur.rowcount == 0:
                return jsonify({'error': 'Child not found'}), 404
        elif 'days' not in data:
            return jsonify({'message': 'Nothing to update'})

        db.commit()
    except psycopg2.errors.ForeignKeyViolation:
        return jsonify({'error': 'That school does not exist'}), 400
    finally:
        db.close()
    return jsonify({'message': 'Child updated'})

@app.route('/api/admin/children/<int:child_id>/reactivate', methods=['POST'])
@jwt_required()
def admin_reactivate_child(child_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    cur = db.execute("UPDATE children SET active = 1 WHERE id = %s RETURNING id", (child_id,))
    row = cur.fetchone()
    db.commit()
    db.close()
    if not row:
        return jsonify({'error': 'Child not found'}), 404
    return jsonify({'message': 'Child reactivated'})

@app.route('/api/admin/children/<int:child_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_child(child_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    db.execute("DELETE FROM children WHERE id = %s", (child_id,))
    db.commit()
    db.close()
    return jsonify({'message': 'Child deleted'})


# ─── PRIVATE ADMIN NOTES ────────────────────────────────────────────────────
# child_notes is its own table rather than another column on `children`, and
# that is the whole feature: `children.notes` already exists and already
# rides along in /api/counselor/roster's SELECT, so a counselor screen that
# started rendering it would leak whatever was written there. Nothing in this
# file ever selects child_notes into a counselor- or parent-facing response —
# every route below is require_admin() — so keeping it structurally apart is
# what makes "admin-only" true rather than merely true today.

@app.route('/api/admin/children/<int:child_id>/notes', methods=['GET'])
@jwt_required()
def admin_list_child_notes(child_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        child = db.execute("SELECT id FROM children WHERE id = %s", (child_id,)).fetchone()
        if not child:
            return jsonify({'error': 'Child not found'}), 404
        rows = db.execute("""
            SELECT id, author_name, body, created_at
              FROM child_notes WHERE child_id = %s
             ORDER BY created_at DESC, id DESC
        """, (child_id,)).fetchall()
    finally:
        db.close()
    return jsonify([{
        'id': r['id'], 'author_name': r['author_name'],
        'body': r['body'], 'created_at': iso_utc(r['created_at']),
    } for r in rows])


@app.route('/api/admin/children/<int:child_id>/notes', methods=['POST'])
@jwt_required()
def admin_add_child_note(child_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    claims = get_jwt()
    user_id = get_jwt_identity()
    author_name = claims.get('name') or 'An admin'
    body = ((request.json or {}).get('body') or '').strip()
    if not body:
        return jsonify({'error': 'Write something first'}), 400

    db = get_db()
    try:
        child = db.execute("SELECT id FROM children WHERE id = %s", (child_id,)).fetchone()
        if not child:
            return jsonify({'error': 'Child not found'}), 404
        note = db.execute("""
            INSERT INTO child_notes (child_id, author_id, author_name, body)
            VALUES (%s, %s, %s, %s)
            RETURNING id, author_name, body, created_at
        """, (child_id, user_id, author_name, body)).fetchone()
        db.commit()
    finally:
        db.close()
    return jsonify({
        'id': note['id'], 'author_name': note['author_name'],
        'body': note['body'], 'created_at': iso_utc(note['created_at']),
    }), 201


@app.route('/api/admin/children/<int:child_id>/notes/<int:note_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_child_note(child_id, note_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        cur = db.execute(
            "DELETE FROM child_notes WHERE id = %s AND child_id = %s",
            (note_id, child_id),
        )
        if cur.rowcount == 0:
            return jsonify({'error': 'Note not found'}), 404
        db.commit()
    finally:
        db.close()
    return jsonify({'message': 'Note deleted'})


@app.route('/api/admin/parents/<int:parent_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_parent(parent_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        cur = db.execute("DELETE FROM users WHERE id = %s AND role = 'parent'", (parent_id,))
        if cur.rowcount == 0:
            return jsonify({'error': 'Parent not found'}), 404
        return jsonify({'message': 'Parent deleted'})
    except psycopg2.errors.ForeignKeyViolation:
        # Audit FKs are supposed to be ON DELETE SET NULL after sql/12 — if we
        # hit this the migration hasn't reached this DB yet.
        return jsonify({'error': 'Cannot delete: linked records still reference this user. Apply sql/12_fix_user_delete_fk_cascade.sql.'}), 409
    finally:
        db.close()

@app.route('/api/admin/parents/<int:parent_id>', methods=['PATCH'])
@jwt_required()
def admin_update_parent(parent_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    updates = {}
    if 'name' in data:
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'Name cannot be empty'}), 400
        updates['name'] = name
    if 'email' in data:
        email = (data.get('email') or '').strip().lower()
        if not email:
            return jsonify({'error': 'Email cannot be empty'}), 400
        updates['email'] = email
    if not updates:
        return jsonify({'error': 'Nothing to update'}), 400

    sets = ', '.join(f"{k} = %s" for k in updates)
    db = get_db()
    try:
        cur = db.execute(
            f"UPDATE users SET {sets} WHERE id = %s AND role = 'parent' RETURNING id",
            (*updates.values(), parent_id),
        )
        if cur.rowcount == 0:
            db.close()
            return jsonify({'error': 'Parent not found'}), 404
        db.commit()
    except psycopg2.IntegrityError:
        db.close()
        return jsonify({'error': 'Email already in use'}), 409
    db.close()
    return jsonify({'message': 'Parent updated'})


@app.route('/api/admin/counselors/<int:counselor_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_counselor(counselor_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        cur = db.execute("DELETE FROM users WHERE id = %s AND role = 'counselor'", (counselor_id,))
        if cur.rowcount == 0:
            return jsonify({'error': 'Counselor not found'}), 404
        return jsonify({'message': 'Counselor deleted'})
    except psycopg2.errors.ForeignKeyViolation:
        return jsonify({'error': 'Cannot delete: linked records still reference this user. Apply sql/12_fix_user_delete_fk_cascade.sql.'}), 409
    finally:
        db.close()

@app.route('/api/admin/counselors/<int:counselor_id>', methods=['PATCH'])
@jwt_required()
def admin_update_counselor(counselor_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    updates = {}
    if 'name' in data:
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'Name cannot be empty'}), 400
        updates['name'] = name
    if 'email' in data:
        email = (data.get('email') or '').strip().lower()
        if not email:
            return jsonify({'error': 'Email cannot be empty'}), 400
        updates['email'] = email
    if not updates:
        return jsonify({'error': 'Nothing to update'}), 400

    sets = ', '.join(f"{k} = %s" for k in updates)
    db = get_db()
    try:
        cur = db.execute(
            f"UPDATE users SET {sets} WHERE id = %s AND role = 'counselor' RETURNING id",
            (*updates.values(), counselor_id),
        )
        if cur.rowcount == 0:
            db.close()
            return jsonify({'error': 'Counselor not found'}), 404
        db.commit()
    except psycopg2.IntegrityError:
        db.close()
        return jsonify({'error': 'Email already in use'}), 409
    db.close()
    return jsonify({'message': 'Counselor updated'})


@app.route('/api/admin/admins', methods=['GET'])
@jwt_required()
def admin_get_admins():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    rows = db.execute("""
        SELECT u.id, u.name, u.email, u.password_set_at, u.created_at,
               u.invited_at AS last_invite_at
        FROM users u
        WHERE u.role = 'admin'
        ORDER BY u.name
    """).fetchall()
    me_id = int(get_jwt_identity())
    result = [{
        'id': r['id'],
        'name': r['name'],
        'email': r['email'],
        'password_set': r['password_set_at'] is not None,
        'last_invite_at': r['last_invite_at'].isoformat() if r['last_invite_at'] else None,
        'created_at': r['created_at'].isoformat() if r['created_at'] else None,
        'is_self': r['id'] == me_id,
    } for r in rows]
    db.close()
    return jsonify(result)

@app.route('/api/admin/admins', methods=['POST'])
@jwt_required()
def admin_add_admin():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    name = data.get('name', '').strip()
    email = data.get('email', '').strip().lower()

    if not name or not email:
        return jsonify({'error': 'Name and email required'}), 400

    random_pw = secrets.token_urlsafe(32)
    pw_hash = bcrypt.hashpw(random_pw.encode(), bcrypt.gensalt(rounds=12)).decode()

    db = get_db()
    try:
        cur = db.execute(
            "INSERT INTO users (email, password_hash, role, name) VALUES (%s, %s, 'admin', %s) RETURNING id",
            (email, pw_hash, name)
        )
        new_id = cur.fetchone()['id']
        setup_url = _issue_setup_token(db, new_id)
        db.commit()
    except psycopg2.IntegrityError:
        db.close()
        return jsonify({'error': 'Email already exists'}), 409

    sent = send_admin_invitation(email, name, setup_url, email_identity(db))
    db.close()
    return jsonify({
        'message': 'Admin added and invitation sent' if sent else 'Admin added (email could not be sent — check SMTP config)',
        'email': email,
        'email_sent': bool(sent),
    })

@app.route('/api/admin/admins/<int:admin_id>/resend-invite', methods=['POST'])
@jwt_required()
def admin_resend_admin_invite(admin_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = %s AND role='admin'", (admin_id,)).fetchone()
    if not user:
        db.close()
        return jsonify({'error': 'Admin not found'}), 404

    random_pw = secrets.token_urlsafe(32)
    pw_hash = bcrypt.hashpw(random_pw.encode(), bcrypt.gensalt(rounds=12)).decode()
    db.execute("UPDATE users SET password_hash = %s, password_set_at = NULL WHERE id = %s", (pw_hash, admin_id))
    db.execute(
        "UPDATE password_reset_tokens SET used=1 WHERE user_id=%s AND used=0",
        (admin_id,)
    )
    setup_url = _issue_setup_token(db, admin_id)
    db.commit()

    sent = send_admin_invitation(user['email'], user['name'], setup_url, email_identity(db))
    db.close()
    return jsonify({
        'message': 'Invitation resent' if sent else 'Token reset, but email could not be sent — check SMTP config',
        'email_sent': bool(sent),
    })

@app.route('/api/admin/admins/<int:admin_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_admin(admin_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    me_id = int(get_jwt_identity())
    if admin_id == me_id:
        return jsonify({'error': "You can't delete your own admin account."}), 400

    db = get_db()
    target = db.execute("SELECT id FROM users WHERE id = %s AND role='admin'", (admin_id,)).fetchone()
    if not target:
        db.close()
        return jsonify({'error': 'Admin not found'}), 404

    remaining = db.execute("SELECT COUNT(*) AS n FROM users WHERE role='admin'").fetchone()['n']
    if remaining <= 1:
        db.close()
        return jsonify({'error': 'Cannot delete the last remaining admin.'}), 400

    db.execute("DELETE FROM users WHERE id = %s AND role = 'admin'", (admin_id,))
    db.commit()
    db.close()
    return jsonify({'message': 'Admin deleted'})


@app.route('/api/admin/admins/<int:admin_id>', methods=['PATCH'])
@jwt_required()
def admin_update_admin(admin_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    updates = {}
    if 'name' in data:
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'Name cannot be empty'}), 400
        updates['name'] = name
    if 'email' in data:
        email = (data.get('email') or '').strip().lower()
        if not email:
            return jsonify({'error': 'Email cannot be empty'}), 400
        updates['email'] = email
    if not updates:
        return jsonify({'error': 'Nothing to update'}), 400

    sets = ', '.join(f"{k} = %s" for k in updates)
    db = get_db()
    try:
        cur = db.execute(
            f"UPDATE users SET {sets} WHERE id = %s AND role = 'admin' RETURNING id",
            (*updates.values(), admin_id),
        )
        if cur.rowcount == 0:
            db.close()
            return jsonify({'error': 'Admin not found'}), 404
        db.commit()
    except psycopg2.IntegrityError:
        db.close()
        return jsonify({'error': 'Email already in use'}), 409
    db.close()
    return jsonify({'message': 'Admin updated'})

# In-memory state for the bulk "send invitations to all parents" job.
#
# History: the original implementation sent each email inline in the request
# thread, which blew past gunicorn's 120s timeout. A follow-up moved the work
# into a background thread but fanned the sends out across an 8-worker pool —
# 8 concurrent SMTP LOGINs to the same Gmail account. Gmail aggressively
# throttles concurrent logins per account ("Too many simultaneous connections"
# / "Try again later"); some sends went through, others were rejected, so a
# subset of parents silently never received their invite.
#
# Now we open ONE persistent SMTP session in the background thread and stream
# every invite through it sequentially. After the initial login the per-email
# cost is just the sendmail roundtrip (tens of ms), so this is actually faster
# than the 8-worker version while staying well under Gmail's per-account
# concurrency limits. We also record the list of addresses that failed so the
# admin UI can show exactly which parents need a manual resend.
# Progress now lives in the database (server/bulk_invites.py) rather than in
# this process, because with more than one gunicorn worker a module-level dict
# lets two workers run the job at once — which sends every parent the
# invitation twice — and lets the progress poll land on a worker that knows
# nothing about the running job.
_invite_org_id = None  # set per worker thread; see _run_all_invites_job


def _record_invite_result(email, ok, reason=None):
    bulk_invites.record_result(_current_invite_org(), email, ok, reason)


def _current_invite_org():
    """The organization the running job belongs to.

    Reads the thread's pin rather than a global, so two organizations sending
    at the same time record against their own job.
    """
    return getattr(database_module._thread_local, 'organization_id', None)


def _mark_user_invited(user_id):
    """Stamp users.invited_at after a successful invitation send. Drives the
    'Invitation sent' vs 'No invite' badge in the admin parents tab."""
    try:
        db = get_db()
        db.execute("UPDATE users SET invited_at = NOW() WHERE id = %s", (user_id,))
        db.close()
    except Exception as e:
        print(f"failed to mark user {user_id} as invited: {e}")


def _send_invitation_via(server, user_email, user_name, setup_url, identity):
    """Send a parent invitation reusing an already-authenticated SMTP session."""
    html = _build_invitation_html(user_name, setup_url, identity)
    msg = MIMEMultipart('alternative')
    msg['Subject'] = parent_invitation_subject(identity)
    msg['From'] = f"{identity.from_name} <{identity.from_email}>"
    msg['To'] = user_email
    if identity.reply_to:
        msg['Reply-To'] = identity.reply_to
    msg.attach(MIMEText(html, 'html'))
    # Envelope sender is the authenticated account — see _send_via_smtp.
    server.sendmail(EMAIL_CONFIG['smtp_user'], user_email, msg.as_string())


def _run_all_invites_via_resend(parent_setups, identity):
    """Bulk path for the Resend HTTPS API. Resend's free tier rate-limits to
    ~2 req/sec, so we throttle to 0.6s/email and retry once on 429."""
    if not EMAIL_CONFIG['resend_api_key']:
        bulk_invites.finish(
            _current_invite_org(),
            error='RESEND_API_KEY is not set on the server.',
            failed_all=[{'email': e, 'reason': 'email provider not configured'}
                        for _, e, _, _ in parent_setups],
        )
        return

    subject = parent_invitation_subject(identity)
    try:
        for user_id, email, name, url in parent_setups:
            if not email or '@' not in email:
                print(f"send-all-invites: skipping invalid email '{email}' for {name}")
                _record_invite_result(email or f'(no-email:{name})', False, 'invalid email format')
                continue
            html = _build_invitation_html(name, url, identity)
            ok, reason = _send_via_resend(email, subject, html, identity)
            if not ok and reason and reason.startswith('rate_limit'):
                time.sleep(2.0)
                ok, reason = _send_via_resend(email, subject, html, identity)
            _record_invite_result(email, ok, reason)
            if ok:
                _mark_user_invited(user_id)
            time.sleep(0.6)
    finally:
        bulk_invites.finish(_current_invite_org())


def _run_all_invites_job(parent_setups, organization_id):
    # This thread outlives the request, so it has no flask.g. Pin it to the
    # organization explicitly or every write lands in the wrong tenant.
    database_module.set_thread_organization(organization_id)
    # Resolved once for the whole run, not per parent: a roster of 400 would
    # otherwise be 400 extra SELECTs against a pool capped at 8. The pin above
    # is what makes this readable here at all — _ambient_org_id() falls back to
    # it precisely because flask.g does not exist in this thread.
    identity = email_identity()
    if EMAIL_CONFIG['provider'] == 'resend':
        _run_all_invites_via_resend(parent_setups, identity)
        return

    if not EMAIL_CONFIG['smtp_user'] or not EMAIL_CONFIG['smtp_pass']:
        bulk_invites.finish(_current_invite_org(), error='SMTP is not configured on the server.', failed_all=[{'email': e, 'reason': 'SMTP not configured'} for _, e, _, _ in parent_setups])
        return

    try:
        with smtplib.SMTP(EMAIL_CONFIG['smtp_host'], EMAIL_CONFIG['smtp_port'], timeout=30) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['smtp_user'], EMAIL_CONFIG['smtp_pass'])
            for user_id, email, name, url in parent_setups:
                if not email or '@' not in email:
                    print(f"send-all-invites: skipping invalid email '{email}' for {name}")
                    _record_invite_result(email or f'(no-email:{name})', False, 'invalid email format')
                    continue
                try:
                    _send_invitation_via(server, email, name, url, identity)
                    _record_invite_result(email, True)
                    _mark_user_invited(user_id)
                except smtplib.SMTPServerDisconnected:
                    # Gmail dropped us; reconnect once and retry this address.
                    try:
                        server = smtplib.SMTP(EMAIL_CONFIG['smtp_host'], EMAIL_CONFIG['smtp_port'], timeout=30)
                        server.starttls()
                        server.login(EMAIL_CONFIG['smtp_user'], EMAIL_CONFIG['smtp_pass'])
                        _send_invitation_via(server, email, name, url, identity)
                        _record_invite_result(email, True)
                        _mark_user_invited(user_id)
                    except Exception as e:
                        print(f"send-all-invites reconnect-retry failed for {email}: {e}")
                        _record_invite_result(email, False, f'SMTP reconnect failed: {e.__class__.__name__}')
                except Exception as e:
                    print(f"send-all-invites error for {email}: {e}")
                    _record_invite_result(email, False, f'SMTP error: {e.__class__.__name__}')
    except Exception as e:
        # Login or socket-level failure — mark every remaining parent as failed.
        print(f"send-all-invites SMTP session error: {e}")
        bulk_invites.finish(_current_invite_org(), error=f'SMTP session failed: {e.__class__.__name__}')


@app.route('/api/admin/parents/send-all-invites', methods=['POST'])
@jwt_required()
def admin_send_all_invites():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    if not PARENT_INVITES_ENABLED:
        return jsonify({
            'error': 'Parent invitations are disabled. Set PARENT_INVITES_ENABLED=1 to enable.'
        }), 403

    org_id = current_org_id()
    db = get_db()
    # Invite every parent who hasn't activated yet — both "No invite"
    # (invited_at IS NULL) and "Pending" (invited_at set, but the parent
    # never created their password). Re-sending acts as a reminder; once
    # the parent finishes setup, password_set_at is stamped and they're
    # excluded from future runs.
    parents = db.execute(
        "SELECT id, email, name FROM users WHERE role='parent' AND password_set_at IS NULL"
    ).fetchall()
    parent_setups = []
    for parent in parents:
        db.execute(
            "UPDATE password_reset_tokens SET used=1 WHERE user_id=%s AND used=0",
            (parent['id'],)
        )
        setup_url = _issue_setup_token(db, parent['id'])
        parent_setups.append((parent['id'], parent['email'], parent['name'], setup_url))
    db.commit()
    db.close()

    # Claiming is a single statement, so two workers racing here cannot both
    # win — that race is what used to double every invitation.
    if not bulk_invites.try_start(org_id, len(parent_setups)):
        return jsonify({
            'error': 'A bulk invite job is already running. Please wait for it to finish.',
            'state': bulk_invites.read(org_id),
        }), 409

    threading.Thread(
        target=_run_all_invites_job, args=(parent_setups, org_id), daemon=True
    ).start()

    return jsonify({
        'started': True,
        'total': len(parent_setups),
        'message': 'Invitations are being sent in the background.',
    })


@app.route('/api/admin/parents/send-all-invites/status', methods=['GET'])
@jwt_required()
def admin_send_all_invites_status():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    return jsonify(bulk_invites.read(current_org_id()))


@app.route('/api/admin/children/second-guardians/send-all-invites', methods=['POST'])
@jwt_required()
def admin_invite_all_second_guardians():
    """Invite every Contact #2 who has an email and no active account yet.

    The bulk twin of POST .../second-guardian/invite (per child) — same
    lookup-or-create-by-email, same _run_all_invites_job worker and the same
    bulk_invites job slot as "Invite every parent", so the two refuse to run
    at once rather than racing each other against the same Resend rate limit.

    Deduped by email in Python, not SQL: two children can share a Contact #2
    (siblings), and that has to become one account invited once, with both
    child_contacts rows linked to it — not two accounts, or two invitations
    to the same person.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    if not PARENT_INVITES_ENABLED:
        return jsonify({
            'error': 'Parent invitations are disabled. Set PARENT_INVITES_ENABLED=1 to enable.'
        }), 403

    org_id = current_org_id()
    db = get_db()
    try:
        # Not yet an active account — unlinked, or linked but still pending.
        # Mirrors admin_send_all_invites' own "no invite or pending" rule.
        rows = db.execute("""
            SELECT cc.id AS contact_id, cc.name, lower(cc.email) AS email,
                   cc.user_id, u.password_set_at
              FROM child_contacts cc
              LEFT JOIN users u ON u.id = cc.user_id
             WHERE cc.priority = 2 AND cc.email IS NOT NULL AND cc.email != ''
               AND (cc.user_id IS NULL OR u.password_set_at IS NULL)
             ORDER BY cc.email
        """).fetchall()

        setups = []
        seen: dict[str, int] = {}  # email -> user_id, one per unique person
        for row in rows:
            email = (row['email'] or '').strip()
            if not email or '@' not in email:
                continue
            if email in seen:
                user_id = seen[email]
                if row['user_id'] != user_id:
                    db.execute(
                        "UPDATE child_contacts SET user_id = %s WHERE id = %s",
                        (user_id, row['contact_id'])
                    )
                continue

            if row['user_id']:
                user_id = row['user_id']
            else:
                existing = db.execute(
                    "SELECT id FROM users WHERE lower(email) = %s", (email,)
                ).fetchone()
                if existing:
                    user_id = existing['id']
                else:
                    pw_hash = bcrypt.hashpw(
                        secrets.token_urlsafe(32).encode(), bcrypt.gensalt()
                    ).decode()
                    user_id = db.execute(
                        "INSERT INTO users (email, password_hash, role, name) "
                        "VALUES (%s, %s, 'parent', %s) RETURNING id",
                        (email, pw_hash, row['name'] or email),
                    ).fetchone()['id']
                db.execute(
                    "UPDATE child_contacts SET user_id = %s WHERE id = %s",
                    (user_id, row['contact_id'])
                )
            seen[email] = user_id

            db.execute(
                "UPDATE password_reset_tokens SET used=1 WHERE user_id=%s AND used=0",
                (user_id,)
            )
            setup_url = _issue_setup_token(db, user_id)
            setups.append((user_id, email, row['name'], setup_url))
        db.commit()
    finally:
        db.close()

    if not bulk_invites.try_start(org_id, len(setups)):
        return jsonify({
            'error': 'A bulk invite job is already running. Please wait for it to finish.',
            'state': bulk_invites.read(org_id),
        }), 409

    threading.Thread(
        target=_run_all_invites_job, args=(setups, org_id), daemon=True
    ).start()

    return jsonify({
        'started': True,
        'total': len(setups),
        'message': 'Invitations are being sent in the background.',
    })


@app.route('/api/admin/children/second-guardians/send-all-invites/status', methods=['GET'])
@jwt_required()
def admin_invite_all_second_guardians_status():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    return jsonify(bulk_invites.read(current_org_id()))


def _user_exists_with_role(db, user_id, role):
    return db.execute(
        "SELECT 1 FROM users WHERE id = %s AND role = %s",
        (user_id, role),
    ).fetchone() is not None


def _mint_setup_link(user_id, role, label):
    """One-time password-setup link for a parent or an admin.

    The same link the invitation email carries, handed back instead of sent.
    Counselors have had this since the mail provider first turned out to be
    optional (admin_counselor_setup_link, above); parents and admins did not,
    which meant that with mail down the only accounts a JCC could actually
    onboard were its counselors.

    NOT GATED BY PARENT_INVITES_ENABLED, AND THAT IS THE POINT.
        That flag exists to stop email going to parents. This route sends no
        email — it hands a link to the administrator standing in front of the
        screen, who then passes it on however they like. Gating it would
        disable the workaround in exactly the situation it exists for: mail
        switched off and a parent who still needs to get in.

    Two things it deliberately does NOT do, both inherited from the counselor
    route for the same reasons:

      * It does not send email.
      * It does not touch password_hash. resend-invite rotates the hash before
        sending, so with mail broken it locks out someone who already had a
        working password and tells nobody. Asking for a link must never be able
        to do that.

    It also does not stamp invited_at. No invitation was sent, and marking one
    as sent would put "Invitation sent" on screen next to a parent whose inbox
    is empty — the exact confusion this whole route exists to end.
    """
    db = get_db()
    try:
        if not _user_exists_with_role(db, user_id, role):
            return jsonify({'error': f'{label} not found'}), 404
        # Only one setup link is live at a time. Handing out a second while the
        # first still works would leave two ways into one account with no way
        # to tell which one is loose.
        db.execute(
            "UPDATE password_reset_tokens SET used = 1 WHERE user_id = %s AND used = 0",
            (user_id,),
        )
        setup_url = _issue_setup_token(db, user_id)
        db.commit()
    except Exception as e:
        print(f"[ERROR] setup link for {role} {user_id}: {e.__class__.__name__}")
        return jsonify({'error': 'Could not create a setup link'}), 500
    finally:
        db.close()
    return jsonify({'setup_url': setup_url, 'expires_in_days': 7})


@app.route('/api/admin/parents/<int:parent_id>/setup-link', methods=['POST'])
@jwt_required()
def admin_parent_setup_link(parent_id):
    """A one-time link that lets this parent set their own password.

    Any admin of this organization can mint one, which is a real power: it sets
    someone else's password. It is the same power they already have by creating
    the account, and RLS keeps it inside the organization — a parent at another
    JCC is a 404 here.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    return _mint_setup_link(parent_id, 'parent', 'Parent')


@app.route('/api/admin/admins/<int:admin_id>/setup-link', methods=['POST'])
@jwt_required()
def admin_admin_setup_link(admin_id):
    """A one-time link that lets a fellow administrator set their own password.

    The superadmin console already does exactly this when it creates a JCC's
    first admin (server/superadmin.py). This is the same thing for the second
    one onwards, which until now depended on mail working.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    return _mint_setup_link(admin_id, 'admin', 'Admin')


@app.route('/api/admin/parents/<int:parent_id>/resend-invite', methods=['POST'])
@jwt_required()
def admin_resend_invite(parent_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    if not PARENT_INVITES_ENABLED:
        return jsonify({
            'error': 'Parent invitations are disabled. Set PARENT_INVITES_ENABLED=1 to enable.'
        }), 403
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = %s AND role='parent'", (parent_id,)).fetchone()
    if not user:
        db.close()
        return jsonify({'error': 'Parent not found'}), 404

    db.execute(
        "UPDATE password_reset_tokens SET used=1 WHERE user_id=%s AND used=0",
        (parent_id,)
    )
    setup_url = _issue_setup_token(db, parent_id)
    db.commit()
    # Resolved before the connection goes back to the pool so it rides along on
    # the one already in hand rather than opening a second.
    identity = email_identity(db)
    db.close()

    ok = send_invitation(user['email'], user['name'], setup_url, identity)
    if not ok:
        return jsonify({
            'error': 'Setup token saved, but the invitation email could not be '
                     'sent. Check the server SMTP configuration / outbound '
                     'network access.'
        }), 502
    _mark_user_invited(parent_id)
    return jsonify({'message': 'Invitation resent'})

def parse_course_option(col_l):
    """Parse column L: JClub|School|Service / Mon-Thu or Friday Addon|Grade||Time"""
    parts = [p.strip() for p in str(col_l).split('|')]
    school = parts[1] if len(parts) > 1 else ''
    service_raw = parts[2] if len(parts) > 2 else ''
    grade_raw = parts[3] if len(parts) > 3 else ''
    # Extract service type: Full Service / Transportation Only / Walkover Only
    service_type = 'Full Service'
    if 'Transportation Only' in service_raw:
        service_type = 'Transportation Only'
    elif 'Walkover Only' in service_raw:
        service_type = 'Walkover Only'
    # Detect Friday Addon
    is_friday = 'Friday Addon' in service_raw
    # Detect release group from grade field
    release_group = None
    if 'Torah' in grade_raw:
        release_group = 'Torah'
    elif 'K-1st' in grade_raw or 'K-1 ' in grade_raw:
        release_group = 'K-1st'
    elif '2nd-5th' in grade_raw or '2-5th' in grade_raw or '2nd–5th' in grade_raw:
        release_group = '2-5th'
    return school, service_type, is_friday, release_group

@app.route('/api/admin/upload-roster', methods=['POST'])
@jwt_required()
def admin_upload_roster():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    filename = file.filename or ''
    results = {'added_parents': 0, 'added_children': 0, 'skipped_children': 0, 'deactivated_children': 0, 'errors': []}

    # Read rows from Excel or CSV
    raw_rows = []
    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if not row[0]:
                    continue
                raw_rows.append({'row_num': i + 2, 'data': row, 'format': 'excel'})
        else:
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            for i, row in enumerate(reader, 2):
                raw_rows.append({'row_num': i, 'data': row, 'format': 'csv'})
    except Exception as e:
        print(f"[ERROR] roster upload read: {e}")
        return jsonify({'error': 'Could not read file'}), 400

    if not raw_rows:
        return jsonify({'error': 'File is empty or has no data rows'}), 400

    # Group Excel rows by (child_name, parent_email) to merge Mon-Thu + Friday Addon
    # Key: (child_name_lower, parent_email_lower) -> {parent_name, parent_email, child_name, school, service_type, days: set}
    children_map = {}

    for item in raw_rows:
        row_num = item['row_num']
        try:
            if item['format'] == 'excel':
                row = item['data']
                child_name = str(row[0]).strip() if row[0] else ''
                parent_name = str(row[5]).strip() if row[5] else ''
                parent_phone = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                parent_email = str(row[7]).strip().lower() if row[7] else ''
                col_l = row[11] if len(row) > 11 else ''
                school_name, service_type, is_friday, release_group = parse_course_option(col_l)
            else:
                row = item['data']
                child_name = row.get('child_name', '').strip()
                parent_name = row.get('parent_name', '').strip()
                parent_phone = row.get('parent_phone', '').strip()
                parent_email = row.get('parent_email', '').strip().lower()
                school_name = row.get('school', '').strip()
                service_type = row.get('service_type', 'Full Service').strip()
                days_raw = row.get('days', '').strip()
                is_friday = 'Friday' in days_raw
                release_group = None  # CSV uploads don't carry grade/group info

            if not all([child_name, parent_name, parent_email, school_name]):
                results['errors'].append(f"Row {row_num}: missing required fields (child='{child_name}', parent='{parent_name}', email='{parent_email}', school='{school_name}')")
                continue

            key = (child_name.lower(), parent_email)
            if key not in children_map:
                children_map[key] = {
                    'parent_name': parent_name,
                    'parent_email': parent_email,
                    'parent_phone': parent_phone,
                    'child_name': child_name,
                    'school_name': school_name,
                    'service_type': service_type,
                    'release_group': release_group,
                    'days': {'Monday', 'Tuesday', 'Wednesday', 'Thursday'},
                }
            else:
                # If a row has a non-null release_group, update the stored one
                if release_group is not None:
                    children_map[key]['release_group'] = release_group
            if is_friday:
                children_map[key]['days'].add('Friday')

        except Exception as e:
            results['errors'].append(f"Row {row_num}: {str(e)}")

    if not children_map:
        return jsonify({'error': 'No valid children found in file. Check that column format is correct.', 'errors': results['errors']}), 400

    # Use a TRANSACTION so if anything fails, we can roll back properly
    # (get_db uses autocommit which makes rollback impossible)
    db = get_db_transaction()

    try:
        # Who is in the program right now. Compared against the same set after
        # the upload to report what *this* file removed — a plain count of
        # inactive rows would also include everyone dropped by earlier uploads,
        # and grow on every import regardless of what the admin just sent.
        was_active = {
            r['id'] for r in db.execute("SELECT id FROM children WHERE active = 1").fetchall()
        }

        # Mark ALL existing children as inactive before processing.
        # Any child that appears in the new upload will be reactivated below.
        # Children NOT in the new upload remain inactive (removed from program).
        db.execute("UPDATE children SET active = 0")

        # Reset the "NEW" pill before this upload. Anything added below sets it
        # back to 1; existing rows matched during this upload stay at 0. The
        # reset is inside the transaction, so a rollback preserves prior flags.
        db.execute("UPDATE users SET is_new = 0 WHERE role = 'parent'")
        db.execute("UPDATE children SET is_new = 0")

        # Track new parents so we can send invitations after commit
        new_parent_invites = []

        # Insert / reactivate children from new upload
        for key, child_data in children_map.items():
            parent_email = child_data['parent_email']
            parent_name = child_data['parent_name']
            parent_phone = child_data.get('parent_phone', '')
            child_name = child_data['child_name']
            school_name = child_data['school_name']
            service_type = child_data['service_type']
            release_group = child_data.get('release_group')
            days = child_data['days']

            # Ensure school exists. New schools get their division_type stamped
            # from the first row's parsed release_group so a brand-new school
            # isn't forced into the 'none' default on its inaugural upload.
            school = db.execute("SELECT id, division_type FROM schools WHERE name = %s", (school_name,)).fetchone()
            if not school:
                inferred_div = 'torah' if release_group == 'Torah' else ('grade' if release_group in ('K-1st', '2-5th') else 'none')
                db.execute("INSERT INTO schools (name, division_type) VALUES (%s, %s)", (school_name, inferred_div))
                school = db.execute("SELECT id, division_type FROM schools WHERE name = %s", (school_name,)).fetchone()

            # Coerce release_group against the school's declared division_type.
            # A dirty roster row (e.g. a Ben Gamla child arriving with 'K-1st'
            # in column L) must not be able to mis-tag the child.
            div = school['division_type']
            if div == 'torah' and release_group != 'Torah':
                if release_group is not None:
                    results['errors'].append(f"{child_name} at {school_name}: ignored '{release_group}' (school uses Torah/Regular split)")
                release_group = None
            elif div == 'grade' and release_group not in ('K-1st', '2-5th'):
                if release_group is not None:
                    results['errors'].append(f"{child_name} at {school_name}: ignored '{release_group}' (school uses grade split)")
                release_group = None
            elif div == 'none':
                release_group = None

            # Ensure parent exists.
            # Parent identity = email, exact case-insensitive match. parent_email
            # is already .strip().lower()-ed on ingest, and users.email is UNIQUE,
            # so this is a deterministic lookup. If a parent's email changes
            # between uploads they will be treated as a new parent.
            parent = db.execute("SELECT id FROM users WHERE email = %s", (parent_email,)).fetchone()
            if not parent:
                temp_pw = secrets.token_urlsafe(10)
                pw_hash = bcrypt.hashpw(temp_pw.encode(), bcrypt.gensalt()).decode()
                cur = db.execute(
                    "INSERT INTO users (email, password_hash, role, name, phone, is_new) VALUES (%s, %s, 'parent', %s, %s, 1) RETURNING id",
                    (parent_email, pw_hash, parent_name, parent_phone or None)
                )
                parent_id = cur.fetchone()['id']
                inv_token = secrets.token_urlsafe(32)
                db.execute("INSERT INTO invitations (user_id, token) VALUES (%s, %s)", (parent_id, inv_token))
                new_parent_invites.append({'email': parent_email, 'name': parent_name, 'temp_pw': temp_pw, 'token': inv_token})
                results['added_parents'] += 1
            else:
                parent_id = parent['id']
                if parent_phone:
                    db.execute("UPDATE users SET phone = %s WHERE id = %s", (parent_phone, parent_id))

            # Check if child already exists for this parent.
            # Child identity = LOWER(name) under the same parent_id. child_name
            # is .strip()-ed on ingest but NOT further normalized: a typo or
            # accent difference ("Jon" vs "John", "María" vs "Maria") will be
            # treated as a different child and flagged NEW. Revisit if fuzzy
            # matching is ever needed.
            existing_child = db.execute(
                "SELECT id FROM children WHERE LOWER(name) = %s AND parent_id = %s",
                (child_name.lower(), parent_id)
            ).fetchone()
            if existing_child:
                # Reactivate and update with latest info from the new roster
                db.execute(
                    "UPDATE children SET active = 1, release_group = %s, service_type = %s, school_id = %s WHERE id = %s",
                    (release_group, service_type, school['id'], existing_child['id'])
                )
                results['skipped_children'] += 1
                continue

            cur = db.execute(
                "INSERT INTO children (name, parent_id, school_id, service_type, release_group, active, is_new) VALUES (%s, %s, %s, %s, %s, 1, 1) RETURNING id",
                (child_name, parent_id, school['id'], service_type, release_group)
            )
            child_id = cur.fetchone()['id']
            for day in days:
                db.execute("INSERT INTO registrations (child_id, day_of_week) VALUES (%s, %s) ON CONFLICT DO NOTHING", (child_id, day))
            results['added_children'] += 1

        # How many children this upload removed: active before, not in the file.
        still_active = {
            r['id'] for r in db.execute("SELECT id FROM children WHERE active = 1").fetchall()
        }
        results['deactivated_children'] = len(was_active - still_active)

        # Everything succeeded — commit the entire transaction
        db.commit()
        db.close()

        # Invitation emails are NOT sent automatically during roster upload.
        # Use the "Send Invites to All Parents" button to send them manually.

        return jsonify(results)

    except Exception as e:
        # Something went wrong — roll back everything (children stay active as before)
        try:
            db.rollback()
        except Exception:
            pass
        db.close()
        print(f"[ERROR] upload-roster transaction failed: {e}")
        return jsonify({'error': 'Upload failed. No changes were made.', 'errors': results['errors']}), 500


# ─── DAILY OPS ROSTER IMPORT ────────────────────────────────────────────────
# The JCCSN importer, gated behind the `daily_ops` module. Distinct from
# /api/admin/upload-roster above, which reads a completely different layout by
# column position and is what every other JCC uses.
#
# Three steps on purpose: upload stages a parse, the admin reviews it and
# resolves any unknown school, and only then does commit touch `children`.
# server/roster_staging.py holds the logic; these are the seams.

@app.route('/api/admin/roster-import', methods=['POST'])
@jwt_required()
def daily_ops_roster_upload():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    filename = file.filename or ''
    if not filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Upload the roster as .xlsx — this importer reads '
                                 'the sheet layout, which CSV does not carry.'}), 400

    # Staging writes a batch and its rows; nothing reaches `children` here. One
    # transaction so a file that fails halfway leaves no half-written preview.
    db = get_db_transaction()
    try:
        result = roster_staging.stage(
            db, filename, io.BytesIO(file.read()), uploaded_by=_current_user_id()
        )
        db.commit()
    except Exception as e:
        db.rollback()
        db.close()
        # The message can carry a spreadsheet's contents, so it is logged
        # without detail and never returned.
        print(f"[ERROR] roster-import stage failed: {type(e).__name__}")
        return jsonify({'error': 'Could not read that file as a roster.'}), 400
    db.close()
    return jsonify(result), 201


@app.route('/api/admin/roster-import', methods=['GET'])
@jwt_required()
def daily_ops_roster_list():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        return jsonify(roster_staging.list_batches(db))
    finally:
        db.close()


@app.route('/api/admin/roster-import/<int:batch_id>', methods=['GET'])
@jwt_required()
def daily_ops_roster_preview(batch_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        result = roster_staging.preview(db, batch_id)
    finally:
        db.close()
    if result is None:
        return jsonify({'error': 'Import not found'}), 404
    return jsonify(result)


@app.route('/api/admin/roster-import/<int:batch_id>/commit', methods=['POST'])
@jwt_required()
def daily_ops_roster_commit(batch_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db_transaction()
    try:
        applied, error = roster_staging.commit(db, batch_id, committed_by=_current_user_id())
        if error:
            db.rollback()
            return jsonify({'error': error}), 409
        db.commit()
    except Exception as e:
        db.rollback()
        db.close()
        print(f"[ERROR] roster-import commit failed: {type(e).__name__}")
        return jsonify({'error': 'The import failed. Nothing was changed.'}), 500
    db.close()
    return jsonify({'message': 'Roster imported', 'applied': applied})


@app.route('/api/admin/roster-import/<int:batch_id>', methods=['DELETE'])
@jwt_required()
def daily_ops_roster_discard(batch_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        ok = roster_staging.discard(db, batch_id)
        db.commit()
    finally:
        db.close()
    if not ok:
        return jsonify({'error': 'No pending import with that id'}), 404
    return jsonify({'message': 'Import discarded'})


@app.route('/api/admin/school-aliases', methods=['GET'])
@jwt_required()
def daily_ops_school_aliases():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        return jsonify(roster_staging.list_school_aliases(db))
    finally:
        db.close()


@app.route('/api/admin/school-aliases/<int:alias_id>', methods=['PUT'])
@jwt_required()
def daily_ops_resolve_school_alias(alias_id):
    """Point one of the roster's school names at a school of yours."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    db = get_db()
    try:
        row = db.execute(
            "SELECT alias_norm FROM school_aliases WHERE id = %s", (alias_id,)
        ).fetchone()
        if not row:
            return jsonify({'error': 'Unknown school alias'}), 404
        alias, error = roster_staging.resolve_school_alias(
            db, row['alias_norm'],
            school_id=data.get('school_id'),
            school_name=data.get('school_name'),
        )
        if error:
            db.rollback()
            return jsonify({'error': error}), 400
        db.commit()
    finally:
        db.close()
    return jsonify(alias)


# ── Rooms — the care rooms children wait in (daily_ops) ─────────────────────
#
# R7 of the spec is explicit that these are named by the admin rather than
# hardcoded: which rooms exist varies per JCC (Ocean, Gym, Playground, WKL,
# MPR, Courts, Pool, Lower Field), and the grade split that fills them moves
# with the headcount. Everything downstream points at a row here — the care
# rules, the staff assignments, and the `Dismiss To` label a counselor reads off
# their phone.

_ROOM_COLUMNS = 'id, name, capacity_hint, sort_order, active'


def _room_payload(data, existing=None):
    """Validate a room body against the table's constraints.

    `existing` supplies the defaults for a PUT, so a body that mentions one
    field leaves the others where they were instead of nulling them.

    Returns (values, error).
    """
    defaults = existing or {
        'name': '', 'capacity_hint': None, 'sort_order': 100, 'active': True,
    }

    name = str(data.get('name', defaults['name']) or '').strip()
    if not name:
        return None, 'A room needs a name'
    if len(name) > 80:
        return None, 'That room name is too long'

    capacity = data.get('capacity_hint', defaults['capacity_hint'])
    if capacity in (None, ''):
        capacity = None
    else:
        try:
            capacity = int(capacity)
        except (TypeError, ValueError):
            return None, 'The capacity hint has to be a whole number'
        if capacity < 1:
            return None, 'The capacity hint has to be at least 1'

    order = data.get('sort_order', defaults['sort_order'])
    try:
        order = int(order)
    except (TypeError, ValueError):
        return None, 'The sort order has to be a whole number'

    return {
        'name': name,
        'capacity_hint': capacity,
        'sort_order': order,
        'active': bool(data.get('active', defaults['active'])),
    }, None


def _room_named(db, name, exclude_id=None):
    """The room already called `name`, case-insensitively, if there is one.

    The table's unique is on the exact name, so "Ocean" and "ocean" would both
    be allowed to exist. Two rooms differing only in case is a data-entry
    mistake rather than a decision, so the endpoints fold case before writing
    and leave the constraint as the backstop.
    """
    sql = f"SELECT {_ROOM_COLUMNS} FROM rooms WHERE lower(name) = lower(%s)"
    params = [name]
    if exclude_id is not None:
        sql += " AND id <> %s"
        params.append(exclude_id)
    return db.execute(sql, tuple(params)).fetchone()


@app.route('/api/admin/rooms', methods=['GET'])
@jwt_required()
def daily_ops_list_rooms():
    """Every room in display order. Archived ones only when asked for.

    care_rule_count travels with each row so the screen can say what a room is
    load-bearing for before anyone archives it.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    include_archived = request.args.get('include_archived') == '1'
    db = get_db()
    try:
        rows = db.execute(
            f"""
            SELECT {_ROOM_COLUMNS},
                   (SELECT COUNT(*) FROM care_assignment_rules c
                     WHERE c.room_id = rooms.id) AS care_rule_count
              FROM rooms
             WHERE (active OR %s)
             ORDER BY sort_order, lower(name)
            """,
            (include_archived,)
        ).fetchall()
    finally:
        db.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/rooms', methods=['POST'])
@jwt_required()
def daily_ops_create_room():
    """Create a room — or bring an archived one back under the same name.

    Typing "Ocean Room" when an archived Ocean Room exists means she wants it
    back. The name is unique per organization so a second row is not an option,
    and refusing would leave her at a dead end with no way out of the screen.

    Reviving keeps the archived row's own name and anything the body leaves out.
    The match is case-insensitive, so "pool" would otherwise silently rename
    "Pool" — and that name is what a counselor reads as `Pool - CARE`. Coming
    back is not a rename; the edit button is.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    values, error = _room_payload(data)
    if error:
        return jsonify({'error': error}), 400

    db = get_db()
    try:
        existing = _room_named(db, values['name'])
        if existing and existing['active']:
            return jsonify({
                'error': f"There is already a room called \"{existing['name']}\""
            }), 409
        if existing:
            revived, error = _room_payload(data, dict(existing))
            if error:
                return jsonify({'error': error}), 400
            row = db.execute(
                f"""
                UPDATE rooms
                   SET capacity_hint = %(capacity_hint)s,
                       sort_order = %(sort_order)s, active = TRUE
                 WHERE id = %(id)s
             RETURNING {_ROOM_COLUMNS}
                """,
                {**revived, 'id': existing['id']}
            ).fetchone()
            return jsonify(dict(row))
        try:
            row = db.execute(
                f"""
                INSERT INTO rooms (name, capacity_hint, sort_order, active)
                     VALUES (%(name)s, %(capacity_hint)s, %(sort_order)s,
                             %(active)s)
                  RETURNING {_ROOM_COLUMNS}
                """,
                values
            ).fetchone()
        except psycopg2.errors.UniqueViolation:
            # Two admins creating the same room at once. The check above lost
            # the race; the constraint caught it.
            return jsonify({
                'error': f"There is already a room called \"{values['name']}\""
            }), 409
    finally:
        db.close()
    return jsonify(dict(row)), 201


@app.route('/api/admin/rooms/<int:room_id>', methods=['PUT'])
@jwt_required()
def daily_ops_update_room(room_id):
    """Rename a room, re-order it, or un-archive it."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        existing = db.execute(
            f"SELECT {_ROOM_COLUMNS} FROM rooms WHERE id = %s", (room_id,)
        ).fetchone()
        if not existing:
            return jsonify({'error': 'Unknown room'}), 404

        values, error = _room_payload(request.json or {}, dict(existing))
        if error:
            return jsonify({'error': error}), 400

        clash = _room_named(db, values['name'], exclude_id=room_id)
        if clash:
            return jsonify({
                'error': f"There is already a room called \"{clash['name']}\""
            }), 409

        row = db.execute(
            f"""
            UPDATE rooms
               SET name = %(name)s, capacity_hint = %(capacity_hint)s,
                   sort_order = %(sort_order)s, active = %(active)s
             WHERE id = %(id)s
         RETURNING {_ROOM_COLUMNS}
            """,
            {**values, 'id': room_id}
        ).fetchone()
    finally:
        db.close()
    return jsonify(dict(row))


@app.route('/api/admin/rooms/<int:room_id>', methods=['DELETE'])
@jwt_required()
def daily_ops_delete_room(room_id):
    """Delete a room, but only while nothing points at it.

    care_assignment_rules.room_id is ON DELETE CASCADE, so deleting a room in
    use would take the grade ranges pointing at it along — and those are what
    decide where a child waits, so they would vanish without anyone being told.
    A room with dependents refuses with 409 and the count; archiving it
    (PUT active=false) is the way to retire one that is in use.

    A room with no dependents is a typo, and a typo should be removable.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        row = db.execute(
            """
            SELECT name,
                   (SELECT COUNT(*) FROM care_assignment_rules c
                     WHERE c.room_id = rooms.id) AS care_rule_count,
                   (SELECT COUNT(*) FROM staff_assignments s
                     WHERE s.room_id = rooms.id) AS staff_count
              FROM rooms WHERE id = %s
            """,
            (room_id,)
        ).fetchone()
        if not row:
            return jsonify({'error': 'Unknown room'}), 404
        # staff_assignments.room_id is ON DELETE CASCADE too, and losing the
        # staffing is the worse half: the roster comes back from a re-import,
        # the counselor lists exist nowhere but here.
        if row['care_rule_count'] or row['staff_count']:
            uses = []
            if row['care_rule_count']:
                uses.append(f"{row['care_rule_count']} care rule(s)")
            if row['staff_count']:
                uses.append(f"{row['staff_count']} staff assignment(s)")
            return jsonify({
                'error': (
                    f"\"{row['name']}\" is used by {' and '.join(uses)}. "
                    "Archive it instead, or remove those first."
                ),
                'care_rule_count': row['care_rule_count'],
                'staff_count': row['staff_count'],
            }), 409
        db.execute("DELETE FROM rooms WHERE id = %s", (room_id,))
    finally:
        db.close()
    return jsonify({'deleted': True})


# ── Class catalogue — where the hours come from (daily_ops) ────────────────
#
# The importer already creates a row per (class name, weekday) from the roster's
# M/T/W/R/F columns, with start_time, end_time and location NULL, because the
# spreadsheet does not carry them. This is the screen that fills them in, and
# until it has been used the routing engine cannot answer a single question:
# `Dismiss To` compares a class's end against the child's dismissal hour, so a
# class with no end has no computable destination.
#
# Identity is (name, weekday): "Crafting" on Monday and "Crafting" on Wednesday
# are two rows, because they can have different hours, rooms and rosters.

_CLASS_COLUMNS = (
    "id, name, day_of_week, to_char(start_time, 'HH24:MI') AS start_time, "
    "to_char(end_time, 'HH24:MI') AS end_time, location, capacity, active"
)


def _parse_clock(value):
    """A "HH:MM" string as a time, or None. Returns (time|None, error|None)."""
    if value in (None, ''):
        return None, None
    text = str(value).strip()
    for fmt in ('%H:%M', '%H:%M:%S'):
        try:
            return datetime.strptime(text, fmt).time(), None
        except ValueError:
            continue
    return None, f'"{text}" is not a time — use 24-hour HH:MM, like 15:30'


def _class_payload(data, existing=None):
    """Validate a class body against the table's constraints.

    `existing` supplies the defaults for a PUT so a partial body leaves the rest
    of the row alone. Returns (values, error).
    """
    defaults = existing or {
        'name': '', 'day_of_week': None, 'start_time': None, 'end_time': None,
        'location': None, 'capacity': None, 'active': True,
    }

    name = str(data.get('name', defaults['name']) or '').strip()
    if not name:
        return None, 'A class needs a name'
    if len(name) > 120:
        return None, 'That class name is too long'

    day = data.get('day_of_week', defaults['day_of_week'])
    if day not in daily_routing.WEEKDAYS:
        return None, ('A class needs a weekday, spelled out: '
                      + ', '.join(daily_routing.WEEKDAYS))

    times = {}
    for field in ('start_time', 'end_time'):
        raw = data.get(field, defaults[field])
        parsed, error = _parse_clock(raw)
        if error:
            return None, error
        times[field] = parsed
    # The table's CHECK enforces this too; catching it here names the field
    # instead of surfacing a constraint name.
    if times['start_time'] and times['end_time'] \
            and times['end_time'] <= times['start_time']:
        return None, 'A class has to end after it starts'

    capacity = data.get('capacity', defaults['capacity'])
    if capacity in (None, ''):
        capacity = None
    else:
        try:
            capacity = int(capacity)
        except (TypeError, ValueError):
            return None, 'The capacity has to be a whole number'
        if capacity < 1:
            return None, 'The capacity has to be at least 1'

    location = data.get('location', defaults['location'])
    location = str(location).strip() if location not in (None, '') else None

    return {
        'name': name,
        'day_of_week': day,
        'start_time': times['start_time'],
        'end_time': times['end_time'],
        'location': location,
        'capacity': capacity,
        'active': bool(data.get('active', defaults['active'])),
    }, None


def _class_row(row):
    """One class as the client sees it, with an hours suggestion if it needs one.

    `suggested_start` / `suggested_end` are only ever a proposal read out of the
    class's own name. Nothing writes them without an admin saying so — see
    daily_routing.parse_class_times for why that matters.
    """
    data = dict(row)
    data['suggested_start'] = None
    data['suggested_end'] = None
    if not data['start_time'] or not data['end_time']:
        suggestion = daily_routing.parse_class_times(data['name'])
        if suggestion:
            start, end = suggestion
            data['suggested_start'] = start.strftime('%H:%M')
            data['suggested_end'] = end.strftime('%H:%M')
    return data


@app.route('/api/admin/class-sessions', methods=['GET'])
@jwt_required()
def daily_ops_list_classes():
    """The catalogue, grouped-ready: ordered by weekday, then start, then name.

    `enrolled_count` counts children who are still active. A withdrawn child
    keeps their class_enrollments rows — the importer's withdraw branch does not
    remove them — so counting rows would report a class as fuller than the room
    it needs.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    day = request.args.get('day')
    if day and day not in daily_routing.WEEKDAYS:
        return jsonify({'error': 'Unknown weekday'}), 400
    include_inactive = request.args.get('include_inactive') == '1'
    db = get_db()
    try:
        rows = db.execute(
            f"""
            SELECT {_CLASS_COLUMNS},
                   (SELECT COUNT(*) FROM class_enrollments e
                      JOIN children c ON c.id = e.child_id
                     WHERE e.class_session_id = class_sessions.id
                       AND c.active = 1) AS enrolled_count
              FROM class_sessions
             WHERE (active OR %s)
               AND (%s IS NULL OR day_of_week = %s)
             ORDER BY array_position(%s::text[], day_of_week),
                      start_time NULLS LAST, lower(name)
            """,
            (include_inactive, day, day, list(daily_routing.WEEKDAYS))
        ).fetchall()
    finally:
        db.close()
    return jsonify([_class_row(r) for r in rows])


@app.route('/api/admin/class-sessions', methods=['POST'])
@jwt_required()
def daily_ops_create_class():
    """Add a class the roster did not mention — a swim lesson, a late signup."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    values, error = _class_payload(request.json or {})
    if error:
        return jsonify({'error': error}), 400
    db = get_db()
    try:
        try:
            row = db.execute(
                f"""
                INSERT INTO class_sessions (name, day_of_week, start_time,
                                            end_time, location, capacity, active)
                     VALUES (%(name)s, %(day_of_week)s, %(start_time)s,
                             %(end_time)s, %(location)s, %(capacity)s, %(active)s)
                  RETURNING {_CLASS_COLUMNS}
                """,
                values
            ).fetchone()
        except psycopg2.errors.UniqueViolation:
            # The start time is part of the class's identity, so the clash is
            # with the class at *that hour*, not with the name and day. Saying
            # "already exists on Tuesday" about a 4pm class when the existing
            # one is at 3pm reads as a bug in the screen: adding the second
            # sitting of HW Club is a thing Heather is allowed to do.
            when = (f" at {values['start_time']}" if values.get('start_time')
                    else " with no start time")
            return jsonify({
                'error': (f"\"{values['name']}\" already exists on "
                          f"{values['day_of_week']}{when}")
            }), 409
    finally:
        db.close()
    return jsonify(_class_row(row)), 201


@app.route('/api/admin/class-sessions/<int:class_id>', methods=['PUT'])
@jwt_required()
def daily_ops_update_class(class_id):
    """Fill in a class's hours, room and capacity — the point of the screen.

    A rename is allowed and comes back with a warning, because it has a
    consequence nothing on the screen would otherwise show: roster_staging
    matches a class by its name against the spreadsheet's own text, so the next
    import recreates the old name as a separate class and splits the enrolments
    between the two.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        existing = db.execute(
            f"SELECT {_CLASS_COLUMNS} FROM class_sessions WHERE id = %s",
            (class_id,)
        ).fetchone()
        if not existing:
            return jsonify({'error': 'Unknown class'}), 404

        values, error = _class_payload(request.json or {}, dict(existing))
        if error:
            return jsonify({'error': error}), 400

        try:
            row = db.execute(
                f"""
                UPDATE class_sessions
                   SET name = %(name)s, day_of_week = %(day_of_week)s,
                       start_time = %(start_time)s, end_time = %(end_time)s,
                       location = %(location)s, capacity = %(capacity)s,
                       active = %(active)s
                 WHERE id = %(id)s
             RETURNING {_CLASS_COLUMNS}
                """,
                {**values, 'id': class_id}
            ).fetchone()
        except psycopg2.errors.UniqueViolation:
            return jsonify({
                'error': (f"\"{values['name']}\" already exists on "
                          f"{values['day_of_week']}")
            }), 409
    finally:
        db.close()

    body = _class_row(row)
    if values['name'] != existing['name']:
        body['warning'] = (
            f"The next roster import will recreate \"{existing['name']}\" as a "
            "separate class, because it matches classes by the name in the "
            "spreadsheet."
        )
    return jsonify(body)


@app.route('/api/admin/class-sessions/<int:class_id>', methods=['DELETE'])
@jwt_required()
def daily_ops_delete_class(class_id):
    """Delete a class, but only while no child is enrolled in it.

    class_enrollments.class_session_id is ON DELETE CASCADE, so deleting a class
    with a roster would silently drop those children's enrolments and with them
    every derived view's idea of where they are at that hour. Archiving
    (PUT active=false) is how a class that ran gets retired.

    The count here is of every enrolment row, including withdrawn children's,
    because that is what the cascade would actually destroy — which is not the
    same as the enrolled_count the list reports.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        row = db.execute(
            """
            SELECT name, day_of_week,
                   (SELECT COUNT(*) FROM class_enrollments e
                     WHERE e.class_session_id = class_sessions.id) AS enrolments,
                   (SELECT COUNT(*) FROM staff_assignments s
                     WHERE s.class_session_id = class_sessions.id) AS staff_count
              FROM class_sessions WHERE id = %s
            """,
            (class_id,)
        ).fetchone()
        if not row:
            return jsonify({'error': 'Unknown class'}), 404
        # Staff counts here for the same reason the children do:
        # staff_assignments.class_session_id is ON DELETE CASCADE, and a class
        # with an instructor but no roster yet is exactly the state a new term
        # starts in.
        if row['enrolments'] or row['staff_count']:
            uses = []
            if row['enrolments']:
                uses.append(f"{row['enrolments']} child(ren) enrolled")
            if row['staff_count']:
                uses.append(f"{row['staff_count']} staff assignment(s)")
            return jsonify({
                'error': (f"\"{row['name']}\" has {' and '.join(uses)}. "
                          'Archive it instead.'),
                'enrolments': row['enrolments'],
                'staff_count': row['staff_count'],
            }), 409
        db.execute("DELETE FROM class_sessions WHERE id = %s", (class_id,))
    finally:
        db.close()
    return jsonify({'deleted': True})


@app.route('/api/admin/class-sessions/apply-time-suggestions', methods=['POST'])
@jwt_required()
def daily_ops_apply_class_time_suggestions():
    """Write the hours read out of the chosen classes' own names.

    This is the accept step, and it is the only reason parse_class_times is
    allowed to exist. Two guardrails make it safe to press:

      • Only the ids in the body. A blanket "fix everything" would be the
        silent default the migration refused to add.
      • Only columns that are still NULL. A class whose hours were typed by hand
        is never overwritten by a guess about its name.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    ids = (request.json or {}).get('ids')
    if not isinstance(ids, list) or not ids:
        return jsonify({'error': 'Choose at least one class'}), 400
    try:
        ids = [int(i) for i in ids]
    except (TypeError, ValueError):
        return jsonify({'error': 'Those are not class ids'}), 400

    applied, skipped = [], []
    db = get_db_transaction()
    try:
        rows = db.execute(
            f"SELECT {_CLASS_COLUMNS} FROM class_sessions WHERE id = ANY(%s)",
            (ids,)
        ).fetchall()
        found = {r['id'] for r in rows}
        for missing in [i for i in ids if i not in found]:
            skipped.append({'id': missing, 'reason': 'unknown'})

        for row in rows:
            if row['start_time'] and row['end_time']:
                skipped.append({'id': row['id'], 'reason': 'already set'})
                continue
            suggestion = daily_routing.parse_class_times(row['name'])
            if not suggestion:
                skipped.append({'id': row['id'], 'reason': 'no times in the name'})
                continue
            start, end = suggestion
            updated = db.execute(
                f"""
                UPDATE class_sessions
                   SET start_time = COALESCE(start_time, %s),
                       end_time = COALESCE(end_time, %s)
                 WHERE id = %s
             RETURNING {_CLASS_COLUMNS}
                """,
                (start, end, row['id'])
            ).fetchone()
            applied.append(_class_row(updated))
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
    return jsonify({'applied': applied, 'skipped': skipped})


# ── Care rules — which grade waits in which room (daily_ops) ───────────────
#
# The parenthesised text in the Care sheet's headers: "3p-4p ... Ocean Room (K)"
# and "Gym (1-4)". Per weekday (or NULL for the standard week) × block × grade
# range → room. The table has existed since sql/29 with no way in.
#
# Overlaps are legal on purpose so a one-Wednesday exception can sit on top of
# the standard week. The precedence that settles them lives in
# daily_routing.resolve_care_room and nowhere else — see there for why a total
# order matters even for two identical rules.

_CARE_COLUMNS = ("id, day_of_week, time_block, grade_min, grade_max, room_id, "
                 "priority")


def _care_payload(data, existing=None):
    """Validate a care-rule body. Returns (values, error)."""
    defaults = existing or {
        'day_of_week': None, 'time_block': None, 'grade_min': None,
        'grade_max': None, 'room_id': None, 'priority': 100,
    }

    day = data.get('day_of_week', defaults['day_of_week'])
    if day in ('', 'all'):
        day = None
    if day is not None and day not in daily_routing.WEEKDAYS:
        return None, ('A rule is either for the whole week or for one weekday: '
                      + ', '.join(daily_routing.WEEKDAYS))

    block = data.get('time_block', defaults['time_block'])
    if block not in daily_routing.TIME_BLOCKS:
        return None, ('Pick a time block: '
                      + ', '.join(daily_routing.TIME_BLOCKS))

    grades = {}
    for field, label in (('grade_min', 'lowest'), ('grade_max', 'highest')):
        raw = data.get(field, defaults[field])
        if raw in (None, ''):
            return None, f'A rule needs a {label} grade'
        try:
            grades[field] = int(raw)
        except (TypeError, ValueError):
            return None, f'The {label} grade has to be a number — K is 0'
        if not 0 <= grades[field] <= 12:
            return None, f'The {label} grade has to be between 0 (K) and 12'
    if grades['grade_min'] > grades['grade_max']:
        return None, 'The lowest grade has to be at or below the highest'

    room_id = data.get('room_id', defaults['room_id'])
    try:
        room_id = int(room_id)
    except (TypeError, ValueError):
        return None, 'A rule needs a room'

    priority = data.get('priority', defaults['priority'])
    try:
        priority = int(priority if priority not in (None, '') else 100)
    except (TypeError, ValueError):
        return None, 'The priority has to be a whole number'

    return {
        'day_of_week': day, 'time_block': block, 'room_id': room_id,
        'priority': priority, **grades,
    }, None


def _care_rules(db, day=None):
    """Every rule that can apply on `day` — its own, plus the whole-week ones.

    With no day, every rule there is. Ordered for display; the precedence that
    decides which one WINS is applied in daily_routing, never by this ORDER BY.
    """
    return [dict(r) for r in db.execute(
        f"""
        SELECT {_CARE_COLUMNS}
          FROM care_assignment_rules
         WHERE (%s IS NULL OR day_of_week IS NULL OR day_of_week = %s)
         ORDER BY array_position(%s::text[], time_block),
                  grade_min, grade_max, priority, id
        """,
        (day, day, list(daily_routing.TIME_BLOCKS))
    ).fetchall()]


def _roster_grades(db, day=None):
    """The grades that actually have an active child registered, low to high.

    This is what makes the screen's columns the program's grades rather than a
    hardcoded K-4, so a JCC with a fifth grade sees a fifth column instead of an
    invisible hole. Children with no grade_num are counted separately: they
    cannot be routed by a grade rule at all, and that is worth saying out loud.
    """
    rows = db.execute(
        """
        SELECT c.grade_num, COUNT(DISTINCT c.id) AS children
          FROM children c
          JOIN registrations r ON r.child_id = c.id
         WHERE c.active = 1
           AND (%s IS NULL OR r.day_of_week = %s)
         GROUP BY c.grade_num
        """,
        (day, day)
    ).fetchall()
    counts = {r['grade_num']: r['children'] for r in rows
              if r['grade_num'] is not None}
    ungraded = sum(r['children'] for r in rows if r['grade_num'] is None)
    return counts, ungraded


@app.route('/api/admin/care-rules', methods=['GET'])
@jwt_required()
def daily_ops_list_care_rules():
    """The rules, the headers they add up to, and who lands in each room.

    Three views of one answer, all derived from resolve_care_room so the screen
    cannot show a room the engine would not send anyone to:

      • `rules`   — the rows themselves, for editing
      • `blocks`  — the Care sheet's own headers, rebuilt: Ocean Room (K),
                    Gym (1-4). Each segment carries its headcount.
      • `matrix`  — block × grade → room, naming every rule that matched so an
                    overlap is something she sees rather than discovers.

    The headcount is an HONEST UPPER BOUND and the client labels it as one: it
    counts children registered that weekday whose grade lands in the segment.
    Subtracting the ones who are in a class for the whole hour needs the class
    hours of piece 2 and the segment logic of piece 4.

    Always about one weekday, defaulting to Monday. A whole-week view would have
    to resolve overlaps without knowing which day's rule applies, and would then
    disagree with every day it claimed to summarise. Whole-week rules show up
    under every day, flagged, which is also how she says it: "on Mondays K is in
    Ocean" is true whether the rule is Monday's or the week's.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    day = request.args.get('day') or daily_routing.WEEKDAYS[0]
    if day not in daily_routing.WEEKDAYS:
        return jsonify({'error': 'Unknown weekday'}), 400

    db = get_db()
    try:
        rules = _care_rules(db, day)
        grade_counts, ungraded = _roster_grades(db, day)
        rooms = {r['id']: r['name'] for r in db.execute(
            "SELECT id, name FROM rooms").fetchall()}
        # Rules can point at a grade nobody is in, and the roster can hold a
        # grade no rule covers. The screen has to show both, so the columns are
        # the union rather than either side alone.
        grades = sorted(set(grade_counts) | {
            g for rule in rules
            for g in range(rule['grade_min'], rule['grade_max'] + 1)
        })
    finally:
        db.close()

    blocks = daily_routing.care_blocks(rules, day, grades)
    for block in blocks:
        for segment in block['segments']:
            segment['room_name'] = rooms.get(segment['room_id'])
            segment['headcount'] = sum(
                grade_counts.get(g, 0) for g in segment['grades'])

    matrix = []
    for block in daily_routing.TIME_BLOCKS:
        cells = []
        for grade in grades:
            matched = daily_routing.care_candidates(rules, day, block, grade)
            winner = matched[0] if matched else None
            cells.append({
                'grade': grade,
                'grade_label': daily_routing.grade_label(grade),
                'children': grade_counts.get(grade, 0),
                'room_id': winner['room_id'] if winner else None,
                'room_name': rooms.get(winner['room_id']) if winner else None,
                'rule_id': winner['id'] if winner else None,
                'overruled': [r['id'] for r in matched[1:]],
            })
        matrix.append({'time_block': block, 'cells': cells})

    return jsonify({
        'day': day,
        'rules': [{**r, 'room_name': rooms.get(r['room_id']),
                   'grade_label': daily_routing.grade_range_label(
                       r['grade_min'], r['grade_max'])}
                  for r in rules],
        'blocks': blocks,
        'matrix': matrix,
        'grades': [{'grade': g,
                    'grade_label': daily_routing.grade_label(g),
                    'children': grade_counts.get(g, 0)} for g in grades],
        'ungraded_children': ungraded,
    })


def _care_conflict(db, values, exclude_id=None):
    """The id of a rule identical to `values`, or None.

    The table has no unique for this. Two identical rows would resolve
    deterministically — the lower id wins — but only by accident, and the pair
    is indistinguishable on screen, so the second one is nothing but a way to
    make the precedence look broken.
    """
    sql = """
        SELECT id FROM care_assignment_rules
         WHERE day_of_week IS NOT DISTINCT FROM %(day_of_week)s
           AND time_block = %(time_block)s
           AND grade_min = %(grade_min)s AND grade_max = %(grade_max)s
           AND room_id = %(room_id)s
    """
    params = dict(values)
    if exclude_id is not None:
        sql += " AND id <> %(exclude)s"
        params['exclude'] = exclude_id
    row = db.execute(sql + " LIMIT 1", params).fetchone()
    return row['id'] if row else None


def _care_room_ok(db, room_id):
    """(name, error). A rule may only point at a room that exists and is live.

    An archived room is a destination no screen will print, so pointing a rule
    at one produces a child whose room is a blank. 400 rather than 404: the id
    is real, the choice is not.
    """
    row = db.execute(
        "SELECT name, active FROM rooms WHERE id = %s", (room_id,)).fetchone()
    if not row:
        return None, 'That room does not exist'
    if not row['active']:
        return None, (f"\"{row['name']}\" is archived. Un-archive it first, or "
                      "pick another room.")
    return row['name'], None


@app.route('/api/admin/care-rules', methods=['POST'])
@jwt_required()
def daily_ops_create_care_rule():
    """Add a grade range to a block."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    values, error = _care_payload(request.json or {})
    if error:
        return jsonify({'error': error}), 400

    db = get_db()
    try:
        _, error = _care_room_ok(db, values['room_id'])
        if error:
            return jsonify({'error': error}), 400
        if _care_conflict(db, values):
            return jsonify({
                'error': 'That exact rule already exists.'
            }), 409
        row = db.execute(
            f"""
            INSERT INTO care_assignment_rules
                        (day_of_week, time_block, grade_min, grade_max,
                         room_id, priority)
                 VALUES (%(day_of_week)s, %(time_block)s, %(grade_min)s,
                         %(grade_max)s, %(room_id)s, %(priority)s)
              RETURNING {_CARE_COLUMNS}
            """,
            values
        ).fetchone()
    finally:
        db.close()
    return jsonify(dict(row)), 201


@app.route('/api/admin/care-rules/<int:rule_id>', methods=['PUT'])
@jwt_required()
def daily_ops_update_care_rule(rule_id):
    """Move a grade range, change its room, or re-prioritise it."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        existing = db.execute(
            f"SELECT {_CARE_COLUMNS} FROM care_assignment_rules WHERE id = %s",
            (rule_id,)
        ).fetchone()
        if not existing:
            return jsonify({'error': 'Unknown care rule'}), 404

        values, error = _care_payload(request.json or {}, dict(existing))
        if error:
            return jsonify({'error': error}), 400
        _, error = _care_room_ok(db, values['room_id'])
        if error:
            return jsonify({'error': error}), 400
        if _care_conflict(db, values, exclude_id=rule_id):
            return jsonify({'error': 'That exact rule already exists.'}), 409

        row = db.execute(
            f"""
            UPDATE care_assignment_rules
               SET day_of_week = %(day_of_week)s, time_block = %(time_block)s,
                   grade_min = %(grade_min)s, grade_max = %(grade_max)s,
                   room_id = %(room_id)s, priority = %(priority)s
             WHERE id = %(id)s
         RETURNING {_CARE_COLUMNS}
            """,
            {**values, 'id': rule_id}
        ).fetchone()
    finally:
        db.close()
    return jsonify(dict(row))


@app.route('/api/admin/care-rules/<int:rule_id>', methods=['DELETE'])
@jwt_required()
def daily_ops_delete_care_rule(rule_id):
    """Remove a rule. This one really deletes, unlike rooms and classes.

    Nothing points at a care rule. The staff assignments of piece 5 hang off
    (room_id, time_block) rather than off a rule id, precisely so that moving a
    grade range does not take the counselors standing in that room with it — she
    moves those ranges to balance the headcount, and that has to stay cheap.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        row = db.execute(
            "DELETE FROM care_assignment_rules WHERE id = %s RETURNING id",
            (rule_id,)
        ).fetchone()
        if not row:
            return jsonify({'error': 'Unknown care rule'}), 404
    finally:
        db.close()
    return jsonify({'deleted': True})


# ── Staff assignments — the counselor lists in the sheet headers (daily_ops) ─
#
# `Ocean Room (K): Katelyn, LaRae & Lila`. `Bball 3:15p-4p Courts: Fischer,
# Mattea, Mollie`. This is the thing the director asked for; rooms, the class
# catalogue and the care rules exist so these rows have something to point at.
#
# One row is one person in one place at one time, never a slot carrying a list,
# so the counselor's own view is a WHERE and not an array filter. The target is
# (room_id, time_block) rather than a care rule id — see sql/37 for why moving a
# grade range must not take the staff with it.

_STAFF_FIELDS = ('id', 'counselor_id', 'day_of_week', 'assignment_date',
                 'class_session_id', 'room_id', 'school_id', 'time_block',
                 'status')
_STAFF_COLUMNS = ', '.join(_STAFF_FIELDS)
#: The same list qualified, for the read that joins users — both tables have an
#: `id`, and an unqualified one is ambiguous rather than merely unclear.
_STAFF_COLUMNS_S = ', '.join(f's.{f}' for f in _STAFF_FIELDS)

#: Who may be given a shift. Parents obviously not; admins yes, because a
#: director covering a room herself is ordinary and refusing it would be an
#: arbitrary rule the sheets do not have.
_STAFFABLE_ROLES = ('counselor', 'admin')


def _staff_payload(data):
    """Validate a staff-assignment body against sql/37's CHECKs.

    Returns (values, error). The CHECKs are the real guarantee; naming the
    problem here is what turns a constraint violation into a sentence.
    """
    try:
        counselor_id = int(data.get('counselor_id'))
    except (TypeError, ValueError):
        return None, 'Pick someone to assign'

    day = data.get('day_of_week') or None
    raw_date = data.get('assignment_date') or None
    if day and raw_date:
        return None, ('An assignment is either part of the weekly pattern or '
                      'for one date, not both')
    if not day and not raw_date:
        return None, 'Say which weekday, or which date'
    if day and day not in daily_routing.WEEKDAYS:
        return None, 'Unknown weekday'

    assignment_date, error = _parse_day(raw_date)
    if error:
        return None, error
    if assignment_date:
        # A date carries its own weekday, and every reader takes it from there
        # rather than from a second column that could disagree.
        if _weekday_name(assignment_date) not in daily_routing.WEEKDAYS:
            return None, 'The program does not run on weekends'

    class_id = data.get('class_session_id') or None
    room_id = data.get('room_id') or None
    school_id = data.get('school_id') or None
    if sum(1 for t in (class_id, room_id, school_id) if t) != 1:
        return None, 'Assign them to a class, a care room, or a bus — not none, not two'

    block = data.get('time_block') or None
    if room_id:
        if block not in daily_routing.TIME_BLOCKS:
            return None, ('A care room shift needs a time block: '
                          + ', '.join(daily_routing.TIME_BLOCKS))
    else:
        # A class already carries its hours, and a bus is one thing for the
        # whole afternoon — either way a second time field on the row would be
        # free to disagree, so it is dropped rather than rejected. The client
        # has no reason to send it and no reason to care.
        block = None

    status = data.get('status') or 'assigned'
    if status not in ('assigned', 'removed'):
        return None, 'Unknown status'
    if status == 'removed' and not assignment_date:
        # Against the template "removed" would mean "never", which is what
        # deleting the row already means.
        return None, ('Removing someone only applies to a single date. To take '
                      'them off the weekly pattern, delete the assignment.')

    return {
        'counselor_id': counselor_id,
        'day_of_week': day,
        'assignment_date': assignment_date,
        'class_session_id': int(class_id) if class_id else None,
        'room_id': int(room_id) if room_id else None,
        'school_id': int(school_id) if school_id else None,
        'time_block': block,
        'status': status,
    }, None


def _staff_target_ok(db, values):
    """(None, error) if the counselor or the target cannot take this shift."""
    who = db.execute(
        "SELECT name, role FROM users WHERE id = %s",
        (values['counselor_id'],)
    ).fetchone()
    if not who:
        return None, 'That person is not in this program'
    if who['role'] not in _STAFFABLE_ROLES:
        return None, f"{who['name']} is a {who['role']}, not staff"

    if values['class_session_id']:
        row = db.execute(
            "SELECT name, day_of_week, active FROM class_sessions WHERE id = %s",
            (values['class_session_id'],)
        ).fetchone()
        if not row:
            return None, 'That class does not exist'
        if not row['active']:
            return None, f"\"{row['name']}\" is archived"
        # The class owns the weekday. Staffing a Monday class on Wednesday is a
        # shift nobody can work, and the table cannot see the contradiction.
        expected = row['day_of_week']
        actual = values['day_of_week'] or _weekday_name(values['assignment_date'])
        if actual != expected:
            return None, (f"\"{row['name']}\" runs on {expected}, "
                          f'not {actual}')
    elif values['room_id']:
        row = db.execute(
            "SELECT name, active FROM rooms WHERE id = %s",
            (values['room_id'],)
        ).fetchone()
        if not row:
            return None, 'That room does not exist'
        if not row['active']:
            return None, f"\"{row['name']}\" is archived"
    else:
        row = db.execute(
            "SELECT name FROM schools WHERE id = %s",
            (values['school_id'],)
        ).fetchone()
        if not row:
            return None, 'That school does not exist'
    return True, None


def _parse_day(value):
    """A "YYYY-MM-DD" string as a date. Returns (date|None, error|None).

    Not parse_date(): that one returns the string it was given and RAISES on a
    bad one, so it can neither be compared to a weekday nor asked politely
    whether the input was a date. Same shape as _parse_clock instead.
    """
    if value in (None, ''):
        return None, None
    try:
        return datetime.strptime(str(value).strip(), '%Y-%m-%d').date(), None
    except ValueError:
        return None, f'"{value}" is not a date — use YYYY-MM-DD'


def _weekday_name(day):
    """A date's weekday as the spelled-out name, or None if there is no date."""
    if day is None:
        return None
    names = ('Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday',
             'Saturday', 'Sunday')
    return names[day.weekday()]


def _staff_by_slot(db, day, on_date=None):
    """Every assignment in effect, keyed by the slot it fills.

    Returns (in_effect, stood_down), both {slot: [rows]} where a slot is
    ('class', id), ('room', id, block) or ('school', id) — a bus, which like a
    class is one thing for the whole afternoon. With `on_date` the weekly
    template is folded together with that date's overrides, which is R8's resolution:
    template, minus the 'removed' rows, plus the added ones. Removing one person
    leaves the colleagues the template put beside them.

    `stood_down` is the people a 'removed' row took out for this date, and it is
    returned rather than discarded because otherwise there is no way back: a
    resolved list simply lacks them, so a removal made by mistake could not be
    undone from the screen that made it.
    """
    rows = db.execute(
        f"""
        SELECT {_STAFF_COLUMNS_S}, u.name AS counselor_name
          FROM staff_assignments s
          JOIN users u ON u.id = s.counselor_id
         WHERE s.day_of_week = %s
            OR (%s::date IS NOT NULL AND s.assignment_date = %s)
         ORDER BY u.name
        """,
        (day, on_date, on_date)
    ).fetchall()

    def slot(row):
        if row['class_session_id']:
            return ('class', row['class_session_id'])
        if row['school_id']:
            return ('school', row['school_id'])
        return ('room', row['room_id'], row['time_block'])

    template, added, stood_down = {}, {}, {}
    for row in rows:
        key = slot(row)
        if row['assignment_date'] is None:
            template.setdefault(key, []).append(dict(row))
        elif row['status'] == 'removed':
            stood_down.setdefault(key, []).append(dict(row))
        else:
            added.setdefault(key, []).append(dict(row))

    out: dict = {}
    for key in set(template) | set(added):
        out_today = {r['counselor_id'] for r in stood_down.get(key, [])}
        people = [r for r in template.get(key, [])
                  if r['counselor_id'] not in out_today]
        seen = {r['counselor_id'] for r in people}
        # An override for someone already in the template is not a second body.
        people += [r for r in added.get(key, [])
                   if r['counselor_id'] not in seen]
        if people:
            out[key] = sorted(people, key=lambda r: (r['counselor_name'] or ''))
    return out, stood_down


def _staff_people(in_effect, counselors, classes, rooms, schools):
    """The same afternoon as `blocks`, read one person at a time.

    This is `M-Staff`, the sheet the per-slot grid deliberately is not: one row
    per counselor, three columns, which is how the office checks that nobody was
    left idle and nobody was put in two places at once. Both questions are about
    a person and neither can be answered by looking at one slot.

    Built by INVERTING the `_staff_by_slot()` result the grid is built from,
    never by querying again. The two views of one weekday disagreeing would be
    worse than either being wrong on its own, and it is the same reason the
    admin board and every counselor's day share a single `_plan_for_day()`.

    Everyone staffable is listed, including whoever has nothing that day: the
    value of the sheet is the empty row. An empty CELL, though, is not flagged —
    see `daily_routing.staff_gaps` for why the edges have to stay quiet.
    """
    by_class = {row['id']: row for row in classes}
    order = {c['id']: index for index, c in enumerate(counselors)}
    known = {c['id']: dict(c) for c in counselors}

    entries: dict[int, list[dict]] = {}
    for slot, rows in in_effect.items():
        for row in rows:
            if slot[0] == 'class':
                session = by_class.get(slot[1])
                if session is None:
                    continue  # archived out from under the assignment
                start = _parse_clock(session['start_time'])[0]
                end = _parse_clock(session['end_time'])[0]
                entry = {'kind': 'class', 'id': slot[1],
                         'label': session['name'], 'time_block': None,
                         'start_time': start, 'end_time': end}
            elif slot[0] == 'school':
                # A bus is one thing for the whole afternoon, with no hours of
                # its own to compare against a class or a room — same reason a
                # class with no hours yet lands in `unscheduled` rather than a
                # block.
                entry = {'kind': 'school', 'id': slot[1],
                         'label': schools.get(slot[1]), 'time_block': None,
                         'start_time': None, 'end_time': None}
            else:
                start, end = daily_routing.BLOCK_BOUNDS[slot[2]]
                entry = {'kind': 'room', 'id': slot[1],
                         'label': rooms.get(slot[1]), 'time_block': slot[2],
                         'start_time': start, 'end_time': end}
            entry['assignment_id'] = row['id']
            entry['only_today'] = row['assignment_date'] is not None
            entries.setdefault(row['counselor_id'], []).append(entry)
            # A row for somebody the counselor query did not return — a role
            # changed under an old assignment. Listing them with the name the
            # row carries beats dropping a shift that is really on the sheet.
            known.setdefault(row['counselor_id'],
                             {'id': row['counselor_id'],
                              'name': row['counselor_name'], 'role': None})

    people = []
    for person in sorted(known.values(),
                         key=lambda c: (order.get(c['id'], len(order)),
                                        (c['name'] or '').lower())):
        mine = entries.get(person['id'], [])
        blocks = {block: [] for block in daily_routing.TIME_BLOCKS}
        unscheduled = []
        bus = []
        for entry in sorted(mine, key=lambda e: (e['start_time'] is None,
                                                 e['start_time'] or
                                                 daily_routing.DAY_START,
                                                 (e['label'] or '').lower())):
            wire = {**entry,
                    'start_time': _clock_str(entry['start_time']),
                    'end_time': _clock_str(entry['end_time'])}
            if entry['kind'] == 'school':
                # A bus route has no hours to sit a block by — it happens
                # before the afternoon's blocks even start — so it is its own
                # list rather than sharing `unscheduled`, which means "a class
                # is missing data", a genuinely different state.
                bus.append(wire)
                continue
            # A class sits in the block its hours START in, as it does in the
            # grid: its end may fall in the next one, but the block is where the
            # children arrive. A class still waiting for its hours has no block
            # to sit in, and is listed apart rather than dropped.
            block = (entry['time_block'] if entry['kind'] == 'room'
                     else daily_routing.block_of(entry['start_time']))
            (blocks[block] if block else unscheduled).append(wire)

        overlaps = [
            {'time_block': daily_routing.block_of(a['start_time']),
             'places': [a['label'], b['label']],
             'assignment_ids': [a['assignment_id'], b['assignment_id']]}
            for a, b in daily_routing.staff_overlaps(mine)
        ]
        people.append({
            'counselor_id': person['id'],
            'name': person['name'],
            'role': person['role'],
            'blocks': blocks,
            'unscheduled': unscheduled,
            'bus': bus,
            'overlaps': overlaps,
            'gaps': daily_routing.staff_gaps(mine),
        })
    return people


@app.route('/api/admin/staff-assignments', methods=['GET'])
@jwt_required()
def daily_ops_list_staff_assignments():
    """One weekday's grid: every slot in it, and who is standing in each.

    Shaped like the sheets rather than like the table — three time blocks, each
    with its classes and its care rooms — because that is how the question gets
    asked. Nobody wonders "which rows does Mollie have"; they wonder "who is in
    the Gym from four to five".

    `?date=` resolves the weekly template against that date's overrides (R8) and
    is what piece 7's screen reads; without it this is the template itself.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    on_date, error = _parse_day(request.args.get('date'))
    if error:
        return jsonify({'error': error}), 400
    if on_date:
        day = _weekday_name(on_date)
        if day not in daily_routing.WEEKDAYS:
            return jsonify({'error': 'The program does not run on weekends'}), 400
    else:
        day = request.args.get('day') or daily_routing.WEEKDAYS[0]
        if day not in daily_routing.WEEKDAYS:
            return jsonify({'error': 'Unknown weekday'}), 400

    db = get_db()
    try:
        staff, stood_down = _staff_by_slot(db, day, on_date)
        classes = db.execute(
            f"""
            SELECT {_CLASS_COLUMNS} FROM class_sessions
             WHERE active AND day_of_week = %s
             ORDER BY start_time NULLS LAST, lower(name)
            """,
            (day,)
        ).fetchall()
        rules = _care_rules(db, day)
        grade_counts, _ungraded = _roster_grades(db, day)
        rooms = {r['id']: r['name'] for r in db.execute(
            "SELECT id, name FROM rooms").fetchall()}
        schools = db.execute("SELECT id, name FROM schools ORDER BY lower(name)").fetchall()
        bus_counts = {r['school_id']: r['n'] for r in db.execute(
            """
            SELECT c.school_id, COUNT(*) AS n
              FROM children c
              JOIN registrations r ON r.child_id = c.id
             WHERE c.active = 1 AND c.bus_rider AND r.day_of_week = %s
             GROUP BY c.school_id
            """,
            (day,)
        ).fetchall()}
        counselors = [dict(r) for r in db.execute(
            """
            SELECT id, name, role FROM users
             WHERE role = ANY(%s) ORDER BY lower(name)
            """,
            (list(_STAFFABLE_ROLES),)
        ).fetchall()]
    finally:
        db.close()

    def people(key):
        return [{'assignment_id': r['id'], 'counselor_id': r['counselor_id'],
                 'counselor_name': r['counselor_name'],
                 'only_today': r['assignment_date'] is not None}
                for r in staff.get(key, [])]

    def out_today(key):
        """Who a 'removed' row took out of this slot for this date.

        Sent so the screen can put them back. A resolved list simply lacks them,
        so without this a removal made by mistake would be unfixable from the
        screen that made it — and standing someone down is exactly the action
        taken in a hurry.
        """
        return [{'assignment_id': r['id'], 'counselor_id': r['counselor_id'],
                 'counselor_name': r['counselor_name']}
                for r in sorted(stood_down.get(key, []),
                                key=lambda r: (r['counselor_name'] or ''))]

    grades = sorted(set(grade_counts) | {
        g for rule in rules
        for g in range(rule['grade_min'], rule['grade_max'] + 1)})
    care = {b['time_block']: b
            for b in daily_routing.care_blocks(rules, day, grades)}

    warnings = []
    blocks = []
    for block in daily_routing.TIME_BLOCKS:
        in_block = []
        for row in classes:
            # A class belongs to the block its hours START in. Its end may fall
            # in the next one, but the shift is where the children arrive.
            start = _parse_clock(row['start_time'])[0]
            if start and daily_routing.block_of(start) == block:
                in_block.append({**_class_row(row),
                                 'staff': people(('class', row['id'])),
                                 'out_today': out_today(('class', row['id']))})
        rooms_here = []
        for segment in care[block]['segments']:
            key = ('room', segment['room_id'], block)
            rooms_here.append({
                'room_id': segment['room_id'],
                'room_name': rooms.get(segment['room_id']),
                'grade_label': segment['grade_label'],
                'headcount': sum(grade_counts.get(g, 0)
                                 for g in segment['grades']),
                'staff': people(key),
                'out_today': out_today(key),
            })
        for entry in in_block:
            if not entry['staff']:
                warnings.append({'code': 'class_without_staff',
                                 'time_block': block,
                                 'label': entry['name']})
        for entry in rooms_here:
            if not entry['staff']:
                warnings.append({'code': 'room_without_staff',
                                 'time_block': block,
                                 'label': entry['room_name']})
        blocks.append({'time_block': block, 'classes': in_block,
                       'rooms': rooms_here})

    # Classes whose hours are still empty have no block to sit in. Listing them
    # apart keeps them visible instead of silently unstaffable.
    unscheduled = [{**_class_row(r), 'staff': people(('class', r['id'])),
                    'out_today': out_today(('class', r['id']))}
                   for r in classes if not r['start_time']]

    # Every school is a candidate bus route, regardless of headcount — unlike a
    # class or a care room, a route is worth setting up before the roster fills
    # in, not only once it has riders.
    bus = [{'school_id': s['id'], 'school_name': s['name'],
            'headcount': bus_counts.get(s['id'], 0),
            'staff': people(('school', s['id'])),
            'out_today': out_today(('school', s['id']))}
           for s in schools]
    for entry in bus:
        if entry['headcount'] and not entry['staff']:
            warnings.append({'code': 'bus_without_staff',
                             'time_block': None, 'label': entry['school_name']})

    # The same assignments, by person. Its warnings are kept in their own array
    # rather than folded into `warnings`: that one counts places with nobody in
    # them, and a banner that says "3 places have nobody assigned" while
    # counting a double-booking among the three is simply wrong.
    roster = _staff_people(staff, counselors, classes, rooms,
                           {s['id']: s['name'] for s in schools})
    staff_warnings = []
    for person in roster:
        for clash in person['overlaps']:
            staff_warnings.append({
                'code': daily_routing.W_STAFF_OVERLAP,
                'time_block': clash['time_block'],
                'counselor_id': person['counselor_id'],
                'label': person['name'],
                'places': clash['places'],
            })
        for block in person['gaps']:
            staff_warnings.append({
                'code': daily_routing.W_STAFF_GAP,
                'time_block': block,
                'counselor_id': person['counselor_id'],
                'label': person['name'],
                'places': [],
            })

    return jsonify({
        'day': day,
        'date': on_date.isoformat() if on_date else None,
        'blocks': blocks,
        'unscheduled_classes': unscheduled,
        'bus': bus,
        'counselors': counselors,
        'people': roster,
        'warnings': warnings,
        'staff_warnings': staff_warnings,
    })


@app.route('/api/admin/staff-assignments', methods=['POST'])
@jwt_required()
def daily_ops_create_staff_assignment():
    """Put someone in a class, a care room, or on a bus."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    values, error = _staff_payload(request.json or {})
    if error:
        return jsonify({'error': error}), 400

    db = get_db()
    try:
        ok, error = _staff_target_ok(db, values)
        if error:
            return jsonify({'error': error}), 400
        try:
            row = db.execute(
                f"""
                INSERT INTO staff_assignments
                            (counselor_id, day_of_week, assignment_date,
                             class_session_id, room_id, school_id, time_block,
                             status, created_by)
                     VALUES (%(counselor_id)s, %(day_of_week)s,
                             %(assignment_date)s, %(class_session_id)s,
                             %(room_id)s, %(school_id)s, %(time_block)s,
                             %(status)s, %(created_by)s)
                  RETURNING {_STAFF_COLUMNS}
                """,
                {**values, 'created_by': _current_user_id()}
            ).fetchone()
        except psycopg2.errors.UniqueViolation:
            return jsonify({
                'error': 'That person is already assigned to this slot.'
            }), 409
    finally:
        db.close()
    return jsonify(dict(row)), 201


@app.route('/api/admin/staff-assignments/<int:assignment_id>',
           methods=['DELETE'])
@jwt_required()
def daily_ops_delete_staff_assignment(assignment_id):
    """Take someone out of a slot.

    A real delete: nothing points at an assignment. Taking a person off the
    weekly pattern is this; taking them out for one day is a 'removed' row for
    that date, which is piece 7 and leaves the pattern alone.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        row = db.execute(
            "DELETE FROM staff_assignments WHERE id = %s RETURNING id",
            (assignment_id,)
        ).fetchone()
        if not row:
            return jsonify({'error': 'Unknown assignment'}), 404
    finally:
        db.close()
    return jsonify({'deleted': True})


# ── The afternoon, resolved once (daily_ops) ───────────────────────────────
#
# Both daily views — the counselor's own day (§6.2, R9) and the admin's board
# (§6.3) — are this same answer, one of them filtered to one person. Deriving it
# twice is how they would come to disagree about the same afternoon without
# either of them looking wrong, so it is derived here and nowhere else.
#
# The columns are read raw rather than through _CLASS_COLUMNS: that one formats
# the times as 'HH:MM' strings for the editor screens, and the engine compares
# them as clock values.

_ENGINE_CLASS_COLUMNS = ('id, name, day_of_week, start_time, end_time, '
                         'location, capacity')


def _plan_for_day(db, day, on_date=None):
    """Run the routing engine over one weekday. Returns its plan, plus lookups.

    `on_date` only affects the absence marks — the plan itself is about the
    weekday, since registrations, classes and care rules are all weekly.
    """
    children = [dict(r) for r in db.execute(
        """
        SELECT c.id, c.name, c.grade_num, c.grade_label, c.allergies,
               c.school_id, s.name AS school_name, r.dismissal_time,
               c.bus_rider
          FROM children c
          JOIN registrations r ON r.child_id = c.id
          LEFT JOIN schools s ON s.id = c.school_id
         WHERE c.active = 1 AND r.day_of_week = %s
         ORDER BY c.name
        """,
        (day,)
    ).fetchall()]
    classes = [dict(r) for r in db.execute(
        f"""
        SELECT {_ENGINE_CLASS_COLUMNS} FROM class_sessions
         WHERE active AND day_of_week = %s
        """,
        (day,)
    ).fetchall()]
    enrollments = [dict(r) for r in db.execute(
        """
        SELECT e.child_id, e.class_session_id
          FROM class_enrollments e
          JOIN class_sessions cs ON cs.id = e.class_session_id
         WHERE cs.active AND cs.day_of_week = %s
        """,
        (day,)
    ).fetchall()]
    rooms = {r['id']: r['name'] for r in db.execute(
        "SELECT id, name FROM rooms").fetchall()}

    plan = daily_routing.plan_day(
        children, classes, enrollments, _care_rules(db, day), day, rooms)

    absent = (absent_child_ids_for_date(db, on_date.isoformat())
              if on_date else set())
    by_id = {c['id']: c for c in children}
    return plan, by_id, rooms, absent


def _child_line(row, by_id, absent):
    """One child as a counselor reads them, on either sheet.

    Allergies travel with every row on purpose: it is the one field that changes
    what the adult standing there does, and a counselor should never have to
    leave the list they are holding to find it.
    """
    child = by_id.get(row['child_id'], {})
    return {
        'child_id': row['child_id'],
        'name': row['child_name'],
        'grade_label': row.get('grade_label'),
        'school': child.get('school_name'),
        'dismissal_time': row.get('dismissal_time'),
        'allergies': child.get('allergies'),
        # Reported absent for the date being viewed. Kept in the list rather
        # than dropped, so a counselor knows not to go looking, and excluded
        # from the count for the same reason.
        'absent': row['child_id'] in absent,
    }


def _by_absent_then_name(line):
    return (line['absent'], (line['name'] or '').lower())


def _where_to_by_child(plan):
    """Each child's first stop after the bus (§3.4's "Where to?"), by child_id.

    Not a new routing rule — just the 3-4 block `_plan_for_day` already
    computed, read once more: the class a child starts in if it begins in
    that block, else the care room they wait in for it. Whichever comes first
    in `plan['classes']`/`plan['care']` wins, so `setdefault` rather than
    overwrite — a child cannot really be in both, but this is the same
    "trust the engine's answer" stance `dismiss_to` itself takes.
    """
    where: dict[int, dict] = {}
    for session in plan['classes']:
        if daily_routing.block_of(session.get('start_time')) != '3-4':
            continue
        for child in session['children']:
            where.setdefault(child['child_id'],
                             {'kind': 'class', 'label': session['name']})
    for group in plan['care']:
        if group['time_block'] != '3-4':
            continue
        for child in group['children']:
            where.setdefault(child['child_id'],
                             {'kind': 'care', 'label': group['room_name']})
    return where


def _clock_str(value):
    """A time as "HH:MM", or None. Not iso_utc(): that one is for timestamps.

    24-hour on the wire, because sorting the day is done on this string and the
    sheets' own "4:45p" does not sort. The client renders it.
    """
    return value.strftime('%H:%M') if value else None


@app.route('/api/admin/daily-board', methods=['GET'])
@jwt_required()
def daily_ops_daily_board():
    """Both attendance sheets for one date — §3.5 and §3.6.

    The same afternoon /api/counselor/my-day serves, without the per-person
    filter: one `_plan_for_day` call, so the board and every counselor's screen
    cannot disagree about where a child goes.

    `{Day} - Classes` is one block per class with its staff and each child's
    `Dismiss To`; `{Day} - Care` is one block per room per hour with the
    From/To column and the bold already worked out. Every block carries its
    headcount, which is the number §6.3 exists to put in front of her: it is what
    she moves a grade range to balance.

    Deliberately not here, so it does not read as an omission: the periodic
    headcount checks of §3.6, `Time Out`/`Initial` from §3.5, the bus manifest and
    the sign-out sheet. The first two need a table of attendance events —
    attendance_records is one row per child per day, not several.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    on_date, error = _parse_day(request.args.get('date') or today_for_org())
    if error:
        return jsonify({'error': error}), 400
    day = _weekday_name(on_date)
    if day not in daily_routing.WEEKDAYS:
        return jsonify({'date': on_date.isoformat(), 'day': day, 'closed': True,
                        'classes': [], 'care': [], 'warnings': []})

    db = get_db()
    try:
        plan, by_id, rooms, absent = _plan_for_day(db, day, on_date)
        staff, stood_down = _staff_by_slot(db, day, on_date)
        rules = _care_rules(db, day)
        grade_counts, ungraded = _roster_grades(db, day)
        capacities = {r['id']: r['capacity'] for r in db.execute(
            "SELECT id, capacity FROM class_sessions WHERE day_of_week = %s",
            (day,)
        ).fetchall()}
    finally:
        db.close()

    def crew(key):
        return [{'counselor_id': r['counselor_id'], 'name': r['counselor_name'],
                 'only_today': r['assignment_date'] is not None}
                for r in staff.get(key, [])]

    warnings = list(plan['warnings'])
    if ungraded:
        warnings.append({'code': daily_routing.W_NO_GRADE, 'count': ungraded,
                         'message': (f'{ungraded} child(ren) have no grade, so no '
                                     'care rule can place them.')})

    # ── {Day} - Classes ─────────────────────────────────────────────────
    sheets_classes = []
    for session in plan['classes']:
        key = ('class', session['id'])
        children = [{**_child_line(r, by_id, absent),
                     'dismiss_to': r['dismiss_to']['label'],
                     'dismiss_kind': r['dismiss_to']['kind'],
                     'chained': r['chained']}
                    for r in session['children']]
        children.sort(key=_by_absent_then_name)
        present = sum(1 for c in children if not c['absent'])
        if not staff.get(key) and children:
            warnings.append({'code': 'class_without_staff',
                             'label': session['name'],
                             'message': f"\"{session['name']}\" has nobody running it."})
        cap = capacities.get(session['id'])
        if cap and present > cap:
            warnings.append({'code': 'class_over_capacity',
                             'label': session['name'],
                             'message': (f"\"{session['name']}\" has {present} "
                                         f'children for {cap} places.')})
        sheets_classes.append({
            'class_id': session['id'],
            'name': session['name'],
            'start_time': _clock_str(session['start_time']),
            'end_time': _clock_str(session['end_time']),
            'location': session['location'],
            'capacity': cap,
            'staff': crew(key),
            'children': children,
            'present_count': present,
        })

    # ── {Day} - Care ────────────────────────────────────────────────────
    grades = sorted(set(grade_counts) | {
        g for rule in rules
        for g in range(rule['grade_min'], rule['grade_max'] + 1)})
    labels = {(b['time_block'], s['room_id']): s
              for b in daily_routing.care_blocks(rules, day, grades)
              for s in b['segments']}

    sheets_care = []
    for group in plan['care']:
        key = ('room', group['room_id'], group['time_block'])
        segment = labels.get((group['time_block'], group['room_id']), {})
        children = [{**_child_line(r, by_id, absent),
                     'partial': r['partial'],
                     'from_class': r['from_class'],
                     'to_class': r['to_class']}
                    for r in group['children']]
        children.sort(key=_by_absent_then_name)
        present = sum(1 for c in children if not c['absent'])
        if not staff.get(key) and children:
            label = group['room_name'] or 'This room'
            warnings.append({'code': 'room_without_staff',
                             'time_block': group['time_block'],
                             'label': group['room_name'],
                             'message': (f'{label} has {present} children and '
                                         f"nobody in it from "
                                         f"{group['time_block'].replace('-', ' to ')}.")})
        sheets_care.append({
            'time_block': group['time_block'],
            'room_id': group['room_id'],
            'room_name': group['room_name'],
            'grade_label': segment.get('grade_label'),
            'staff': crew(key),
            'children': children,
            'present_count': present,
            'stood_down': [r['counselor_name']
                           for r in stood_down.get(key, [])],
        })

    return jsonify({
        'date': on_date.isoformat(),
        'day': day,
        'closed': False,
        'classes': sheets_classes,
        'care': sheets_care,
        'warnings': warnings,
    })


@app.route('/api/counselor/my-day', methods=['GET'])
@jwt_required()
def counselor_my_day():
    """One counselor's own afternoon, in the order it happens.

    This replaces R9: today the director prints all four sheets for every
    counselor and highlights each person's blocks by hand with a marker. Here a
    counselor sees only their own — their classes with each child's `Dismiss To`,
    and their care rooms with who is waiting in each and for how much of the
    hour.

    Filtered by the counselor_id in the token, never by a parameter: one
    counselor must not be able to read another's day by changing a number.
    """
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403
    me = int(get_jwt_identity())

    on_date, error = _parse_day(request.args.get('date') or today_for_org())
    if error:
        return jsonify({'error': error}), 400
    day = _weekday_name(on_date)
    if day not in daily_routing.WEEKDAYS:
        return jsonify({'date': on_date.isoformat(), 'day': day,
                        'blocks': [], 'closed': True})

    db = get_db()
    try:
        in_effect, _stood_down = _staff_by_slot(db, day, on_date)
        mine = {slot: rows for slot, rows in in_effect.items()
                if any(r['counselor_id'] == me for r in rows)}
        if not mine:
            return jsonify({'date': on_date.isoformat(), 'day': day,
                            'blocks': [], 'closed': False})
        plan, by_id, rooms, absent = _plan_for_day(db, day, on_date)
        my_schools = {slot[1] for slot in mine if slot[0] == 'school'}
        school_names = ({r['id']: r['name'] for r in db.execute(
            "SELECT id, name FROM schools WHERE id = ANY(%s)",
            (list(my_schools),))}
                        if my_schools else {})
    finally:
        db.close()

    def alongside(slot):
        """The colleagues in this slot — who to hand over to, and who to find."""
        return [{'counselor_id': r['counselor_id'], 'name': r['counselor_name']}
                for r in mine[slot] if r['counselor_id'] != me]

    blocks = []
    for session in plan['classes']:
        slot = ('class', session['id'])
        if slot not in mine:
            continue
        blocks.append({
            'kind': 'class',
            'class_id': session['id'],
            'title': session['name'],
            'start_time': _clock_str(session['start_time']),
            'end_time': _clock_str(session['end_time']),
            'location': session['location'],
            'with': alongside(slot),
            'children': [
                {**_child_line(r, by_id, absent),
                 'dismiss_to': r['dismiss_to']['label'],
                 'dismiss_kind': r['dismiss_to']['kind'],
                 # R3's own convention: she marks a chained child with `**`.
                 'chained': r['chained']}
                for r in session['children']
            ],
        })

    for group in plan['care']:
        slot = ('room', group['room_id'], group['time_block'])
        if slot not in mine:
            continue
        blocks.append({
            'kind': 'care',
            'room_id': group['room_id'],
            'title': group['room_name'],
            'time_block': group['time_block'],
            'start_time': _clock_str(
                daily_routing.BLOCK_BOUNDS[group['time_block']][0]),
            'end_time': _clock_str(
                daily_routing.BLOCK_BOUNDS[group['time_block']][1]),
            'with': alongside(slot),
            'children': [
                {**_child_line(r, by_id, absent),
                 # The Care sheet's BOLD: here for only part of the hour.
                 'partial': r['partial'],
                 'from_class': r['from_class'],
                 'to_class': r['to_class']}
                for r in group['children']
            ],
        })

    if my_schools:
        # "Where to?" (§3.4): the same 3-4 block the routing engine already
        # worked out, read once more rather than re-derived — see
        # _where_to_by_child.
        where_to = _where_to_by_child(plan)
        for school_id in my_schools:
            slot = ('school', school_id)
            roster = sorted(
                (c for c in by_id.values()
                 if c['school_id'] == school_id and c['bus_rider']),
                key=lambda c: (c['name'] or '').lower())
            blocks.append({
                'kind': 'bus',
                'school_id': school_id,
                'title': school_names.get(school_id),
                'start_time': None,
                'end_time': None,
                'with': alongside(slot),
                'children': [
                    {**_child_line({'child_id': c['id'], 'child_name': c['name'],
                                    'grade_label': c['grade_label'],
                                    'dismissal_time': c['dismissal_time']},
                                   by_id, absent),
                     'dismiss_to': where_to.get(c['id'], {}).get('label'),
                     'dismiss_kind': where_to.get(c['id'], {}).get('kind', 'unknown'),
                     'chained': False}
                    for c in roster
                ],
            })

    # In the order the afternoon happens — the bus first, since it happens
    # before the first of the afternoon's own blocks. A class with no hours
    # yet has no place in that order, so it goes last rather than being
    # dropped.
    blocks.sort(key=lambda b: (b['kind'] != 'bus', b['start_time'] is None,
                               b['start_time'] or '', b['title'] or ''))
    for block in blocks:
        block['children'].sort(key=_by_absent_then_name)
        block['present_count'] = sum(
            1 for c in block['children'] if not c['absent'])

    return jsonify({
        'date': on_date.isoformat(),
        'day': day,
        'closed': False,
        'blocks': blocks,
        # Only the warnings about this counselor's own children, so the screen
        # is not a list of somebody else's data problems.
        'warnings': [
            w for w in plan['warnings']
            if w.get('child_id') in {c['child_id'] for b in blocks
                                     for c in b['children']}
        ],
    })


def _my_slots(db, me, day, on_date):
    """The blocks this counselor is standing in, as (kind, id, time_block).

    Same resolution as my-day: the weekly template folded together with the
    date's overrides. A counselor confirms children only where they are, so
    every read and every write below is bounded by this set rather than by a
    block id the client sent.
    """
    in_effect, _stood_down = _staff_by_slot(db, day, on_date)
    slots = set()
    for slot, rows in in_effect.items():
        if not any(r['counselor_id'] == me for r in rows):
            continue
        slots.add(slot if len(slot) == 3 else (slot[0], slot[1], None))
    return slots


def _block_check_target(payload):
    """The (kind, id, time_block) a write is aimed at, or (None, error)."""
    block = payload.get('block')
    if not isinstance(block, dict):
        return None, 'block is required'
    kind = block.get('kind')
    if kind not in ('class', 'room', 'school'):
        return None, "block.kind must be 'class', 'room' or 'school'"
    try:
        ref = int(block.get('id'))
    except (TypeError, ValueError):
        return None, 'block.id must be an integer'
    time_block = block.get('time_block')
    # A room is staffed and counted by the hour; a class carries its own hours
    # and a second time field on the same row is free to disagree with them.
    if kind == 'room':
        if time_block not in daily_routing.TIME_BLOCKS:
            return None, 'block.time_block must be 3-4, 4-5 or 5-6'
    else:
        time_block = None
    return (kind, ref, time_block), None


@app.route('/api/counselor/block-checks', methods=['GET'])
@jwt_required()
def counselor_block_checks():
    """Who has already been confirmed, in this counselor's own blocks.

    Deliberately not folded into /api/counselor/my-day: the plan there is
    expensive to derive and changes once a day, while this changes on every tap.
    Two queries means a confirmation invalidates the cheap one and leaves the
    afternoon's plan cached.

    Co-staffed blocks share their confirmations. Two counselors working the same
    care room are answering one question about one group of children, and a
    headcount that disagreed between their two tablets would be worse than
    useless — it is the number they use to decide whether anyone is missing.
    """
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403
    me = int(get_jwt_identity())

    on_date, error = _parse_day(request.args.get('date') or today_for_org())
    if error:
        return jsonify({'error': error}), 400
    day = _weekday_name(on_date)
    if day not in daily_routing.WEEKDAYS:
        return jsonify({'date': on_date.isoformat(), 'checks': []})

    db = get_db()
    try:
        slots = _my_slots(db, me, day, on_date)
        if not slots:
            return jsonify({'date': on_date.isoformat(), 'checks': []})
        class_ids = [s[1] for s in slots if s[0] == 'class']
        school_ids = [s[1] for s in slots if s[0] == 'school']
        room_slots = [(s[1], s[2]) for s in slots if s[0] == 'room']
        rows = db.execute(
            """
            SELECT child_id, class_session_id, room_id, school_id, time_block
              FROM block_checks
             WHERE check_date = %s
               AND (class_session_id = ANY(%s)
                    OR school_id = ANY(%s)
                    OR (room_id, time_block) IN (
                        SELECT * FROM unnest(%s::int[], %s::text[])))
            """,
            (on_date, class_ids, school_ids,
             [r[0] for r in room_slots], [r[1] for r in room_slots])
        ).fetchall()
    finally:
        db.close()

    return jsonify({
        'date': on_date.isoformat(),
        'checks': [
            {'kind': ('class' if r['class_session_id']
                      else 'school' if r['school_id'] else 'room'),
             'id': r['class_session_id'] or r['school_id'] or r['room_id'],
             'time_block': r['time_block'],
             'child_id': r['child_id']}
            for r in rows
        ],
    })


@app.route('/api/counselor/block-checks', methods=['POST'])
@jwt_required()
def counselor_set_block_checks():
    """Confirm — or un-confirm — children in one block.

    Takes a list rather than a single child so that "route all 5" is one round
    trip and not five, and so a whole group cannot end up half written.

    A stray child_id is not validated against the block's roster, and does not
    need to be: every read is bounded by the counselor's own slots, and the
    screen intersects those rows with the roster it is drawing. A row for a
    child who is not in the block is inert — it can never reach a headcount,
    because the headcount is counted over the block's children, not over these
    rows. Re-deriving the whole afternoon's plan on every tap to prove that
    would cost the one thing a counselor tapping down a line does not have.
    """
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403
    me = int(get_jwt_identity())

    payload = request.get_json(silent=True) or {}
    on_date, error = _parse_day(payload.get('date') or today_for_org())
    if error:
        return jsonify({'error': error}), 400
    target, error = _block_check_target(payload)
    if error:
        return jsonify({'error': error}), 400
    kind, ref, time_block = target

    child_ids = payload.get('child_ids')
    if not isinstance(child_ids, list) or not child_ids:
        return jsonify({'error': 'child_ids must be a non-empty list'}), 400
    try:
        child_ids = [int(c) for c in child_ids]
    except (TypeError, ValueError):
        return jsonify({'error': 'child_ids must be integers'}), 400
    present = bool(payload.get('present', True))

    day = _weekday_name(on_date)
    if day not in daily_routing.WEEKDAYS:
        return jsonify({'error': 'The program does not run that day'}), 400

    db = get_db_transaction()
    try:
        if (kind, ref, time_block) not in _my_slots(db, me, day, on_date):
            # Not 404: the block exists, this counselor is simply not in it.
            # Saying so is what stops one counselor from quietly writing
            # headcounts into somebody else's room.
            db.rollback()
            return jsonify({'error': 'That block is not yours today'}), 403

        column = ('class_session_id' if kind == 'class'
                  else 'school_id' if kind == 'school' else 'room_id')
        if present:
            # One statement over the whole list, not a loop: DbConnection has no
            # executemany, and "route all 5" has to be atomic anyway — a group
            # that half-confirmed would show a headcount nobody produced.
            #
            # ON CONFLICT DO NOTHING with no target, so it rests on whichever of
            # the two partial uniques init_db() built for this kind of block. A
            # second tap on an already-confirmed child is a no-op rather than an
            # error: two counselors in one care room will both tap the same one.
            db.execute(
                f"""
                INSERT INTO block_checks
                    (child_id, check_date, {column}, time_block, marked_by)
                SELECT c, %s, %s, %s, %s FROM unnest(%s::int[]) AS c
                ON CONFLICT DO NOTHING
                """,
                (on_date, ref, time_block, me, child_ids)
            )
        else:
            db.execute(
                f"""
                DELETE FROM block_checks
                 WHERE check_date = %s AND {column} = %s
                   AND time_block IS NOT DISTINCT FROM %s
                   AND child_id = ANY(%s)
                """,
                (on_date, ref, time_block, child_ids)
            )
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    return jsonify({'ok': True, 'count': len(child_ids)})


@app.route('/api/admin/absences', methods=['GET'])
@jwt_required()
def admin_get_absences():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    date_str = request.args.get('date', today_for_org())
    try:
        parse_date(date_str)
    except ValueError:
        return jsonify({'error': 'Invalid date'}), 400
    db = get_db()
    absent_ids = absent_child_ids_for_date(db, date_str)
    if not absent_ids:
        db.close()
        return jsonify([])
    absences = db.execute("""
        SELECT c.id as child_id, c.name as child_name, u.name as parent_name,
               s.name as school, %s::text as absence_date
        FROM children c
        JOIN users u ON c.parent_id = u.id
        JOIN schools s ON c.school_id = s.id
        WHERE c.id = ANY(%s)
        ORDER BY s.name, c.name
    """, (date_str, list(absent_ids))).fetchall()
    db.close()
    return jsonify([dict(a) for a in absences])


@app.route('/api/admin/absences', methods=['POST'])
@jwt_required()
def admin_mark_absence():
    """Record an absence on a parent's behalf — a phone call, not a report.

    Same effect as /api/parent/absences POST, and deliberately no signature or
    confirmation step: unlike a release, nobody has to be handed anything
    back, so there is no accountability gap for a write this reversible to
    close. Scoped to the org by RLS on `children`, not by parent_id, since the
    caller is staff rather than the family.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.json or {}
    child_id = data.get('child_id')
    absence_date = data.get('date')
    user_id = get_jwt_identity()

    if not child_id or not absence_date:
        return jsonify({'error': 'Missing child_id or date'}), 400
    try:
        absence_date = parse_date(absence_date)
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    db = get_db()
    child = db.execute("SELECT id FROM children WHERE id = %s", (child_id,)).fetchone()
    if not child:
        db.close()
        return jsonify({'error': 'Child not found'}), 404

    try:
        # Same override as the parent's own POST: a one-off absence beats any
        # prior "attending this day" exception for the same date.
        db.execute(
            "DELETE FROM absence_exceptions WHERE child_id = %s AND exception_date = %s",
            (child_id, absence_date)
        )
        db.execute(
            "INSERT INTO absences (child_id, absence_date, marked_by) VALUES (%s, %s, %s) ON CONFLICT DO NOTHING",
            (child_id, absence_date, user_id)
        )
        db.commit()
    except Exception as e:
        db.close()
        print(f"[ERROR] admin_mark_absence: {e}")
        return jsonify({'error': 'Failed to mark absence'}), 400

    db.close()
    return jsonify({'message': 'Absence marked'})


@app.route('/api/admin/absences', methods=['DELETE'])
@jwt_required()
def admin_remove_absence():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.json or {}
    child_id = data.get('child_id')
    absence_date = data.get('date')
    user_id = get_jwt_identity()

    if not child_id or not absence_date:
        return jsonify({'error': 'Missing child_id or date'}), 400
    try:
        absence_date = parse_date(absence_date)
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    db = get_db()
    child = db.execute("SELECT id FROM children WHERE id = %s", (child_id,)).fetchone()
    if not child:
        db.close()
        return jsonify({'error': 'Child not found'}), 404

    # Same as the parent's own DELETE: "remove" means "the child WILL attend
    # this date", so a recurring rule for this weekday needs an exception too,
    # not just deleting the one-off row.
    db.execute("DELETE FROM absences WHERE child_id = %s AND absence_date = %s", (child_id, absence_date))
    db.execute(
        "INSERT INTO absence_exceptions (child_id, exception_date, marked_by) "
        "VALUES (%s, %s, %s) ON CONFLICT DO NOTHING",
        (child_id, absence_date, user_id)
    )
    db.commit()
    db.close()
    return jsonify({'message': 'Absence removed'})

# ─── LIVE HEADCOUNT ─────────────────────────────────────────────────────────
# How many children are in the building right now, per school. Published on
# every attendance change and every pickup release, over the same per-organization
# LISTEN/NOTIFY bus the pickup queue uses — a second bus would have the same
# fan-out problem across workers that the first one was built to solve.

def _headcount_snapshot(db, date_str):
    """Present / expected / released / absent / unscheduled per school, one day.

    `expected` is the denominator on the admin's tile: children enrolled on
    that weekday who are not accounted for as absent. It is the same set the
    counselor roster shows, so the two can't drift apart.

    Three things this deliberately does NOT do, each of which it used to:

      * It does not drop a child who is marked absent but standing in the
        building. Being reported absent is a prediction a parent made in the
        morning; a counselor checking the child in is an observation. The
        observation wins, in both the numerator and the denominator, or the
        tile claims a child who is here is not.
      * It does not hide a school. The join to registrations is a LEFT JOIN,
        so a school whose children are all off today comes back with
        expected 0 instead of vanishing — "nobody scheduled" and "school not
        returned" looked identical before, and only one of them is true.
      * It does not silently drop the children who are not in today. They are
        reported as `unscheduled` — enrolled on some other weekday, or on no
        weekday at all, which is what a roster row whose class signups have
        not happened yet looks like from here.

    The three buckets partition every active child, so
    expected + absent + unscheduled summed over the schools is the same
    number as the "Children" tile. That is deliberate: the gap between that
    tile and the headcount denominator is the question the dashboard kept
    provoking, and now the board itself answers it.
    """
    try:
        day_name = datetime.strptime(date_str, '%Y-%m-%d').strftime('%A')
    except ValueError:
        return []
    absent_ids = list(absent_child_ids_for_date(db, date_str)) or None
    rows = db.execute("""
        WITH day AS (
            SELECT c.id, c.school_id,
                   r.id IS NOT NULL                        AS scheduled,
                   COALESCE(ar.on_bus, 0) = 1              AS marked_in,
                   ar.checked_out_at IS NOT NULL           AS released,
                   (%s::int[] IS NOT NULL
                    AND c.id = ANY(%s::int[]))             AS reported_absent
              FROM children c
              LEFT JOIN registrations r
                     ON r.child_id = c.id AND r.day_of_week = %s
              LEFT JOIN attendance_records ar
                     ON ar.child_id = c.id AND ar.attendance_date = %s
             WHERE c.active = 1
        )
        SELECT s.id AS school_id, s.name AS school,
               COUNT(*) FILTER (
                   WHERE d.scheduled AND (NOT d.reported_absent OR d.marked_in)
               ) AS expected,
               COUNT(*) FILTER (WHERE d.marked_in AND NOT d.released) AS present,
               COUNT(*) FILTER (WHERE d.released)                     AS released,
               COUNT(*) FILTER (
                   WHERE d.scheduled AND d.reported_absent AND NOT d.marked_in
               ) AS absent,
               COUNT(*) FILTER (WHERE NOT d.scheduled)                AS unscheduled
          FROM day d
          JOIN schools s ON s.id = d.school_id
         GROUP BY s.id, s.name
         ORDER BY s.name
    """, (absent_ids, absent_ids, day_name, date_str)).fetchall()
    return [dict(r) for r in rows]


def _publish_headcount(date_str):
    """Fan the current headcount out to the organization's admins."""
    try:
        db = get_db()
        try:
            schools = _headcount_snapshot(db, date_str)
        finally:
            db.close()
        pickup_events.publish(
            'headcount', {'date': date_str, 'schools': schools}, current_org_id()
        )
    except Exception as e:
        # A headcount that fails to publish must never fail the attendance
        # write that triggered it — the tile can lag, the record cannot.
        print(f"[WARN] _publish_headcount: {e}")


# ─── THE AFTERNOON, PER SCHOOL ──────────────────────────────────────────────
# The headcount tile answers "how many are here". Running the afternoon needs
# the other half: who is still out there, who is covering that school, and —
# the reason this exists — which children a counselor could not find.
#
# `not_found` and `issue` were addable from the counselor's phone and visible
# nowhere else. A counselor could report that a child was not at the gate and
# the director would learn about it from a phone call.


def _operations_snapshot(db, date_str):
    """Per school: who is covering it, and where every expected child stands.

    The four states partition `expected` exactly —
    in_building + checked_out + problems + waiting = expected — so the board
    can be read as "N still to collect" without the number being a guess.
    `waiting` is the remainder rather than its own count, which is what makes
    that hold even for a status this function has not heard of yet.

    Children outside that set are still reported, because leaving them out is
    how a screen ends up quietly disagreeing with the roster: `absent` for the
    ones a parent called in, `unscheduled` for the ones who do not come today.
    """
    try:
        day_name = datetime.strptime(date_str, '%Y-%m-%d').strftime('%A')
    except ValueError:
        return []
    absent_ids = list(absent_child_ids_for_date(db, date_str)) or None

    rows = db.execute("""
        WITH day AS (
            SELECT c.id, c.school_id,
                   r.id IS NOT NULL                  AS scheduled,
                   COALESCE(ar.on_bus, 0) = 1        AS marked_in,
                   ar.checked_out_at IS NOT NULL     AS released,
                   ar.status                         AS status,
                   (%s::int[] IS NOT NULL
                    AND c.id = ANY(%s::int[]))       AS reported_absent
              FROM children c
              LEFT JOIN registrations r
                     ON r.child_id = c.id AND r.day_of_week = %s
              LEFT JOIN attendance_records ar
                     ON ar.child_id = c.id AND ar.attendance_date = %s
             WHERE c.active = 1
        ), scoped AS (
            SELECT *,
                   (scheduled AND (NOT reported_absent OR marked_in)) AS expected
              FROM day
        )
        SELECT s.id AS school_id, s.name AS school,
               COUNT(*) FILTER (WHERE d.expected)                        AS expected,
               COUNT(*) FILTER (WHERE d.expected AND d.marked_in
                                  AND NOT d.released)                    AS in_building,
               COUNT(*) FILTER (WHERE d.expected AND d.released)          AS checked_out,
               COUNT(*) FILTER (WHERE d.expected AND NOT d.marked_in
                                  AND NOT d.released
                                  AND d.status IN ('not_found', 'issue')) AS problems,
               COUNT(*) FILTER (WHERE d.scheduled AND d.reported_absent
                                  AND NOT d.marked_in)                    AS absent,
               COUNT(*) FILTER (WHERE NOT d.scheduled)                     AS unscheduled
          FROM scoped d
          JOIN schools s ON s.id = d.school_id
         GROUP BY s.id, s.name
         ORDER BY s.name
    """, (absent_ids, absent_ids, day_name, date_str)).fetchall()

    # Who is covering each school today. Same window rule as everywhere else:
    # a cover that has expired does not put someone's name on a board.
    covering = db.execute(f"""
        SELECT cs.school_id, u.id, u.name,
               cs.effective_from IS NULL AND cs.effective_to IS NULL AS permanent
          FROM counselor_schools cs
          JOIN users u ON u.id = cs.counselor_id
         WHERE u.role = 'counselor' AND {_ASSIGNMENT_IN_EFFECT}
         ORDER BY u.name
    """, (date_str, date_str)).fetchall()
    by_school = {}
    for c in covering:
        by_school.setdefault(c['school_id'], []).append(
            {'id': c['id'], 'name': c['name'], 'permanent': c['permanent']})

    # The children someone has to do something about, by name. Small by
    # definition — an afternoon with twenty of these has bigger problems than
    # a payload size.
    flagged = db.execute("""
        SELECT ar.child_id, c.name, c.school_id, ar.status, ar.status_note,
               ar.status_at, u.name AS reported_by
          FROM attendance_records ar
          JOIN children c ON c.id = ar.child_id
          LEFT JOIN users u ON u.id = ar.status_by
         WHERE ar.attendance_date = %s
           AND ar.status IN ('not_found', 'issue')
           AND ar.checked_out_at IS NULL
           AND COALESCE(ar.on_bus, 0) = 0
         ORDER BY c.name
    """, (date_str,)).fetchall()
    flags_by_school = {}
    for f in flagged:
        flags_by_school.setdefault(f['school_id'], []).append({
            'child_id': f['child_id'],
            'name': f['name'],
            'status': f['status'],
            'note': f['status_note'],
            'at': iso_utc(f['status_at']),
            'reported_by': f['reported_by'],
        })

    result = []
    for r in rows:
        row = dict(r)
        counselors = by_school.get(row['school_id'], [])
        row['counselors'] = counselors
        # An empty list is not "nobody covers this school" — a counselor with
        # no assignment at all covers every school (see counselor_school_ids),
        # so the board must not print "unstaffed" over a JCC that simply never
        # divided its people up.
        row['waiting'] = (row['expected'] - row['in_building']
                          - row['checked_out'] - row['problems'])
        row['flagged'] = flags_by_school.get(row['school_id'], [])
        result.append(row)
    return result


@app.route('/api/admin/operations', methods=['GET'])
@jwt_required()
def admin_operations():
    """The afternoon board: every school, who covers it, and who is still out."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    date_str = request.args.get('date', today_for_org())
    try:
        parse_date(date_str)
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400
    db = get_db()
    try:
        schools = _operations_snapshot(db, date_str)
        # True when no counselor anywhere has an assignment, which is the
        # fail-open state rather than an unstaffed program.
        unassigned_org = not db.execute(
            "SELECT 1 FROM counselor_schools LIMIT 1").fetchone()
    finally:
        db.close()
    return jsonify({
        'date': date_str,
        'every_counselor_covers_every_school': bool(unassigned_org),
        'schools': schools,
    })


@app.route('/api/attendance/headcount', methods=['GET'])
@jwt_required()
def attendance_headcount():
    claims = get_jwt()
    if claims.get('role') not in ('admin', 'superadmin'):
        return jsonify({'error': 'Unauthorized'}), 403

    # A headcount is a question about one JCC, so it is answered inside one
    # JCC. A superadmin's token carries is_superadmin, which satisfies
    # org_isolation for every row — without this the platform owner would get
    # every organization's children summed into one board, grouped by school
    # names that collide across JCCs. The SSE route next door already pins
    # this; the polled one did not, and the two must not disagree.
    from flask import g
    g.is_superadmin = False

    date_str = request.args.get('date', today_for_org())
    db = get_db()
    try:
        return jsonify({'date': date_str, 'schools': _headcount_snapshot(db, date_str)})
    finally:
        db.close()


@app.route('/api/attendance/stream', methods=['GET'])
def attendance_stream():
    """Server-Sent Events feed for the live headcount.

    Same shape as pickups_stream: EventSource can't set an Authorization
    header, so the JWT arrives as a query parameter and is validated here.
    """
    token = request.args.get('token', '')
    try:
        decoded = decode_token(token)
    except Exception:
        return jsonify({'error': 'Invalid token'}), 401

    role = decoded.get('role') or (decoded.get('user_claims') or {}).get('role')
    if role not in ('admin', 'superadmin'):
        return jsonify({'error': 'Unauthorized'}), 403

    from flask import g
    org_id = decoded.get('org') or (decoded.get('user_claims') or {}).get('org')
    if org_id is None:
        return jsonify({'error': 'Token carries no organization'}), 403
    g.organization_id = org_id
    g.is_superadmin = False

    # The token came in the query string, so _enforce_module_access saw no
    # organization and let this through. Apply the same rule now.
    if not tenancy.module_enabled(_request_modules(), 'check_in_out'):
        return jsonify({'error': 'This feature is not enabled for your organization'}), 403

    date_str = request.args.get('date', today_for_org())
    q = pickup_events.subscribe(org_id)
    db = get_db()
    try:
        snapshot = _headcount_snapshot(db, date_str)
    finally:
        db.close()

    def generate():
        try:
            yield pickup_events.format_sse(
                {'type': 'headcount', 'data': {'date': date_str, 'schools': snapshot}}
            )
            while True:
                try:
                    # 15s matches pickups_stream: under the ~30s idle-close
                    # window common on hosted proxies.
                    event = q.get(timeout=15)
                except Exception:
                    yield pickup_events.format_sse({'type': 'heartbeat', 'data': {}})
                    continue
                # One bus per organization carries every event type. Events
                # this stream doesn't speak still get a heartbeat rather than
                # being dropped silently: during carpool the bus is busy with
                # pickup events, and a connection with no bytes on the wire is
                # what hosted proxies idle-close.
                if event.get('type') == 'headcount':
                    yield pickup_events.format_sse(event)
                else:
                    yield pickup_events.format_sse({'type': 'heartbeat', 'data': {}})
        finally:
            pickup_events.unsubscribe(q, org_id)

    return Response(generate(), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no',
        'Connection': 'keep-alive',
    })


# ─── COUNSELOR ATTENDANCE SUBMIT ────────────────────────────────────────────

@app.route('/api/counselor/attendance', methods=['POST'])
@jwt_required()
def counselor_submit_attendance():
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    data = request.json
    attendance_date = data.get('date')
    records = data.get('records', [])  # [{child_id, on_bus}]

    if not attendance_date:
        return jsonify({'error': 'Date required'}), 400
    try:
        attendance_date = parse_date(attendance_date)
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    try:
        db = get_db()
        # Scoped to the date being marked, not to today: a counselor filling
        # in yesterday's attendance is judged by who covered that school
        # yesterday. None means no assignment at all, so every school is
        # theirs; an empty set means their assignments do not reach this date.
        assigned = counselor_school_ids(db, user_id, attendance_date)
        assigned_school_ids = None if assigned is None else set(assigned)
        touched = []
        for rec in records:
            child_id = rec.get('child_id')
            on_bus = 1 if rec.get('on_bus') else 0
            child = db.execute("SELECT school_id FROM children WHERE id = %s", (child_id,)).fetchone()
            if not child:
                continue
            if assigned_school_ids is not None and child['school_id'] not in assigned_school_ids:
                print(f"[WARN] counselor {user_id} attempted attendance for child {child_id} in unassigned school {child['school_id']} — skipped")
                continue
            # checked_in_at is stamped on the first tap of the day and left
            # alone afterwards — re-tapping a child who is already marked must
            # not move their arrival time. Clearing the mark clears it, so an
            # unmark-then-remark reads as a fresh arrival rather than keeping a
            # stale one.
            #
            # Clearing the mark does NOT clear checked_out_at any more. It
            # used to, and that quietly detached a signed pickup_releases row
            # — signature, authorized person, timestamp — from an attendance
            # row that then read "never here". A child who has been released
            # cannot be un-released by untapping a checkbox; taking that back
            # is a decision with a paper trail, not a side effect.
            #
            # status rides along, derived from the tap. A JCC whose counselors
            # only ever mark present/absent gets a status for free and never
            # sees the word.
            status = STATUS_PICKED_UP if on_bus else STATUS_WAITING
            db.execute("""
                INSERT INTO attendance_records
                    (child_id, school_id, attendance_date, on_bus, submitted_by,
                     checked_in_at, status, status_at, status_by)
                VALUES (%s, %s, %s, %s, %s,
                        CASE WHEN %s = 1 THEN CURRENT_TIMESTAMP END,
                        %s, CURRENT_TIMESTAMP, %s)
                ON CONFLICT(child_id, attendance_date) DO UPDATE SET
                    on_bus = excluded.on_bus,
                    submitted_by = excluded.submitted_by,
                    submitted_at = CURRENT_TIMESTAMP,
                    checked_in_at = CASE
                        WHEN excluded.on_bus = 0 THEN NULL
                        ELSE COALESCE(attendance_records.checked_in_at, CURRENT_TIMESTAMP)
                    END,
                    -- A released child stays released. Untapping them leaves
                    -- checked_out_at and the checked_out status alone.
                    status = CASE
                        WHEN attendance_records.checked_out_at IS NOT NULL
                            THEN attendance_records.status
                        ELSE excluded.status
                    END,
                    status_at = CASE
                        WHEN attendance_records.checked_out_at IS NOT NULL
                            THEN attendance_records.status_at
                        ELSE CURRENT_TIMESTAMP
                    END,
                    status_by = CASE
                        WHEN attendance_records.checked_out_at IS NOT NULL
                            THEN attendance_records.status_by
                        ELSE excluded.status_by
                    END
            """, (child_id, child['school_id'], attendance_date, on_bus, user_id,
                  on_bus, status, user_id))
            touched.append(child_id)
        db.commit()
        db.close()
        # The admin headcount tile listens for this. Published after the commit
        # so a subscriber that reacts by re-reading sees the new state.
        if touched:
            _publish_headcount(attendance_date)
        return jsonify({'message': 'Attendance submitted'})
    except Exception as e:
        print(f"[ERROR] counselor_submit_attendance: {e}")
        return jsonify({'error': 'Failed to submit attendance'}), 500

@app.route('/api/counselor/attendance', methods=['GET'])
@jwt_required()
def counselor_get_attendance():
    claims = get_jwt()
    # Also admin — see counselor_get_roster just above.
    if claims.get('role') not in ('counselor', 'admin'):
        return jsonify({'error': 'Unauthorized'}), 403

    date_str = request.args.get('date', today_for_org())
    db = get_db()
    records = db.execute("""
        SELECT child_id, on_bus, status, status_note, checked_in_at, checked_out_at
          FROM attendance_records WHERE attendance_date = %s
    """, (date_str,)).fetchall()
    db.close()
    # This used to return {child_id: on_bus} and nothing else, which is why a
    # child released an hour ago still showed as checked in on the any-date
    # roster: that screen had no way to know they had gone. The tick state
    # stays `on_bus` so the checkbox behaves the same; everything else is
    # additive.
    #
    # The clock times stay behind check_in_out, matching the roster endpoint —
    # they are what that module sells. `checked_out_at` is not one of them:
    # whether a child has already left is a correctness fact the screen needs
    # in order not to lie, and it is only ever non-null at an organization
    # that bought secure_pickup anyway.
    times_visible = module_on('check_in_out')
    return jsonify({
        str(r['child_id']): {
            'on_bus': bool(r['on_bus']),
            'status': derive_status(r['status'], r['on_bus'], r['checked_out_at']),
            'note': r['status_note'],
            'checked_in_at': iso_utc(r['checked_in_at']) if times_visible else None,
            'checked_out_at': iso_utc(r['checked_out_at']),
        } for r in records
    })

def _child_status_payload(data):
    """Validate a child-status write shared by the counselor and admin
    endpoints. Returns (child_id, status, note, date_str, error_response).
    """
    child_id = data.get('child_id')
    status = data.get('status')
    note = (data.get('note') or '').strip() or None

    if not child_id:
        return None, None, None, None, (jsonify({'error': 'child_id required'}), 400)
    settable = [s for s in CHILD_STATUSES if s != STATUS_CHECKED_OUT]
    if status not in settable:
        return None, None, None, None, (jsonify({
            'error': f"status must be one of: {', '.join(settable)}"
        }), 400)
    if status == STATUS_ISSUE and not note:
        # An "issue" with no description is a flag nobody can act on.
        return None, None, None, None, (jsonify({
            'error': 'An issue needs a note saying what it is'
        }), 400)

    date_str = data.get('date') or today_for_org()
    try:
        date_str = parse_date(date_str)
    except ValueError:
        return None, None, None, None, (jsonify({
            'error': 'Invalid date format. Use YYYY-MM-DD.'
        }), 400)

    return child_id, status, note, date_str, None


def _write_child_status(db, child_id, school_id, date_str, status, note, actor_id):
    """Record where one child is, on one date. Shared INSERT for the
    counselor's and the admin's own child-status endpoints — one write path
    so the two can never come to record this differently.
    """
    on_bus = _ON_BUS_FOR_STATUS[status]
    db.execute("""
        INSERT INTO attendance_records
            (child_id, school_id, attendance_date, on_bus, submitted_by,
             checked_in_at, status, status_at, status_by, status_note)
        VALUES (%s, %s, %s, %s, %s,
                CASE WHEN %s = 1 THEN CURRENT_TIMESTAMP END,
                %s, CURRENT_TIMESTAMP, %s, %s)
        ON CONFLICT(child_id, attendance_date) DO UPDATE SET
            on_bus = excluded.on_bus,
            submitted_by = excluded.submitted_by,
            submitted_at = CURRENT_TIMESTAMP,
            checked_in_at = CASE
                WHEN excluded.on_bus = 0 THEN NULL
                ELSE COALESCE(attendance_records.checked_in_at, CURRENT_TIMESTAMP)
            END,
            status = excluded.status,
            status_at = CURRENT_TIMESTAMP,
            status_by = excluded.status_by,
            status_note = excluded.status_note
    """, (child_id, school_id, date_str, on_bus, actor_id,
          on_bus, status, actor_id, note))


@app.route('/api/counselor/child-status', methods=['POST'])
@jwt_required()
def counselor_set_child_status():
    """Mark where one child is.

    The richer half of the attendance tap: the tap can only say present or
    not, and the afternoon has states that matter and are not either —
    `not_found` when a child is not at the school gate, `issue` when something
    needs the director. Both are reported rather than guessed at, which is the
    point of having them.

    `checked_out` is deliberately NOT settable here. Releasing a child to a
    parent takes a signature on the counselor's own device (spec R6) and goes
    through /api/counselor/pickup/release, which records who collected them.
    Letting this endpoint set it would be a way to sign a child out without
    signing for them.
    """
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    data = request.json or {}
    child_id, status, note, date_str, error = _child_status_payload(data)
    if error:
        return error

    db = get_db()
    try:
        if not counselor_covers_child(db, user_id, child_id, date_str):
            return jsonify({'error': 'Child not in your schools'}), 403
        child = db.execute("SELECT school_id FROM children WHERE id = %s",
                           (child_id,)).fetchone()
        if not child:
            return jsonify({'error': 'Child not found'}), 404

        released = db.execute("""
            SELECT checked_out_at FROM attendance_records
             WHERE child_id = %s AND attendance_date = %s
        """, (child_id, date_str)).fetchone()
        if released and released['checked_out_at']:
            # Same rule as the attendance tap: a signed release is not undone
            # by a status button.
            return jsonify({
                'error': 'This child has already been released to a parent.'
            }), 409

        _write_child_status(db, child_id, child['school_id'], date_str,
                            status, note, user_id)
        db.commit()
    except Exception as e:
        print(f"[ERROR] counselor_set_child_status: {e}")
        return jsonify({'error': 'Could not save that status'}), 500
    finally:
        db.close()

    _publish_headcount(date_str)
    return jsonify({'child_id': child_id, 'status': status, 'date': date_str})


@app.route('/api/admin/child-status', methods=['POST'])
@jwt_required()
def admin_set_child_status():
    """Mark where one child is — the admin's own version of the counselor
    tap, for the child an admin resolves from the office rather than the
    school gate: a parent calls to say their kid showed up after all, or a
    `not_found`/`issue` flag on the Live Board gets sorted out.

    No `counselor_covers_child` check: an admin is not scoped to a set of
    schools the way a counselor is, and RLS already keeps this inside their
    own organization. Otherwise the same rules as the counselor endpoint —
    same settable statuses, same note requirement for `issue`, same refusal
    once a child has been released to a parent.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = _current_user_id()
    data = request.json or {}
    child_id, status, note, date_str, error = _child_status_payload(data)
    if error:
        return error

    db = get_db()
    try:
        child = db.execute("SELECT school_id FROM children WHERE id = %s",
                           (child_id,)).fetchone()
        if not child:
            return jsonify({'error': 'Child not found'}), 404

        released = db.execute("""
            SELECT checked_out_at FROM attendance_records
             WHERE child_id = %s AND attendance_date = %s
        """, (child_id, date_str)).fetchone()
        if released and released['checked_out_at']:
            return jsonify({
                'error': 'This child has already been released to a parent.'
            }), 409

        _write_child_status(db, child_id, child['school_id'], date_str,
                            status, note, user_id)
        db.commit()
    except Exception as e:
        print(f"[ERROR] admin_set_child_status: {e}")
        return jsonify({'error': 'Could not save that status'}), 500
    finally:
        db.close()

    _publish_headcount(date_str)
    return jsonify({'child_id': child_id, 'status': status, 'date': date_str})


# ─── COUNSELOR NOTIFICATIONS ─────────────────────────────────────────────────

@app.route('/api/counselor/notify-parent', methods=['POST'])
@jwt_required()
def counselor_notify_parent():
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    data = request.json
    child_id = data.get('child_id')
    notification_date = data.get('date', today_for_org())
    try:
        notification_date = parse_date(notification_date)
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    db = get_db()
    child = db.execute("""
        SELECT c.name as child_name, c.parent_id
        FROM children c
        WHERE c.id = %s
    """, (child_id,)).fetchone()

    if not child:
        db.close()
        return jsonify({'error': 'Child not found'}), 404

    message = f"Is {child['child_name']} coming to Kikar Afterschool today ({notification_date})? Please confirm in your parent portal."

    # Check if notification already sent today
    existing = db.execute(
        "SELECT id FROM parent_notifications WHERE child_id = %s AND notification_date = %s",
        (child_id, notification_date)
    ).fetchone()
    if existing:
        notif_row = db.execute(
            "UPDATE parent_notifications SET message=%s, response=NULL, responded_at=NULL, sent_by=%s WHERE id=%s RETURNING id",
            (message, user_id, existing['id'])
        ).fetchone()
    else:
        notif_row = db.execute(
            "INSERT INTO parent_notifications (child_id, parent_id, sent_by, notification_date, message) VALUES (%s,%s,%s,%s,%s) RETURNING id",
            (child_id, child['parent_id'], user_id, notification_date, message)
        ).fetchone()
    db.commit()

    # Both linked accounts hear about this, not just whichever one the
    # parent_notifications row itself is stamped with — parent_get_notifications
    # and parent_respond_notification already read/answer it by child, not by
    # that stamp, so either parent sees the card and either may answer it.
    owners = db.execute(
        "SELECT id, name, email FROM users WHERE id = ANY(%s)",
        (child_owner_ids(db, [child_id]),)
    ).fetchall()

    identity = email_identity(db)
    date_display = datetime.strptime(notification_date, '%Y-%m-%d').strftime('%A, %B %d')
    for owner in owners:
        # Fan the event out to any parent portals that are open RIGHT NOW so
        # the "Is your child coming?" card appears without needing a refresh.
        # The stream filters by parent_id server-side.
        pickup_events.publish('parent-notification', {
            'id': notif_row['id'],
            'parent_id': owner['id'],
            'child_id': child_id,
            'child_name': child['child_name'],
            'notification_date': str(notification_date),
            'message': message,
        }, current_org_id())

        if owner['email']:
            html = _email_shell(
                identity, f"{identity.org_name} – Attendance Confirmation", f"""
                <p>Hello {_escape(owner['name'])},</p>
                <p>Your child <strong>{_escape(child['child_name'])}</strong> has not been confirmed for {_escape(identity.org_name)} on <strong>{date_display}</strong>.</p>
                <p>Please log in to your parent portal to confirm if they are attending or mark an absence.</p>
                {_cta(identity, f"{EMAIL_CONFIG['base_url']}/app/home", 'Open Parent Portal')}
            """)
            # Off the request thread — see send_email_async / send_push_to_user_async
            # comments. The 'parent-notification' SSE event already fired above, so
            # parents with the portal open see the card without waiting on SMTP/FCM.
            send_email_async(
                owner['email'],
                f"{identity.org_name} – Please confirm {child['child_name']}'s attendance",
                html,
                identity,
            )
        notify_parent(owner['id'], 'attendance_check',
                      f"📋 {identity.org_name} – Attendance Check",
                      f"Is {child['child_name']} coming today? Tap to confirm.", '/app/home')
    db.close()
    return jsonify({'message': 'Notification sent'})


@app.route('/api/counselor/notification-responses', methods=['GET'])
@jwt_required()
def counselor_get_notification_responses():
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403

    date_str = request.args.get('date', today_for_org())
    user_id = get_jwt_identity()

    db = get_db()
    # Get all notifications sent by this counselor (or any counselor) for today
    notifs = db.execute("""
        SELECT n.child_id, n.response, n.responded_at
        FROM parent_notifications n
        WHERE n.notification_date = %s AND n.response IS NOT NULL
    """, (date_str,)).fetchall()
    db.close()

    # Return as {child_id: response} map
    result = {}
    for n in notifs:
        result[str(n['child_id'])] = n['response']
    return jsonify(result)

# ─── PARENT NOTIFICATIONS ─────────────────────────────────────────────────────

@app.route('/api/parent/notifications', methods=['GET'])
@jwt_required()
def parent_get_notifications():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    db = get_db()
    # Any child this account reaches — as Contact #1 or a linked Contact #2 —
    # not just the one parent_id the row was created under. Either parent can
    # answer; see parent_respond_notification below for the same broadening.
    notifs = db.execute(f"""
        SELECT n.id, n.child_id, c.name as child_name, n.notification_date, n.message, n.response, n.responded_at
        FROM parent_notifications n
        JOIN children c ON n.child_id = c.id
        WHERE {_CHILD_BELONGS_TO_USER} AND n.response IS NULL
        ORDER BY n.created_at DESC
    """, (user_id, user_id)).fetchall()
    db.close()
    return jsonify([dict(n) for n in notifs])

@app.route('/api/parent/notifications/<int:notif_id>/respond', methods=['POST'])
@jwt_required()
def parent_respond_notification(notif_id):
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    response = request.json.get('response')  # 'attending' or 'absent'

    if response not in ('attending', 'absent'):
        return jsonify({'error': 'Invalid response'}), 400

    db = get_db()
    # Whoever answers first settles it for the family — same rule as marking
    # an absence, which either linked parent can already do.
    notif = db.execute(f"""
        SELECT n.* FROM parent_notifications n
          JOIN children c ON c.id = n.child_id
         WHERE n.id = %s AND {_CHILD_BELONGS_TO_USER}
    """, (notif_id, user_id, user_id)).fetchone()
    if not notif:
        db.close()
        return jsonify({'error': 'Notification not found'}), 404

    db.execute(
        "UPDATE parent_notifications SET response=%s, responded_at=CURRENT_TIMESTAMP WHERE id=%s",
        (response, notif_id)
    )
    # If absent, auto-mark absence (and clear any prior "attending this day"
    # exception so the absence sticks even when a recurring rule was overridden).
    if response == 'absent':
        db.execute(
            "DELETE FROM absence_exceptions WHERE child_id = %s AND exception_date = %s",
            (notif['child_id'], notif['notification_date'])
        )
        db.execute(
            "INSERT INTO absences (child_id, absence_date, marked_by) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
            (notif['child_id'], notif['notification_date'], user_id)
        )
    db.commit()
    db.close()

    pickup_events.publish('parent-response', {
        'child_id': notif['child_id'],
        'notification_date': str(notif['notification_date']),
        'response': response,
    }, current_org_id())

    return jsonify({'message': 'Response saved'})

# ─── ADMIN ATTENDANCE RECORDS ─────────────────────────────────────────────────

@app.route('/api/admin/attendance', methods=['GET'])
@jwt_required()
def admin_get_attendance():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    date_str = request.args.get('date', today_for_org())
    school_id = request.args.get('school_id')
    db = get_db()
    query = """
        SELECT ar.attendance_date, c.name as child_name, c.service_type,
               s.name as school, ar.on_bus,
               COALESCE(u.name, '(deleted user)') as submitted_by, ar.submitted_at
        FROM attendance_records ar
        JOIN children c ON ar.child_id = c.id
        JOIN schools s ON ar.school_id = s.id
        LEFT JOIN users u ON ar.submitted_by = u.id
        WHERE ar.attendance_date = %s
    """
    params = [date_str]
    if school_id:
        query += " AND ar.school_id = %s"
        params.append(school_id)
    query += " ORDER BY s.name, c.name"
    records = db.execute(query, params).fetchall()
    db.close()
    return jsonify([dict(r) for r in records])

# ─── PUSH NOTIFICATIONS ─────────────────────────────────────────────────────

def send_push_to_user(user_id, title, body, url='/app/', tag=None, pickup_id=None):
    if not PUSH_AVAILABLE or not VAPID_PRIVATE_KEY:
        print(f"[PUSH SKIP] PUSH_AVAILABLE={PUSH_AVAILABLE}, has_key={bool(VAPID_PRIVATE_KEY)}")
        return
    # Say which of the two zeroes this is.
    #
    # push_subscriptions is behind RLS, so a connection with no organization
    # reads none of them — indistinguishable, in the log, from a user who never
    # enabled notifications. That ambiguity is what hid a bug where every
    # background push delivered to nobody. If there is no tenant in scope here,
    # that is a caller that forgot to pin its thread, and it should look like a
    # fault rather than like a quiet user.
    if _ambient_org_id() is None:
        print(f"[PUSH BUG] No organization in scope sending to user {user_id}: "
              f"push_subscriptions is behind RLS, so this would read zero rows. "
              f"The caller must pin the thread — see send_push_to_user_async.")
        return
    db = get_db()
    subs = db.execute("SELECT * FROM push_subscriptions WHERE user_id=%s", (user_id,)).fetchall()
    print(f"[PUSH] Sending to user {user_id}: {len(subs)} subscription(s) found")
    payload = {"title": title, "body": body, "url": url}
    if tag:
        payload["tag"] = tag
    if pickup_id is not None:
        payload["pickup_id"] = pickup_id
    stale_ids = []
    for sub in subs:
        try:
            webpush(
                subscription_info={"endpoint": sub['endpoint'], "keys": {"p256dh": sub['p256dh'], "auth": sub['auth']}},
                data=json.dumps(payload),
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims={"sub": "mailto:admin@kikarafterschool.com"}
            )
            print(f"[PUSH] Sent OK to subscription {sub['id']}")
        except WebPushException as e:
            print(f"[PUSH] WebPush error for sub {sub['id']}: {e}")
            # Remove expired/invalid subscriptions (410 Gone, 404 Not Found)
            if hasattr(e, 'response') and e.response is not None and e.response.status_code in (404, 410):
                stale_ids.append(sub['id'])
        except Exception as e:
            print(f"[PUSH] Error for sub {sub['id']}: {e}")
    # Clean up stale subscriptions
    for sid in stale_ids:
        db.execute("DELETE FROM push_subscriptions WHERE id=%s", (sid,))
        print(f"[PUSH] Removed stale subscription {sid}")
    if stale_ids:
        db.commit()
    db.close()

def _org_notification_prefs(organization_id):
    """This organization's parent-notification switches, or {} if unreadable.

    Never raises and never blocks a send on its own failure: if the row cannot
    be read, tenancy.notification_on() falls back to the catalogue defaults,
    which is the behaviour every organization had before there was anything to
    switch off.
    """
    if organization_id is None:
        return {}
    try:
        db = get_db()
        try:
            row = db.execute(
                "SELECT prefs FROM notification_settings WHERE organization_id = %s",
                (organization_id,),
            ).fetchone()
        finally:
            db.close()
    except Exception as e:
        print(f"[PUSH] could not read notification prefs for org "
              f"{organization_id}: {e.__class__.__name__}")
        return {}
    return (row and row['prefs']) or {}


def notify_parent(user_id, kind, title, body, url='/app/', tag=None, pickup_id=None):
    """Push to a parent, unless their JCC switched this kind of thing off.

    THE ONE DOOR. Every parent-facing push goes through here rather than
    calling send_push_to_user_async directly, because a preference that half
    the events consult is worse than no preference: the admin unticks
    "child was picked up", the pushes keep arriving, and now the screen is
    lying to them.

    `kind` must be a key in tenancy.PARENT_NOTIFICATIONS. An unknown key sends —
    see notification_on() for why an unrecognised event fails open while an
    unrecognised module fails closed.

    Staff notifications do not come through here. A counselor being told about
    a pickup is the product working, not a courtesy a JCC tunes.

    Email is the second channel, gated by the organization switch above AND
    the parent's own opt-in (users.email_notifications, sql/59) — push needs
    no such opt-in check because a browser subscription already is one.
    """
    if not tenancy.notification_on(_org_notification_prefs(_ambient_org_id()), kind):
        print(f"[PUSH] {kind} is switched off for this organization; not sending")
        return
    send_push_to_user_async(user_id, title, body, url, tag, pickup_id)

    db = get_db()
    try:
        row = db.execute(
            "SELECT email, email_notifications FROM users WHERE id = %s",
            (user_id,)
        ).fetchone()
        if row and row['email_notifications'] and row['email']:
            identity = email_identity(db)
            send_email_async(row['email'], title,
                             _notification_email_html(identity, title, body, url),
                             identity)
    finally:
        db.close()


def send_push_to_user_async(user_id, title, body, url='/app/', tag=None, pickup_id=None):
    """Fire-and-forget variant of send_push_to_user. Keeps HTTP handlers snappy
    so the SSE realtime path isn't gated on pywebpush round-trips to FCM/APNs,
    which can take seconds per subscription and pile up when every staff member
    has multiple devices.

    THE ORGANIZATION IS CAPTURED HERE AND PINNED INSIDE THE THREAD, AND WITHOUT
    THAT NONE OF THIS WORKED.

    `push_subscriptions` is a tenant table with RLS. A bare threading.Thread has
    no flask.g and no thread pin, so database._resolve_organization() returned
    None, get_db() set app.organization_id to '', and the org_isolation policy
    matched nothing: every async push read zero subscriptions and sent nothing.

    It failed in the worst possible way — silently and plausibly. The log line
    it printed was "0 subscription(s) found", which reads as "that user has not
    enabled notifications" rather than "we cannot see their rows". Every push
    on this path was affected: messages from the office, attendance checks,
    pickup-claimed, day-off decisions.
    """
    organization_id = _ambient_org_id()

    def run():
        # Same move _run_all_invites_job makes, for the same reason.
        database_module.set_thread_organization(organization_id)
        send_push_to_user(user_id, title, body, url, tag, pickup_id)

    threading.Thread(
        target=run,
        daemon=True,
    ).start()


def iso_utc(dt):
    """Serialize a datetime from the DB as an ISO 8601 UTC string with 'Z'.

    Our TIMESTAMP columns are naive but the DB session is UTC (Supabase
    default), so NOW()/CURRENT_TIMESTAMP rows come back as naive UTC datetimes.
    Without the trailing 'Z', browsers (Chrome at minimum) parse the string
    as LOCAL time and the pickup card shows a time 4+ hours off for users in
    Miami. Tagging UTC lets new Date(...) convert it correctly to the
    viewer's local zone (Eastern), so counselors see the moment the parent
    actually tapped "I'm here."
    """
    if dt is None:
        return None
    if isinstance(dt, datetime):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)
        return dt.isoformat().replace('+00:00', 'Z')
    return str(dt)


def local_clock(dt, tzname):
    """A naive-UTC timestamp as a local wall clock time, e.g. "3:42 PM".

    The SPA never needs this: it gets UTC from iso_utc and the browser converts
    to the viewer's zone. A spreadsheet has no browser, so a time written into
    a cell has to already be local or it reads four hours off for every JCC on
    the east coast — and an export of when children left is exactly the place
    that error would go unnoticed.
    """
    if dt is None:
        return ''
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    try:
        dt = dt.astimezone(ZoneInfo(tzname or DEFAULT_TIMEZONE))
    except Exception:
        # An unknown zone in the database is not worth failing an export over.
        dt = dt.astimezone(timezone.utc)
    # %-I is glibc-only; lstrip keeps this working wherever it runs.
    return dt.strftime('%I:%M %p').lstrip('0')


def pickup_row_to_dict(row):
    """Convert a pickup_notifications row into a JSON-ready dict with UTC
    timestamps tagged. Use this everywhere we send a pickup over the wire
    so the client never has to guess the timezone."""
    d = dict(row)
    if 'created_at' in d:
        d['created_at'] = iso_utc(d['created_at'])
    if 'claimed_at' in d:
        d['claimed_at'] = iso_utc(d['claimed_at'])
    return d


def send_email_async(to_email, subject, html_body, identity):
    """Fire-and-forget SMTP send. SMTP login + send can take 1-3s per message
    on Gmail; doing it inline blocks the request thread (and with gunicorn's
    bounded thread pool, blocks everyone else's requests too). The critical
    parent-notification SSE event is published BEFORE this is called, so the
    parent portal sees the card instantly either way.

    `identity` must be resolved by the caller, in request context. The thread
    started here has neither flask.g nor an organization pin, so resolving it
    on the other side would silently produce Kikar's branding for a JCC's
    parents — the one failure mode of this design that nothing would report.
    """
    threading.Thread(
        target=send_email,
        args=(to_email, subject, html_body, identity),
        daemon=True,
    ).start()


@app.route('/api/push/vapid-public-key', methods=['GET'])
def get_vapid_public_key():
    return jsonify({'public_key': VAPID_PUBLIC_KEY or ''})

@app.route('/api/push/subscribe', methods=['POST'])
@jwt_required()
def push_subscribe():
    user_id = get_jwt_identity()
    data = request.json
    sub = data.get('subscription', {})
    endpoint = sub.get('endpoint', '')
    keys = sub.get('keys', {})
    p256dh = keys.get('p256dh', '')
    auth = keys.get('auth', '')
    if not endpoint or not p256dh or not auth:
        return jsonify({'error': 'Invalid subscription'}), 400
    db = get_db()
    db.execute(
        "INSERT INTO push_subscriptions (user_id, endpoint, p256dh, auth) VALUES (%s,%s,%s,%s) ON CONFLICT (user_id, endpoint) DO UPDATE SET p256dh = EXCLUDED.p256dh, auth = EXCLUDED.auth",
        (user_id, endpoint, p256dh, auth)
    )
    db.commit()
    total = db.execute("SELECT COUNT(*) as n FROM push_subscriptions WHERE user_id=%s", (user_id,)).fetchone()['n']
    db.close()
    print(f"[PUSH] User {user_id} subscribed. Total subscriptions for user: {total}. Endpoint: {endpoint[:60]}...")
    return jsonify({'message': 'Subscribed', 'subscription_count': total})


@app.route('/api/push/test', methods=['POST'])
@jwt_required()
def push_test():
    """Send a test push notification to yourself."""
    user_id = get_jwt_identity()
    db = get_db()
    count = db.execute("SELECT COUNT(*) as n FROM push_subscriptions WHERE user_id=%s", (user_id,)).fetchone()['n']
    db.close()
    if count == 0:
        return jsonify({'error': 'No push subscriptions found for your account. Enable notifications first.', 'subscriptions': 0}), 400
    send_push_to_user(user_id, 'Kikar Afterschool test', 'Push notifications are working!', '/app/today')
    return jsonify({'message': f'Test notification sent to {count} device(s)', 'subscriptions': count})

@app.route('/api/push/unsubscribe', methods=['POST'])
@jwt_required()
def push_unsubscribe():
    user_id = get_jwt_identity()
    endpoint = request.json.get('endpoint', '')
    db = get_db()
    db.execute("DELETE FROM push_subscriptions WHERE user_id=%s AND endpoint=%s", (user_id, endpoint))
    db.commit()
    db.close()
    return jsonify({'message': 'Unsubscribed'})

# ─── ADMIN → PARENT MESSAGES (INBOX) ────────────────────────────────────────

def _resolve_message_recipients(db, audience_type, school_ids):
    """Return DISTINCT user ids that should receive a message given the
    audience filter. Always restricted to parents who have at least one
    active child in the program — as Contact #1 (parent_id) or a linked
    Contact #2 (child_contacts.user_id), so a broadcast reaches both."""
    if audience_type == 'all':
        rows = db.execute("""
            SELECT DISTINCT parent_id AS user_id FROM children WHERE active = 1
            UNION
            SELECT DISTINCT cc.user_id
              FROM child_contacts cc JOIN children c ON c.id = cc.child_id
             WHERE c.active = 1 AND cc.user_id IS NOT NULL
        """).fetchall()
    else:
        if not school_ids:
            return []
        rows = db.execute("""
            SELECT DISTINCT parent_id AS user_id FROM children
             WHERE active = 1 AND school_id = ANY(%s)
            UNION
            SELECT DISTINCT cc.user_id
              FROM child_contacts cc JOIN children c ON c.id = cc.child_id
             WHERE c.active = 1 AND c.school_id = ANY(%s) AND cc.user_id IS NOT NULL
        """, (list(school_ids), list(school_ids))).fetchall()
    return [r['user_id'] for r in rows]


@app.route('/api/admin/messages/preview-count', methods=['POST'])
@jwt_required()
def admin_messages_preview_count():
    """Returns how many parents would receive a message with the given
    audience filter. Used by the admin compose UI to show a live count
    before sending."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    audience_type = data.get('audience_type', 'schools')
    school_ids = data.get('school_ids') or []
    if audience_type not in ('all', 'schools'):
        return jsonify({'error': 'Invalid audience_type'}), 400
    db = get_db()
    parent_ids = _resolve_message_recipients(db, audience_type, school_ids)
    # Parents an audience can never include, because every audience — even
    # "everyone" — is resolved through `children`. A parent added by hand and
    # not yet given a child can be written to individually and will never
    # receive an announcement, which from the admin's side looks like the
    # broadcast is broken rather than like a roster gap.
    orphaned = db.execute("""
        SELECT count(*) AS n FROM users u
         WHERE u.role = 'parent'
           AND NOT EXISTS (SELECT 1 FROM children c
                            WHERE c.parent_id = u.id AND c.active = 1)
    """).fetchone()
    school_count = len(school_ids) if audience_type == 'schools' else None
    db.close()
    return jsonify({
        'recipient_count': len(parent_ids),
        'school_count': school_count,
        'parents_without_children': orphaned['n'] if orphaned else 0,
    })


@app.route('/api/admin/messages', methods=['POST'])
@jwt_required()
def admin_send_message():
    """Compose and send a broadcast message to parents."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    subject = (data.get('subject') or '').strip()
    body = (data.get('body') or '').strip()
    audience_type = data.get('audience_type', 'schools')
    school_ids = data.get('school_ids') or []

    if not subject or len(subject) > 200:
        return jsonify({'error': 'Subject is required (max 200 chars)'}), 400
    if not body or len(body) > 4000:
        return jsonify({'error': 'Body is required (max 4000 chars)'}), 400
    if audience_type not in ('all', 'schools'):
        return jsonify({'error': 'Invalid audience_type'}), 400
    if audience_type == 'schools' and not school_ids:
        return jsonify({'error': 'Pick at least one school or use audience_type=all'}), 400

    sender_id = get_jwt_identity()

    db = get_db_transaction()
    try:
        msg_id = db.execute(
            "INSERT INTO admin_messages (sender_id, subject, body, audience_type) VALUES (%s,%s,%s,%s) RETURNING id",
            (sender_id, subject, body, audience_type)
        ).fetchone()['id']

        if audience_type == 'schools':
            for sid in school_ids:
                db.execute(
                    "INSERT INTO admin_message_schools (admin_message_id, school_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                    (msg_id, int(sid))
                )

        parent_ids = _resolve_message_recipients(db, audience_type, school_ids)

        # Insert per-recipient copies. Use executemany-style loop; volume is bounded
        # by the size of the parent roster (a few hundred at most for this program),
        # so a tight Python loop is fine and keeps the SQL simple.
        parent_message_rows = []
        for pid in parent_ids:
            row = db.execute(
                "INSERT INTO parent_messages (admin_message_id, parent_id) VALUES (%s,%s) RETURNING id",
                (msg_id, pid)
            ).fetchone()
            parent_message_rows.append((pid, row['id']))

        db.commit()
    except Exception as e:
        db.rollback()
        db.close()
        print(f"[ERROR] admin_send_message: {e}")
        return jsonify({'error': 'Failed to send message'}), 500
    db.close()

    # Fire push notifications outside the transaction.
    push_title = '📬 New message from Kikar Afterschool'
    push_body = f"{subject}" if len(subject) <= 100 else subject[:97] + '...'

    # Pre-count parents who actually have at least one push subscription, so the
    # admin can see in the response how many were reachable by push (vs only
    # in-app inbox). Helps diagnose "I sent a message but parent X got no push"
    # which is almost always because parent X never enabled notifications.
    diag_db = get_db()
    parent_ids_only = [pid for pid, _ in parent_message_rows]
    if parent_ids_only:
        sub_count_row = diag_db.execute(
            "SELECT COUNT(DISTINCT user_id) AS n FROM push_subscriptions WHERE user_id = ANY(%s)",
            (parent_ids_only,)
        ).fetchone()
        push_eligible = sub_count_row['n'] if sub_count_row else 0
    else:
        push_eligible = 0
    diag_db.close()
    print(f"[MSG] Broadcast id={msg_id} subject={subject!r} recipients={len(parent_message_rows)} with_push={push_eligible}")

    # Run all pushes in ONE background thread instead of spawning a thread per
    # recipient. With maxconn=8 in the pool, fanning out 100+ daemon threads
    # at once made some of them queue behind the pool and silently lose work
    # if the gunicorn worker recycled mid-flight. A single sequential worker
    # uses one connection at a time, completes deterministically, and logs
    # progress so we can see exactly what happened in Render's logs.
    if parent_message_rows:
        push_jobs = [
            (pid, pm_id)
            for pid, pm_id in parent_message_rows
        ]
        threading.Thread(
            target=_run_broadcast_pushes,
            # The organization travels with the job. Without it this worker
            # reads push_subscriptions with no tenant pinned, RLS returns
            # nothing, and the broadcast reports "attempted=N failed=0" having
            # delivered to nobody.
            #
            # subject/body travel too, separately from push_title/push_body:
            # email has no 100-character reason to truncate the message, so
            # the worker sends the full thing there while push keeps the
            # short version.
            args=(msg_id, push_jobs, push_title, push_body, subject, body,
                  current_org_id()),
            daemon=True,
        ).start()

    return jsonify({
        'id': msg_id,
        'recipient_count': len(parent_message_rows),
        'push_eligible_count': push_eligible,
    })


def _run_broadcast_pushes(broadcast_id, jobs, title, body, subject, full_body,
                          organization_id):
    """Sequentially deliver push AND email for one admin broadcast.

    Runs in a single daemon thread, calling the synchronous
    send_push_to_user (which itself loops over a recipient's subscriptions
    and handles webpush errors / stale-subscription cleanup). Each recipient
    is wrapped in try/except so one failure doesn't kill the whole batch.

    `organization_id` is not optional. This thread outlives the request, so it
    has no flask.g; without the pin below, every read of push_subscriptions is
    filtered by RLS to nothing and the whole broadcast delivers to nobody while
    reporting success. See send_push_to_user_async.

    Email does not go through notify_parent() — broadcast never has, on
    purpose (see the module docstring above this function's neighbours) — but
    it is the same second channel, gated by the same organization switch
    checked once below plus each parent's own email_notifications opt-in,
    checked here in one query rather than one per recipient.
    """
    database_module.set_thread_organization(organization_id)
    # Checked once for the whole batch rather than per recipient: it is one
    # organization's setting and the answer cannot differ between parents.
    if not tenancy.notification_on(_org_notification_prefs(organization_id), 'broadcast'):
        print(f"[MSG_PUSH] Broadcast {broadcast_id}: announcements are switched "
              f"off for this organization; message saved, no push sent")
        return
    print(f"[MSG_PUSH] Broadcast {broadcast_id}: worker starting for {len(jobs)} recipients")
    sent = 0
    failed = 0
    for pid, pm_id in jobs:
        try:
            send_push_to_user(
                pid, title, body,
                url='/app/inbox',
                tag=f'msg-{pm_id}',
            )
            sent += 1
        except Exception as e:
            failed += 1
            print(f"[MSG_PUSH] Broadcast {broadcast_id}: error pushing to user {pid}: {type(e).__name__}: {e}")
    print(f"[MSG_PUSH] Broadcast {broadcast_id}: worker done — attempted={sent} failed={failed}")

    pids = [pid for pid, _pm_id in jobs]
    db = get_db()
    try:
        opted_in = db.execute(
            "SELECT id, email FROM users WHERE id = ANY(%s) AND email_notifications",
            (pids,)
        ).fetchall()
        if opted_in:
            # One identity for the whole batch — it does not vary by
            # recipient, and email_identity() must run in this pinned thread
            # rather than inside send_email_async's own unpinned one.
            identity = email_identity(db)
            html = _notification_email_html(
                identity, f"📬 {subject}", full_body, '/app/inbox')
            e_sent = 0
            for row in opted_in:
                if row['email']:
                    send_email_async(row['email'], subject, html, identity)
                    e_sent += 1
            print(f"[MSG_EMAIL] Broadcast {broadcast_id}: queued {e_sent} email(s)")
    finally:
        db.close()


def _run_attendance_check(organization_id):
    """Ask every unconfirmed family whether their child is coming today.

    The scheduled twin of /api/counselor/notify-parent, which does one child at
    a time. Returns how many parents were asked.

    WHO IS ASKED, AND WHO IS NOT
      * only children active and registered for today's weekday — asking about
        a child who is not expected is how a parent learns to ignore this;
      * only families who have not already answered today. A parent who marked
        an absence this morning has said everything the question is for.

    Runs on the scheduler's thread, which is already pinned to this
    organization (server/scheduler.py), so every query below is tenant-scoped
    the same way a request would be.
    """
    # today_for_org() hands back a 'YYYY-MM-DD' string and _weekday_name wants
    # a date. Parsed once here rather than called twice: two calls could also
    # straddle midnight and register the question against a different day than
    # the weekday it filtered on.
    check_date = today_for_org()
    day_name = _weekday_name(datetime.strptime(check_date, '%Y-%m-%d').date())

    db = get_db_transaction()
    try:
        pending = db.execute("""
            SELECT c.id AS child_id, c.name AS child_name, c.parent_id
              FROM children c
              JOIN registrations r ON r.child_id = c.id AND r.day_of_week = %s
             WHERE c.active = 1
               AND c.parent_id IS NOT NULL
               AND NOT EXISTS (
                     SELECT 1 FROM parent_notifications n
                      WHERE n.child_id = c.id
                        AND n.notification_date = %s
                        AND n.response IS NOT NULL)
               AND NOT EXISTS (
                     SELECT 1 FROM absences a
                      WHERE a.child_id = c.id AND a.absence_date = %s)
             GROUP BY c.id, c.name, c.parent_id
        """, (day_name, check_date, check_date)).fetchall()

        rows = []
        for child in pending:
            message = (f"Is {child['child_name']} coming today ({check_date})? "
                       f"Please confirm in your parent portal.")
            existing = db.execute(
                "SELECT id FROM parent_notifications "
                " WHERE child_id = %s AND notification_date = %s",
                (child['child_id'], check_date)).fetchone()
            if existing:
                notif = db.execute(
                    "UPDATE parent_notifications SET message = %s "
                    " WHERE id = %s RETURNING id",
                    (message, existing['id'])).fetchone()
            else:
                notif = db.execute(
                    "INSERT INTO parent_notifications "
                    "  (child_id, parent_id, sent_by, notification_date, message) "
                    "VALUES (%s, %s, NULL, %s, %s) RETURNING id",
                    (child['child_id'], child['parent_id'], check_date, message)).fetchone()
            rows.append((notif['id'], child, message))
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    # Published and pushed after the commit, never inside it: a parent whose
    # portal is open should only be shown a card that actually exists.
    #
    # Batched rather than one notify_parent() call per owner: this can be a
    # hundred-plus children, each with up to two owners, and notify_parent()
    # opens its own connection to check the email preference every time it is
    # called. See child_owner_map() for why that shape emptied the pool.
    org_prefs = _org_notification_prefs(organization_id)
    push_on = tenancy.notification_on(org_prefs, 'attendance_check')
    identity = email_identity()
    check_db = get_db()
    try:
        owners_by_child = child_owner_map(check_db, [c['child_id'] for _i, c, _m in rows])
        all_owner_ids = sorted({oid for ids in owners_by_child.values() for oid in ids})
        owner_rows = check_db.execute(
            "SELECT id, email, email_notifications FROM users WHERE id = ANY(%s)",
            (all_owner_ids,)
        ).fetchall() if all_owner_ids else []
        owners = {r['id']: r for r in owner_rows}
    finally:
        check_db.close()

    if not push_on:
        print(f"[PUSH] attendance_check is switched off for org {organization_id}; not sending")
        return len(rows)

    for notif_id, child, message in rows:
        title = f"📋 {identity.org_name} – Attendance Check"
        body = f"Is {child['child_name']} coming today? Tap to confirm."
        # Both linked accounts, not just whichever one parent_id happens to
        # name — see counselor_notify_parent for the same broadening.
        for owner_id in owners_by_child.get(child['child_id'], []):
            pickup_events.publish('parent-notification', {
                'id': notif_id,
                'parent_id': owner_id,
                'child_id': child['child_id'],
                'child_name': child['child_name'],
                'notification_date': str(check_date),
                'message': message,
            }, organization_id)
            send_push_to_user_async(owner_id, title, body, '/app/home')
            owner = owners.get(owner_id)
            if owner and owner['email_notifications'] and owner['email']:
                send_email_async(
                    owner['email'], title,
                    _notification_email_html(identity, title, body, '/app/home'),
                    identity,
                )
    return len(rows)


@app.route('/api/admin/notification-settings', methods=['GET'])
@jwt_required()
def admin_get_notification_settings():
    """What this JCC's parents are notified about, and when the daily ask goes.

    The catalogue travels with the values so the screen renders from the
    server's list rather than a copy in the client that drifts — same shape the
    superadmin module toggles use.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        # LEFT JOIN because a JCC that has never opened this screen has no row,
        # and that is the normal state rather than an error: every key falls
        # back to its catalogue default.
        row = db.execute("""
            SELECT s.prefs, s.attendance_check_at, o.timezone
              FROM organizations o
              LEFT JOIN notification_settings s ON s.organization_id = o.id
             WHERE o.id = %s
        """, (current_org_id(),)).fetchone()
    finally:
        db.close()
    if not row:
        return jsonify({'error': 'Organization not found'}), 404
    prefs = dict(tenancy.default_notifications())
    prefs.update(row['prefs'] or {})
    return jsonify({
        'catalogue': [
            {'key': key, 'label': label, 'description': description,
             'default': default}
            for key, (label, description, default) in tenancy.PARENT_NOTIFICATIONS.items()
        ],
        'notifications': prefs,
        'attendance_check_at': row['attendance_check_at'],
        'timezone': row['timezone'],
    })


@app.route('/api/admin/notification-settings', methods=['PUT'])
@jwt_required()
def admin_set_notification_settings():
    """An organization's own admin owns these. See PARENT_NOTIFICATIONS for why
    that differs from modules, which only a superadmin may change."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    updates = {}

    if 'notifications' in data:
        incoming = data.get('notifications')
        if not isinstance(incoming, dict):
            return jsonify({'error': 'notifications must be an object of key -> boolean'}), 400
        unknown = [k for k in incoming if k not in tenancy.PARENT_NOTIFICATIONS]
        if unknown:
            # Rejected rather than ignored, for the same reason the module
            # toggles are: a typo that silently does nothing is how someone
            # ends up certain they turned something off.
            return jsonify({'error': f"Unknown notifications: {', '.join(sorted(unknown))}"}), 400
        updates['prefs'] = json.dumps({k: bool(v) for k, v in incoming.items()})

    if 'attendance_check_at' in data:
        at = (data.get('attendance_check_at') or '').strip()
        if at and not re.match(r'^([01][0-9]|2[0-3]):[0-5][0-9]$', at):
            return jsonify({'error': 'Time must be HH:MM on a 24-hour clock'}), 400
        # Empty means off, and off is a real setting rather than an unset one:
        # NULL is what the scheduler reads as "this JCC does not do this".
        updates['attendance_check_at'] = at or None

    if not updates:
        return jsonify({'error': 'Nothing to update'}), 400

    # An upsert, not an update: most organizations have no row until the first
    # time someone touches this screen. One statement rather than a read and a
    # branch, so two admins saving at the same moment cannot both decide the
    # row is missing and insert one — notification_settings_org_unique makes
    # the second one conflict into the DO UPDATE instead.
    #
    # Only the keys that were sent are overwritten; a save of just the time
    # leaves the switches alone.
    columns = ', '.join(['organization_id', *updates])
    placeholders = ', '.join(
        '%s::jsonb' if k == 'prefs' else '%s' for k in updates
    )
    assignments = ', '.join(f"{k} = EXCLUDED.{k}" for k in updates)
    db = get_db()
    try:
        row = db.execute(f"""
            INSERT INTO notification_settings ({columns})
            VALUES (%s, {placeholders})
            ON CONFLICT (organization_id) DO UPDATE
               SET {assignments}, updated_at = CURRENT_TIMESTAMP
            RETURNING prefs, attendance_check_at
        """, (current_org_id(), *updates.values())).fetchone()
        db.commit()
    finally:
        db.close()
    if not row:
        return jsonify({'error': 'Organization not found'}), 404
    prefs = dict(tenancy.default_notifications())
    prefs.update(row['prefs'] or {})
    return jsonify({
        'notifications': prefs,
        'attendance_check_at': row['attendance_check_at'],
    })


@app.route('/api/admin/notification-settings/run-now', methods=['POST'])
@jwt_required()
def admin_run_attendance_check_now():
    """Send the attendance check immediately, for this organization only.

    Exists because a setting whose effect is invisible until 1:00 PM tomorrow
    is a setting nobody trusts. Bypasses the daily claim on purpose — this is a
    person asking for it, not the schedule firing — so it can be used twice; the
    second run only re-asks families who still have not answered.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        asked = _run_attendance_check(current_org_id())
    except Exception as e:
        print(f"[ERROR] manual attendance check: {e.__class__.__name__}: {e}")
        return jsonify({'error': 'Could not send the attendance check'}), 500
    return jsonify({'asked': asked})


@app.route('/api/cron/run-due', methods=['POST'])
def cron_run_due():
    """Run whatever daily job is due, for an external scheduler.

    Unauthenticated by JWT on purpose — a cron has no user — and therefore
    guarded by a shared secret instead. With CRON_SECRET unset the route is
    closed rather than open: an endpoint that fans out pushes to every parent
    in every organization is not something to leave ungated by default.

    Safe to call as often as you like. The claim in scheduler.run_due means a
    second call on the same day for the same organization does nothing.
    """
    secret = os.environ.get('CRON_SECRET') or ''
    if not secret:
        return jsonify({'error': 'CRON_SECRET is not configured'}), 503
    supplied = request.headers.get('X-Cron-Secret') or ''
    if not secrets.compare_digest(supplied, secret):
        return jsonify({'error': 'Unauthorized'}), 403
    return jsonify({'ran': scheduler.run_due(_run_attendance_check)})


@app.route('/api/admin/messages', methods=['GET'])
@jwt_required()
def admin_list_messages():
    """List broadcast messages sent by any admin, most recent first."""
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    rows = db.execute("""
        SELECT m.id, m.subject, m.body, m.audience_type, m.created_at,
               u.name AS sender_name,
               (SELECT COUNT(*) FROM parent_messages pm WHERE pm.admin_message_id = m.id) AS recipient_count,
               (SELECT COUNT(*) FROM parent_messages pm WHERE pm.admin_message_id = m.id AND pm.read_at IS NOT NULL) AS read_count,
               COALESCE(
                 (SELECT json_agg(s.name ORDER BY s.name)
                  FROM admin_message_schools ams
                  JOIN schools s ON s.id = ams.school_id
                  WHERE ams.admin_message_id = m.id), '[]'
               ) AS schools
        FROM admin_messages m
        LEFT JOIN users u ON u.id = m.sender_id
        ORDER BY m.created_at DESC
        LIMIT 100
    """).fetchall()
    db.close()
    out = []
    for r in rows:
        d = dict(r)
        d['created_at'] = iso_utc(d['created_at'])
        if isinstance(d['schools'], str):
            d['schools'] = json.loads(d['schools'])
        out.append(d)
    return jsonify(out)


@app.route('/api/parent/messages', methods=['GET'])
@jwt_required()
def parent_list_messages():
    """Inbox list for the logged-in parent. Excludes soft-deleted rows."""
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    db = get_db()
    rows = db.execute("""
        SELECT pm.id, pm.read_at, pm.created_at AS received_at,
               m.subject, m.body, m.created_at AS sent_at,
               u.name AS sender_name, u.role AS sender_role
        FROM parent_messages pm
        JOIN admin_messages m ON m.id = pm.admin_message_id
        LEFT JOIN users u ON u.id = m.sender_id
        WHERE pm.parent_id = %s AND pm.deleted_at IS NULL
        ORDER BY m.created_at DESC
        LIMIT 200
    """, (user_id,)).fetchall()
    db.close()
    out = []
    for r in rows:
        d = dict(r)
        d['received_at'] = iso_utc(d['received_at'])
        d['sent_at'] = iso_utc(d['sent_at'])
        d['read_at'] = iso_utc(d['read_at'])
        out.append(d)
    return jsonify(out)


@app.route('/api/parent/messages/unread-count', methods=['GET'])
@jwt_required()
def parent_unread_count():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    db = get_db()
    row = db.execute("""
        SELECT COUNT(*) AS n
        FROM parent_messages
        WHERE parent_id = %s AND read_at IS NULL AND deleted_at IS NULL
    """, (user_id,)).fetchone()
    db.close()
    return jsonify({'unread': row['n']})


@app.route('/api/parent/messages/mark-all-read', methods=['POST'])
@jwt_required()
def parent_mark_all_read():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    db = get_db()
    db.execute("""
        UPDATE parent_messages
        SET read_at = CURRENT_TIMESTAMP
        WHERE parent_id = %s AND read_at IS NULL AND deleted_at IS NULL
    """, (user_id,))
    db.commit()
    db.close()
    return jsonify({'message': 'All messages marked as read'})


@app.route('/api/parent/messages/<int:pm_id>', methods=['DELETE'])
@jwt_required()
def parent_delete_message(pm_id):
    """Soft-delete: only affects this parent's copy. The admin's record
    in admin_messages is untouched."""
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    db = get_db()
    cur = db.execute("""
        UPDATE parent_messages
        SET deleted_at = CURRENT_TIMESTAMP
        WHERE id = %s AND parent_id = %s AND deleted_at IS NULL
    """, (pm_id, user_id))
    db.commit()
    affected = cur.rowcount
    db.close()
    if affected == 0:
        return jsonify({'error': 'Message not found'}), 404
    return jsonify({'message': 'Deleted'})


# ─── PARENT ↔ ADMIN CONVERSATIONS ───────────────────────────────────────────
# The `parent_messaging` module. One thread per family, both directions.
# Distinct from the broadcast above: that one fans an announcement out to many
# parents and can't be replied to.

def _thread_for_parent(db, parent_id, create=False):
    row = db.execute(
        "SELECT * FROM parent_threads WHERE parent_id = %s", (parent_id,)
    ).fetchone()
    if row or not create:
        return row
    # ON CONFLICT rather than a plain INSERT: two tabs sending the first
    # message at once would otherwise race on the unique index.
    return db.execute("""
        INSERT INTO parent_threads (parent_id) VALUES (%s)
        ON CONFLICT (parent_id) DO UPDATE SET parent_id = excluded.parent_id
        RETURNING *
    """, (parent_id,)).fetchone()


def _message_json(row):
    """One message, with its timestamp tagged UTC.

    Flask's own datetime serializer emits RFC 1123, which browsers happen to
    read as UTC, but every other endpoint here goes through iso_utc() and the
    two formats side by side is how a naive-local-time bug gets reintroduced.
    """
    m = dict(row)
    m['created_at'] = iso_utc(m['created_at'])
    return m


def _thread_messages(db, thread_id):
    rows = db.execute("""
        SELECT id, sender_role, sender_name, body, created_at
          FROM thread_messages WHERE thread_id = %s
         ORDER BY created_at, id
    """, (thread_id,)).fetchall()
    return [_message_json(r) for r in rows]


@app.route('/api/parent/conversation', methods=['GET'])
@jwt_required()
def parent_get_conversation():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    db = get_db()
    try:
        thread = _thread_for_parent(db, user_id)
        if not thread:
            return jsonify({'messages': [], 'unread': 0})
        messages = _thread_messages(db, thread['id'])
        # Opening the thread is the read receipt, the same rule the broadcast
        # inbox already uses.
        if thread['parent_unread']:
            db.execute(
                "UPDATE parent_threads SET parent_unread = 0 WHERE id = %s",
                (thread['id'],),
            )
            db.commit()
    finally:
        db.close()
    return jsonify({'messages': messages, 'unread': 0})


@app.route('/api/parent/conversation', methods=['POST'])
@jwt_required()
def parent_send_message():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    sender_name = claims.get('name') or 'A parent'
    body = ((request.json or {}).get('body') or '').strip()
    if not body:
        return jsonify({'error': 'Write a message first'}), 400

    db = get_db_transaction()
    try:
        thread = _thread_for_parent(db, user_id, create=True)
        msg = db.execute("""
            INSERT INTO thread_messages (thread_id, sender_id, sender_role, sender_name, body)
            VALUES (%s, %s, 'parent', %s, %s)
            RETURNING id, sender_role, sender_name, body, created_at
        """, (thread['id'], user_id, sender_name, body)).fetchone()
        db.execute("""
            UPDATE parent_threads
               SET admin_unread = admin_unread + 1,
                   last_message_at = CURRENT_TIMESTAMP
             WHERE id = %s
        """, (thread['id'],))
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"[ERROR] parent_send_message: {e}")
        return jsonify({'error': 'Could not send the message'}), 500
    finally:
        db.close()

    payload = _message_json(msg)
    # Before the push, for the same reason every other publisher here does it:
    # anyone with the thread open should see the message land, and a push that
    # takes a second to reach a phone should not be what a screen waits on.
    pickup_events.publish(
        'conversation-message',
        {'parent_id': int(user_id), 'message': payload},
        current_org_id(),
    )

    db = get_db()
    try:
        admins = db.execute("SELECT id FROM users WHERE role = 'admin'").fetchall()
    finally:
        db.close()
    for a in admins:
        send_push_to_user_async(
            a['id'], '💬 New message',
            f'{sender_name}: {body[:80]}', '/app/conversations',
        )
    return jsonify(payload), 201


@app.route('/api/admin/conversations', methods=['GET'])
@jwt_required()
def admin_list_conversations():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        rows = db.execute("""
            SELECT t.id, t.parent_id, t.last_message_at, t.admin_unread,
                   u.name AS parent_name, u.email AS parent_email,
                   (SELECT body FROM thread_messages m
                     WHERE m.thread_id = t.id
                     ORDER BY m.created_at DESC, m.id DESC LIMIT 1) AS last_body
              FROM parent_threads t
              JOIN users u ON u.id = t.parent_id
             ORDER BY t.last_message_at DESC NULLS LAST
        """).fetchall()
    finally:
        db.close()
    return jsonify([
        {**dict(r), 'last_message_at': iso_utc(r['last_message_at'])}
        for r in rows
    ])


@app.route('/api/admin/conversations/unread-count', methods=['GET'])
@jwt_required()
def admin_conversations_unread_count():
    """How many parent messages nobody at the office has read yet.

    Feeds the nav badge. Without it an admin only learns a family wrote in by
    opening the screen, which is the same hole the make-up badge was added to
    close. Sums admin_unread so the number matches the per-thread pills in the
    list rather than counting families.

    Declared before /<int:parent_id> for readability only — the int converter
    would not match this path either way.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        row = db.execute(
            "SELECT COALESCE(SUM(admin_unread), 0) AS count FROM parent_threads"
        ).fetchone()
    finally:
        db.close()
    return jsonify({'count': int(row['count'] or 0)})


@app.route('/api/admin/conversations/<int:parent_id>', methods=['GET'])
@jwt_required()
def admin_get_conversation(parent_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        parent = db.execute("""
            SELECT id, name, email, phone, password_set_at
              FROM users WHERE id = %s AND role = 'parent'
        """, (parent_id,)).fetchone()
        if not parent:
            return jsonify({'error': 'Parent not found'}), 404
        # Who this family actually is. The thread only knows a parent_id, but
        # the office thinks in children — without this the admin is answering
        # a name with no idea which kid it belongs to.
        children = db.execute("""
            SELECT c.id, c.name, c.grade_label, c.active, s.name AS school
              FROM children c
              LEFT JOIN schools s ON s.id = c.school_id
             WHERE c.parent_id = %s
             ORDER BY c.name
        """, (parent_id,)).fetchall()
        # Contact 2 has no portal login (child_contacts.user_id is only wired
        # for contact 1), so it is shown as a phone number to call, not as
        # someone else to message.
        others = db.execute("""
            SELECT DISTINCT cc.name, cc.phone, cc.email
              FROM child_contacts cc
              JOIN children c ON c.id = cc.child_id
             WHERE c.parent_id = %s AND cc.priority = 2
             ORDER BY cc.name
        """, (parent_id,)).fetchall()
        thread = _thread_for_parent(db, parent_id)
        messages = _thread_messages(db, thread['id']) if thread else []
        if thread and thread['admin_unread']:
            db.execute(
                "UPDATE parent_threads SET admin_unread = 0 WHERE id = %s",
                (thread['id'],),
            )
            db.commit()
    finally:
        db.close()
    return jsonify({
        'parent': {
            'id': parent['id'],
            'name': parent['name'],
            'email': parent['email'],
            'phone': parent['phone'] or '',
            'password_set': parent['password_set_at'] is not None,
        },
        'children': [{
            'id': c['id'],
            'name': c['name'],
            'grade_label': c['grade_label'],
            'school': c['school'],
            'active': bool(c['active']),
        } for c in children],
        'other_contacts': [dict(o) for o in others],
        'messages': messages,
    })


@app.route('/api/admin/conversations/<int:parent_id>', methods=['POST'])
@jwt_required()
def admin_reply_conversation(parent_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    claims = get_jwt()
    user_id = get_jwt_identity()
    sender_name = claims.get('name') or 'The office'
    body = ((request.json or {}).get('body') or '').strip()
    if not body:
        return jsonify({'error': 'Write a message first'}), 400

    db = get_db_transaction()
    try:
        # Scoped to role = 'parent' so an admin can't open a thread against a
        # counselor or another admin, which nothing in the UI can read.
        parent = db.execute(
            "SELECT id FROM users WHERE id = %s AND role = 'parent'", (parent_id,)
        ).fetchone()
        if not parent:
            db.rollback()
            return jsonify({'error': 'Parent not found'}), 404
        thread = _thread_for_parent(db, parent_id, create=True)
        msg = db.execute("""
            INSERT INTO thread_messages (thread_id, sender_id, sender_role, sender_name, body)
            VALUES (%s, %s, 'admin', %s, %s)
            RETURNING id, sender_role, sender_name, body, created_at
        """, (thread['id'], user_id, sender_name, body)).fetchone()
        db.execute("""
            UPDATE parent_threads
               SET parent_unread = parent_unread + 1,
                   last_message_at = CURRENT_TIMESTAMP
             WHERE id = %s
        """, (thread['id'],))
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"[ERROR] admin_reply_conversation: {e}")
        return jsonify({'error': 'Could not send the message'}), 500
    finally:
        db.close()

    payload = _message_json(msg)
    pickup_events.publish(
        'conversation-message',
        {'parent_id': parent_id, 'message': payload},
        current_org_id(),
    )

    notify_parent(
        parent_id, 'office_message',
        '💬 Message from the office', body[:80], '/app/inbox',
    )
    return jsonify(payload), 201


# ─── STAFF MESSAGING ────────────────────────────────────────────────────────
# The `staff_messaging` module. Same shape as parent_messaging above — one
# thread per person, both directions — but the person is a counselor and the
# other side is always the admin. A counselor can write to the admin; there is
# no thread between two counselors and no way to reach a parent from here.

def _thread_for_counselor(db, counselor_id, create=False):
    row = db.execute(
        "SELECT * FROM staff_threads WHERE counselor_id = %s", (counselor_id,)
    ).fetchone()
    if row or not create:
        return row
    # ON CONFLICT rather than a plain INSERT: two tabs sending the first
    # message at once would otherwise race on the unique index.
    return db.execute("""
        INSERT INTO staff_threads (counselor_id) VALUES (%s)
        ON CONFLICT (counselor_id) DO UPDATE SET counselor_id = excluded.counselor_id
        RETURNING *
    """, (counselor_id,)).fetchone()


def _staff_thread_messages(db, thread_id):
    rows = db.execute("""
        SELECT id, sender_role, sender_name, body, created_at
          FROM staff_thread_messages WHERE thread_id = %s
         ORDER BY created_at, id
    """, (thread_id,)).fetchall()
    return [_message_json(r) for r in rows]


@app.route('/api/counselor/conversation/unread-count', methods=['GET'])
@jwt_required()
def counselor_conversation_unread_count():
    """How many office replies this counselor hasn't opened yet.

    Feeds the nav badge (CounselorShell), same reason the admin's Conversations
    badge exists. A separate route from the GET above rather than folding the
    count into it: that one clears counselor_unread as its read receipt, and a
    nav badge polling every screen would clear it just by rendering.
    """
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    db = get_db()
    try:
        thread = _thread_for_counselor(db, user_id)
    finally:
        db.close()
    return jsonify({'count': thread['counselor_unread'] if thread else 0})


@app.route('/api/counselor/conversation', methods=['GET'])
@jwt_required()
def counselor_get_conversation():
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    db = get_db()
    try:
        thread = _thread_for_counselor(db, user_id)
        if not thread:
            return jsonify({'messages': [], 'unread': 0})
        messages = _staff_thread_messages(db, thread['id'])
        if thread['counselor_unread']:
            db.execute(
                "UPDATE staff_threads SET counselor_unread = 0 WHERE id = %s",
                (thread['id'],),
            )
            db.commit()
    finally:
        db.close()
    return jsonify({'messages': messages, 'unread': 0})


@app.route('/api/counselor/conversation', methods=['POST'])
@jwt_required()
def counselor_send_message():
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    sender_name = claims.get('name') or 'A staff member'
    body = ((request.json or {}).get('body') or '').strip()
    if not body:
        return jsonify({'error': 'Write a message first'}), 400

    db = get_db_transaction()
    try:
        thread = _thread_for_counselor(db, user_id, create=True)
        msg = db.execute("""
            INSERT INTO staff_thread_messages (thread_id, sender_id, sender_role, sender_name, body)
            VALUES (%s, %s, 'counselor', %s, %s)
            RETURNING id, sender_role, sender_name, body, created_at
        """, (thread['id'], user_id, sender_name, body)).fetchone()
        db.execute("""
            UPDATE staff_threads
               SET admin_unread = admin_unread + 1,
                   last_message_at = CURRENT_TIMESTAMP
             WHERE id = %s
        """, (thread['id'],))
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"[ERROR] counselor_send_message: {e}")
        return jsonify({'error': 'Could not send the message'}), 500
    finally:
        db.close()

    payload = _message_json(msg)
    pickup_events.publish(
        'staff-message',
        {'counselor_id': int(user_id), 'message': payload},
        current_org_id(),
    )

    db = get_db()
    try:
        admins = db.execute("SELECT id FROM users WHERE role = 'admin'").fetchall()
    finally:
        db.close()
    for a in admins:
        send_push_to_user_async(
            a['id'], '💬 New staff message',
            f'{sender_name}: {body[:80]}', '/app/conversations',
        )
    return jsonify(payload), 201


@app.route('/api/admin/staff-conversations', methods=['GET'])
@jwt_required()
def admin_list_staff_conversations():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        rows = db.execute("""
            SELECT t.id, t.counselor_id, t.last_message_at, t.admin_unread,
                   u.name AS counselor_name, u.email AS counselor_email,
                   (SELECT body FROM staff_thread_messages m
                     WHERE m.thread_id = t.id
                     ORDER BY m.created_at DESC, m.id DESC LIMIT 1) AS last_body
              FROM staff_threads t
              JOIN users u ON u.id = t.counselor_id
             ORDER BY t.last_message_at DESC NULLS LAST
        """).fetchall()
    finally:
        db.close()
    return jsonify([
        {**dict(r), 'last_message_at': iso_utc(r['last_message_at'])}
        for r in rows
    ])


@app.route('/api/admin/staff-conversations/unread-count', methods=['GET'])
@jwt_required()
def admin_staff_conversations_unread_count():
    """How many staff messages nobody at the office has read yet.

    Feeds the nav badge, same reason /api/admin/conversations/unread-count
    does. Declared before /<int:counselor_id> for readability only — the int
    converter would not match this path either way.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        row = db.execute(
            "SELECT COALESCE(SUM(admin_unread), 0) AS count FROM staff_threads"
        ).fetchone()
    finally:
        db.close()
    return jsonify({'count': int(row['count'] or 0)})


@app.route('/api/admin/staff-conversations/<int:counselor_id>', methods=['GET'])
@jwt_required()
def admin_get_staff_conversation(counselor_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        # Scoped to role = 'counselor' so an admin can't open a thread against
        # a parent or another admin by guessing an id.
        counselor = db.execute("""
            SELECT id, name, email FROM users
             WHERE id = %s AND role = 'counselor'
        """, (counselor_id,)).fetchone()
        if not counselor:
            return jsonify({'error': 'Counselor not found'}), 404
        thread = _thread_for_counselor(db, counselor_id)
        messages = _staff_thread_messages(db, thread['id']) if thread else []
        if thread and thread['admin_unread']:
            db.execute(
                "UPDATE staff_threads SET admin_unread = 0 WHERE id = %s",
                (thread['id'],),
            )
            db.commit()
    finally:
        db.close()
    return jsonify({
        'counselor': {
            'id': counselor['id'], 'name': counselor['name'], 'email': counselor['email'],
        },
        'messages': messages,
    })


@app.route('/api/admin/staff-conversations/<int:counselor_id>', methods=['POST'])
@jwt_required()
def admin_reply_staff_conversation(counselor_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    claims = get_jwt()
    user_id = get_jwt_identity()
    sender_name = claims.get('name') or 'The office'
    body = ((request.json or {}).get('body') or '').strip()
    if not body:
        return jsonify({'error': 'Write a message first'}), 400

    db = get_db_transaction()
    try:
        counselor = db.execute(
            "SELECT id FROM users WHERE id = %s AND role = 'counselor'", (counselor_id,)
        ).fetchone()
        if not counselor:
            db.rollback()
            return jsonify({'error': 'Counselor not found'}), 404
        thread = _thread_for_counselor(db, counselor_id, create=True)
        msg = db.execute("""
            INSERT INTO staff_thread_messages (thread_id, sender_id, sender_role, sender_name, body)
            VALUES (%s, %s, 'admin', %s, %s)
            RETURNING id, sender_role, sender_name, body, created_at
        """, (thread['id'], user_id, sender_name, body)).fetchone()
        db.execute("""
            UPDATE staff_threads
               SET counselor_unread = counselor_unread + 1,
                   last_message_at = CURRENT_TIMESTAMP
             WHERE id = %s
        """, (thread['id'],))
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"[ERROR] admin_reply_staff_conversation: {e}")
        return jsonify({'error': 'Could not send the message'}), 500
    finally:
        db.close()

    payload = _message_json(msg)
    pickup_events.publish(
        'staff-message',
        {'counselor_id': counselor_id, 'message': payload},
        current_org_id(),
    )
    # Direct push rather than notify_parent(): that helper is specifically for
    # a parent recipient's notification preferences (CLAUDE.md §3), and a
    # counselor is not one. Mirrors the direct send counselor_send_message
    # above already uses for the admin side of this same conversation.
    send_push_to_user_async(
        counselor_id, '💬 Message from the office',
        body[:80], '/app/office',
    )
    return jsonify(payload), 201


# ─── DAILY PHOTOS ───────────────────────────────────────────────────────────
# The `photos` module. Counselors upload and tag; each parent sees only the
# photos their own children appear in.
#
# RLS separates one JCC from another. It does not separate one family from
# another inside a JCC — as far as the policies are concerned every parent in
# an organization can read that organization's rows. The parent_id join in
# parent_list_photos is the only thing between one family and another family's
# photographs. Treat it accordingly.

MAX_PHOTO_BYTES = 8 * 1024 * 1024

# How many photos any one gallery request returns.
#
# Bounded because "all photos" is now a view someone can open: a JCC a year in
# has thousands, and neither the grid nor the signing call should scale with
# the whole history. The client says so when it gets a full page rather than
# quietly showing a truncated year.
PHOTO_PAGE_LIMIT = 200


def _photo_payload(row, url_seconds=3600, url=_UNSET):
    """One photo for the client.

    `url` is passed in by list endpoints, which sign a whole page in one call
    (photo_storage.signed_urls). Left out, it signs on its own — correct for
    the single photo a POST hands back, and a round trip per row anywhere else.
    """
    return {
        'id': row['id'],
        'photo_date': row['photo_date'],
        'caption': row['caption'],
        'created_at': row['created_at'],
        'url': (photo_storage.signed_url(row['storage_path'], url_seconds)
                if url is _UNSET else url),
    }


@app.route('/api/counselor/photos', methods=['POST'])
@jwt_required()
def counselor_upload_photo():
    claims = get_jwt()
    if claims.get('role') not in ('counselor', 'admin'):
        return jsonify({'error': 'Unauthorized'}), 403
    if not photo_storage.is_configured():
        return jsonify({'error': 'Photo storage is not configured yet.'}), 503

    user_id = get_jwt_identity()
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file uploaded'}), 400

    # getlist: the form sends one child_ids entry per tagged child.
    try:
        child_ids = [int(c) for c in request.form.getlist('child_ids') if c]
    except ValueError:
        return jsonify({'error': 'Invalid child_ids'}), 400
    if not child_ids:
        return jsonify({'error': 'Tag at least one child'}), 400

    caption = (request.form.get('caption') or '').strip() or None
    photo_date = request.form.get('date') or today_for_org()
    try:
        photo_date = parse_date(photo_date)
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    data = file.read()
    if not data:
        return jsonify({'error': 'That file is empty'}), 400
    if len(data) > MAX_PHOTO_BYTES:
        return jsonify({
            'error': 'That photo is too large. Please pick one under 8 MB.'
        }), 413

    # A transaction, not autocommit: the photo row and its tags have to land
    # together. Committing the photo and then failing on a tag would leave an
    # image nobody is tagged in, which no parent can see and no counselor can
    # find — visible only as storage cost.
    db = get_db_transaction()
    try:
        # Every tagged child must be in a school this counselor covers. Without
        # this a counselor could tag any child in the JCC and push the photo
        # into a family's gallery they have nothing to do with.
        if claims.get('role') == 'counselor':
            allowed = db.execute("""
                SELECT c.id FROM children c
                  JOIN counselor_schools cs
                    ON cs.school_id = c.school_id AND cs.counselor_id = %s
                 WHERE c.id = ANY(%s)
            """, (user_id, child_ids)).fetchall()
        else:
            allowed = db.execute(
                "SELECT id FROM children WHERE id = ANY(%s)", (child_ids,)
            ).fetchall()
        allowed_ids = [r['id'] for r in allowed]
        if len(allowed_ids) != len(set(child_ids)):
            db.rollback()
            return jsonify({'error': 'One of those children is not on your roster'}), 403

        org_id = current_org_id()
        try:
            path = photo_storage.build_path(org_id, photo_date, file.filename)
            photo_storage.upload(path, data, file.filename)
        except photo_storage.StorageError as e:
            db.rollback()
            return jsonify({'error': str(e)}), 400

        try:
            photo = db.execute("""
                INSERT INTO photos (storage_path, uploaded_by, photo_date, caption)
                VALUES (%s, %s, %s, %s)
                RETURNING id, storage_path, photo_date, caption, created_at
            """, (path, user_id, photo_date, caption)).fetchone()
            for cid in allowed_ids:
                db.execute(
                    "INSERT INTO photo_tags (photo_id, child_id) VALUES (%s, %s) "
                    "ON CONFLICT DO NOTHING",
                    (photo['id'], cid),
                )
            # Read the families to notify before committing, so the whole
            # handler is one transaction and the connection goes back to the
            # pool clean rather than mid-snapshot. Both linked accounts for
            # each tagged child, not just Contact #1.
            owner_ids = child_owner_ids(db, allowed_ids)
            db.commit()
        except Exception:
            # The bytes are already in the bucket and the rows just rolled
            # back. Without this the object would linger with nothing pointing
            # at it and no way to find it again.
            db.rollback()
            photo_storage.delete(path)
            raise
    except photo_storage.StorageError as e:
        db.rollback()
        print(f"[ERROR] counselor_upload_photo storage: {e}")
        return jsonify({'error': 'Could not store that photo'}), 502
    except Exception as e:
        db.rollback()
        print(f"[ERROR] counselor_upload_photo: {e}")
        return jsonify({'error': 'Could not save that photo'}), 500
    finally:
        db.close()

    for owner_id in owner_ids:
        notify_parent(
            owner_id, 'new_photos', '📸 New photos',
            'A new photo of your child was shared today.', '/app/photos',
        )
    return jsonify(_photo_payload(photo)), 201


@app.route('/api/counselor/photos', methods=['GET'])
@jwt_required()
def counselor_list_photos():
    claims = get_jwt()
    if claims.get('role') not in ('counselor', 'admin'):
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()

    # `date=all` drops the day filter.
    #
    # Without it this screen opens on today and, on any day nobody posted,
    # shows an empty grid and a date picker with no clue which days have
    # anything in them. Photos are not posted every day, so "empty" was the
    # normal state and looked like a broken feature rather than a quiet
    # Tuesday.
    #
    # Kept opt-in rather than becoming the default: a counselor posting this
    # afternoon's photos wants today, and paging the whole year to find it
    # would be the wrong screen for the more common job.
    date_str = request.args.get('date', today_for_org())
    show_all = date_str == 'all'

    # An admin sees the whole day, a counselor only their own uploads. The
    # delete below already lets an admin remove anyone's photo; without this
    # they could delete one but had no way to find it. Moderating a photo of
    # someone's child has to be possible without hunting down whoever posted
    # it. RLS still scopes both to the organization.
    is_admin = claims.get('role') == 'admin'

    where, params = [], []
    if not show_all:
        where.append('p.photo_date = %s')
        params.append(date_str)
    if not is_admin:
        where.append('p.uploaded_by = %s')
        params.append(user_id)
    clause = f"WHERE {' AND '.join(where)}" if where else ''

    db = get_db()
    try:
        rows = db.execute(f"""
            SELECT p.id, p.storage_path, p.photo_date, p.caption, p.created_at,
                   u.name AS uploaded_by_name
              FROM photos p
              LEFT JOIN users u ON u.id = p.uploaded_by
              {clause}
             ORDER BY p.photo_date DESC, p.created_at DESC
             LIMIT %s
        """, (*params, PHOTO_PAGE_LIMIT)).fetchall()
        tags = db.execute("""
            SELECT t.photo_id, c.name
              FROM photo_tags t
              JOIN children c ON c.id = t.child_id
             WHERE t.photo_id = ANY(%s)
        """, ([r['id'] for r in rows] or [0],)).fetchall()
    finally:
        db.close()
    named = {}
    for t in tags:
        named.setdefault(t['photo_id'], []).append(t['name'])
    # One signing call for the page instead of one per photo. See signed_urls.
    urls = photo_storage.signed_urls([r['storage_path'] for r in rows])
    return jsonify([
        {
            **_photo_payload(r, url=urls.get(r['storage_path'])),
            'children': named.get(r['id'], []),
            'uploaded_by_name': r['uploaded_by_name'],
        }
        for r in rows
    ])


@app.route('/api/counselor/photos/<int:photo_id>', methods=['DELETE'])
@jwt_required()
def counselor_delete_photo(photo_id):
    claims = get_jwt()
    if claims.get('role') not in ('counselor', 'admin'):
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    db = get_db()
    try:
        # A counselor deletes their own uploads; an admin can delete any of the
        # organization's, which is what makes a mistaken photo removable
        # without hunting down whoever posted it.
        if claims.get('role') == 'counselor':
            row = db.execute(
                "DELETE FROM photos WHERE id = %s AND uploaded_by = %s "
                "RETURNING storage_path",
                (photo_id, user_id),
            ).fetchone()
        else:
            row = db.execute(
                "DELETE FROM photos WHERE id = %s RETURNING storage_path",
                (photo_id,),
            ).fetchone()
        db.commit()
    finally:
        db.close()
    if not row:
        return jsonify({'error': 'Photo not found'}), 404
    photo_storage.delete(row['storage_path'])
    return jsonify({'message': 'Deleted'})


@app.route('/api/parent/photos', methods=['GET'])
@jwt_required()
def parent_list_photos():
    """Photos of this parent's own children, newest first.

    The join to children is the access control. RLS keeps other JCCs out;
    nothing but this clause keeps the family next door out. "This parent's
    own children" reaches a child as Contact #1 (parent_id) or a linked
    Contact #2 (child_contacts.user_id) — see _CHILD_BELONGS_TO_USER.
    """
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    db = get_db()
    try:
        rows = db.execute(f"""
            SELECT DISTINCT p.id, p.storage_path, p.photo_date, p.caption,
                            p.created_at
              FROM photos p
              JOIN photo_tags t ON t.photo_id = p.id
              JOIN children c ON c.id = t.child_id
             WHERE {_CHILD_BELONGS_TO_USER}
             ORDER BY p.photo_date DESC, p.created_at DESC
             LIMIT %s
        """, (user_id, user_id, PHOTO_PAGE_LIMIT)).fetchall()
        # Which of *their* children are in each photo. Deliberately filtered
        # the same way again: without it this would name every child in the
        # frame, including other families'.
        mine = db.execute(f"""
            SELECT t.photo_id, c.name
              FROM photo_tags t
              JOIN children c ON c.id = t.child_id
             WHERE {_CHILD_BELONGS_TO_USER} AND t.photo_id = ANY(%s)
        """, (user_id, user_id, [r['id'] for r in rows] or [0])).fetchall()
    finally:
        db.close()
    named = {}
    for t in mine:
        named.setdefault(t['photo_id'], []).append(t['name'])
    # This gallery has always been unfiltered by date and capped at 200, so it
    # was the endpoint most likely to be signing two hundred URLs one at a
    # time. One call now.
    urls = photo_storage.signed_urls([r['storage_path'] for r in rows])
    return jsonify([
        {**_photo_payload(r, url=urls.get(r['storage_path'])),
         'children': named.get(r['id'], [])}
        for r in rows
    ])


# ─── FORGOT / RESET PASSWORD ─────────────────────────────────────────────────

@app.route('/api/auth/forgot-password', methods=['POST'])
def forgot_password():
    client_ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()
    if _is_rate_limited(client_ip):
        return jsonify({'message': 'If that email exists, a reset link has been sent.'})
    _record_login_attempt(client_ip)

    email = (request.json or {}).get('email', '').strip().lower()
    db = get_db()
    # auth_lookup() rather than a plain SELECT, for the same reason sign-in
    # uses it: this request carries no token, so no organization is pinned, and
    # under RLS `SELECT * FROM users` returns nothing for every address. That
    # is not a visible failure — the route answers "if that email exists…"
    # either way — so it read as "no such user" for every user on the platform.
    user = db.execute("SELECT * FROM auth_lookup(%s)", (email,)).fetchone()
    if not user:
        db.close()
        return jsonify({'message': 'If that email exists, a reset link has been sent.'})

    # The address checked out, so binding the request to that user's
    # organization is legitimate now — and necessary, because the insert below
    # takes its organization_id from the connection. Same move sign-in makes.
    from flask import g
    g.organization_id = user['organization_id']
    g.is_superadmin = (user['role'] == 'superadmin')
    db.close()
    db = get_db()

    token = secrets.token_urlsafe(32)
    expires = (datetime.now() + timedelta(hours=2)).strftime('%Y-%m-%d %H:%M:%S')
    db.execute("INSERT INTO password_reset_tokens (user_id, token, expires_at) VALUES (%s,%s,%s)",
               (user['id'], token, expires))
    db.commit()
    # Resolved on this connection, before it is returned: g was pinned to the
    # user's organization above, so the org_self policy lets the row be read.
    # A superadmin has no organization and correctly falls back to Kikar's.
    identity = email_identity(db)
    db.close()
    base_url = os.environ.get('BASE_URL', 'http://localhost:5001').rstrip('/')
    reset_url = f"{base_url}/reset-password?token={token}"
    html = _email_shell(identity, f"{identity.org_name} Password Reset", f"""
            <p>Hello {_escape(user['name'])},</p>
            <p>Click the button below to reset your password. This link expires in 2 hours.</p>
            {_cta(identity, reset_url, 'Reset My Password')}
            <p style="color:#64748b;font-size:0.85rem;">If you didn't request this, ignore this email.</p>
    """)
    send_email(email, f"{identity.org_name} – Reset Your Password", html, identity)
    return jsonify({'message': 'If that email exists, a reset link has been sent.'})

@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    token = request.json.get('token', '').strip()
    new_password = request.json.get('new_password', '')
    pw_error = validate_password(new_password)
    if pw_error:
        return jsonify({'error': pw_error}), 400
    db = get_db()
    # Same shape as forgot-password and sign-in: this request carries a token
    # and nothing else, so no organization is pinned and a plain SELECT on
    # password_reset_tokens is hidden by RLS — which made every reset link and
    # every invitation email in the platform unredeemable. auth_reset_lookup()
    # is the narrow SECURITY DEFINER window that turns the secret into the
    # organization; everything after this line runs under the usual policies.
    row = db.execute("SELECT * FROM auth_reset_lookup(%s)", (token,)).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Invalid or expired reset link'}), 400

    from flask import g
    g.organization_id = row['organization_id']
    # A superadmin's token has no organization — nothing to pin, so the write
    # below needs the superadmin flag instead. It comes from the role on the
    # token's own row, so a caller cannot ask to be one.
    g.is_superadmin = (row['role'] == 'superadmin')
    db.close()
    db = get_db()

    pw_hash = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
    db.execute("UPDATE users SET password_hash=%s, password_set_at = COALESCE(password_set_at, NOW()) WHERE id=%s", (pw_hash, row['user_id']))
    db.execute("UPDATE password_reset_tokens SET used=1 WHERE id=%s", (row['id'],))
    db.commit()
    user = db.execute("SELECT role FROM users WHERE id=%s", (row['user_id'],)).fetchone()
    db.close()
    return jsonify({'message': 'Password reset successfully', 'role': user['role'] if user else None})

# ─── EXPORT TO EXCEL ─────────────────────────────────────────────────────────

def style_header_row(ws, row_num, fill_color="005F73"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')

@app.route('/api/admin/export/attendance', methods=['GET'])
@jwt_required()
def export_attendance():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    date_str = request.args.get('date', today_for_org())
    school_id = request.args.get('school_id')
    db = get_db()
    query = """SELECT ar.attendance_date, c.name as child_name, c.service_type,
               s.name as school, ar.on_bus,
               COALESCE(u.name, '(deleted user)') as submitted_by, ar.submitted_at
               FROM attendance_records ar
               JOIN children c ON ar.child_id = c.id
               JOIN schools s ON ar.school_id = s.id
               LEFT JOIN users u ON ar.submitted_by = u.id
               WHERE ar.attendance_date = %s"""
    params = [date_str]
    if school_id:
        query += " AND ar.school_id = %s"
        params.append(school_id)
    query += " ORDER BY s.name, c.name"
    records = db.execute(query, params).fetchall()
    db.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Date", "School", "Child Name", "Service Type", "On Bus", "Submitted By", "Submitted At"])
    style_header_row(ws, 1)
    for r in records:
        ws.append([r['attendance_date'], r['school'], r['child_name'], r['service_type'],
                   'Yes' if r['on_bus'] else 'No', r['submitted_by'], r['submitted_at']])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"kikar_attendance_{date_str}.xlsx"
    return Response(out.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={"Content-Disposition": f"attachment;filename={fname}"})

@app.route('/api/admin/export/absences', methods=['GET'])
@jwt_required()
def export_absences():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    now = now_for_org()
    start = request.args.get('start', now.strftime('%Y-%m-01'))
    end = request.args.get('end', now.strftime('%Y-%m-%d'))
    try:
        start_d = datetime.strptime(start, '%Y-%m-%d').date()
        end_d = datetime.strptime(end, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date range'}), 400
    if end_d < start_d:
        return jsonify({'error': 'end must be >= start'}), 400

    db = get_db()
    children_meta = {row['id']: row for row in db.execute("""
        SELECT c.id, c.name AS child_name, c.service_type,
               s.name AS school, u.name AS parent_name, u.email AS parent_email
        FROM children c
        JOIN schools s ON c.school_id = s.id
        JOIN users u ON c.parent_id = u.id
    """).fetchall()}

    rows = []
    cur = start_d
    while cur <= end_d:
        ds = cur.strftime('%Y-%m-%d')
        for cid in absent_child_ids_for_date(db, ds):
            meta = children_meta.get(cid)
            if not meta:
                continue
            rows.append({
                'absence_date': ds,
                'child_name': meta['child_name'],
                'school': meta['school'],
                'service_type': meta['service_type'],
                'parent_name': meta['parent_name'],
                'parent_email': meta['parent_email'],
            })
        cur += timedelta(days=1)
    db.close()

    rows.sort(key=lambda r: (r['absence_date'], r['school'], r['child_name']))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Absences"
    ws.append(["Date", "Child Name", "School", "Service Type", "Parent Name", "Parent Email"])
    style_header_row(ws, 1)
    for r in rows:
        ws.append([r['absence_date'], r['child_name'], r['school'], r['service_type'],
                   r['parent_name'], r['parent_email']])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"kikar_absences_{start}_to_{end}.xlsx"
    return Response(out.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={"Content-Disposition": f"attachment;filename={fname}"})

@app.route('/api/admin/export/roster', methods=['GET'])
@jwt_required()
def export_roster():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    rows = db.execute("""SELECT u.name as parent_name, u.email as parent_email,
               c.name as child_name, s.name as school, c.service_type,
               STRING_AGG(r.day_of_week, ', ') as days
               FROM children c JOIN users u ON c.parent_id=u.id
               JOIN schools s ON c.school_id=s.id
               LEFT JOIN registrations r ON r.child_id=c.id
               WHERE c.active = 1
               -- All three ids, not just c.id: Postgres derives a table's other
               -- columns from its primary key, but only for tables actually
               -- named in GROUP BY. Grouping by c.id alone leaves u.name,
               -- u.email and s.name ungrouped and the export fails outright.
               GROUP BY c.id, u.id, s.id ORDER BY s.name, c.name""").fetchall()
    db.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Full Roster"
    ws.append(["Parent Name", "Parent Email", "Child Name", "School", "Service Type", "Registered Days"])
    style_header_row(ws, 1)
    for r in rows:
        ws.append([r['parent_name'], r['parent_email'], r['child_name'],
                   r['school'], r['service_type'], r['days'] or ''])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(out.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={"Content-Disposition": "attachment;filename=kikar_full_roster.xlsx"})

# ─── DASHBOARD CHART STATS ────────────────────────────────────────────────────

@app.route('/api/admin/stats/charts', methods=['GET'])
@jwt_required()
def admin_chart_stats():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    # Always return a valid payload — never 500 here, since this endpoint
    # only powers dashboard visualizations. Each query is isolated; if the
    # database is completely unreachable we return empty arrays so the UI
    # can render a friendly "no data" state instead of a scary error.
    daily, absences_daily, by_school, by_service = [], [], [], []
    db = None
    try:
        db = get_db()
    except Exception as e:
        print(f"[ERROR] admin_chart_stats - get_db: {e}")

    if db is not None:
        now = now_for_org(db)
        thirty_days_ago = (now - timedelta(days=29)).strftime('%Y-%m-%d')
        today_str = now.strftime('%Y-%m-%d')

        # Attendance last 30 days
        try:
            daily = db.execute("""SELECT attendance_date, COUNT(DISTINCT child_id) as cnt
                       FROM attendance_records WHERE attendance_date BETWEEN %s AND %s
                       GROUP BY attendance_date ORDER BY attendance_date""",
                       (thirty_days_ago, today_str)).fetchall() or []
        except Exception as e:
            print(f"[ERROR] chart stats - attendance query: {e}")
            daily = []

        # Absences last 30 days. Counts only one-off absences (the absences
        # table); recurring rules are intentionally NOT expanded here — these
        # charts are an approximation and re-resolving every day in the range
        # adds query cost without changing the trend the dashboard surfaces.
        try:
            absences_daily = db.execute("""SELECT absence_date, COUNT(*) as cnt
                       FROM absences WHERE absence_date BETWEEN %s AND %s
                       GROUP BY absence_date ORDER BY absence_date""",
                       (thirty_days_ago, today_str)).fetchall() or []
        except Exception as e:
            print(f"[ERROR] chart stats - absences query: {e}")
            absences_daily = []

        # By school (enrolled children)
        try:
            by_school = db.execute("""SELECT s.name, COUNT(c.id) as cnt
                       FROM children c JOIN schools s ON c.school_id=s.id
                       WHERE c.active = 1
                       GROUP BY s.id, s.name ORDER BY cnt DESC""").fetchall() or []
        except Exception as e:
            print(f"[ERROR] chart stats - by_school query: {e}")
            by_school = []

        # By service type
        try:
            by_service = db.execute("""SELECT service_type, COUNT(*) as cnt
                       FROM children WHERE active = 1 GROUP BY service_type""").fetchall() or []
        except Exception as e:
            print(f"[ERROR] chart stats - by_service query: {e}")
            by_service = []

        try: db.close()
        except Exception: pass

    def _serialize(rows, *fields):
        out = []
        for r in rows or []:
            try:
                out.append({k: (str(r[src]) if src.endswith('_date') else r[src]) for k, src in fields})
            except Exception as e:
                print(f"[WARN] chart stats - row serialize: {e}")
        return out

    return jsonify({
        'attendance_daily': _serialize(daily, ('date', 'attendance_date'), ('count', 'cnt')),
        'absences_daily': _serialize(absences_daily, ('date', 'absence_date'), ('count', 'cnt')),
        'by_school': _serialize(by_school, ('school', 'name'), ('count', 'cnt')),
        'by_service': _serialize(by_service, ('type', 'service_type'), ('count', 'cnt')),
    })

# ─── ATTENDANCE WEEK / MONTH VIEW ─────────────────────────────────────────────

@app.route('/api/admin/attendance/week', methods=['GET'])
@jwt_required()
def admin_attendance_week():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    date_str = request.args.get('date', today_for_org())
    d = datetime.strptime(date_str, '%Y-%m-%d')
    week_start = d - timedelta(days=d.weekday())  # Monday
    week_dates = [(week_start + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(5)]
    db = get_db()
    result = {}
    for wd in week_dates:
        count = db.execute("SELECT COUNT(DISTINCT child_id) as cnt FROM attendance_records WHERE attendance_date=%s", (wd,)).fetchone()['cnt']
        absences = db.execute("SELECT COUNT(*) as cnt FROM absences WHERE absence_date=%s", (wd,)).fetchone()['cnt']
        result[wd] = {'attendance': count, 'absences': absences}
    # also get per-school breakdown for the week
    school_rows = db.execute("""SELECT ar.attendance_date, s.name as school,
               COUNT(DISTINCT ar.child_id) as cnt
               FROM attendance_records ar JOIN schools s ON ar.school_id=s.id
               WHERE ar.attendance_date BETWEEN %s AND %s
               GROUP BY ar.attendance_date, s.id""",
               (week_dates[0], week_dates[-1])).fetchall()
    db.close()
    by_school = {}
    for r in school_rows:
        if r['school'] not in by_school:
            by_school[r['school']] = {}
        by_school[r['school']][r['attendance_date']] = r['cnt']
    return jsonify({'dates': week_dates, 'daily': result, 'by_school': by_school})

@app.route('/api/admin/attendance/month', methods=['GET'])
@jwt_required()
def admin_attendance_month():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    now = now_for_org()
    year = int(request.args.get('year', now.year))
    month = int(request.args.get('month', now.month))
    import calendar
    days_in_month = calendar.monthrange(year, month)[1]
    start = f"{year}-{month:02d}-01"
    end = f"{year}-{month:02d}-{days_in_month:02d}"
    db = get_db()
    rows = db.execute("""SELECT attendance_date, COUNT(DISTINCT child_id) as cnt
               FROM attendance_records WHERE attendance_date BETWEEN %s AND %s
               GROUP BY attendance_date""", (start, end)).fetchall()
    abs_rows = db.execute("""SELECT absence_date, COUNT(*) as cnt
               FROM absences WHERE absence_date BETWEEN %s AND %s
               GROUP BY absence_date""", (start, end)).fetchall()
    db.close()
    attendance_map = {r['attendance_date']: r['cnt'] for r in rows}
    absence_map = {r['absence_date']: r['cnt'] for r in abs_rows}
    result = []
    for d in range(1, days_in_month + 1):
        ds = f"{year}-{month:02d}-{d:02d}"
        result.append({'date': ds, 'attendance': attendance_map.get(ds, 0), 'absences': absence_map.get(ds, 0)})
    return jsonify({'year': year, 'month': month, 'days': result})

# ─── CALENDAR EVENTS ────────────────────────────────────────────────────────

@app.route('/api/admin/calendar', methods=['GET'])
@jwt_required()
def admin_get_calendar():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    year = request.args.get('year', now_for_org().strftime('%Y'))
    try:
        db = get_db()
        events = db.execute(
            "SELECT id, event_date, event_type, title, description FROM calendar_events WHERE event_date LIKE %s ORDER BY event_date",
            (f'{year}-%',)
        ).fetchall()
        db.close()
        return jsonify([dict(e) for e in events])
    except Exception as e:
        print(f"[ERROR] admin_get_calendar: {e}")
        return jsonify({'error': 'Failed to load calendar'}), 500

@app.route('/api/admin/calendar', methods=['POST'])
@jwt_required()
def admin_create_calendar_event():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    data = request.json
    event_date = data.get('event_date', '').strip()
    event_type = data.get('event_type', 'special_event')
    title = data.get('title', '').strip()
    description = data.get('description', '').strip()

    if not event_date or not title:
        return jsonify({'error': 'Date and title are required'}), 400
    if event_type not in ('closed', 'mini_camp', 'early_dismissal', 'special_event'):
        return jsonify({'error': 'Invalid event type'}), 400

    try:
        db = get_db()
        cur = db.execute(
            "INSERT INTO calendar_events (event_date, event_type, title, description, created_by) VALUES (%s,%s,%s,%s,%s) RETURNING id",
            (event_date, event_type, title, description, user_id)
        )
        event_id = cur.fetchone()['id']
        db.close()
        return jsonify({'id': event_id, 'message': 'Event created'})
    except Exception as e:
        print(f"[ERROR] admin_create_calendar_event: {e}")
        return jsonify({'error': 'Failed to create event'}), 500

@app.route('/api/admin/calendar/upload', methods=['POST'])
@jwt_required()
def admin_upload_calendar():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    filename = file.filename.lower()
    added = 0
    errors = []

    TYPE_MAP = {
        'closed': 'closed', 'close': 'closed', 'no service': 'closed',
        'mini camp': 'mini_camp', 'mini_camp': 'mini_camp', 'minicamp': 'mini_camp',
        'early dismissal': 'early_dismissal', 'early_dismissal': 'early_dismissal',
        '4pm': 'early_dismissal', 'closes at 4': 'early_dismissal', 'early': 'early_dismissal',
        'special event': 'special_event', 'special_event': 'special_event',
        'event': 'special_event', 'jcc event': 'special_event', 'other': 'special_event',
    }
    VALID_TYPES = {'closed', 'mini_camp', 'early_dismissal', 'special_event'}

    rows = []
    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # skip header
                if not row or not row[0]:
                    continue
                rows.append(row)
        elif filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content))
            next(reader, None)
            for row in reader:
                if row:
                    rows.append(row)
        else:
            return jsonify({'error': 'Unsupported format. Use .xlsx or .csv'}), 400
    except Exception as e:
        print(f"[ERROR] calendar upload read: {e}")
        return jsonify({'error': 'Failed to read file'}), 400

    db = get_db()
    for i, row in enumerate(rows):
        try:
            raw_date = row[0]
            raw_type = str(row[1]).strip().lower() if len(row) > 1 and row[1] else 'special_event'
            title = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            description = str(row[3]).strip() if len(row) > 3 and row[3] else ''

            if not title:
                errors.append(f"Row {i+2}: Title (column C) is required")
                continue

            # Parse date
            if hasattr(raw_date, 'strftime'):
                event_date = raw_date.strftime('%Y-%m-%d')
            else:
                event_date = None
                for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y', '%d/%m/%Y', '%m/%d/%y'):
                    try:
                        event_date = datetime.strptime(str(raw_date).strip(), fmt).strftime('%Y-%m-%d')
                        break
                    except ValueError:
                        continue
                if not event_date:
                    errors.append(f"Row {i+2}: Cannot parse date '{raw_date}'")
                    continue

            event_type = raw_type if raw_type in VALID_TYPES else TYPE_MAP.get(raw_type, 'special_event')

            db.execute(
                "INSERT INTO calendar_events (event_date, event_type, title, description, created_by) VALUES (%s,%s,%s,%s,%s)",
                (event_date, event_type, title, description, user_id)
            )
            added += 1
        except Exception as e:
            errors.append(f"Row {i+2}: {str(e)}")

    db.commit()
    db.close()
    return jsonify({'added': added, 'errors': errors})

@app.route('/api/admin/calendar/<int:event_id>', methods=['PUT'])
@jwt_required()
def admin_update_calendar_event(event_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}

    updates = {}
    if 'event_date' in data:
        event_date = (data.get('event_date') or '').strip()
        if not event_date:
            return jsonify({'error': 'Date cannot be empty'}), 400
        updates['event_date'] = event_date
    if 'event_type' in data:
        event_type = data.get('event_type')
        if event_type not in ('closed', 'mini_camp', 'early_dismissal', 'special_event'):
            return jsonify({'error': 'Invalid event type'}), 400
        updates['event_type'] = event_type
    if 'title' in data:
        title = (data.get('title') or '').strip()
        if not title:
            return jsonify({'error': 'Title cannot be empty'}), 400
        updates['title'] = title
    if 'description' in data:
        updates['description'] = (data.get('description') or '').strip() or None
    if not updates:
        return jsonify({'error': 'Nothing to update'}), 400

    sets = ', '.join(f"{k} = %s" for k in updates)
    db = get_db()
    cur = db.execute(
        f"UPDATE calendar_events SET {sets} WHERE id = %s RETURNING id",
        (*updates.values(), event_id),
    )
    if cur.rowcount == 0:
        db.close()
        return jsonify({'error': 'Event not found'}), 404
    db.commit()
    db.close()
    return jsonify({'message': 'Event updated'})

@app.route('/api/admin/calendar/<int:event_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_calendar_event(event_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    db.execute("DELETE FROM calendar_events WHERE id=%s", (event_id,))
    db.commit()
    db.close()
    return jsonify({'message': 'Event deleted'})

@app.route('/api/parent/calendar', methods=['GET'])
@jwt_required()
def parent_get_calendar():
    now = now_for_org()
    today_str = now.strftime('%Y-%m-%d')
    end_str = (now + timedelta(days=90)).strftime('%Y-%m-%d')
    db = get_db()
    events = db.execute(
        "SELECT id, event_date, event_type, title, description FROM calendar_events WHERE event_date >= %s AND event_date <= %s ORDER BY event_date",
        (today_str, end_str)
    ).fetchall()
    db.close()
    return jsonify([dict(e) for e in events])

@app.route('/api/counselor/calendar', methods=['GET'])
@jwt_required()
def counselor_get_calendar():
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403
    now = now_for_org()
    today_str = now.strftime('%Y-%m-%d')
    end_str = (now + timedelta(days=90)).strftime('%Y-%m-%d')
    db = get_db()
    events = db.execute(
        "SELECT id, event_date, event_type, title, description FROM calendar_events WHERE event_date >= %s AND event_date <= %s ORDER BY event_date",
        (today_str, end_str)
    ).fetchall()
    db.close()
    return jsonify([dict(e) for e in events])

# ─── SECURE PICKUP ─────────────────────────────────────────────────────────
# The `secure_pickup` module: a parent-maintained list of who may collect each
# child, a counselor confirming which of them actually did, and that person's
# signature.
#
# Independent of the announce-then-claim queue below. Both can run at once;
# neither knows about the other. What ties this one to the rest of the app is
# that a release also closes the child's attendance for the day.

# A signature is a few hundred strokes of monochrome PNG — 5-10 KB in practice.
# The cap is far above that and exists so this endpoint cannot be used as a
# file host: it accepts an image from an unattended phone in a carpool line.
MAX_SIGNATURE_BYTES = 64 * 1024
_SIGNATURE_PREFIX = 'data:image/png;base64,'
_PNG_MAGIC = b'\x89PNG\r\n\x1a\n'


def decode_signature(value):
    """Turn the client's data URL into PNG bytes, or raise ValueError.

    Deliberately strict about the one format SignaturePad produces. Checking
    the magic bytes rather than trusting the data URL's own prefix is the point
    — the prefix is just a string the caller chose, and the bytes are what get
    stored and rendered back to an admin years from now.
    """
    if not isinstance(value, str) or not value.startswith(_SIGNATURE_PREFIX):
        raise ValueError('A signature is required')
    try:
        raw = base64.b64decode(value[len(_SIGNATURE_PREFIX):], validate=True)
    except (binascii.Error, ValueError):
        raise ValueError('That signature could not be read')
    if not raw.startswith(_PNG_MAGIC):
        raise ValueError('That signature could not be read')
    if len(raw) > MAX_SIGNATURE_BYTES:
        raise ValueError('That signature is too large')
    return raw


@app.route('/api/parent/authorized-pickups', methods=['GET'])
@jwt_required()
def parent_list_authorized_pickups():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    db = get_db()
    try:
        rows = db.execute(f"""
            SELECT p.id, p.child_id, p.name, p.relationship, c.name AS child_name
              FROM authorized_pickup_people p
              JOIN children c ON c.id = p.child_id
             WHERE {_CHILD_BELONGS_TO_USER} AND c.active = 1
             ORDER BY c.name, p.name
        """, (user_id, user_id)).fetchall()
    finally:
        db.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/parent/authorized-pickups', methods=['POST'])
@jwt_required()
def parent_add_authorized_pickup():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    data = request.json or {}
    child_id = data.get('child_id')
    name = (data.get('name') or '').strip()
    relationship = (data.get('relationship') or '').strip() or None

    if not child_id or not name:
        return jsonify({'error': 'A child and a name are required'}), 400

    db = get_db()
    try:
        # The child has to belong to this parent — as Contact #1 or a linked
        # Contact #2. Without this check a parent could authorize themselves
        # onto another family's child.
        child = db.execute(
            f"SELECT c.id FROM children c "
            f"WHERE c.id = %s AND {_CHILD_BELONGS_TO_USER} AND c.active = 1",
            (child_id, user_id, user_id),
        ).fetchone()
        if not child:
            return jsonify({'error': 'Child not found'}), 404

        # One list per family, not per child. A parent authorizing a
        # grandmother means she collects their children — plural — and making
        # them retype her name once per child produced busywork and two lists
        # that drifted apart. Somebody who may genuinely collect only one child
        # is removed from the other with the × that is already there, which is
        # one tap against the many this saves. "Their children" now includes
        # any child this account reaches as Contact #1 or Contact #2.
        rows = db.execute(f"""
            INSERT INTO authorized_pickup_people (child_id, name, relationship, created_by)
            SELECT c.id, %s, %s, %s
              FROM children c
             WHERE {_CHILD_BELONGS_TO_USER} AND c.active = 1
            ON CONFLICT (child_id, name) DO UPDATE SET relationship = excluded.relationship
            RETURNING id, child_id, name, relationship
        """, (name, relationship, user_id, user_id, user_id)).fetchall()
        db.commit()
    except Exception as e:
        print(f"[ERROR] parent_add_authorized_pickup: {e}")
        return jsonify({'error': 'Could not add that person'}), 400
    finally:
        db.close()
    return jsonify([dict(r) for r in rows]), 201


@app.route('/api/parent/authorized-pickups/<int:person_id>', methods=['PATCH'])
@jwt_required()
def parent_edit_authorized_pickup(person_id):
    """Correct a name or a relationship already on the list.

    Separate from DELETE-then-POST because the two are not equivalent: a
    delete would break the person_id link from any release this person already
    signed for. The release keeps its own copy of the name either way, so past
    records read the same before and after — but the live link is worth
    keeping.
    """
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    data = request.json or {}
    name = (data.get('name') or '').strip()
    relationship = (data.get('relationship') or '').strip() or None
    if not name:
        return jsonify({'error': 'A name is required'}), 400

    db = get_db()
    try:
        # Same ownership join as the delete below: the row has to hang off a
        # child of this parent's, or a person_id from another family would be
        # editable by anyone who guessed the number.
        current = db.execute(f"""
            SELECT p.id, p.name
              FROM authorized_pickup_people p
              JOIN children c ON c.id = p.child_id
             WHERE p.id = %s AND {_CHILD_BELONGS_TO_USER}
        """, (person_id, user_id, user_id)).fetchone()
        if not current:
            return jsonify({'error': 'Not found'}), 404

        # The edit fans out the same way the add does. This person is one entry
        # in the family's list that happens to be stored per child, so fixing a
        # misspelled name fixes it everywhere rather than leaving the other
        # child with the typo — and a counselor reading two spellings of the
        # same grandmother has no way to tell they are one person.
        updated = db.execute(f"""
            UPDATE authorized_pickup_people p
               SET name = %s, relationship = %s
              FROM children c
             WHERE c.id = p.child_id AND {_CHILD_BELONGS_TO_USER} AND p.name = %s
            RETURNING p.id, p.child_id, p.name, p.relationship
        """, (name, relationship, user_id, user_id, current['name'])).fetchall()
        db.commit()
    except psycopg2.errors.UniqueViolation:
        # UNIQUE (child_id, name) — they renamed onto someone already listed.
        db.rollback()
        return jsonify({
            'error': f'{name} is already on the list for this child.'
        }), 409
    except Exception as e:
        db.rollback()
        print(f"[ERROR] parent_edit_authorized_pickup: {e}")
        return jsonify({'error': 'Could not save that change'}), 400
    finally:
        db.close()
    return jsonify([dict(r) for r in updated])


@app.route('/api/parent/authorized-pickups/<int:person_id>', methods=['DELETE'])
@jwt_required()
def parent_remove_authorized_pickup(person_id):
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    db = get_db()
    try:
        # Deleting the person does not touch pickup_releases: those rows carry
        # their own copy of the name precisely so removing someone from the
        # list never rewrites the history of who collected a child.
        deleted = db.execute(f"""
            DELETE FROM authorized_pickup_people p
             USING children c
             WHERE p.id = %s AND c.id = p.child_id AND {_CHILD_BELONGS_TO_USER}
            RETURNING p.id
        """, (person_id, user_id, user_id)).fetchone()
        db.commit()
    finally:
        db.close()
    if not deleted:
        return jsonify({'error': 'Not found'}), 404
    return jsonify({'message': 'Removed'})


@app.route('/api/counselor/authorized-pickups', methods=['GET'])
@jwt_required()
def counselor_list_authorized_pickups():
    """Who may collect one child, for the counselor about to release them."""
    claims = get_jwt()
    if claims.get('role') not in ('counselor', 'admin'):
        return jsonify({'error': 'Unauthorized'}), 403

    child_id = request.args.get('child_id', type=int)
    if not child_id:
        return jsonify({'error': 'child_id is required'}), 400

    user_id = get_jwt_identity()
    db = get_db()
    try:
        if claims.get('role') == 'counselor':
            # Same scoping rule as the roster and the attendance writes.
            if not counselor_covers_child(db, user_id, child_id):
                return jsonify({'error': 'Child not in your schools'}), 403

        # The registered parent heads the list without having to add
        # themselves. They are the most common answer to "who is collecting
        # this child", and requiring them to authorize themselves left the
        # counselor with nobody to pick when a parent walked up.
        child = db.execute("""
            SELECT c.id, u.name AS parent_name
              FROM children c
              JOIN users u ON u.id = c.parent_id
             WHERE c.id = %s
        """, (child_id,)).fetchone()
        if not child:
            return jsonify({'error': 'Child not found'}), 404

        rows = db.execute("""
            SELECT id, name, relationship
              FROM authorized_pickup_people
             WHERE child_id = %s
             ORDER BY name
        """, (child_id,)).fetchall()
        released = db.execute("""
            SELECT person_name, person_relationship, counselor_name, released_at,
                   signature IS NOT NULL AS has_signature
              FROM pickup_releases
             WHERE child_id = %s AND release_date = %s
        """, (child_id, today_for_org(db))).fetchone()
    finally:
        db.close()

    parent_name = child['parent_name']
    people = [{
        'id': None,
        'name': parent_name,
        'relationship': 'Parent',
        'is_parent': True,
    }]
    # A parent who already added themselves by name would otherwise appear
    # twice, and the counselor would have to guess which one to tap.
    for r in rows:
        if r['name'].strip().lower() == (parent_name or '').strip().lower():
            continue
        people.append({
            'id': r['id'],
            'name': r['name'],
            'relationship': r['relationship'],
            'is_parent': False,
        })

    return jsonify({
        'people': people,
        'released': dict(released) if released else None,
    })


@app.route('/api/admin/authorized-pickups', methods=['POST'])
@jwt_required()
def admin_add_authorized_pickup():
    """The office's equivalent of a parent's own POST, for a child registered
    by hand whose parent has no account yet to add the list themselves.

    Same one-list-per-family fan-out as parent_add_authorized_pickup: this
    person is authorized to collect every active child of the same parent,
    not just the one the admin had open, so the two entry points don't leave
    a sibling's list one name short of the other's.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.json or {}
    child_id = data.get('child_id')
    name = (data.get('name') or '').strip()
    relationship = (data.get('relationship') or '').strip() or None

    if not child_id or not name:
        return jsonify({'error': 'A child and a name are required'}), 400

    admin_id = get_jwt_identity()
    db = get_db()
    try:
        child = db.execute(
            "SELECT parent_id FROM children WHERE id = %s AND active = 1",
            (child_id,),
        ).fetchone()
        if not child:
            return jsonify({'error': 'Child not found'}), 404

        rows = db.execute("""
            INSERT INTO authorized_pickup_people (child_id, name, relationship, created_by)
            SELECT c.id, %s, %s, %s
              FROM children c
             WHERE c.parent_id = %s AND c.active = 1
            ON CONFLICT (child_id, name) DO UPDATE SET relationship = excluded.relationship
            RETURNING id, child_id, name, relationship
        """, (name, relationship, admin_id, child['parent_id'])).fetchall()
        db.commit()
    except Exception as e:
        print(f"[ERROR] admin_add_authorized_pickup: {e}")
        return jsonify({'error': 'Could not add that person'}), 400
    finally:
        db.close()
    return jsonify([dict(r) for r in rows]), 201


@app.route('/api/admin/authorized-pickups/<int:person_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_authorized_pickup(person_id):
    """Undo one row — a mistaken import, a person who is no longer approved.

    Admin-only. A parent already has their own DELETE at
    /api/parent/authorized-pickups/<id>, scoped to their own children; this
    is the equivalent for the office, with no parent_id check because the
    caller isn't the family — RLS is what keeps this inside one organization.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    try:
        cur = db.execute(
            "DELETE FROM authorized_pickup_people WHERE id = %s", (person_id,)
        )
        if cur.rowcount == 0:
            return jsonify({'error': 'Not found'}), 404
        db.commit()
    finally:
        db.close()
    return jsonify({'message': 'Removed'})


# ─── AUTHORIZED PICKUP IMPORT ───────────────────────────────────────────────
# Bulk-loads authorized_pickup_people from the membership system's export,
# instead of every family adding their own list by hand one at a time. Purely
# additive — never deletes or deactivates a row — so unlike roster import
# there is no staging/preview/commit ceremony: matched rows write in the same
# request, and the response reports exactly what happened so a bad row is
# fixable by fixing the file and re-uploading, safely, because the write is
# idempotent (ON CONFLICT keeps re-running the same file a no-op).

def _pickup_child_index(db) -> dict:
    """Every active child in this organization, indexed by (last, first) and,
    as a fallback, by full name — same two-key shape roster_staging.py's own
    _existing_children uses to match a spreadsheet row to a child, for the
    same reason: a membership export's name columns are the only handle it
    gives you.
    """
    index: dict[tuple, list[int]] = {}
    rows = db.execute(
        "SELECT id, name, first_name, last_name FROM children WHERE active = 1"
    ).fetchall()
    for r in rows:
        if r['last_name'] and r['first_name']:
            key = ('n', r['last_name'].strip().casefold(), r['first_name'].strip().casefold())
            index.setdefault(key, []).append(r['id'])
        elif r['name']:
            key = ('d', re.sub(r'\s+', ' ', r['name']).strip().casefold())
            index.setdefault(key, []).append(r['id'])
    return index


@app.route('/api/admin/authorized-pickup-import', methods=['POST'])
@jwt_required()
def admin_authorized_pickup_import():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    filename = file.filename or ''
    if not filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Upload the list as .xlsx — this importer reads '
                                 'the column headers, which CSV does not carry.'}), 400

    try:
        rows, problems = pickup_import.parse_pickup_workbook(io.BytesIO(file.read()))
    except Exception as e:
        print(f"[ERROR] authorized-pickup-import parse failed: {type(e).__name__}")
        return jsonify({'error': 'Could not read that file.'}), 400
    if problems:
        return jsonify({'error': problems[0]}), 400

    user_id = _current_user_id()
    db = get_db_transaction()
    try:
        index = _pickup_child_index(db)
        matched = 0
        added = 0
        updated = 0
        unmatched = []
        ambiguous = []
        for row in rows:
            key = ('n', row.last_name.casefold(), row.first_name.casefold())
            candidates = index.get(key)
            if not candidates:
                key = ('d', f'{row.first_name} {row.last_name}'.strip().casefold())
                candidates = index.get(key)
            if not candidates:
                unmatched.append({
                    'row': row.row_number, 'name': f'{row.first_name} {row.last_name}'.strip(),
                })
                continue
            if len(candidates) > 1:
                ambiguous.append({
                    'row': row.row_number, 'name': f'{row.first_name} {row.last_name}'.strip(),
                    'count': len(candidates),
                })
                continue

            matched += 1
            child_id = candidates[0]
            for person in row.people:
                result = db.execute("""
                    INSERT INTO authorized_pickup_people (child_id, name, relationship, created_by)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (child_id, name) DO UPDATE SET relationship = excluded.relationship
                    RETURNING (xmax = 0) AS inserted
                """, (child_id, person.name, person.relationship, user_id)).fetchone()
                if result['inserted']:
                    added += 1
                else:
                    updated += 1
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"[ERROR] authorized-pickup-import write failed: {type(e).__name__}")
        return jsonify({'error': 'Could not import that file.'}), 500
    finally:
        db.close()

    return jsonify({
        'rows_read': len(rows),
        'matched': matched,
        'people_added': added,
        'people_updated': updated,
        'unmatched': unmatched,
        'ambiguous': ambiguous,
    }), 201


@app.route('/api/counselor/pickup/release', methods=['POST'])
@jwt_required()
def counselor_release_child():
    """Confirm who collected a child, take their signature, and close the day.

    Everything happens together or not at all: the authorization is checked,
    the release and its signature are recorded, attendance is closed out, and
    everyone watching is told. They share one transaction because a release
    that recorded the person but left the child marked present would read, on
    every other screen, as a child still in the building — and because a
    release without its signature is the one thing this record cannot be.
    """
    claims = get_jwt()
    if claims.get('role') not in ('counselor', 'admin'):
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    counselor_name = claims.get('name') or 'Staff'
    data = request.json or {}
    child_id = data.get('child_id')
    # person_id is None when the registered parent is collecting: they are on
    # the list implicitly and have no authorized_pickup_people row. The name is
    # resolved from the database below, never taken from the request.
    person_id = data.get('person_id')
    if not child_id:
        return jsonify({'error': 'child_id is required'}), 400

    try:
        signature = decode_signature(data.get('signature'))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    release_date = today_for_org()
    db = get_db_transaction()
    try:
        if claims.get('role') == 'counselor':
            if not counselor_covers_child(db, user_id, child_id, release_date):
                db.rollback()
                return jsonify({'error': 'Child not in your schools'}), 403

        child = db.execute("""
            SELECT c.id, c.name, c.parent_id, u.name AS parent_name
              FROM children c
              JOIN users u ON u.id = c.parent_id
             WHERE c.id = %s
        """, (child_id,)).fetchone()
        if not child:
            db.rollback()
            return jsonify({'error': 'Child not found'}), 404

        if person_id is None:
            person = {
                'id': None,
                'name': child['parent_name'],
                'relationship': 'Parent',
            }
        else:
            # The person must be authorized *for this child*. Existing is not
            # enough — that would let anyone on any family's list collect
            # anyone.
            person = db.execute("""
                SELECT id, name, relationship FROM authorized_pickup_people
                 WHERE id = %s AND child_id = %s
            """, (person_id, child_id)).fetchone()
            if not person:
                db.rollback()
                return jsonify({'error': 'That person is not authorized for this child'}), 403

        # DO NOTHING rather than DO UPDATE: two counselors tapping at once in
        # the carpool line is ordinary, and the second tap must not overwrite
        # the record of who actually took the child.
        release = db.execute("""
            INSERT INTO pickup_releases
                (child_id, person_id, person_name, person_relationship,
                 counselor_id, counselor_name, release_date,
                 signature, signature_mime)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'image/png')
            ON CONFLICT (child_id, release_date) DO NOTHING
            RETURNING id, released_at
        """, (child_id, person['id'], person['name'], person['relationship'],
              user_id, counselor_name, release_date,
              psycopg2.Binary(signature))).fetchone()

        if release is None:
            existing = db.execute("""
                SELECT person_name, counselor_name, released_at
                  FROM pickup_releases WHERE child_id = %s AND release_date = %s
            """, (child_id, release_date)).fetchone()
            db.rollback()
            return jsonify({
                'error': f"{child['name']} was already released to "
                         f"{existing['person_name']} by {existing['counselor_name']}.",
                'released': dict(existing),
            }), 409

        # Close the day. The row may not exist if the child was never marked
        # present — releasing someone who was never checked in is a real thing
        # (a parent collecting a child who arrived late), so create it rather
        # than failing.
        db.execute("""
            INSERT INTO attendance_records
                (child_id, school_id, attendance_date, on_bus, submitted_by,
                 checked_out_at, status, status_at, status_by)
            SELECT %s, c.school_id, %s, 1, %s, CURRENT_TIMESTAMP,
                   'checked_out', CURRENT_TIMESTAMP, %s
              FROM children c WHERE c.id = %s
            ON CONFLICT (child_id, attendance_date)
            DO UPDATE SET checked_out_at = CURRENT_TIMESTAMP,
                          status = 'checked_out',
                          status_at = CURRENT_TIMESTAMP,
                          status_by = excluded.status_by
        """, (child_id, release_date, user_id, user_id, child_id))
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"[ERROR] counselor_release_child: {e}")
        return jsonify({'error': 'Could not record the pickup'}), 500
    finally:
        db.close()

    payload = {
        'child_id': child_id,
        'child_name': child['name'],
        'person_name': person['name'],
        'counselor_name': counselor_name,
        'released_at': release['released_at'],
    }
    pickup_events.publish('child-released', payload, current_org_id())
    _publish_headcount(release_date)

    # The parent asked for this record; tell them without making them look.
    # Both linked accounts, not just whichever one is children.parent_id.
    notify_db = get_db()
    try:
        owner_ids = child_owner_ids(notify_db, [child_id])
    finally:
        notify_db.close()
    for owner_id in owner_ids:
        notify_parent(
            owner_id,
            'pickup_released',
            '👋 Picked up',
            f"{child['name']} was picked up by {person['name']}.",
            '/app/home',
        )
    return jsonify(payload), 201


# The admin's permanent record: which child left, with whom, their
# relationship, at what time, and the signature that person gave.

# One month of a mid-sized JCC is a few thousand rows. The cap is a backstop
# against a year-wide range returning everything at once; the response says
# when it bit rather than quietly serving a partial history as if it were the
# whole thing.
MAX_RELEASE_ROWS = 2000


@app.route('/api/admin/pickup-releases', methods=['GET'])
@jwt_required()
def admin_list_pickup_releases():
    """The release log for a date range.

    Deliberately does not return the signature bytes. A month of releases with
    an image inline would be megabytes of base64 for a table that shows none of
    it — the bytes come from the endpoint below, one at a time, when an admin
    actually opens a record.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    now = now_for_org()
    end = request.args.get('end', now.strftime('%Y-%m-%d'))
    start = request.args.get(
        'start',
        (now - timedelta(days=30)).strftime('%Y-%m-%d'),
    )
    try:
        start = parse_date(start)
        end = parse_date(end)
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400
    if end < start:
        return jsonify({'error': 'end must be >= start'}), 400

    school_id = request.args.get('school_id', type=int)

    db = get_db()
    try:
        # release_date is TEXT in YYYY-MM-DD, so a string BETWEEN orders the
        # same way a date one would.
        # dismissal_time is per weekday, so each row needs the registration for
        # the weekday IT was released on, not today's — to_char gives the same
        # weekday name registrations.day_of_week is written in (see the
        # generate_series pattern in the activity-roster and time-off queries).
        sql = """
            SELECT r.id, r.child_id, c.name AS child_name, s.name AS school,
                   s.id AS school_id,
                   r.person_name, r.person_relationship,
                   r.counselor_name, r.release_date, r.released_at,
                   a.checked_in_at,
                   reg.dismissal_time,
                   r.signature IS NOT NULL AS has_signature
              FROM pickup_releases r
              JOIN children c ON c.id = r.child_id
              JOIN schools s ON s.id = c.school_id
              LEFT JOIN attendance_records a
                     ON a.child_id = r.child_id
                    AND a.attendance_date = r.release_date
              LEFT JOIN registrations reg
                     ON reg.child_id = r.child_id
                    AND reg.day_of_week = to_char(r.release_date::date, 'FMDay')
             WHERE r.release_date BETWEEN %s AND %s
        """
        params = [start, end]
        if school_id:
            sql += " AND c.school_id = %s"
            params.append(school_id)
        sql += " ORDER BY r.released_at DESC LIMIT %s"
        params.append(MAX_RELEASE_ROWS + 1)
        rows = db.execute(sql, tuple(params)).fetchall()
    finally:
        db.close()

    truncated = len(rows) > MAX_RELEASE_ROWS
    rows = rows[:MAX_RELEASE_ROWS]

    return jsonify({
        'releases': [{
            'id': r['id'],
            'child_id': r['child_id'],
            'child_name': r['child_name'],
            'school': r['school'],
            'school_id': r['school_id'],
            'person_name': r['person_name'],
            'person_relationship': r['person_relationship'],
            'counselor_name': r['counselor_name'],
            'release_date': r['release_date'],
            'released_at': iso_utc(r['released_at']),
            'checked_in_at': iso_utc(r['checked_in_at']),
            # NULL is not 6pm — see dismissal_clock. An organization that
            # doesn't use dismissal times, or a child released on a weekday
            # they aren't registered for, has nothing to compare against, and
            # the client has to say so rather than guess.
            'dismissal_time': r['dismissal_time'],
            'has_signature': r['has_signature'],
        } for r in rows],
        'truncated': truncated,
        'start': start,
        'end': end,
    })


@app.route('/api/admin/pickup-releases/<int:release_id>/signature', methods=['GET'])
@jwt_required()
def admin_get_pickup_signature(release_id):
    """One signature, as a data URL inside JSON.

    Not served as raw image bytes on purpose. An <img src> cannot send the
    Authorization header, so a route that returned the image directly would
    have to be readable without a token — which would put every family's
    signature behind nothing but an guessable integer.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    db = get_db()
    try:
        row = db.execute("""
            SELECT r.signature, r.signature_mime, r.person_name,
                   r.person_relationship, r.counselor_name, r.released_at,
                   c.name AS child_name
              FROM pickup_releases r
              JOIN children c ON c.id = r.child_id
             WHERE r.id = %s
        """, (release_id,)).fetchone()
    finally:
        db.close()

    if not row:
        return jsonify({'error': 'Not found'}), 404

    signature = None
    if row['signature'] is not None:
        mime = row['signature_mime'] or 'image/png'
        encoded = base64.b64encode(bytes(row['signature'])).decode('ascii')
        signature = f'data:{mime};base64,{encoded}'

    return jsonify({
        'id': release_id,
        'child_name': row['child_name'],
        'person_name': row['person_name'],
        'person_relationship': row['person_relationship'],
        'counselor_name': row['counselor_name'],
        'released_at': iso_utc(row['released_at']),
        # None for releases recorded before signatures existed. The client says
        # so rather than rendering an empty frame.
        'signature': signature,
    })


@app.route('/api/admin/export/pickup-releases', methods=['GET'])
@jwt_required()
def export_pickup_releases():
    """The release log as a spreadsheet.

    Carries the metadata only. Signatures stay in the app: an image floating
    over a cell detaches from its row the moment anyone sorts, which is the
    opposite of what a record like this is for.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    now = now_for_org()
    end = request.args.get('end', now.strftime('%Y-%m-%d'))
    start = request.args.get(
        'start',
        (now - timedelta(days=30)).strftime('%Y-%m-%d'),
    )
    try:
        start = parse_date(start)
        end = parse_date(end)
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400
    if end < start:
        return jsonify({'error': 'end must be >= start'}), 400

    db = get_db()
    try:
        rows = db.execute("""
            SELECT c.name AS child_name, s.name AS school,
                   r.person_name, r.person_relationship, r.counselor_name,
                   r.release_date, r.released_at, a.checked_in_at,
                   reg.dismissal_time,
                   r.signature IS NOT NULL AS has_signature
              FROM pickup_releases r
              JOIN children c ON c.id = r.child_id
              JOIN schools s ON s.id = c.school_id
              LEFT JOIN attendance_records a
                     ON a.child_id = r.child_id
                    AND a.attendance_date = r.release_date
              LEFT JOIN registrations reg
                     ON reg.child_id = r.child_id
                    AND reg.day_of_week = to_char(r.release_date::date, 'FMDay')
             WHERE r.release_date BETWEEN %s AND %s
             ORDER BY r.release_date, s.name, c.name
        """, (start, end)).fetchall()
        tzname = org_timezone(db)
    finally:
        db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pickups"
    ws.append(["Date", "Child Name", "School", "Picked Up By", "Relationship",
               "Checked In", "Expected", "Checked Out", "Late", "Confirmed By", "Signed"])
    style_header_row(ws, 1)
    for r in rows:
        # Same conversion local_clock does, kept alongside it rather than
        # factored out for a time() comparison that string formatting alone
        # can't give — see PickupLog's isLate() for the client-side twin.
        expected = daily_routing.dismissal_clock(r['dismissal_time'])
        late = ''
        if expected and r['released_at']:
            released = r['released_at']
            if released.tzinfo is None:
                released = released.replace(tzinfo=timezone.utc)
            released = released.astimezone(ZoneInfo(tzname or DEFAULT_TIMEZONE))
            late = 'Yes' if released.time() > expected else 'No'
        ws.append([
            r['release_date'], r['child_name'], r['school'],
            r['person_name'], r['person_relationship'] or '',
            local_clock(r['checked_in_at'], tzname),
            expected.strftime('%I:%M %p').lstrip('0') if expected else '',
            local_clock(r['released_at'], tzname),
            late,
            r['counselor_name'], 'Yes' if r['has_signature'] else 'No',
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or '')) for c in col) + 4
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"kikar_pickups_{start}_to_{end}.xlsx"
    return Response(
        out.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment;filename={fname}"},
    )


# ─── PICKUP NOTIFICATIONS ──────────────────────────────────────────────────

@app.route('/api/parent/pickup', methods=['POST'])
@jwt_required()
def parent_pickup_notify():
    claims = get_jwt()
    if claims.get('role') != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    parent_name = claims.get('name', 'A parent')
    data = request.json
    child_id = data.get('child_id')

    if not child_id:
        return jsonify({'error': 'Missing child_id'}), 400

    db = get_db()

    # Verify the child belongs to this parent
    child = db.execute(
        f"SELECT c.id, c.name FROM children c "
        f"WHERE c.id = %s AND {_CHILD_BELONGS_TO_USER} AND c.active = 1",
        (child_id, user_id, user_id)
    ).fetchone()
    if not child:
        db.close()
        return jsonify({'error': 'Child not found'}), 404

    child_name = child['name']

    # Idempotency window: if the parent already has an unclaimed pickup for this
    # child in the last 30 minutes, return success without inserting a new row,
    # re-firing pushes, or re-publishing on the SSE bus. Stops a double-tap (or
    # the parent reopening the app and tapping again) from creating duplicate
    # cards on the counselor side.
    existing = db.execute(
        """
        SELECT id FROM pickup_notifications
        WHERE child_id = %s AND parent_id = %s AND claimed_by IS NULL
          AND created_at > NOW() - INTERVAL '30 minutes'
        LIMIT 1
        """,
        (child_id, user_id)
    ).fetchone()
    if existing:
        db.close()
        return jsonify({'message': f'Staff have already been notified that you are here to pick up {child_name}.'})

    # Record the pickup notification and grab the new row id so we can publish
    # it on the realtime bus.
    new_row = db.execute(
        "INSERT INTO pickup_notifications (child_id, parent_id, parent_name, child_name) VALUES (%s, %s, %s, %s) RETURNING id, created_at",
        (child_id, user_id, parent_name, child_name)
    ).fetchone()
    db.commit()

    pickup_events.publish('created', {
        'id': new_row['id'],
        'child_id': child_id,
        'parent_name': parent_name,
        'child_name': child_name,
        'created_at': iso_utc(new_row['created_at']),
        'claimed_by': None,
        'claimed_by_name': None,
        'claimed_at': None,
    }, current_org_id())

    # Send push notifications to all admins and counselors
    staff = db.execute(
        "SELECT id FROM users WHERE role IN ('admin', 'counselor')"
    ).fetchall()
    db.close()

    notification_title = 'Carpool Pickup Alert'
    notification_body = f'{parent_name} has arrived to pick up {child_name}.'

    # Fire push in background threads — the SSE 'created' event above already
    # reached every open counselor portal instantly. Blocking the parent's
    # response here (one pywebpush call per staff subscription) is what made
    # the "I'm here" button feel slow and made pickups look delayed.
    # Stable tag per pickup so dedupe / future "clear" pushes can target it.
    pickup_tag = f'pickup-{new_row["id"]}'
    for staff_member in staff:
        send_push_to_user_async(
            staff_member['id'],
            notification_title,
            notification_body,
            '/app/today',
            tag=pickup_tag,
            pickup_id=new_row['id'],
        )

    return jsonify({'message': f'Staff have been notified that you are here to pick up {child_name}.'})


@app.route('/api/counselor/pickup-alerts', methods=['GET'])
@jwt_required()
def counselor_get_pickup_alerts():
    """Return every unclaimed pickup, earliest first.

    Visibility IS the accountability signal: if a pickup shows up here, nobody
    has claimed it yet. Once any counselor/admin claims it, it disappears from
    this feed for everyone.
    """
    claims = get_jwt()
    if claims.get('role') not in ('counselor', 'admin'):
        return jsonify({'error': 'Unauthorized'}), 403

    db = get_db()
    pickups = db.execute("""
        SELECT id, child_id, parent_name, child_name, created_at,
               claimed_by, claimed_by_name, claimed_at
        FROM pickup_notifications
        WHERE claimed_by IS NULL
        ORDER BY created_at ASC
    """).fetchall()
    db.close()
    return jsonify([pickup_row_to_dict(p) for p in pickups])


@app.route('/api/pickups/<int:pickup_id>/claim', methods=['POST'])
@jwt_required()
def claim_pickup(pickup_id):
    """Atomically claim an unclaimed pickup. Race-safe via WHERE claimed_by IS NULL."""
    claims = get_jwt()
    if claims.get('role') not in ('counselor', 'admin'):
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    user_name = claims.get('name', 'Staff')

    db = get_db()
    # The WHERE clause guarantees only one caller wins — second caller gets 0 rows.
    row = db.execute("""
        UPDATE pickup_notifications
        SET claimed_by = %s, claimed_by_name = %s, claimed_at = NOW()
        WHERE id = %s AND claimed_by IS NULL
        RETURNING id, child_id, parent_id, parent_name, child_name, created_at,
                  claimed_by, claimed_by_name, claimed_at
    """, (user_id, user_name, pickup_id)).fetchone()

    if not row:
        # Already claimed — surface the current claimer so the UI can say who.
        existing = db.execute("""
            SELECT claimed_by, claimed_by_name
            FROM pickup_notifications WHERE id = %s
        """, (pickup_id,)).fetchone()
        db.close()
        if not existing:
            return jsonify({'error': 'Pickup not found'}), 404
        return jsonify({
            'error': 'Already claimed',
            'claimed_by': existing['claimed_by'],
            'claimed_by_name': existing['claimed_by_name'],
        }), 409

    db.execute("""
        INSERT INTO pickup_claim_audit (pickup_id, user_id, user_name, action)
        VALUES (%s, %s, %s, 'claim')
    """, (pickup_id, user_id, user_name))
    db.commit()
    db.close()

    payload = pickup_row_to_dict(row)
    pickup_events.publish('claimed', payload, current_org_id())

    # Notify the parent that a specific counselor is now bringing their child.
    # Separate event type so /api/parent/stream can route it to that parent
    # instead of having every parent listen to the staff-facing 'claimed' feed
    # — "that parent" now being whichever of the two linked accounts, not
    # only whoever tapped "I'm here" (payload['parent_id']).
    notify_db = get_db()
    try:
        owner_ids = child_owner_ids(notify_db, [payload['child_id']])
    finally:
        notify_db.close()
    for owner_id in owner_ids:
        pickup_events.publish('pickup-claimed', {
            'pickup_id': payload['id'],
            'child_id': payload['child_id'],
            'child_name': payload['child_name'],
            'parent_id': owner_id,
            'counselor_name': payload['claimed_by_name'],
            'claimed_at': payload['claimed_at'],
        }, current_org_id())
        notify_parent(
            owner_id,
            'pickup_claimed',
            title=f"{payload['claimed_by_name']} is on the way",
            body=f"{payload['claimed_by_name']} is bringing {payload['child_name']} to you now",
            url='/app/home',
            tag=f"pickup-claimed-{payload['id']}",
            pickup_id=payload['id'],
        )
    return jsonify(payload)


@app.route('/api/pickups/<int:pickup_id>/unclaim', methods=['POST'])
@jwt_required()
def unclaim_pickup(pickup_id):
    """Release a claim so the pickup reappears for everyone. Only the current
    claimer may unclaim."""
    claims = get_jwt()
    if claims.get('role') not in ('counselor', 'admin'):
        return jsonify({'error': 'Unauthorized'}), 403

    user_id = get_jwt_identity()
    user_name = claims.get('name', 'Staff')

    db = get_db()
    row = db.execute("""
        UPDATE pickup_notifications
        SET claimed_by = NULL, claimed_by_name = NULL, claimed_at = NULL
        WHERE id = %s AND claimed_by = %s
        RETURNING id, child_id, parent_id, parent_name, child_name, created_at
    """, (pickup_id, user_id)).fetchone()

    if not row:
        db.close()
        return jsonify({'error': 'Not your claim'}), 403

    db.execute("""
        INSERT INTO pickup_claim_audit (pickup_id, user_id, user_name, action)
        VALUES (%s, %s, %s, 'unclaim')
    """, (pickup_id, user_id, user_name))
    db.commit()
    db.close()

    payload = pickup_row_to_dict(row)
    payload.update({'claimed_by': None, 'claimed_by_name': None, 'claimed_at': None})
    pickup_events.publish('unclaimed', payload, current_org_id())

    # Tell the parent's portal to revert the "is bringing your child" UI back
    # to the generic "Staff have been notified" state. No push — a silent undo
    # would be confusing if it landed on the lock screen. Both linked
    # accounts, same as the claim event above.
    notify_db = get_db()
    try:
        owner_ids = child_owner_ids(notify_db, [payload['child_id']])
    finally:
        notify_db.close()
    for owner_id in owner_ids:
        pickup_events.publish('pickup-unclaimed', {
            'pickup_id': payload['id'],
            'child_id': payload['child_id'],
            'child_name': payload['child_name'],
            'parent_id': owner_id,
        }, current_org_id())
    return jsonify(payload)


@app.route('/api/pickups/stream', methods=['GET'])
def pickups_stream():
    """Server-Sent Events feed for pickup changes.

    EventSource can't set Authorization headers, so the JWT is passed as a
    query parameter. The token is validated manually below.
    """
    token = request.args.get('token', '')
    try:
        decoded = decode_token(token)
    except Exception:
        return jsonify({'error': 'Invalid token'}), 401

    role = decoded.get('role') or (decoded.get('user_claims') or {}).get('role')
    if role not in ('counselor', 'admin'):
        return jsonify({'error': 'Unauthorized'}), 403

    # EventSource can't send headers, so before_request never saw this token
    # and no organization is bound yet. Bind it by hand before touching the
    # database, or the snapshot below reads nothing.
    from flask import g
    org_id = decoded.get('org') or (decoded.get('user_claims') or {}).get('org')
    if org_id is None:
        return jsonify({'error': 'Token carries no organization'}), 403
    g.organization_id = org_id
    g.is_superadmin = False

    # _enforce_module_access saw no organization on this request (the token was
    # in the query string), so it let the request through. Now that the token
    # is decoded, apply the same rule by hand.
    if not tenancy.module_enabled(_request_modules(), 'pickups'):
        return jsonify({'error': 'This feature is not enabled for your organization'}), 403

    # Grab initial snapshot BEFORE subscribing so we don't miss any events
    # published between the snapshot read and subscription.
    q = pickup_events.subscribe(org_id)
    try:
        db = get_db()
        snapshot = db.execute("""
            SELECT id, child_id, parent_name, child_name, created_at,
                   claimed_by, claimed_by_name, claimed_at
            FROM pickup_notifications
            WHERE claimed_by IS NULL
            ORDER BY created_at ASC
        """).fetchall()
        db.close()
        snapshot_list = [pickup_row_to_dict(p) for p in snapshot]
    except Exception:
        pickup_events.unsubscribe(q, org_id)
        raise

    def generate():
        try:
            yield pickup_events.format_sse({'type': 'snapshot', 'data': snapshot_list})
            while True:
                try:
                    # 15s gives a comfortable margin under the 30s idle-close
                    # most proxies (Render/Cloudflare) apply, so the connection
                    # stays warm and events get flushed promptly.
                    event = q.get(timeout=15)
                    yield pickup_events.format_sse(event)
                except Exception:
                    yield pickup_events.format_sse({'type': 'heartbeat', 'data': {}})
        finally:
            pickup_events.unsubscribe(q, org_id)

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection': 'keep-alive',
        },
    )


@app.route('/api/parent/stream', methods=['GET'])
def parent_stream():
    """Server-Sent Events feed for one parent.

    Delivers the 'parent-notification' event (counselor asked "is your child
    coming today?") the instant it's published, so parents with the portal
    open don't need to refresh to see the attendance-confirmation card. The
    same single pickup_events bus is used, with server-side filtering by
    parent_id so no parent ever sees another parent's events.
    """
    token = request.args.get('token', '')
    try:
        decoded = decode_token(token)
    except Exception:
        return jsonify({'error': 'Invalid token'}), 401

    role = decoded.get('role') or (decoded.get('user_claims') or {}).get('role')
    if role != 'parent':
        return jsonify({'error': 'Unauthorized'}), 403

    try:
        parent_id = int(decoded.get('sub'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid token'}), 401

    # Same as the pickups stream: the token arrived as a query parameter, so
    # the organization has to be read off it by hand.
    org_id = decoded.get('org') or (decoded.get('user_claims') or {}).get('org')
    if org_id is None:
        return jsonify({'error': 'Token carries no organization'}), 403

    q = pickup_events.subscribe(org_id)

    def generate():
        try:
            # Prime the connection so EventSource's readyState flips to OPEN
            # immediately — otherwise the client waits on the first event.
            yield pickup_events.format_sse({'type': 'ready', 'data': {}})
            while True:
                try:
                    # See pickups_stream: 15s keeps the connection under the
                    # ~30s idle-close window common on hosted proxies.
                    event = q.get(timeout=15)
                except Exception:
                    yield pickup_events.format_sse({'type': 'heartbeat', 'data': {}})
                    continue
                # Bus events for other parents / other event types still need
                # to keep this connection warm — emit a heartbeat instead of
                # silently dropping, otherwise during busy carpool periods
                # (lots of 'created'/'claimed' events) the proxy can idle-
                # close the SSE stream because no bytes flow on the wire.
                data = event.get('data') or {}
                # A payload too large for NOTIFY comes through as 'resync' with
                # no parent_id, so it cannot be filtered by recipient. Forward
                # it to everyone: it carries nothing, and its whole job is to
                # tell a client to refetch. Dropping it would leave the one
                # case that needed it — a long message — as the one that does
                # not arrive.
                if event.get('type') == 'resync':
                    yield pickup_events.format_sse(event)
                    continue
                # Forward only events targeted at THIS parent. Everything else
                # (other parents' notifications, staff-only events like
                # 'created'/'claimed') gets a heartbeat to keep the SSE warm.
                allowed_types = ('parent-notification', 'pickup-claimed',
                                 'pickup-unclaimed', 'conversation-message')
                if event.get('type') not in allowed_types or data.get('parent_id') != parent_id:
                    yield pickup_events.format_sse({'type': 'heartbeat', 'data': {}})
                    continue
                yield pickup_events.format_sse(event)
        finally:
            pickup_events.unsubscribe(q, org_id)

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection': 'keep-alive',
        },
    )


@app.route('/api/admin/conversations/stream', methods=['GET'])
def admin_conversations_stream():
    """Server-Sent Events feed for the admin inbox.

    Its own endpoint rather than a share of /api/pickups/stream, because that
    one is gated on the `pickups` module: an organization that bought parent
    messaging and not the pickup queue would be refused its own inbox.

    WHY THIS IS THE ONLY NEW CONNECTION IN THIS CHANGE. Every open SSE holds a
    gunicorn thread for its entire life and there are 64 of them, so a stream
    per parent would be the expensive half of live messaging — parents are the
    many. Parents get their messages down /api/parent/stream, which already
    exists and cost nothing to extend. Admins are a handful per JCC, and only
    while the inbox is on screen, so a connection each is affordable.

    Unlike the parent feed this does not filter by recipient: an admin can read
    every thread in their organization from this screen anyway, so a message to
    any of them belongs on all of their screens.

    Per §8 of CLAUDE.md: EventSource cannot send an Authorization header, so
    _bind_tenant and _enforce_module_access never saw a token. The decode, the
    organization pin and the module check are all repeated by hand below, and a
    stream added after this one has to repeat them too.
    """
    token = request.args.get('token', '')
    try:
        decoded = decode_token(token)
    except Exception:
        return jsonify({'error': 'Invalid token'}), 401

    role = decoded.get('role') or (decoded.get('user_claims') or {}).get('role')
    if role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    from flask import g
    org_id = decoded.get('org') or (decoded.get('user_claims') or {}).get('org')
    if org_id is None:
        return jsonify({'error': 'Token carries no organization'}), 403
    g.organization_id = org_id
    g.is_superadmin = False

    # Either conversation module is enough to open this connection: it now
    # carries both parent and staff threads, and an organization can buy one
    # without the other.
    if not (tenancy.module_enabled(_request_modules(), 'parent_messaging')
            or tenancy.module_enabled(_request_modules(), 'staff_messaging')):
        return jsonify({'error': 'This feature is not enabled for your organization'}), 403

    q = pickup_events.subscribe(org_id)

    def generate():
        try:
            # Prime the connection so EventSource flips to OPEN immediately
            # rather than waiting on the first real message.
            yield pickup_events.format_sse({'type': 'ready', 'data': {}})
            while True:
                try:
                    # 15s, under the ~30s idle-close most proxies apply.
                    event = q.get(timeout=15)
                except Exception:
                    yield pickup_events.format_sse({'type': 'heartbeat', 'data': {}})
                    continue
                # This organization's bus also carries pickups and headcounts.
                # They keep the connection warm as a heartbeat rather than
                # being forwarded to a screen that has no use for them.
                if event.get('type') not in ('conversation-message', 'staff-message', 'resync'):
                    yield pickup_events.format_sse({'type': 'heartbeat', 'data': {}})
                    continue
                yield pickup_events.format_sse(event)
        finally:
            pickup_events.unsubscribe(q, org_id)

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection': 'keep-alive',
        },
    )


# ─── END OF SCHOOL YEAR ─────────────────────────────────────────────────────

@app.route('/api/admin/end-of-year', methods=['POST'])
@jwt_required()
def admin_end_of_year():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    confirm = request.json.get('confirm', '')
    if confirm != 'END OF YEAR':
        return jsonify({'error': 'Invalid confirmation text'}), 400

    db = get_db()
    # Count what will be cleared for the summary
    children_count = db.execute("SELECT COUNT(*) as cnt FROM children").fetchone()['cnt']
    absences_count = db.execute("SELECT COUNT(*) as cnt FROM absences").fetchone()['cnt']
    attendance_count = db.execute("SELECT COUNT(*) as cnt FROM attendance_records").fetchone()['cnt']

    # Delete all children (cascades to registrations, absences, recurring_absences,
    # absence_exceptions, attendance_records, parent_notifications) — keeps users,
    # schools, counselor_schools, calendar_events
    db.execute("DELETE FROM attendance_records")
    db.execute("DELETE FROM absence_exceptions")
    db.execute("DELETE FROM recurring_absences")
    db.execute("DELETE FROM absences")
    db.execute("DELETE FROM parent_notifications")
    db.execute("DELETE FROM pickup_notifications")
    db.execute("DELETE FROM registrations")
    db.execute("DELETE FROM children")
    db.commit()
    db.close()

    return jsonify({
        'message': 'School year ended successfully. All children and registrations have been cleared.',
        'cleared': {
            'children': children_count,
            'absences': absences_count,
            'attendance_records': attendance_count
        }
    })

# ─── WIPE ALL DATA (TESTING) ────────────────────────────────────────────────

@app.route('/api/admin/wipe-all-data', methods=['POST'])
@jwt_required()
def admin_wipe_all_data():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.json or {}
    confirm = data.get('confirm', '')
    password = data.get('password', '')

    if confirm != 'WIPE ALL DATA':
        return jsonify({'error': "Invalid confirmation text. Type 'WIPE ALL DATA' exactly."}), 400

    admin_id = get_jwt_identity()

    # Verify the admin's password before wiping anything
    db_check = get_db()
    admin_row = db_check.execute(
        "SELECT password_hash FROM users WHERE id=%s AND role='admin'",
        (admin_id,)
    ).fetchone()
    db_check.close()

    if not admin_row or not bcrypt.checkpw(password.encode(), admin_row['password_hash'].encode()):
        return jsonify({'error': 'Invalid admin password'}), 401

    db = get_db_transaction()
    try:
        counts = {
            'parents': db.execute("SELECT COUNT(*) AS c FROM users WHERE role='parent'").fetchone()['c'],
            'counselors': db.execute("SELECT COUNT(*) AS c FROM users WHERE role='counselor'").fetchone()['c'],
            'children': db.execute("SELECT COUNT(*) AS c FROM children").fetchone()['c'],
            'activities': db.execute("SELECT COUNT(*) AS c FROM activities").fetchone()['c'],
            'calendar_events': db.execute("SELECT COUNT(*) AS c FROM calendar_events").fetchone()['c'],
        }

        db.execute("DELETE FROM attendance_records")
        db.execute("DELETE FROM absence_exceptions")
        db.execute("DELETE FROM recurring_absences")
        db.execute("DELETE FROM absences")
        db.execute("DELETE FROM parent_notifications")
        db.execute("DELETE FROM pickup_claim_audit")
        db.execute("DELETE FROM pickup_notifications")
        db.execute("DELETE FROM registrations")
        db.execute("DELETE FROM children")
        db.execute("DELETE FROM activity_roster")
        db.execute("DELETE FROM activity_schedules")
        db.execute("DELETE FROM activities")
        db.execute("DELETE FROM calendar_events")
        db.execute("DELETE FROM counselor_schools")
        db.execute("DELETE FROM invitations")
        db.execute("DELETE FROM password_reset_tokens")
        db.execute("DELETE FROM push_subscriptions WHERE user_id NOT IN (SELECT id FROM users WHERE role='admin')")
        db.execute("DELETE FROM user_totp WHERE user_id NOT IN (SELECT id FROM users WHERE role='admin')")
        db.execute("DELETE FROM users WHERE role <> 'admin'")
        db.execute("DELETE FROM login_attempts")

        db.commit()
        db.close()

        return jsonify({
            'message': 'All testing data wiped. Admin accounts and schools preserved.',
            'cleared': counts
        })
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        db.close()
        print(f"[ERROR] wipe-all-data transaction failed: {e}")
        return jsonify({'error': 'Wipe failed. No changes were made.'}), 500

# ─── ACTIVITY ROSTERS ───────────────────────────────────────────────────────

import re as _re

def _normalize_action(s):
    if not s: return None
    s = str(s).strip().lower()
    if s.startswith('drop'): return 'Drop Off'
    if s.startswith('pick'): return 'Pick Up'
    return None

def _normalize_days(s):
    """Return list of canonical day names from a messy day string."""
    if not s: return []
    typos = {'thrusday': 'Thursday', 'thurday': 'Thursday', 'wenesday': 'Wednesday'}
    canonical = {'monday','tuesday','wednesday','thursday','friday'}
    parts = _re.split(r'[;,\s\-\(\)]+', str(s).strip())
    result = []
    for p in parts:
        p = p.strip().lower()
        p = typos.get(p, p)
        if p in canonical:
            day = p.capitalize()
            if day not in result:
                result.append(day)
    return result

def _extract_days_from_pipe_parts(parts):
    """Days may be in parts[4] (canonical) or embedded in parts[3] (class name
    like 'Jazz I - 1st grade - TUESDAY'). Try both."""
    days_raw = parts[4] if len(parts) > 4 else ''
    days = _normalize_days(days_raw)
    if not days and len(parts) > 3:
        days = _normalize_days(parts[3])
    return days

def _normalize_option(s):
    if not s: return 'Full Service'
    if 'Transportation' in str(s): return 'Transportation Only'
    if 'Walkover' in str(s): return 'Walkover Only'
    return 'Full Service'

def _normalize_category(s):
    if not s: return s
    fixes = {'swmming': 'Swimming', 'swimning': 'Swimming', 'basketbal': 'Basketball'}
    return fixes.get(str(s).strip().lower(), str(s).strip())

def _time_to_str(t):
    """Convert datetime.time or HH:MM:SS string to HH:MM string."""
    if t is None: return None
    if hasattr(t, 'strftime'): return t.strftime('%H:%M')
    s = str(t).strip()
    if len(s) >= 5: return s[:5]
    return s

def _parse_time_12h(s):
    """Convert '4:15 PM' or '4:15PM' to '16:15'. Returns None if unparseable."""
    if not s: return None
    s = str(s).strip()
    m = _re.match(r'(\d{1,2}):(\d{2})\s*(AM|PM)', s, _re.IGNORECASE)
    if not m: return None
    h, mn, ampm = int(m.group(1)), int(m.group(2)), m.group(3).upper()
    if ampm == 'PM' and h != 12: h += 12
    elif ampm == 'AM' and h == 12: h = 0
    return f'{h:02d}:{mn:02d}'

def _try_auto_assign(db, counselor_name):
    """Try to find a counselor user whose name starts with the given first name."""
    if not counselor_name: return None
    first = counselor_name.strip().split()[0].lower()
    users = db.execute("SELECT id, name FROM users WHERE role='counselor'").fetchall()
    for u in users:
        if u['name'].strip().lower().startswith(first):
            return u['id']
    return None


@app.route('/api/admin/activity-roster/upload', methods=['POST'])
@jwt_required()
def admin_upload_activity_roster():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"[ERROR] activity roster upload read: {e}")
        return jsonify({'error': 'Could not read file'}), 400

    db = get_db()
    results = {'rows_imported': 0, 'activities_found': 0, 'kept_unchanged': 0,
               'need_counselor': 0, 'removed': 0, 'added': 0, 'updated': 0, 'errors': []}

    # ── Detect file format ──────────────────────────────────────────────────
    # New JCC export format: col[1]=child name, col[11]=program name, col[12]=pipe-delimited option
    # Old custom format: col[0]=child name, col[7]=activity, col[8]=days, col[9]=action, col[10]=time
    header_row = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    is_new_format = (
        len(header_row) >= 13 and
        'Full Name' in header_row[1] and
        'Program Name' in header_row[11]
    )
    is_sf_report_format = (
        not is_new_format and
        len(header_row) >= 4 and
        'Full Name' in str(header_row[1]) and
        'Full Course Option Name' in str(header_row[3])
    )
    # Compact Salesforce export: only Full Name | Phone | Full Course Option Name
    is_sf_compact_format = (
        not is_new_format and not is_sf_report_format and
        len(header_row) >= 3 and
        'Full Name' in str(header_row[0]) and
        'Full Course Option Name' in str(header_row[2])
    )

    # Parse all rows from Excel
    parsed = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        try:
            if is_new_format:
                # ── NEW JCC EXPORT FORMAT ────────────────────────────────────
                # col[1]=child name, col[5]=school, col[11]=program name, col[12]=pipe option
                child_name = str(row[1]).strip() if row[1] else ''
                if not child_name:
                    continue
                program_name = str(row[11]).strip() if row[11] else ''
                if not program_name or program_name.upper() == 'JCLUB':
                    continue  # Skip JClub attendance rows and empty rows
                school = str(row[5]).strip() if row[5] else ''
                full_option = str(row[12]).strip() if row[12] else ''

                # ── FILTER 1: skip programs JClub doesn't supervise ──────────
                SKIP_PROGRAMS = {
                    'hebraica', 'nitzanim', 'early childhood academy',
                    'special needs (or gadol)', 'the hub', 'jclub'
                }
                act_name = _normalize_category(program_name)
                if act_name.lower() in SKIP_PROGRAMS:
                    continue

                # ── FILTER 2: child must be active in JClub ──────────────────
                child_row = db.execute("""
                    SELECT c.service_type FROM children c
                    WHERE LOWER(TRIM(c.name)) = LOWER(TRIM(%s)) AND c.active = 1
                    LIMIT 1
                """, (child_name,)).fetchone()
                if not child_row:
                    continue  # Child not in JClub roster — skip

                service_type = child_row['service_type']

                # Parse pipe-delimited: Program|Subcategory|Season|ClassName|Days; |Time
                parts = [p.strip() for p in full_option.split('|')]
                sub_type = parts[1] if len(parts) > 1 else ''
                season   = parts[2] if len(parts) > 2 else ''
                time_raw = parts[5] if len(parts) > 5 else ''

                days = _extract_days_from_pipe_parts(parts)
                if not days:
                    continue  # Skip tournament/travel rows with no scheduled days

                time_str = _parse_time_12h(time_raw)

                # ── Transportation Only / Walkover Only ───────────────────────
                # JClub DROPS them off only — no Pick Up.
                # Only show if activity time is at or before 4:30 PM.
                if service_type in ('Transportation Only', 'Walkover Only'):
                    if not time_str or time_str > '16:30':
                        continue  # Activity after 4:30 PM — parents handle it directly
                    for day in days:
                        parsed.append({
                            'child_name': child_name, 'school': school,
                            'option_type': service_type, 'grade_group': '',
                            'category': act_name, 'sub_type': sub_type,
                            'season': season, 'act_name': act_name,
                            'day': day, 'action': 'Drop Off',
                            'time': time_str, 'counselor_name': '',
                        })
                    continue  # Done with this child — no Pick Up entries

                # ── Full Service ──────────────────────────────────────────────
                # ── FILTER 3: skip activities ending after 5:30 PM ───────────
                if time_str and time_str > '17:30':
                    continue

                # For each day, create Pick Up entry (file time = end-of-activity)
                # Also create Drop Off if Activity Schedules defines one for this activity+day
                for day in days:
                    sched = db.execute("""
                        SELECT dropoff_time, pickup_time FROM activity_schedules
                        WHERE %s ILIKE '%%' || activity_name_pattern || '%%' AND day_of_week=%s
                        LIMIT 1
                    """, (act_name, day)).fetchone()

                    pu_time = (sched['pickup_time'] if sched else None) or time_str
                    do_time = sched['dropoff_time'] if sched else None

                    if pu_time and pu_time > '17:30':
                        pu_time = None

                    parsed.append({
                        'child_name': child_name, 'school': school,
                        'option_type': 'Full Service', 'grade_group': '',
                        'category': act_name, 'sub_type': sub_type,
                        'season': season, 'act_name': act_name,
                        'day': day, 'action': 'Pick Up',
                        'time': pu_time, 'counselor_name': '',
                    })
                    if do_time:
                        parsed.append({
                            'child_name': child_name, 'school': school,
                            'option_type': 'Full Service', 'grade_group': '',
                            'category': act_name, 'sub_type': sub_type,
                            'season': season, 'act_name': act_name,
                            'day': day, 'action': 'Drop Off',
                            'time': do_time, 'counselor_name': '',
                        })
            elif is_sf_report_format:
                # ── SALESFORCE REPORT EXCEL FORMAT ───────────────────────────
                # col[1]=child name, col[3]=pipe-delimited option, col[6]=school
                # col[12]=start time (12h string), col[13]=end time ("17:15:00.000")
                child_name = str(row[1]).strip() if row[1] else ''
                if not child_name or child_name == 'None':
                    continue

                full_option = str(row[3]).strip() if row[3] else ''
                if not full_option:
                    continue

                school = str(row[6]).strip() if row[6] else ''

                parts = [p.strip() for p in full_option.split('|')]
                act_name = _normalize_category(parts[0]) if parts[0] else ''
                if not act_name:
                    continue

                SKIP_SF = {'hebraica', 'nitzanim', 'early childhood academy',
                           'special needs (or gadol)', 'the hub', 'jclub',
                           'bamachol costume fees', 'costume'}
                if act_name.lower() in SKIP_SF or any(s in act_name.lower() for s in ['fee', 'deposit', 'costume']):
                    continue

                sub_type  = parts[1] if len(parts) > 1 else ''
                season    = parts[2] if len(parts) > 2 else ''
                pipe_time = parts[5] if len(parts) > 5 else ''

                days = _extract_days_from_pipe_parts(parts)
                if not days:
                    continue

                child_row = db.execute("""
                    SELECT c.service_type FROM children c
                    WHERE LOWER(TRIM(c.name)) = LOWER(TRIM(%s)) AND c.active = 1
                    LIMIT 1
                """, (child_name,)).fetchone()
                if not child_row:
                    continue

                service_type = child_row['service_type']

                start_raw = row[12]
                end_raw   = row[13]

                start_str = (_parse_time_12h(str(start_raw)) if start_raw else None) or \
                            (_time_to_str(start_raw) if start_raw else None) or \
                            _parse_time_12h(pipe_time) or None

                if end_raw:
                    end_str = str(end_raw).strip()[:5] if isinstance(end_raw, str) else _time_to_str(end_raw)
                else:
                    end_str = None

                if service_type in ('Transportation Only', 'Walkover Only'):
                    if not start_str or start_str > '16:30':
                        continue
                    for day in days:
                        parsed.append({
                            'child_name': child_name, 'school': school,
                            'option_type': service_type, 'grade_group': '',
                            'category': act_name, 'sub_type': sub_type,
                            'season': season, 'act_name': act_name,
                            'day': day, 'action': 'Drop Off',
                            'time': start_str, 'counselor_name': '',
                        })
                else:
                    for day in days:
                        sched = db.execute("""
                            SELECT dropoff_time, pickup_time FROM activity_schedules
                            WHERE %s ILIKE '%%' || activity_name_pattern || '%%' AND day_of_week=%s
                            LIMIT 1
                        """, (act_name, day)).fetchone()

                        # Prefer schedule override; fall back to end time from file
                        pu_time = (sched['pickup_time'] if sched else None) or end_str
                        do_time = (sched['dropoff_time'] if sched else None) or start_str

                        # If pickup is past 5:30 PM and no schedule override, skip Pick Up —
                        # JClub only drops off; activity ends too late for pickup
                        if pu_time and pu_time > '17:30' and not (sched and sched['pickup_time']):
                            pu_time = None

                        if pu_time:
                            parsed.append({
                                'child_name': child_name, 'school': school,
                                'option_type': 'Full Service', 'grade_group': '',
                                'category': act_name, 'sub_type': sub_type,
                                'season': season, 'act_name': act_name,
                                'day': day, 'action': 'Pick Up',
                                'time': pu_time, 'counselor_name': '',
                            })
                        if do_time:
                            parsed.append({
                                'child_name': child_name, 'school': school,
                                'option_type': 'Full Service', 'grade_group': '',
                                'category': act_name, 'sub_type': sub_type,
                                'season': season, 'act_name': act_name,
                                'day': day, 'action': 'Drop Off',
                                'time': do_time, 'counselor_name': '',
                            })
            elif is_sf_compact_format:
                # ── COMPACT SALESFORCE EXPORT ────────────────────────────────
                # col[0]=child name, col[1]=phone, col[2]=pipe-delimited option
                # No school column and no separate start/end time columns —
                # rely on time embedded in the pipe (parts[5]).
                child_name = str(row[0]).strip() if row[0] else ''
                if not child_name or child_name == 'None':
                    continue

                full_option = str(row[2]).strip() if row[2] else ''
                if not full_option:
                    continue

                parts = [p.strip() for p in full_option.split('|')]
                act_name = _normalize_category(parts[0]) if parts[0] else ''
                if not act_name:
                    continue

                SKIP_SF = {'hebraica', 'nitzanim', 'early childhood academy',
                           'special needs (or gadol)', 'the hub', 'jclub',
                           'bamachol costume fees', 'costume'}
                lower_full = full_option.lower()
                if act_name.lower() in SKIP_SF or any(
                    s in lower_full for s in ['costume fee', 'deposit', 'costume - ']
                ):
                    continue

                sub_type  = parts[1] if len(parts) > 1 else ''
                season    = parts[2] if len(parts) > 2 else ''
                pipe_time = parts[5] if len(parts) > 5 else ''

                days = _extract_days_from_pipe_parts(parts)
                if not days:
                    continue

                child_row = db.execute("""
                    SELECT c.service_type FROM children c
                    WHERE LOWER(TRIM(c.name)) = LOWER(TRIM(%s)) AND c.active = 1
                    LIMIT 1
                """, (child_name,)).fetchone()
                if not child_row:
                    continue

                service_type = child_row['service_type']
                start_str = _parse_time_12h(pipe_time)

                if service_type in ('Transportation Only', 'Walkover Only'):
                    if not start_str or start_str > '16:30':
                        continue
                    for day in days:
                        parsed.append({
                            'child_name': child_name, 'school': '',
                            'option_type': service_type, 'grade_group': '',
                            'category': act_name, 'sub_type': sub_type,
                            'season': season, 'act_name': act_name,
                            'day': day, 'action': 'Drop Off',
                            'time': start_str, 'counselor_name': '',
                        })
                    continue

                if start_str and start_str > '17:30':
                    continue

                for day in days:
                    sched = db.execute("""
                        SELECT dropoff_time, pickup_time FROM activity_schedules
                        WHERE %s ILIKE '%%' || activity_name_pattern || '%%' AND day_of_week=%s
                        LIMIT 1
                    """, (act_name, day)).fetchone()

                    pu_time = (sched['pickup_time'] if sched else None) or start_str
                    do_time = sched['dropoff_time'] if sched else None

                    if pu_time and pu_time > '17:30':
                        pu_time = None

                    if pu_time:
                        parsed.append({
                            'child_name': child_name, 'school': '',
                            'option_type': 'Full Service', 'grade_group': '',
                            'category': act_name, 'sub_type': sub_type,
                            'season': season, 'act_name': act_name,
                            'day': day, 'action': 'Pick Up',
                            'time': pu_time, 'counselor_name': '',
                        })
                    if do_time:
                        parsed.append({
                            'child_name': child_name, 'school': '',
                            'option_type': 'Full Service', 'grade_group': '',
                            'category': act_name, 'sub_type': sub_type,
                            'season': season, 'act_name': act_name,
                            'day': day, 'action': 'Drop Off',
                            'time': do_time, 'counselor_name': '',
                        })
            else:
                # ── OLD CUSTOM FORMAT ────────────────────────────────────────
                # col[0]=child, col[1]=school, col[7]=activity, col[8]=days,
                # col[9]=action, col[10]=time, col[11]=counselor
                if not row[0]: continue
                child_name  = str(row[0]).strip()
                school      = str(row[1]).strip() if row[1] else ''
                option_type = _normalize_option(row[2])
                grade_group = str(row[3]).strip() if row[3] else ''
                category    = _normalize_category(row[4]) if row[4] else ''
                sub_type    = str(row[5]).strip() if row[5] else ''
                season      = str(row[6]).strip() if row[6] else ''
                act_name    = str(row[7]).strip() if row[7] else ''
                days_raw    = row[8]
                action_raw  = row[9]
                time_raw    = row[10]
                counselor_raw = str(row[11]).strip() if row[11] else ''

                action = _normalize_action(action_raw)
                if not action:
                    results['errors'].append(f'Row {i}: unknown action "{action_raw}"')
                    continue

                days = _normalize_days(days_raw)
                if not days:
                    results['errors'].append(f'Row {i}: no valid days in "{days_raw}"')
                    continue

                if option_type in ('Transportation Only', 'Walkover Only') and action == 'Pick Up':
                    continue

                time_str = _time_to_str(time_raw)
                for day in days:
                    parsed.append({
                        'child_name': child_name, 'school': school,
                        'option_type': option_type, 'grade_group': grade_group,
                        'category': category, 'sub_type': sub_type,
                        'season': season, 'act_name': act_name,
                        'day': day, 'action': action,
                        'time': time_str, 'counselor_name': counselor_raw,
                    })
        except Exception as e:
            results['errors'].append(f'Row {i}: {str(e)}')

    # Mark all existing excel-sourced rows for potential deletion
    db.execute("UPDATE activity_roster SET source='excel_pending' WHERE source='excel'")
    db.commit()

    activity_cache = {}
    for entry in parsed:
        act_name = entry['act_name']
        if act_name not in activity_cache:
            existing = db.execute("SELECT id FROM activities WHERE name=%s", (act_name,)).fetchone()
            if existing:
                activity_cache[act_name] = existing['id']
            else:
                cur = db.execute(
                    "INSERT INTO activities (name, category, sub_type, season) VALUES (%s,%s,%s,%s) RETURNING id",
                    (act_name, entry['category'], entry['sub_type'], entry['season'])
                )
                activity_cache[act_name] = cur.fetchone()['id']
                results['activities_found'] += 1

        activity_id = activity_cache[act_name]

        # Apply schedule override if time is missing
        time_str = entry['time']
        if not time_str:
            if entry['action'] == 'Drop Off':
                sched = db.execute("""
                    SELECT dropoff_time AS t FROM activity_schedules
                    WHERE %s ILIKE '%%' || activity_name_pattern || '%%' AND day_of_week=%s
                    LIMIT 1
                """, (act_name, entry['day'])).fetchone()
            else:
                sched = db.execute("""
                    SELECT pickup_time AS t FROM activity_schedules
                    WHERE %s ILIKE '%%' || activity_name_pattern || '%%' AND day_of_week=%s
                    LIMIT 1
                """, (act_name, entry['day'])).fetchone()
            if sched and sched['t']:
                time_str = sched['t']

        # Try auto-assign counselor from name
        auto_id = _try_auto_assign(db, entry['counselor_name'])

        # Check if this row already existed (match on child+activity+day+action)
        existing_row = db.execute("""
            SELECT id, assigned_counselor_id FROM activity_roster
            WHERE child_name=%s AND activity_id=%s AND day_of_week=%s AND action=%s
              AND source IN ('excel','excel_pending')
        """, (entry['child_name'], activity_id, entry['day'], entry['action'])).fetchone()

        if existing_row:
            # Keep existing counselor assignment if already set; otherwise try auto-assign
            keep_counselor = existing_row['assigned_counselor_id'] or auto_id
            db.execute("""
                UPDATE activity_roster
                SET school=%s, option_type=%s, grade_group=%s, action_time=%s,
                    counselor_name=%s, assigned_counselor_id=%s, season=%s, source='excel'
                WHERE id=%s
            """, (entry['school'], entry['option_type'], entry['grade_group'],
                  time_str, entry['counselor_name'], keep_counselor, entry['season'],
                  existing_row['id']))
            results['updated'] += 1
            if not keep_counselor:
                results['need_counselor'] += 1
        else:
            db.execute("""
                INSERT INTO activity_roster
                (activity_id, child_name, school, option_type, grade_group,
                 action, day_of_week, action_time, counselor_name, assigned_counselor_id,
                 season, source)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'excel')
            """, (activity_id, entry['child_name'], entry['school'],
                  entry['option_type'], entry['grade_group'],
                  entry['action'], entry['day'], time_str,
                  entry['counselor_name'], auto_id, entry['season']))
            results['added'] += 1
            results['rows_imported'] += 1
            if not auto_id:
                results['need_counselor'] += 1

        db.commit()

    # Remove rows still marked pending (no longer in Excel)
    removed = db.execute("SELECT COUNT(*) as cnt FROM activity_roster WHERE source='excel_pending'").fetchone()['cnt']
    db.execute("DELETE FROM activity_roster WHERE source='excel_pending'")
    # Clean up activities that now have 0 roster entries (filtered out or removed)
    db.execute("""
        DELETE FROM activities
        WHERE id NOT IN (SELECT DISTINCT activity_id FROM activity_roster)
    """)
    db.commit()
    results['removed'] = removed
    results['activities_found'] = len(activity_cache)
    results['unassigned'] = results['need_counselor']

    db.close()
    return jsonify(results)


@app.route('/api/admin/activities', methods=['GET'])
@jwt_required()
def admin_get_activities():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    acts = db.execute("""
        SELECT a.id, a.name, a.category, a.sub_type, a.season,
               COUNT(ar.id) as total,
               SUM(CASE WHEN ar.assigned_counselor_id IS NULL THEN 1 ELSE 0 END) as unassigned
        FROM activities a
        JOIN activity_roster ar ON ar.activity_id = a.id
        GROUP BY a.id
        HAVING COUNT(ar.id) > 0
        ORDER BY a.name
    """).fetchall()
    db.close()
    return jsonify([dict(a) for a in acts])


@app.route('/api/admin/activity-roster/<int:activity_id>', methods=['GET'])
@jwt_required()
def admin_get_activity_roster(activity_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    day = request.args.get('day', '')
    action = request.args.get('action', '')
    counselor = request.args.get('counselor', '')

    query = """
        SELECT ar.id, ar.child_name, ar.school, ar.option_type, ar.grade_group,
               ar.action, ar.day_of_week, ar.action_time, ar.counselor_name,
               ar.assigned_counselor_id, ar.season, ar.source,
               to_char(ar.specific_date, 'YYYY-MM-DD') AS specific_date,
               u.name as assigned_counselor_name
        FROM activity_roster ar
        LEFT JOIN users u ON u.id = ar.assigned_counselor_id
        WHERE ar.activity_id = %s
          AND (ar.specific_date IS NULL OR ar.specific_date >= CURRENT_DATE)
    """
    params = [activity_id]
    if day:
        query += " AND ar.day_of_week = %s"; params.append(day)
    if action:
        query += " AND ar.action = %s"; params.append(action)
    if counselor == 'unassigned':
        query += " AND ar.assigned_counselor_id IS NULL"
    elif counselor:
        query += " AND ar.assigned_counselor_id = %s"; params.append(int(counselor))
    query += " ORDER BY ar.day_of_week, ar.action_time, ar.action, ar.child_name"

    db = get_db()
    rows = db.execute(query, params).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


def _cascade_makeup_counselor(db, recurring_ids, counselor_id):
    """When admin reassigns one or more recurring activity_roster rows,
    propagate the new counselor to every future parent-submitted make-up
    row that matches the same slot (activity, weekday, action, time) and
    isn't admin-locked."""
    if not recurring_ids:
        return 0
    cur = db.execute("""
        WITH slots AS (
          SELECT DISTINCT activity_id, day_of_week, action,
                          COALESCE(action_time,'') AS at
            FROM activity_roster
           WHERE id = ANY(%s) AND specific_date IS NULL
        )
        UPDATE activity_roster ar
           SET assigned_counselor_id = %s
          FROM slots s
         WHERE ar.source = 'parent_makeup'
           AND ar.counselor_locked = FALSE
           AND ar.specific_date >= CURRENT_DATE
           AND ar.activity_id = s.activity_id
           AND ar.day_of_week = s.day_of_week
           AND ar.action = s.action
           AND COALESCE(ar.action_time,'') = s.at
    """, (recurring_ids, counselor_id))
    return cur.rowcount


@app.route('/api/admin/activity-roster/<int:entry_id>/assign', methods=['PUT'])
@jwt_required()
def admin_assign_counselor(entry_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    counselor_id = request.json.get('counselor_id')  # None = unassign
    db = get_db()
    db.execute("UPDATE activity_roster SET assigned_counselor_id=%s WHERE id=%s",
               (counselor_id, entry_id))
    _cascade_makeup_counselor(db, [entry_id], counselor_id)
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/admin/activity-roster/bulk-assign', methods=['POST'])
@jwt_required()
def admin_bulk_assign_counselor():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    body = request.json or {}
    entry_ids = body.get('entry_ids')
    counselor_id = body.get('counselor_id')  # None = unassign
    if not isinstance(entry_ids, list) or not entry_ids:
        return jsonify({'error': 'entry_ids must be a non-empty list'}), 400
    try:
        ids = [int(x) for x in entry_ids]
    except (TypeError, ValueError):
        return jsonify({'error': 'entry_ids must contain integers'}), 400
    db = get_db()
    cur = db.execute(
        "UPDATE activity_roster SET assigned_counselor_id=%s WHERE id = ANY(%s)",
        (counselor_id, ids)
    )
    updated = cur.rowcount
    _cascade_makeup_counselor(db, ids, counselor_id)
    db.commit()
    db.close()
    return jsonify({'ok': True, 'updated': updated})


@app.route('/api/admin/counselor-roster', methods=['GET'])
@jwt_required()
def admin_counselor_roster():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    counselor_id = request.args.get('counselor_id', type=int)
    day = request.args.get('day', '').strip()
    if not counselor_id:
        return jsonify({'error': 'counselor_id required'}), 400
    db = get_db()
    # Same generate_series + override pattern as /api/counselor/activities so
    # admin sees the effective counselor for the upcoming week, including any
    # time-off reassignments. When ?day=Tuesday is set, the wd join filters
    # to the single matching date in the next 7 days.
    # Each activity_roster row matches at most one week_dates row (the two
    # branches of the JOIN's OR are mutually exclusive: a recurring row has
    # specific_date IS NULL, a one-shot row has it set), so no DISTINCT is
    # needed. DISTINCT was incompatible with the CASE expression in ORDER BY
    # — Postgres requires DISTINCT's ORDER BY terms to appear verbatim in the
    # select list.
    sql = """
        WITH week_dates AS (
          SELECT d::date AS d, to_char(d, 'FMDay') AS dow
            FROM generate_series(CURRENT_DATE,
                                 CURRENT_DATE + INTERVAL '6 days',
                                 INTERVAL '1 day') AS d
        )
        SELECT ar.id, ar.child_name, ar.school, ar.grade_group,
               ar.action, ar.day_of_week, ar.action_time,
               to_char(ar.specific_date, 'YYYY-MM-DD') AS specific_date,
               a.name AS activity_name, a.category,
               (ac.id IS NOT NULL) AS completed
        FROM week_dates wd
        JOIN activity_roster ar
          ON ((ar.specific_date IS NULL AND ar.day_of_week = wd.dow)
              OR ar.specific_date = wd.d)
        JOIN activities a ON a.id = ar.activity_id
        LEFT JOIN activity_roster_overrides o
               ON o.roster_entry_id = ar.id AND o.override_date = wd.d
        LEFT JOIN activity_completions ac
               ON ac.roster_entry_id = ar.id
              AND ac.completion_date = to_char(CURRENT_DATE, 'YYYY-MM-DD')
        WHERE COALESCE(o.assigned_counselor_id, ar.assigned_counselor_id) = %s
          AND (ar.specific_date IS NULL OR ar.specific_date >= CURRENT_DATE)
    """
    params = [counselor_id]
    if day:
        sql += " AND ar.day_of_week = %s AND wd.dow = %s"
        params.extend([day, day])
    sql += """
        ORDER BY CASE ar.day_of_week
                   WHEN 'Monday' THEN 1 WHEN 'Tuesday' THEN 2 WHEN 'Wednesday' THEN 3
                   WHEN 'Thursday' THEN 4 WHEN 'Friday' THEN 5 ELSE 6 END,
                 ar.action_time, ar.action, ar.child_name
    """
    rows = db.execute(sql, params).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/child-search', methods=['GET'])
@jwt_required()
def admin_child_search():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    q = request.args.get('q', '').strip()
    day = request.args.get('day', '').strip()
    if len(q) < 2:
        return jsonify([])
    # Escape LIKE wildcards in user input so % and _ are matched literally.
    q_safe = q.replace('\\', '\\\\').replace('%', '\\%').replace('_', '\\_')
    pattern = f'%{q_safe}%'
    db = get_db()
    sql = """
        SELECT ar.id, ar.child_name, ar.school, ar.grade_group,
               ar.action, ar.day_of_week, ar.action_time,
               to_char(ar.specific_date, 'YYYY-MM-DD') AS specific_date,
               a.name AS activity_name,
               u.name AS counselor_name,
               ar.assigned_counselor_id
        FROM activity_roster ar
        JOIN activities a ON a.id = ar.activity_id
        LEFT JOIN users u ON u.id = ar.assigned_counselor_id
        WHERE ar.child_name ILIKE %s ESCAPE '\\'
          AND (ar.specific_date IS NULL OR ar.specific_date >= CURRENT_DATE)
    """
    params = [pattern]
    if day:
        sql += " AND ar.day_of_week = %s"
        params.append(day)
    sql += """
        ORDER BY CASE ar.day_of_week
                   WHEN 'Monday' THEN 1 WHEN 'Tuesday' THEN 2 WHEN 'Wednesday' THEN 3
                   WHEN 'Thursday' THEN 4 WHEN 'Friday' THEN 5 ELSE 6 END,
                 ar.action_time, ar.action, ar.child_name
        LIMIT 100
    """
    rows = db.execute(sql, params).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


# ── Activity Schedules ──────────────────────────────────────────────────────

@app.route('/api/admin/activity-schedules', methods=['GET'])
@jwt_required()
def admin_get_schedules():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    rows = db.execute("SELECT * FROM activity_schedules ORDER BY activity_name_pattern, day_of_week").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/admin/activity-schedules', methods=['POST'])
@jwt_required()
def admin_add_schedule():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    d = request.json
    db = get_db()
    try:
        cur = db.execute(
            "INSERT INTO activity_schedules (activity_name_pattern, day_of_week, dropoff_time, pickup_time) VALUES (%s,%s,%s,%s) RETURNING *",
            (d['activity_name_pattern'], d['day_of_week'], d.get('dropoff_time'), d.get('pickup_time'))
        )
        row = cur.fetchone()
        db.close()
        return jsonify(dict(row)), 201
    except Exception as e:
        db.close()
        print(f"[ERROR] admin_add_schedule: {e}")
        return jsonify({'error': 'Failed to add schedule'}), 400

@app.route('/api/admin/activity-schedules/<int:sid>', methods=['PUT'])
@jwt_required()
def admin_update_schedule(sid):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    d = request.json
    db = get_db()
    db.execute("UPDATE activity_schedules SET dropoff_time=%s, pickup_time=%s WHERE id=%s",
               (d.get('dropoff_time'), d.get('pickup_time'), sid))
    db.commit()
    db.close()
    return jsonify({'ok': True})

@app.route('/api/admin/activity-schedules/<int:sid>', methods=['DELETE'])
@jwt_required()
def admin_delete_schedule(sid):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    db.execute("DELETE FROM activity_schedules WHERE id=%s", (sid,))
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── Manual Entries (e.g. Swimming) ─────────────────────────────────────────

@app.route('/api/admin/activity-roster/manual', methods=['GET'])
@jwt_required()
def admin_get_manual_entries():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    # Lazy GC of expired make-up class entries: removes one-day entries whose
    # date has already passed so they "disappear" the day after.
    db.execute("""
        DELETE FROM activity_roster
         WHERE source = 'manual'
           AND specific_date IS NOT NULL
           AND specific_date < CURRENT_DATE
    """)
    rows = db.execute("""
        SELECT ar.id, ar.child_name, ar.school, ar.grade_group, ar.option_type,
               ar.action, ar.day_of_week, ar.action_time, ar.assigned_counselor_id,
               to_char(ar.specific_date, 'YYYY-MM-DD') AS specific_date,
               a.name as activity_name, a.category,
               u.name as assigned_counselor_name
        FROM activity_roster ar
        JOIN activities a ON a.id = ar.activity_id
        LEFT JOIN users u ON u.id = ar.assigned_counselor_id
        WHERE ar.source = 'manual'
        ORDER BY ar.specific_date NULLS LAST, a.name, ar.day_of_week, ar.action_time, ar.child_name
    """).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/admin/activity-roster/manual', methods=['POST'])
@jwt_required()
def admin_add_manual_entry():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    d = request.json
    act_name = d.get('activity_name', '').strip()
    if not act_name:
        return jsonify({'error': 'activity_name required'}), 400

    # Optional make-up class (one-day-only) entry. When specific_date is set,
    # the entry's day_of_week is derived from the date and the row is auto-
    # cleaned up the day after on the next manual-entries fetch.
    specific_date = (d.get('specific_date') or '').strip() or None
    day_of_week = d.get('day_of_week')
    if specific_date:
        try:
            parsed = datetime.strptime(specific_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'specific_date must be YYYY-MM-DD'}), 400
        if parsed < now_for_org().date():
            return jsonify({'error': 'specific_date cannot be in the past'}), 400
        day_of_week = parsed.strftime('%A')
    if not day_of_week:
        return jsonify({'error': 'day_of_week required'}), 400

    db = get_db()
    # Get or create activity
    act = db.execute("SELECT id FROM activities WHERE name=%s", (act_name,)).fetchone()
    if act:
        activity_id = act['id']
    else:
        cur = db.execute("INSERT INTO activities (name, category, sub_type) VALUES (%s,%s,%s) RETURNING id",
                         (act_name, d.get('category', act_name), d.get('category', act_name)))
        activity_id = cur.fetchone()['id']

    cur = db.execute("""
        INSERT INTO activity_roster
        (activity_id, child_name, school, grade_group, option_type,
         action, day_of_week, action_time, assigned_counselor_id, source, specific_date)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'manual',%s) RETURNING id
    """, (activity_id, d['child_name'], d.get('school',''), d.get('grade_group',''),
          d.get('option_type','Full Service'), d['action'], day_of_week,
          d.get('action_time'), d.get('assigned_counselor_id'), specific_date))
    entry_id = cur.fetchone()['id']
    db.close()
    return jsonify({'id': entry_id}), 201

@app.route('/api/admin/activity-roster/manual/<int:entry_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_manual_entry(entry_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    db.execute("DELETE FROM activity_roster WHERE id=%s AND source='manual'", (entry_id,))
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── Admin: Parent-submitted make-up requests ───────────────────────────────

@app.route('/api/admin/makeup-requests', methods=['GET'])
@jwt_required()
def admin_list_makeup_requests():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    rows = db.execute("""
        SELECT ar.id, ar.child_name, ar.school, ar.action, ar.action_time,
               ar.day_of_week,
               to_char(ar.specific_date, 'YYYY-MM-DD') AS specific_date,
               ar.parent_will_pickup, ar.assigned_counselor_id,
               ar.counselor_locked, ar.admin_seen_at, ar.created_at,
               a.id AS activity_id, a.name AS activity_name,
               u.name AS counselor_name,
               p.id AS parent_id, p.name AS parent_name, p.email AS parent_email
          FROM activity_roster ar
          JOIN activities a ON a.id = ar.activity_id
          LEFT JOIN users u ON u.id = ar.assigned_counselor_id
          LEFT JOIN users p ON p.id = ar.requested_by_parent_id
         WHERE ar.source = 'parent_makeup'
           AND ar.specific_date >= CURRENT_DATE
         ORDER BY ar.specific_date, ar.action_time, ar.child_name
    """).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/makeup-requests/unseen-count', methods=['GET'])
@jwt_required()
def admin_makeup_unseen_count():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    row = db.execute("""
        SELECT COUNT(*) AS n FROM activity_roster
         WHERE source='parent_makeup'
           AND admin_seen_at IS NULL
           AND specific_date >= CURRENT_DATE
    """).fetchone()
    db.close()
    return jsonify({'count': row['n']})


@app.route('/api/admin/makeup-requests/mark-seen', methods=['POST'])
@jwt_required()
def admin_makeup_mark_seen():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    db.execute("""
        UPDATE activity_roster
           SET admin_seen_at = CURRENT_TIMESTAMP
         WHERE source='parent_makeup'
           AND admin_seen_at IS NULL
           AND specific_date >= CURRENT_DATE
    """)
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/admin/makeup-requests/<int:entry_id>/assign', methods=['PUT'])
@jwt_required()
def admin_makeup_override_counselor(entry_id):
    # counselor_locked=TRUE so the cascade in admin_assign_counselor / bulk-assign
    # won't overwrite a manually-overridden make-up.
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    body = request.json or {}
    counselor_id = body.get('counselor_id')  # None = unassign
    db = get_db()
    row = db.execute(
        "SELECT id FROM activity_roster WHERE id=%s AND source='parent_makeup'",
        (entry_id,)
    ).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Not found'}), 404
    db.execute("""
        UPDATE activity_roster
           SET assigned_counselor_id=%s, counselor_locked=TRUE
         WHERE id=%s
    """, (counselor_id, entry_id))
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/admin/makeup-requests/<int:entry_id>', methods=['DELETE'])
@jwt_required()
def admin_makeup_delete(entry_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    db.execute(
        "DELETE FROM activity_roster WHERE id=%s AND source='parent_makeup'",
        (entry_id,)
    )
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── Counselor: My Activities ────────────────────────────────────────────────

@app.route('/api/counselor/activities', methods=['GET'])
@jwt_required()
def counselor_get_activities():
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    day_filter = request.args.get('day', '')  # empty = today

    if not day_filter:
        day_filter = now_for_org().strftime('%A')

    db = get_db()
    # Filter rules for one-day make-up entries (specific_date IS NOT NULL):
    #   - "today" view: only show make-ups whose date is exactly today, so a
    #     future make-up scheduled on this same weekday doesn't leak in.
    #   - "week" view: hide expired make-ups (older than today); upcoming
    #     make-ups appear on their day_of_week column.
    # Apply per-date overrides from activity_roster_overrides so day-off
    # reassignments (admin-approved time_off) move entries between counselors
    # for a single date without mutating the base recurring rows.
    if day_filter == 'week':
        # generate_series gives us the 7 upcoming dates; the LEFT JOIN matches
        # each entry against its override for the matching weekday in that range.
        rows = db.execute("""
            WITH week_dates AS (
              SELECT d::date AS d, to_char(d, 'FMDay') AS dow
                FROM generate_series(CURRENT_DATE,
                                     CURRENT_DATE + INTERVAL '6 days',
                                     INTERVAL '1 day') AS d
            )
            SELECT DISTINCT ar.id, ar.child_name, ar.school, ar.grade_group,
                   ar.action, ar.day_of_week, ar.action_time,
                   to_char(ar.specific_date, 'YYYY-MM-DD') AS specific_date,
                   a.name as activity_name, a.category,
                   (ac.id IS NOT NULL) AS completed
            FROM week_dates wd
            JOIN activity_roster ar
              ON ((ar.specific_date IS NULL AND ar.day_of_week = wd.dow)
                  OR ar.specific_date = wd.d)
            JOIN activities a ON a.id = ar.activity_id
            LEFT JOIN activity_roster_overrides o
                   ON o.roster_entry_id = ar.id AND o.override_date = wd.d
            LEFT JOIN activity_completions ac
                   ON ac.roster_entry_id = ar.id
                  AND ac.completion_date = to_char(CURRENT_DATE, 'YYYY-MM-DD')
            WHERE COALESCE(o.assigned_counselor_id, ar.assigned_counselor_id) = %s
              AND (ar.specific_date IS NULL OR ar.specific_date >= CURRENT_DATE)
            ORDER BY ar.day_of_week, ar.action_time, ar.action, ar.child_name
        """, (user_id,)).fetchall()
    else:
        rows = db.execute("""
            SELECT ar.id, ar.child_name, ar.school, ar.grade_group,
                   ar.action, ar.day_of_week, ar.action_time,
                   to_char(ar.specific_date, 'YYYY-MM-DD') AS specific_date,
                   a.name as activity_name, a.category,
                   (ac.id IS NOT NULL) AS completed
            FROM activity_roster ar
            JOIN activities a ON a.id = ar.activity_id
            LEFT JOIN activity_roster_overrides o
                   ON o.roster_entry_id = ar.id AND o.override_date = CURRENT_DATE
            LEFT JOIN activity_completions ac
                   ON ac.roster_entry_id = ar.id
                  AND ac.completion_date = to_char(CURRENT_DATE, 'YYYY-MM-DD')
            WHERE COALESCE(o.assigned_counselor_id, ar.assigned_counselor_id) = %s
              AND ar.day_of_week = %s
              AND (ar.specific_date IS NULL OR ar.specific_date = CURRENT_DATE)
            ORDER BY ar.action_time, ar.action, ar.child_name
        """, (user_id, day_filter)).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/counselor/activities/<int:roster_id>/complete', methods=['POST'])
@jwt_required()
def counselor_toggle_complete(roster_id):
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()

    body = request.json or {}
    completed = bool(body.get('completed'))

    db = get_db()
    row = db.execute(
        "SELECT assigned_counselor_id FROM activity_roster WHERE id=%s",
        (roster_id,)
    ).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Not found'}), 404
    if str(row['assigned_counselor_id']) != str(user_id):
        db.close()
        return jsonify({'error': 'Not your entry'}), 403

    if completed:
        db.execute("""
            INSERT INTO activity_completions (roster_entry_id, completion_date, counselor_id)
            VALUES (%s, to_char(CURRENT_DATE, 'YYYY-MM-DD'), %s)
            ON CONFLICT (roster_entry_id, completion_date) DO NOTHING
        """, (roster_id, user_id))
    else:
        db.execute("""
            DELETE FROM activity_completions
            WHERE roster_entry_id=%s
              AND completion_date = to_char(CURRENT_DATE, 'YYYY-MM-DD')
        """, (roster_id,))
    db.commit()
    db.close()
    return jsonify({'ok': True, 'completed': completed})


# ─── COUNSELOR DAY-OFF REQUESTS ─────────────────────────────────────────────
# Counselor submits a day-off request; admin approves/rejects. On approval,
# the admin chooses substitute counselors per affected roster entry; we record
# those choices in activity_roster_overrides so the base recurring rows in
# activity_roster stay untouched.

def _serialize_time_off(row):
    return {
        'id': row['id'],
        'counselor_id': row['counselor_id'],
        'counselor_name': row.get('counselor_name'),
        'off_date': str(row['off_date']),
        'reason': row['reason'],
        'status': row['status'],
        'reviewed_by': row['reviewed_by'],
        'reviewed_by_name': row.get('reviewed_by_name'),
        'reviewed_at': row['reviewed_at'].isoformat() if row['reviewed_at'] else None,
        'review_note': row['review_note'],
        'created_at': row['created_at'].isoformat() if row['created_at'] else None,
        'affected_count': row.get('affected_count', 0),
    }


def _affected_roster_entries(db, counselor_id, off_date):
    """Roster entries currently effective for counselor_id on off_date.

    Looks at activity_roster (recurring + one-shot by specific_date) joined
    with activity_roster_overrides so an entry already overridden away from
    this counselor is NOT returned, and an entry overridden TO this counselor
    IS returned even when ar.assigned_counselor_id points elsewhere.
    """
    dow = off_date.strftime('%A')
    return db.execute("""
        SELECT ar.id, ar.activity_id, ar.child_name, ar.school, ar.grade_group,
               ar.action, ar.day_of_week, ar.action_time,
               a.name AS activity_name, a.category,
               COALESCE(o.assigned_counselor_id, ar.assigned_counselor_id) AS effective_counselor_id,
               o.id AS override_id
          FROM activity_roster ar
          JOIN activities a ON a.id = ar.activity_id
          LEFT JOIN activity_roster_overrides o
                 ON o.roster_entry_id = ar.id AND o.override_date = %s
         WHERE COALESCE(o.assigned_counselor_id, ar.assigned_counselor_id) = %s
           AND ((ar.specific_date IS NULL AND ar.day_of_week = %s)
                OR ar.specific_date = %s)
         ORDER BY ar.action_time, ar.action, ar.child_name
    """, (off_date, counselor_id, dow, off_date)).fetchall()


def _rank_substitute_candidates(db, entry, off_date, exclude_counselor_id):
    """Return ranked list of substitute counselors for one roster entry.

    Score = 5*exact_activity_matches (cap 25) + 2*category_matches (cap 10)
          + 3 (same-school bonus) − day_load
    Candidates: counselors at the same school (via counselor_schools)
    excluding the requester and anyone with an approved time-off on off_date.
    """
    if entry['school']:
        candidates = db.execute("""
            SELECT u.id, u.name
              FROM users u
              JOIN counselor_schools cs ON cs.counselor_id = u.id
              JOIN schools s ON s.id = cs.school_id
             WHERE u.role = 'counselor'
               AND u.id <> %s
               AND s.name = %s
               AND NOT EXISTS (
                     SELECT 1 FROM counselor_time_off cto
                      WHERE cto.counselor_id = u.id
                        AND cto.off_date = %s
                        AND cto.status = 'approved'
                   )
        """, (exclude_counselor_id, entry['school'], off_date)).fetchall()
    else:
        candidates = db.execute("""
            SELECT u.id, u.name FROM users u
             WHERE u.role = 'counselor'
               AND u.id <> %s
               AND NOT EXISTS (
                     SELECT 1 FROM counselor_time_off cto
                      WHERE cto.counselor_id = u.id
                        AND cto.off_date = %s
                        AND cto.status = 'approved'
                   )
        """, (exclude_counselor_id, off_date)).fetchall()

    if not candidates:
        return []
    cand_ids = [c['id'] for c in candidates]

    exact_rows = db.execute("""
        SELECT ac.counselor_id, COUNT(*) AS n
          FROM activity_completions ac
          JOIN activity_roster ar2 ON ar2.id = ac.roster_entry_id
         WHERE ac.counselor_id = ANY(%s)
           AND ar2.activity_id = %s
           AND ac.completion_date >= to_char(CURRENT_DATE - INTERVAL '90 days','YYYY-MM-DD')
         GROUP BY ac.counselor_id
    """, (cand_ids, entry['activity_id'])).fetchall()
    exact_map = {r['counselor_id']: r['n'] for r in exact_rows}

    category_map = {}
    if entry['category']:
        cat_rows = db.execute("""
            SELECT ac.counselor_id, COUNT(*) AS n
              FROM activity_completions ac
              JOIN activity_roster ar2 ON ar2.id = ac.roster_entry_id
              JOIN activities a2 ON a2.id = ar2.activity_id
             WHERE ac.counselor_id = ANY(%s)
               AND a2.category = %s
               AND ar2.activity_id <> %s
               AND ac.completion_date >= to_char(CURRENT_DATE - INTERVAL '90 days','YYYY-MM-DD')
             GROUP BY ac.counselor_id
        """, (cand_ids, entry['category'], entry['activity_id'])).fetchall()
        category_map = {r['counselor_id']: r['n'] for r in cat_rows}

    dow = off_date.strftime('%A')
    load_rows = db.execute("""
        SELECT eff_counselor AS counselor_id, COUNT(*) AS n,
               array_agg(action_time) AS times
          FROM (
            SELECT COALESCE(o.assigned_counselor_id, ar.assigned_counselor_id) AS eff_counselor,
                   ar.action_time
              FROM activity_roster ar
              LEFT JOIN activity_roster_overrides o
                     ON o.roster_entry_id = ar.id AND o.override_date = %s
             WHERE ((ar.specific_date IS NULL AND ar.day_of_week = %s)
                    OR ar.specific_date = %s)
          ) e
         WHERE eff_counselor = ANY(%s)
         GROUP BY eff_counselor
    """, (off_date, dow, off_date, cand_ids)).fetchall()
    load_map = {r['counselor_id']: (r['n'], r['times'] or []) for r in load_rows}

    results = []
    for c in candidates:
        cid = c['id']
        ex = exact_map.get(cid, 0)
        cat = category_map.get(cid, 0)
        day_n, times = load_map.get(cid, (0, []))
        time_conflict = bool(entry.get('action_time') and entry['action_time'] in (times or []))
        score = min(ex, 5) * 5 + min(cat, 5) * 2 + 3 - day_n
        results.append({
            'id': cid,
            'name': c['name'],
            'score': score,
            'exact_matches': ex,
            'category_matches': cat,
            'day_load': day_n,
            'time_conflict': time_conflict,
        })
    results.sort(key=lambda x: (-x['score'], (x['name'] or '').lower()))
    return results


@app.route('/api/counselor/time-off', methods=['POST'])
@jwt_required()
def counselor_create_time_off():
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    data = request.json or {}
    off_date = (data.get('off_date') or '').strip()
    reason = (data.get('reason') or '').strip() or None
    try:
        parse_date(off_date)
    except ValueError:
        return jsonify({'error': 'Invalid off_date (expected YYYY-MM-DD)'}), 400
    today_str = today_for_org()
    if off_date < today_str:
        return jsonify({'error': 'Cannot request a day off in the past'}), 400

    db = get_db()
    try:
        existing = db.execute("""
            SELECT id, status FROM counselor_time_off
             WHERE counselor_id = %s AND off_date = %s AND status IN ('pending','approved')
        """, (user_id, off_date)).fetchone()
        if existing:
            db.close()
            return jsonify({'error': 'You already have an active request for this date'}), 409

        row = db.execute("""
            INSERT INTO counselor_time_off (counselor_id, off_date, reason)
            VALUES (%s, %s, %s)
            RETURNING id, counselor_id, off_date, reason, status,
                      reviewed_by, reviewed_at, review_note, created_at
        """, (user_id, off_date, reason)).fetchone()
        db.commit()

        counselor = db.execute("SELECT name FROM users WHERE id=%s", (user_id,)).fetchone()
        admin_ids = [r['id'] for r in db.execute("SELECT id FROM users WHERE role='admin'").fetchall()]
    finally:
        db.close()

    counselor_name = counselor['name'] if counselor else 'A counselor'
    push_title = '🗓️ New day-off request'
    push_body = f'{counselor_name} requested day off on {off_date}.'
    for aid in admin_ids:
        send_push_to_user_async(aid, push_title, push_body,
                                url='/app/timeoff', tag='time-off-request')

    result = _serialize_time_off(dict(row))
    return jsonify(result), 201


@app.route('/api/counselor/time-off', methods=['GET'])
@jwt_required()
def counselor_list_time_off():
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()
    db = get_db()
    rows = db.execute("""
        SELECT cto.id, cto.counselor_id, cto.off_date, cto.reason, cto.status,
               cto.reviewed_by, cto.reviewed_at, cto.review_note, cto.created_at,
               r.name AS reviewed_by_name
          FROM counselor_time_off cto
          LEFT JOIN users r ON r.id = cto.reviewed_by
         WHERE cto.counselor_id = %s
         ORDER BY cto.off_date DESC, cto.created_at DESC
         LIMIT 100
    """, (user_id,)).fetchall()
    db.close()
    return jsonify([_serialize_time_off(dict(r)) for r in rows])


@app.route('/api/counselor/time-off/<int:request_id>', methods=['DELETE'])
@jwt_required()
def counselor_cancel_time_off(request_id):
    """Counselor cancels their own request.

    - Pending  → just flip status to 'cancelled'.
    - Approved → flip status, delete the overrides this request created
                 (the original counselor is restored automatically via the
                 COALESCE in roster reads), and notify admins + substitutes
                 so they aren't surprised.
    - Other states (rejected, cancelled) → 409.
    """
    claims = get_jwt()
    if claims.get('role') != 'counselor':
        return jsonify({'error': 'Unauthorized'}), 403
    user_id = get_jwt_identity()

    db = get_db_transaction()
    removed = []
    prior_status = None
    off_date = None
    counselor_name = ''
    try:
        row = db.execute("""
            UPDATE counselor_time_off
               SET status = 'cancelled'
             WHERE id = %s AND counselor_id = %s AND status IN ('pending','approved')
            RETURNING id, off_date
        """, (request_id, user_id)).fetchone()
        if not row:
            db.rollback(); db.close()
            return jsonify({'error': 'Cannot cancel (already decided or not yours)'}), 409
        off_date = row['off_date']

        # Discover whether this used to be approved by checking if any overrides
        # were tied to it. (RETURNING above can't easily give us the old status,
        # so we use the side-effect: only approved requests have overrides.)
        removed = db.execute("""
            DELETE FROM activity_roster_overrides
             WHERE source_time_off_id = %s
            RETURNING assigned_counselor_id
        """, (request_id,)).fetchall()
        prior_status = 'approved' if removed else 'pending'

        nm = db.execute("SELECT name FROM users WHERE id=%s", (user_id,)).fetchone()
        counselor_name = nm['name'] if nm else 'A counselor'
        admin_ids = [r['id'] for r in db.execute("SELECT id FROM users WHERE role='admin'").fetchall()]
        db.commit()
    except Exception as e:
        db.rollback(); db.close()
        print(f"[ERROR] counselor_cancel_time_off: {e}")
        return jsonify({'error': 'An internal error occurred'}), 500
    db.close()

    # Notify only when the cancellation undid an approved request — pending
    # cancels are routine and don't need to interrupt anyone.
    if prior_status == 'approved':
        off_date_str = str(off_date)
        admin_title = '↩️ Day off cancelled by counselor'
        admin_body = f'{counselor_name} canceled their approved day off on {off_date_str}; their activities were restored.'
        for aid in admin_ids:
            send_push_to_user_async(aid, admin_title, admin_body,
                                    url='/app/timeoff', tag='time-off-request')
        notified = set()
        for r in removed:
            sid = r['assigned_counselor_id']
            if sid and sid not in notified:
                notified.add(sid)
                send_push_to_user_async(
                    sid, '↩️ Assignment reverted',
                    f'Activities reassigned to you on {off_date_str} were reverted; {counselor_name} will work that day.',
                    url=f'/app/my-day?date={off_date_str}',
                    tag='time-off-reassign'
                )

    return jsonify({'ok': True, 'overrides_removed': len(removed)})


@app.route('/api/admin/time-off', methods=['GET'])
@jwt_required()
def admin_list_time_off():
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    status_filter = (request.args.get('status') or 'pending').lower()
    valid = {'pending', 'approved', 'rejected', 'cancelled', 'all'}
    if status_filter not in valid:
        return jsonify({'error': 'Invalid status filter'}), 400
    db = get_db()
    if status_filter == 'all':
        rows = db.execute("""
            SELECT cto.id, cto.counselor_id, cto.off_date, cto.reason, cto.status,
                   cto.reviewed_by, cto.reviewed_at, cto.review_note, cto.created_at,
                   c.name AS counselor_name, r.name AS reviewed_by_name
              FROM counselor_time_off cto
              JOIN users c ON c.id = cto.counselor_id
              LEFT JOIN users r ON r.id = cto.reviewed_by
             ORDER BY (cto.status = 'pending') DESC, cto.off_date DESC, cto.created_at DESC
             LIMIT 200
        """).fetchall()
    else:
        rows = db.execute("""
            SELECT cto.id, cto.counselor_id, cto.off_date, cto.reason, cto.status,
                   cto.reviewed_by, cto.reviewed_at, cto.review_note, cto.created_at,
                   c.name AS counselor_name, r.name AS reviewed_by_name
              FROM counselor_time_off cto
              JOIN users c ON c.id = cto.counselor_id
              LEFT JOIN users r ON r.id = cto.reviewed_by
             WHERE cto.status = %s
             ORDER BY cto.off_date ASC, cto.created_at ASC
             LIMIT 200
        """, (status_filter,)).fetchall()

    # Attach affected count per request for richer list display.
    out = []
    for r in rows:
        d = dict(r)
        try:
            affected = _affected_roster_entries(db, d['counselor_id'], d['off_date'])
            d['affected_count'] = len(affected)
        except Exception as e:
            print(f"[WARN] admin_list_time_off affected count failed: {e}")
            d['affected_count'] = 0
        out.append(_serialize_time_off(d))
    db.close()
    return jsonify(out)


@app.route('/api/admin/time-off/<int:request_id>/preview', methods=['GET'])
@jwt_required()
def admin_preview_time_off(request_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    req = db.execute("""
        SELECT cto.id, cto.counselor_id, cto.off_date, cto.reason, cto.status,
               c.name AS counselor_name
          FROM counselor_time_off cto
          JOIN users c ON c.id = cto.counselor_id
         WHERE cto.id = %s
    """, (request_id,)).fetchone()
    if not req:
        db.close()
        return jsonify({'error': 'Not found'}), 404

    entries = _affected_roster_entries(db, req['counselor_id'], req['off_date'])
    out_entries = []
    for e in entries:
        ed = dict(e)
        ranked = _rank_substitute_candidates(db, ed, req['off_date'], req['counselor_id'])
        out_entries.append({
            'roster_entry_id': ed['id'],
            'activity_id': ed['activity_id'],
            'activity_name': ed['activity_name'],
            'category': ed['category'],
            'school': ed['school'],
            'grade_group': ed['grade_group'],
            'action': ed['action'],
            'action_time': ed['action_time'],
            'child_name': ed['child_name'],
            'day_of_week': ed['day_of_week'],
            'has_existing_override': ed.get('override_id') is not None,
            'candidates': ranked[:5],
        })
    db.close()

    return jsonify({
        'request': {
            'id': req['id'],
            'counselor_id': req['counselor_id'],
            'counselor_name': req['counselor_name'],
            'off_date': str(req['off_date']),
            'reason': req['reason'],
            'status': req['status'],
        },
        'entries': out_entries,
    })


@app.route('/api/admin/time-off/<int:request_id>/approve', methods=['POST'])
@jwt_required()
def admin_approve_time_off(request_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    admin_id = get_jwt_identity()
    body = request.json or {}
    assignments = body.get('assignments') or []
    if not isinstance(assignments, list):
        return jsonify({'error': 'assignments must be a list'}), 400

    db = get_db_transaction()
    try:
        # Atomic state transition: only flip pending → approved exactly once.
        req = db.execute("""
            UPDATE counselor_time_off
               SET status = 'approved', reviewed_by = %s, reviewed_at = CURRENT_TIMESTAMP
             WHERE id = %s AND status = 'pending'
            RETURNING id, counselor_id, off_date
        """, (admin_id, request_id)).fetchone()
        if not req:
            db.rollback(); db.close()
            return jsonify({'error': 'Request is not pending'}), 409

        off_date = req['off_date']
        counselor_id = req['counselor_id']

        # Recompute affected entries server-side so the client cannot smuggle
        # in arbitrary roster_entry_ids that the counselor was not actually
        # assigned to on that date.
        affected = _affected_roster_entries(db, counselor_id, off_date)
        affected_ids = {row['id']: dict(row) for row in affected}

        # Index incoming assignments by roster_entry_id; missing entries get
        # an override with NULL counselor (= "uncovered, needs assignment").
        by_id = {}
        for a in assignments:
            try:
                rid = int(a.get('roster_entry_id'))
            except (TypeError, ValueError):
                continue
            if rid not in affected_ids:
                continue
            sub = a.get('assigned_counselor_id')
            sub_id = None
            if sub is not None:
                try:
                    sub_id = int(sub)
                except (TypeError, ValueError):
                    sub_id = None
            by_id[rid] = sub_id

        # Validate substitute counselors: must be counselors, must not be the
        # requester, must not themselves have an approved time-off that day.
        sub_ids = {v for v in by_id.values() if v is not None}
        if sub_ids:
            valid_subs = db.execute("""
                SELECT u.id FROM users u
                 WHERE u.id = ANY(%s) AND u.role = 'counselor' AND u.id <> %s
                   AND NOT EXISTS (
                         SELECT 1 FROM counselor_time_off cto
                          WHERE cto.counselor_id = u.id
                            AND cto.off_date = %s
                            AND cto.status = 'approved'
                            AND cto.id <> %s
                       )
            """, (list(sub_ids), counselor_id, off_date, request_id)).fetchall()
            valid_sub_ids = {r['id'] for r in valid_subs}
            for rid in list(by_id.keys()):
                if by_id[rid] is not None and by_id[rid] not in valid_sub_ids:
                    by_id[rid] = None  # silently null out invalid pick

        # Upsert one override per affected entry.
        created_for = {}  # counselor_id → list of entry summaries (for push)
        for rid, entry in affected_ids.items():
            sub_id = by_id.get(rid, None)
            db.execute("""
                INSERT INTO activity_roster_overrides
                    (roster_entry_id, override_date, assigned_counselor_id,
                     reason, source_time_off_id, created_by)
                VALUES (%s, %s, %s, 'time_off', %s, %s)
                ON CONFLICT (roster_entry_id, override_date) DO UPDATE
                   SET assigned_counselor_id = EXCLUDED.assigned_counselor_id,
                       reason = 'time_off',
                       source_time_off_id = EXCLUDED.source_time_off_id,
                       created_by = EXCLUDED.created_by
            """, (rid, off_date, sub_id, request_id, admin_id))
            if sub_id is not None:
                created_for.setdefault(sub_id, []).append(entry)

        counselor = db.execute("SELECT name FROM users WHERE id=%s", (counselor_id,)).fetchone()
        db.commit()
    except Exception as e:
        db.rollback(); db.close()
        print(f"[ERROR] admin_approve_time_off: {e}")
        return jsonify({'error': 'An internal error occurred'}), 500
    db.close()

    # Notify the requester and each substitute (best-effort, async).
    off_date_str = str(off_date)
    send_push_to_user_async(counselor_id, '✅ Day off approved',
                            f'Your request for {off_date_str} was approved.',
                            url='/app/schedule', tag='time-off-decision')
    for sub_id, entries_for_sub in created_for.items():
        if len(entries_for_sub) == 1:
            e = entries_for_sub[0]
            body_msg = f"{e['activity_name']} on {off_date_str} at {e.get('action_time') or 'TBD'}."
        else:
            body_msg = f"{len(entries_for_sub)} activities reassigned to you on {off_date_str}."
        send_push_to_user_async(sub_id, '📋 New activity assignment', body_msg,
                                url=f'/app/my-day?date={off_date_str}',
                                tag='time-off-reassign')

    return jsonify({
        'ok': True,
        'request_id': request_id,
        'overrides_created': len(affected_ids),
    })


@app.route('/api/admin/time-off/<int:request_id>/reject', methods=['POST'])
@jwt_required()
def admin_reject_time_off(request_id):
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    admin_id = get_jwt_identity()
    body = request.json or {}
    note = (body.get('review_note') or '').strip() or None

    db = get_db()
    try:
        row = db.execute("""
            UPDATE counselor_time_off
               SET status = 'rejected', reviewed_by = %s,
                   reviewed_at = CURRENT_TIMESTAMP, review_note = %s
             WHERE id = %s AND status = 'pending'
            RETURNING id, counselor_id, off_date
        """, (admin_id, note, request_id)).fetchone()
        db.commit()
    finally:
        db.close()
    if not row:
        return jsonify({'error': 'Request is not pending'}), 409

    send_push_to_user_async(
        row['counselor_id'], '❌ Day off rejected',
        f'Your request for {row["off_date"]} was rejected.' + (f' Note: {note}' if note else ''),
        url='/app/schedule', tag='time-off-decision'
    )
    return jsonify({'ok': True})


@app.route('/api/admin/time-off/<int:request_id>/revert', methods=['POST'])
@jwt_required()
def admin_revert_time_off(request_id):
    """Undo an approval: counselor will work that day after all.
    Marks the request as cancelled and deletes the overrides it created.
    Notifies the original counselor and any substitutes that were assigned.
    """
    if not require_admin():
        return jsonify({'error': 'Unauthorized'}), 403
    db = get_db_transaction()
    try:
        req = db.execute("""
            UPDATE counselor_time_off
               SET status = 'cancelled'
             WHERE id = %s AND status = 'approved'
            RETURNING id, counselor_id, off_date
        """, (request_id,)).fetchone()
        if not req:
            db.rollback(); db.close()
            return jsonify({'error': 'Request is not approved'}), 409

        removed = db.execute("""
            DELETE FROM activity_roster_overrides
             WHERE source_time_off_id = %s
            RETURNING assigned_counselor_id
        """, (request_id,)).fetchall()
        db.commit()
    except Exception as e:
        db.rollback(); db.close()
        print(f"[ERROR] admin_revert_time_off: {e}")
        return jsonify({'error': 'An internal error occurred'}), 500
    db.close()

    off_date_str = str(req['off_date'])
    send_push_to_user_async(req['counselor_id'], '↩️ Day off reverted',
                            f'Your approved day off on {off_date_str} was reverted by admin.',
                            url='/app/schedule', tag='time-off-decision')
    notified = set()
    for r in removed:
        sid = r['assigned_counselor_id']
        if sid and sid not in notified:
            notified.add(sid)
            send_push_to_user_async(sid, '↩️ Assignment reverted',
                                    f'Activities reassigned to you on {off_date_str} were reverted.',
                                    url=f'/app/my-day?date={off_date_str}',
                                    tag='time-off-reassign')
    return jsonify({'ok': True, 'overrides_removed': len(removed)})


# ─── STATIC FILE SERVING ────────────────────────────────────────────────────

@app.route('/sw.js')
def serve_sw():
    response = send_from_directory('../public', 'sw.js')
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Service-Worker-Allowed'] = '/'
    return response

@app.route('/manifest.json')
def serve_manifest():
    # The name and short_name are the one part of this file that is the JCC's,
    # not the platform's — everything else (icons, colours, scope) is fixed
    # regardless of which organization is asking. Read from the database
    # rather than hardcoding a name here, same reasoning as /api/public/branding.
    from flask import g
    with open(os.path.join(os.path.dirname(__file__), '..', 'public', 'manifest.json'), encoding='utf-8') as fh:
        manifest = json.load(fh)
    g.is_superadmin = True
    db = get_db()
    try:
        row = db.execute('SELECT name FROM organizations ORDER BY id LIMIT 1').fetchone()
    finally:
        db.close()
    if row and row['name']:
        manifest['name'] = row['name']
        manifest['short_name'] = row['name'][:12]
        manifest['description'] = f"{row['name']} attendance and pickup"
    response = jsonify(manifest)
    response.headers['Cache-Control'] = 'no-cache'
    return response

def _no_cache_html(response):
    # SPA HTML must never be cached aggressively, otherwise users keep
    # running stale JavaScript after a deploy.
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# / used to be a marketing landing page pitching the product to other
# organizations. Single-org now, so there is nothing to pitch — straight to
# sign-in. The old parent and counselor portals redirect into the SPA rather
# than lingering with stale branding and a second login. /admin stays
# reachable: roster upload, activity-roster management and the end-of-year
# tools have not been rebuilt yet.
@app.route('/')
def index():
    return redirect('/app/login')


@app.route('/parent')
@app.route('/parent/')
@app.route('/counselor')
@app.route('/counselor/')
def legacy_portal_redirect():
    return redirect('/app/', code=302)

# ─── KIKAR SPA ────────────────────────────────────────────────────────────
# The React app is built to public/app and owns every route under /app/*.
# Its hashed assets are served by Flask's static handler; everything else
# falls through to index.html so client-side routing survives a hard refresh.
@app.route('/app')
@app.route('/app/')
@app.route('/app/<path:subpath>')
def kikar_app(subpath=''):
    app_dir = os.path.join(os.path.dirname(__file__), '..', 'public', 'app')
    if subpath:
        candidate = os.path.join(app_dir, subpath)
        if os.path.isfile(candidate):
            return send_from_directory(app_dir, subpath)
    return _no_cache_html(send_from_directory(app_dir, 'index.html'))

@app.route('/admin')
@app.route('/admin/')
def admin_portal():
    return _no_cache_html(send_from_directory('../public/admin', 'index.html'))

@app.route('/reset-password')
@app.route('/reset-password/')
def reset_password_page():
    return _no_cache_html(send_from_directory('../public/reset-password', 'index.html'))

# Started here, at the bottom, rather than next to init_db(): the ticker calls
# _run_attendance_check, which has to exist by the time the first tick lands.
#
# Under gunicorn this runs once per worker and that is fine — the day's claim
# in scheduled_runs means only one of them does the work. Guarded so a test or
# a management script that imports app.py does not quietly acquire a thread
# that starts sending push notifications to real parents.
if scheduler.enabled() and not os.environ.get('KIKAR_NO_SCHEDULER'):
    scheduler.start(_run_attendance_check)

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    print(f"\n Kikar Afterschool server running at http://localhost:{port}")
    print(f"   Admin portal:     http://localhost:{port}/admin")
    print(f"   Parent portal:    http://localhost:{port}/parent")
    print(f"   Counselor portal: http://localhost:{port}/counselor\n")
    app.run(host='0.0.0.0', port=port, debug=False)
