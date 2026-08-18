"""Parse the "who may pick up this child" export into rows ready to import.

Pure, like roster_import.py: no database, no Flask, no request context. Takes
an .xlsx and returns dataclasses; server/app.py does the child-matching and
the write to `authorized_pickup_people`.

WHY A SEPARATE MODULE FROM roster_import.py
    That one reads the JCC's own master roster workbook — school, grade,
    dismissal time, one sheet per term. This reads a membership system's
    export of who may collect each child, which is flat (one row per child,
    headers on row one) and has nothing structurally in common. Different
    file, different shape, its own parser.

COLUMNS ARE RESOLVED BY HEADER NAME, NEVER POSITION
    Same reasoning as roster_import.py: a membership system reorders its
    export columns between runs more than the file changes underneath them.
    `Approved Person #N Full Name` / `... Relationship to Child` repeat for
    as many people as the export carries — the parser reads whatever N
    appears in the header rather than assuming a fixed count of seven, so a
    file with more or fewer approved-person columns still reads correctly.

WHAT IS DELIBERATELY NOT IMPORTED
    `Parent/Guardian #1` is the account already registered as the child's
    parent — already implicitly allowed to collect them (see
    /api/counselor/authorized-pickups), so importing it would just be a
    second row naming the same person. `Confirmed` is not read at all: it
    does not gate whether a row imports.
"""

from __future__ import annotations

import re
from dataclasses import dataclass

try:
    from server.roster_import import normalize_header
except ImportError:  # running from inside server/
    from roster_import import normalize_header


@dataclass(frozen=True)
class ApprovedPerson:
    name: str
    relationship: str | None


@dataclass(frozen=True)
class PickupRow:
    row_number: int  # 1-based spreadsheet row, so a fix points at the right line
    membership_no: str | None
    last_name: str
    first_name: str
    people: tuple[ApprovedPerson, ...]


_APPROVED_NAME = re.compile(r'^approved person (\d+) full name$')
_APPROVED_REL = re.compile(r'^approved person (\d+) relationship to child$')

_LAST_NAME_HEADERS = {'participants last name', 'participant s last name', 'last name'}
_FIRST_NAME_HEADERS = {'participants first name', 'participant s first name', 'first name'}
_MEMBERSHIP_HEADERS = {'membership', 'membership no', 'membership number'}
_GUARDIAN2_HEADERS = {'parent guardian 2 full name'}


def _resolve_headers(header_row) -> tuple[dict, dict, dict]:
    """Map column index to meaning. Unknown columns are simply ignored,
    the same tolerance roster_import.py extends to a workbook with extra
    columns nobody asked for.
    """
    fields: dict[str, int] = {}
    approved_name: dict[int, int] = {}
    approved_rel: dict[int, int] = {}
    for i, raw in enumerate(header_row or []):
        token = normalize_header(raw)
        if not token:
            continue
        if token in _LAST_NAME_HEADERS:
            fields['last_name'] = i
        elif token in _FIRST_NAME_HEADERS:
            fields['first_name'] = i
        elif token in _MEMBERSHIP_HEADERS:
            fields['membership_no'] = i
        elif token in _GUARDIAN2_HEADERS:
            fields['guardian2'] = i
        elif (m := _APPROVED_NAME.match(token)):
            approved_name[int(m.group(1))] = i
        elif (m := _APPROVED_REL.match(token)):
            approved_rel[int(m.group(1))] = i
    return fields, approved_name, approved_rel


def _cell(row, idx) -> str | None:
    if idx is None or idx >= len(row):
        return None
    value = row[idx]
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_sheet(rows: list[list]) -> list[PickupRow] | None:
    """None means this sheet does not look like the pickup export at all —
    the caller tries the next one, the same way roster_import.py skips a
    sheet with no School column rather than reading it as an empty roster.
    """
    if not rows:
        return None
    fields, approved_name, approved_rel = _resolve_headers(rows[0])
    if 'last_name' not in fields or 'first_name' not in fields:
        return None

    parsed = []
    for row_number, row in enumerate(rows[1:], start=2):
        last = _cell(row, fields['last_name'])
        first = _cell(row, fields['first_name'])
        if not last and not first:
            continue  # a blank spacer row, not a child

        people = []
        seen = set()  # (name.casefold()) — a person named twice on one row
                       # (Guardian #2 re-listed as Approved Person #1, say)
                       # should not become two rows racing on the same
                       # (child_id, name) unique key.
        for n in sorted(approved_name):
            name = _cell(row, approved_name[n])
            if not name or name.casefold() in seen:
                continue
            seen.add(name.casefold())
            people.append(ApprovedPerson(name=name, relationship=_cell(row, approved_rel.get(n))))

        guardian2 = _cell(row, fields.get('guardian2'))
        if guardian2 and guardian2.casefold() not in seen:
            seen.add(guardian2.casefold())
            people.append(ApprovedPerson(name=guardian2, relationship='Parent/Guardian'))

        parsed.append(PickupRow(
            row_number=row_number,
            membership_no=_cell(row, fields.get('membership_no')),
            last_name=last or '',
            first_name=first or '',
            people=tuple(people),
        ))
    return parsed


def parse_pickup_workbook(source) -> tuple[list[PickupRow], list[str]]:
    """`source` is a path or a file-like object (the same thing
    roster_import.parse_workbook accepts).

    Returns (rows, problems). A non-empty `problems` means no sheet in the
    file could be read as this export at all — nothing was found named
    "Participant's Last Name" and "Participant's First Name" anywhere — and
    the caller should refuse the whole upload rather than import zero rows
    silently.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        for title in workbook.sheetnames:
            rows = [list(row) for row in workbook[title].iter_rows(values_only=True)]
            parsed = _parse_sheet(rows)
            if parsed is not None:
                return parsed, []
    finally:
        workbook.close()

    return [], [
        "Could not find \"Participant's Last Name\" and \"Participant's First "
        "Name\" columns on any sheet."
    ]
