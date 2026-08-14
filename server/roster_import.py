"""Parse the JCC's master roster workbook into rows the importer can stage.

This module is deliberately pure: it takes cell values and returns dataclasses.
No database, no Flask, no request context, no logging. Every decision it makes
is reproducible from its inputs, which is what lets the import preview an admin
approves be byte-identically what commits — the endpoint parses once, stores the
result in `roster_import_rows`, and applies that stored parse later.

WHY A SECOND IMPORTER
    `/api/admin/upload-roster` (server/app.py:2522) reads Excel by position —
    row[0] is the child, row[5] the contact, row[11] the course option. That
    layout has nothing to do with this JCC's file, where row[0] is a billing
    annotation and row[5] is a class column that is empty until signups run.
    Pointed at this roster it would import the billing notes as children's
    names. It is left untouched because other organizations depend on it.

COLUMNS ARE RESOLVED BY HEADER NAME, NEVER BY INDEX
    Two independent reasons, both observed in the real file:

      • The spec's column map is off by one from column U onward — it places
        `complete` at AC when the file has it at AD. A positional reader would
        take `complete` as Contact #1 and lose Email #2 entirely, and it would
        not fail: it would import every child with one plausible-looking bad
        contact.
      • The `No Longer` sheet is a different shape: 34 columns instead of 36,
        starting at `Notes` instead of `?? for parents`, with the day columns
        headed `M T W R F` instead of `Mon Tue Wed Thur Fri`, and
        `Pick up (pg 3)` where the main sheet says `(pg 2)`.

    By name both problems disappear, and so does the next reordering.

WHICH SHEETS ARE READ
    Any sheet with a School column and a name column, decided by looking at it.
    This was once decided by the sheet's *title*, needing the word `roster` in
    it, with a fallback to "read everything" when nothing matched. Heather's
    workbook names its sheets `Spring 2025-26` and `No Longer`: the second
    matched, so the filter came back non-empty, so the fallback never fired, and
    the whole roster was dropped without a word. A hundred and ten children
    became zero, and the screen said the file was empty rather than that a sheet
    had been skipped. The title still says *withdrawn or active*; it no longer
    says whether to look.

CLASSES CARRY THEIR HOURS
    The `X - Class` columns hold `Soccer 3:15p-4p`, not `Soccer`.
    parse_class_cell splits the two, so class_sessions gets start and end times
    instead of a name with a time buried in it — which is what R2 needs to route
    a child when their class ends, and what lets `HW Club` at 3 and `HW Club` at
    4 on the same Tuesday stay the two different classes they are. Cells that
    are not classes at all (`NO CLASS`, `BUS ONLY - EHS`) keep the child's day
    and create none.

WHAT COUNTS AS A CHILD
    The roster is two blocks stacked in one sheet. Below the real roster there
    is a scratch area — names typed without a comma, no date of birth, no
    contact — that is where every duplicated name in the file lives. It is
    separated by exactly one thing: whether the `School` column has a value.

    But `School` alone is not enough. Two rows in the upper block are legend
    lines that happen to have something in `School` (`REG` / `BUS ONLY - EHS`
    and `NUMBERS` / `Bball 3:15p-4p`). They are excluded by a second clause:
    no comma in the name *and* no contact *and* no date of birth. `????` as a
    school passes that test and is kept, because it is a real child whose
    school needs resolving — not a junk row.

NOTHING IS DROPPED SILENTLY
    Every row the parser refuses, every cell it cannot read and every value it
    normalises away leaves a Notice carrying the source row number, so the
    review screen can say "row 143 has no dismissal time" about the file
    Heather still has open. Notices at level 'error' mean the row must not
    commit as-is; 'warning' means it commits with something worth a look.

PRIVACY
    Notices and parsed rows carry children's names, dates of birth and contact
    details — that is the point of a review screen — so they belong in the
    response to an authenticated admin of that organization and nowhere else.
    Do not log them, and do not put them in an error message that escapes the
    tenant.
"""

import re
from dataclasses import dataclass, field
from datetime import date, datetime

WEEKDAYS = ('Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday')

# The four dismissal hours the whole feature is built on. The spec is explicit
# that there are exactly these; anything else in a day column is a notice.
DISMISSAL_HOURS = (3, 4, 5, 6)


# ─────────────────────────────────────────────────────────────────────────────
# Notices
# ─────────────────────────────────────────────────────────────────────────────

ERROR = 'error'
WARNING = 'warning'
INFO = 'info'


@dataclass(frozen=True)
class Notice:
    """Something the reviewer should see about one row, or about the sheet.

    `row` is the spreadsheet's own 1-based row number, not an index into
    anything the parser built, so it can be read straight off the screen and
    typed into Excel's go-to box.
    """

    code: str
    message: str
    level: str = WARNING
    sheet: str | None = None
    row: int | None = None
    column: str | None = None
    value: str | None = None

    def as_dict(self) -> dict:
        return {
            'code': self.code,
            'message': self.message,
            'level': self.level,
            'sheet': self.sheet,
            'row': self.row,
            'column': self.column,
            'value': self.value,
        }


# ─────────────────────────────────────────────────────────────────────────────
# Header resolution
# ─────────────────────────────────────────────────────────────────────────────

def normalize_header(value) -> str:
    """Reduce a header cell to a comparable token.

    Apostrophes are deleted rather than replaced so that `Child's Name` becomes
    `childs name` instead of `child s name`, and a parenthesised page reference
    is removed outright so that `Pick up (pg 2)` and `Pick up (pg 3)` — the same
    form, numbered differently on the two sheets — land on the same token.
    """
    text = '' if value is None else str(value)
    text = text.replace('’', "'").replace("'", '')
    text = re.sub(r'\(\s*pg\.?\s*\d+\s*\)', ' ', text, flags=re.I)
    text = re.sub(r'[^0-9a-z]+', ' ', text.lower())
    return text.strip()


def column_letter(index: int) -> str:
    """0-based column index to its spreadsheet letter, for talking to a human."""
    letters = ''
    index += 1
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord('A') + remainder) + letters
    return letters


# Plain per-child fields. The spare spellings are there because the withdrawn
# sheet and the weekly attendance workbook name the same things differently.
FIELD_HEADERS = {
    'for parents': 'roster_flag',
    'school': 'school',
    'childs name': 'name',
    'child name': 'name',
    'students name': 'name',
    'student name': 'name',
    'student': 'name',
    'name': 'name',
    'grade': 'grade',
    'bus': 'bus',
    'sex': 'sex',
    'gender': 'sex',
    'notes': 'notes',
    'dob': 'dob',
    'date of birth': 'dob',
    'birthdate': 'dob',
    'allergies': 'allergies',
    'allergy': 'allergies',
}

# Day columns hold the dismissal hour, not an enrolled/not-enrolled flag. The
# single letters are the withdrawn sheet's spelling; R is Thursday.
DAY_HEADERS = {
    'mon': 'Monday', 'monday': 'Monday', 'm': 'Monday',
    'tue': 'Tuesday', 'tues': 'Tuesday', 'tuesday': 'Tuesday', 't': 'Tuesday',
    'wed': 'Wednesday', 'weds': 'Wednesday', 'wednesday': 'Wednesday', 'w': 'Wednesday',
    'thu': 'Thursday', 'thur': 'Thursday', 'thurs': 'Thursday',
    'thursday': 'Thursday', 'r': 'Thursday',
    'fri': 'Friday', 'friday': 'Friday', 'f': 'Friday',
}

CLASS_HEADERS = {}
for _token, _day in DAY_HEADERS.items():
    CLASS_HEADERS[f'{_token} class'] = _day
del _token, _day

# The paperwork checklist. Values map onto child_compliance.item, whose CHECK
# constraint is the authority — a header that lands outside this map becomes an
# unknown column rather than a new de-facto item.
COMPLIANCE_HEADERS = {
    'regist rcvd': 'registration',
    'registration received': 'registration',
    'registration': 'registration',
    'regist fee chged': 'registration_fee',
    'registration fee': 'registration_fee',
    'member y n sent': 'membership',
    'member': 'membership',
    'membership': 'membership',
    'pick up': 'pickup_form',
    'pickup': 'pickup_form',
    'first aid': 'first_aid',
    'medication': 'medication',
    'photo': 'photo',
    'physical': 'physical',
    'imm': 'immunization',
    'immunization': 'immunization',
    'immunizations': 'immunization',
    'complete': 'complete',
    'b day card': 'bday_card',
    'bday card': 'bday_card',
    'birthday card': 'bday_card',
}


@dataclass
class ContactColumns:
    priority: int
    name: int | None = None
    phone: int | None = None
    email: int | None = None


@dataclass
class ColumnMap:
    """Where each thing lives in this particular sheet."""

    fields: dict = field(default_factory=dict)       # 'school'   -> column index
    days: dict = field(default_factory=dict)         # 'Monday'   -> column index
    classes: dict = field(default_factory=dict)      # 'Monday'   -> column index
    compliance: dict = field(default_factory=dict)   # 'first_aid'-> column index
    contacts: list = field(default_factory=list)     # [ContactColumns, …]
    unknown: list = field(default_factory=list)      # [(letter, header text), …]

    def get(self, cells, key):
        """Read a mapped field out of a row, tolerating a short row."""
        index = self.fields.get(key)
        if index is None or index >= len(cells):
            return None
        return cells[index]

    def at(self, cells, index):
        if index is None or index >= len(cells):
            return None
        return cells[index]


def resolve_columns(header_cells, sheet=None):
    """Build a ColumnMap from a header row. Returns (ColumnMap, [Notice])."""
    columns = ColumnMap()
    notices = []
    seen = {}
    open_contact = None

    for index, raw in enumerate(header_cells):
        token = normalize_header(raw)
        if not token:
            continue
        letter = column_letter(index)

        # A repeated header is real: the withdrawn sheet carries `Notes` in both
        # column A and column Q. Deciding by position would be a guess, so the
        # first occurrence wins — deterministic — and the duplicate is reported
        # so a human can rename the column if the guess was the wrong one.
        if token in seen:
            notices.append(Notice(
                'duplicate_header',
                f"Column {letter} repeats the header {raw!r} already found in "
                f"column {seen[token]}; column {seen[token]} was used.",
                INFO, sheet=sheet, column=letter, value=str(raw),
            ))
            continue
        seen[token] = letter

        if token in DAY_HEADERS:
            columns.days[DAY_HEADERS[token]] = index
        elif token in CLASS_HEADERS:
            columns.classes[CLASS_HEADERS[token]] = index
        elif token in FIELD_HEADERS:
            columns.fields[FIELD_HEADERS[token]] = index
        elif token in COMPLIANCE_HEADERS:
            columns.compliance[COMPLIANCE_HEADERS[token]] = index
        elif token.startswith('contact'):
            open_contact = ContactColumns(
                priority=_trailing_number(token, len(columns.contacts) + 1),
                name=index,
            )
            columns.contacts.append(open_contact)
        elif token.startswith('phone') or token.startswith('email'):
            # Attach to the contact block currently open. A phone or email
            # before any `Contact` header opens an implicit block rather than
            # being dropped — losing a parent's phone number silently is the
            # one failure mode this file cannot afford.
            if open_contact is None:
                open_contact = ContactColumns(priority=len(columns.contacts) + 1)
                columns.contacts.append(open_contact)
                notices.append(Notice(
                    'contact_column_orphan',
                    f"Column {letter} ({raw!r}) came before any Contact column; "
                    f"read as contact {open_contact.priority}.",
                    WARNING, sheet=sheet, column=letter, value=str(raw),
                ))
            setattr(open_contact, 'phone' if token.startswith('phone') else 'email', index)
        else:
            columns.unknown.append((letter, str(raw)))

    for key in ('school', 'name'):
        if key not in columns.fields:
            notices.append(Notice(
                'header_missing',
                f"No {key!r} column found in the header row.",
                ERROR, sheet=sheet,
            ))

    # child_contacts.priority is CHECK (1, 2). A third block would fail on
    # insert, so it is refused here where the message can name the column.
    for extra in columns.contacts[2:]:
        notices.append(Notice(
            'contact_extra',
            f"Contact {extra.priority} is beyond the two the app stores; ignored.",
            WARNING, sheet=sheet,
            column=column_letter(extra.name) if extra.name is not None else None,
        ))
    columns.contacts = columns.contacts[:2]
    for position, block in enumerate(columns.contacts, start=1):
        block.priority = position

    if columns.unknown:
        listed = ', '.join(f"{letter} ({text!r})" for letter, text in columns.unknown)
        notices.append(Notice(
            'header_unknown',
            f"{len(columns.unknown)} column(s) not recognised and left unread: {listed}.",
            INFO, sheet=sheet,
        ))

    return columns, notices


def _trailing_number(token: str, default: int) -> int:
    match = re.search(r'(\d+)\s*$', token)
    return int(match.group(1)) if match else default


def find_header_row(rows, limit: int = 10):
    """Locate the header row. Returns (index, cells) or (None, None).

    The header is row 1 in every sheet seen so far, but a title row above it is
    a normal thing for someone to add, and the failure mode — importing the
    header as a child — is silent.
    """
    for index, cells in enumerate(rows[:limit]):
        tokens = {normalize_header(cell) for cell in cells}
        has_school = 'school' in tokens
        has_name = bool(tokens & {t for t, f in FIELD_HEADERS.items() if f == 'name'})
        if has_school and has_name:
            return index, cells
    return None, None


# ─────────────────────────────────────────────────────────────────────────────
# Cell readers
#
# Each returns (value, code): the reading, and None or a notice code the caller
# turns into a Notice with row and column context. Keeping the context out of
# here is what makes them testable one cell at a time.
# ─────────────────────────────────────────────────────────────────────────────

# What Heather types in a continuation row to mean "same as the line above".
# Every column of the second row of a two-row child carries one. Kept to marks
# that cannot be a value in their own right — `same` is not in here, because a
# Notes column saying "same" is a sentence about the child, not a ditto.
_DITTO = {'--', '---', '–', '—', '"', 'ditto'}


def cell_text(value):
    """Cell to a stripped string, or None if it is empty.

    Excel hands back 5 as int on one row and 5.0 as float on the next; both
    have to read as "5" or the day columns split into two vocabularies.

    A ditto mark reads as empty, which is what it means. Read literally it is a
    value that parses as nothing, and the roster's 44 continuation rows carry
    one in every column they repeat — around 750 notices saying a date of birth
    of `--` is unreadable, on a screen whose whole job is to show the handful
    that matter. Empty is also the reading merge_children already wants: it
    fills a blank from a later row and leaves a filled one alone, so the first
    row keeps winning exactly as before.
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, datetime):
        # Excel stores a date as a midnight timestamp. Rendering the time back
        # puts `2026-05-14T00:00:00` in front of a reviewer who wrote 5/14.
        value = value.date() if value.time() == datetime.min.time() else value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value).strip()
    if text.lower() in _DITTO:
        return None
    return text or None


def parse_person_name(value):
    """Split `"Last, First"`. Returns ((last, first), code).

    A name without a comma is kept whole as the last name rather than guessed
    at: in this file every such row in the roster block turned out to be a
    legend line, and the ones that are not are worth a human's eye.
    """
    text = cell_text(value)
    if not text:
        return (None, None), 'child_name_missing'
    text = re.sub(r'\s+', ' ', text)
    if ',' in text:
        last, _, first = text.partition(',')
        last, first = last.strip(), first.strip()
        if last and first:
            return (last, first), None
        return (last or first or text, None), 'name_incomplete'
    return (text, None), 'name_not_last_first'


def display_name(last, first) -> str:
    """How the child's name reads on screen, everywhere outside this file."""
    return ' '.join(part for part in (first, last) if part)


# `Brown - Drop Off` is Brown with the parent driving; `D.O.` on the attendance
# sheets is drop-off with no school named at all.
_DROPOFF_SUFFIX = re.compile(r'\s*[-–]\s*drop\s*off\s*$', re.I)
_DROPOFF_ONLY = {'do', 'd o', 'drop off', 'dropoff'}


def parse_school(value):
    """Split the School cell into (token, arrival_mode) plus a code."""
    text = cell_text(value)
    if not text:
        return (None, None), 'school_missing'

    arrival = None
    stripped = _DROPOFF_SUFFIX.sub('', text).strip()
    if stripped != text:
        arrival = 'dropoff'
        text = stripped

    flattened = re.sub(r'[^0-9a-z]+', ' ', text.lower()).strip()
    if flattened in _DROPOFF_ONLY:
        # Drop-off is how the child arrives, not where they come from.
        return (None, 'dropoff'), None
    if not flattened:
        # `????` and friends: a real child whose school still has to be picked.
        return (text, arrival), 'school_placeholder'
    return (text, arrival), None


def normalize_school_token(value) -> str | None:
    """The form school_aliases.alias_norm stores, so lookups are case-blind.

    A token made entirely of punctuation — `????` is one of the ten values in
    the real School column — normalises to the empty string, and returning None
    there would drop it out of the list of schools to resolve. It is exactly
    the one that most needs resolving, so it keeps its own spelling instead.
    """
    text = cell_text(value)
    if not text:
        return None
    flattened = re.sub(r'[^0-9a-z]+', ' ', text.lower()).strip()
    return flattened or text.lower()


_GRADE_WORDS = {
    'k': ('K', 0), 'kg': ('K', 0), 'kinder': ('K', 0), 'kindergarten': ('K', 0),
}


def parse_grade(value):
    """Returns ((label, number), code). The label keeps the file's spelling.

    K and 0 both become 0 so that a care rule written as "grades 0–1" catches
    kindergarten however it was typed, while grade_label keeps which one was
    written — the collapse stays reversible if it turns out this JCC means two
    different things by them.
    """
    text = cell_text(value)
    if not text:
        return (None, None), 'grade_missing'
    token = text.strip()
    flattened = re.sub(r'[^0-9a-z]+', '', token.lower())
    if flattened in _GRADE_WORDS:
        label, number = _GRADE_WORDS[flattened]
        return (token, number), None
    if re.fullmatch(r'\d{1,2}', flattened):
        number = int(flattened)
        if 0 <= number <= 12:
            return (token, number), None
    if re.fullmatch(r'(\d{1,2})(st|nd|rd|th)', flattened):
        return (token, int(re.match(r'\d+', flattened).group())), None
    return (token, None), 'grade_unrecognized'


# Billing notes Heather appends to a cell — to a class name and to a dismissal
# hour alike. Anchored to the end and spelled out rather than "strip any
# parenthesis", because parentheses are also part of real class names:
# `Swim T (Mini White)`, `(Mini Green)` and `(Mini Black)` are three different
# swim levels, and `KIM (Hip Hop)` and `KIM (Dance Combo)` are two different
# classes. A general strip merges them into one.
_ANNOTATION = re.compile(
    r'\s*\(\s*(?:pro-?\s*rate|no\s+refund|as\s+of|remove|removed|cancell?ed)\b[^)]*\)\s*$',
    re.I,
)
# `REMOVE A&C 3:15p-4p (No refund)` — the same note, written in front instead.
_LEADING_ANNOTATION = re.compile(r'^\s*(?:remove|removed|cancell?ed)\s+', re.I)


def strip_annotation(text: str) -> str:
    """Take Heather's billing note off a cell, leaving what the cell is about.

    The note records what the office has to do about a child — pro-rate them,
    charge a change fee — and never changes what the cell says: a class with
    `(pro-rate as of 5/1)` on it is still that class, and `5 (remove as of 5/1)`
    is still a five o'clock dismissal. Kept in the notice, never in the value.
    """
    return _LEADING_ANNOTATION.sub('', _ANNOTATION.sub('', text or '').strip()).strip()


_DISMISSAL = re.compile(r'^(\d{1,2})(?::0{2})?\s*(?:p|pm)?$', re.I)


def parse_dismissal(value):
    """Read a day cell. Returns ((hour, starred), code).

    The cell is not a yes/no: it is the hour the child leaves. `**` is Heather's
    mark for a child with more than one class that day — kept as a hint, never
    as data, because the app derives it by counting class enrollments and the
    file itself misses two of them on Thursday.

    The same billing notes that appear on class names appear here —
    `5 (remove as of 5/1 - charge change fee)` — and read literally they cost
    the child their dismissal hour, which is the one field the daily board
    cannot work without.
    """
    text = cell_text(value)
    if not text:
        return (None, False), None
    starred = '*' in text
    token = text.replace('*', '').strip()
    if not token:
        return (None, starred), 'dismissal_unreadable'

    code = None
    stripped = strip_annotation(token)
    if stripped != token:
        token, code = stripped, 'dismissal_annotation_stripped'
    if not token:
        return (None, starred), 'dismissal_unreadable'

    match = _DISMISSAL.match(token)
    if not match:
        return (None, starred), 'dismissal_unreadable'
    hour = int(match.group(1))
    if hour in DISMISSAL_HOURS:
        return (hour, starred), code
    return (None, starred), 'dismissal_out_of_range'


# ─────────────────────────────────────────────────────────────────────────────
# Class cells
#
# The `M - Class` … `F - Class` columns. What they hold is a class name with its
# hours written into it — `Soccer 3:15p-4p`, `HW Club 4p-5p`.
#
# sql/35 and roster_staging both state the opposite: that the roster names the
# classes but never gives their hours, so class_sessions.start_time and
# .end_time have to be typed in afterwards. That was true of the file available
# when they were written and is not true of Heather's. Reading the hours matters
# beyond tidiness — R2 routes a child by comparing the class end against their
# dismissal time, so a class with no end time has no computable destination.
# ─────────────────────────────────────────────────────────────────────────────

# Not classes. `NO CLASS` says the child comes that day and takes nothing;
# `BUS ONLY - EHS` says they are only on the bus. Both belong on the day's
# registration, which is kept — the child is still in care for the block — but
# neither is a thing to put in the catalog, give a room, or staff a counselor to.
_NOT_A_CLASS = re.compile(r'^(no\s*class|bus\s*only\b.*|none|n\s*/?\s*a)$', re.I)

# `3:15p-4p`, `4p-5p`, `3p-3:30p`. The start's am/pm is often left off and
# inherited from the end, which is the only reading that makes `4p-5p` mean
# 4pm to 5pm rather than 4am.
_CLASS_TIME_RANGE = re.compile(
    r'(\d{1,2})(?::(\d{2}))?\s*([ap])?\s*[-–—]\s*(\d{1,2})(?::(\d{2}))?\s*([ap])\b',
    re.I,
)


def _clock(hour, minute, meridiem):
    """`(4, None, 'p')` → `'16:00'`. Returns None if the hour is not a clock."""
    hour, minute = int(hour), int(minute or 0)
    if not (1 <= hour <= 12) or minute > 59:
        return None
    if meridiem and meridiem.lower() == 'p' and hour != 12:
        hour += 12
    elif meridiem and meridiem.lower() == 'a' and hour == 12:
        hour = 0
    return f'{hour:02d}:{minute:02d}'


@dataclass
class ParsedClass:
    """One class on one weekday. `raw` is what the cell actually said.

    The times are strings (`'15:15'`) and not `datetime.time`, because this
    whole parse is serialised into roster_import_rows as JSON and read back at
    commit; a shape that survives that round trip unchanged is what makes the
    preview an admin approves the same thing that commits.
    """

    name: str
    start_time: str | None = None
    end_time: str | None = None
    raw: str | None = None

    @property
    def key(self):
        return (self.name.casefold(), self.start_time)

    def as_dict(self) -> dict:
        return {'name': self.name, 'start_time': self.start_time,
                'end_time': self.end_time, 'raw': self.raw}


def parse_class_cell(value):
    """Read one `X - Class` cell. Returns ((ParsedClass or None), code).

    Order matters: the annotation comes off before the times are read, because
    `Soccer 3:15p-4p (pro-rate as of 5/1)` ends in a date that would otherwise
    look like the tail of a range.
    """
    raw = cell_text(value)
    if not raw:
        return None, None
    if _NOT_A_CLASS.match(raw):
        return None, 'class_not_a_class'

    code = None
    text = re.sub(r'\s+', ' ', raw).strip()
    cleaned = strip_annotation(text)
    if cleaned != text:
        code = 'class_annotation_stripped'
    if not cleaned:
        return None, 'class_not_a_class'

    match = _CLASS_TIME_RANGE.search(cleaned)
    if not match:
        # A name with no hours in it is the shape sql/35 was written for, and
        # it stays legal: the class is created and an admin fills the times in.
        return ParsedClass(name=cleaned, raw=raw), code or 'class_time_missing'

    end_meridiem = match.group(6)
    start = _clock(match.group(1), match.group(2), match.group(3) or end_meridiem)
    end = _clock(match.group(4), match.group(5), end_meridiem)
    name = cleaned[:match.start()].strip(' -–—·,')
    if not name or start is None or end is None or start >= end:
        # An unreadable clock or a backwards range is a typo in the file. Keep
        # the whole cell as the name rather than inventing hours from it: a
        # wrong end time routes a child to a room their parent is not standing in.
        return ParsedClass(name=cleaned, raw=raw), 'class_time_unreadable'

    return ParsedClass(name=name, start_time=start, end_time=end, raw=raw), code


_TRUE_WORDS = {'x', 'yes', 'y', 'true', '1', 'bus'}
_FALSE_WORDS = {'no', 'n', 'false', '0', 'none', 'na', 'n a'}


def parse_bus(value):
    """`x` rides the JCC bus, `NO` does not, `??` is nobody's answer yet."""
    text = cell_text(value)
    if not text:
        return None, None
    flattened = re.sub(r'[^0-9a-z]+', ' ', text.lower()).strip()
    if flattened in _TRUE_WORDS:
        return True, None
    if flattened in _FALSE_WORDS:
        return False, None
    return None, 'bus_unreadable'


def parse_sex(value):
    text = cell_text(value)
    if not text:
        return None, None
    letter = text.strip()[:1].upper()
    if letter in ('M', 'F'):
        return letter, None
    return None, 'sex_unreadable'


# Heather types dates as dates most of the time and as text the rest of it.
_DATE_FORMATS = (
    '%m/%d/%Y', '%m/%d/%y', '%m-%d-%Y', '%m-%d-%y',
    '%Y-%m-%d', '%Y/%m/%d', '%d %b %Y', '%b %d %Y', '%B %d %Y',
)


def parse_date_cell(value):
    """Returns (date, code). Two-digit years are read by strptime's own rule."""
    if value is None:
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    text = cell_text(value)
    if not text:
        return None, None
    cleaned = text.replace(',', ' ').replace('.', ' ')
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(cleaned, fmt).date(), None
        except ValueError:
            continue
    return None, 'date_unreadable'


def parse_dob(value):
    parsed, code = parse_date_cell(value)
    return parsed, ('dob_unreadable' if code else None)


_COMPLIANCE_DONE = {'x', 'xx', 'y', 'yes', 'done', 'complete', 'completed', 'ok', 'rcvd', 'received'}
_COMPLIANCE_NA = {'na', 'n a', 'none', 'n'}
_COMPLIANCE_MISSING = {'missing', 'no', 'needed', 'need'}
_COMPLIANCE_WAIVED = {'waived', 'waive', 'exempt', 'exempted'}
_COMPLIANCE_PENDING_WORDS = ('request', 'sent', 'pending', 'wait', 'updat', 'follow')


def parse_compliance(value, item=None):
    """Read one checklist cell. Returns ((status, recorded_on, raw), code).

    An empty cell returns None: on a paper form "not filled in" and "we do not
    know" are the same thing, and neither of them is `MISSING`, which is a word
    the staff writes on purpose. Whatever the cell said is kept verbatim in
    raw_value, so a status of 'other' is never a dead end.

    `item` matters for exactly one column. `Medication (pg 4)` is not a tick
    box: where a form is on file the staff write what the medication is, and
    two of the real roster's rows name one. Read as an unrecognised value those
    two children come back on a "who still owes paperwork" list they do not
    belong on, so a named medication counts as on file — and still raises a
    notice, because a drug name arriving through the compliance column is worth
    a human seeing.
    """
    raw = cell_text(value)
    if raw is None:
        return None, None

    recorded_on, date_code = parse_date_cell(value)
    if recorded_on is not None:
        return ('done', recorded_on, raw), None

    flattened = re.sub(r'[^0-9a-z]+', ' ', raw.lower()).strip()
    if flattened in _COMPLIANCE_DONE:
        return ('done', None, raw), None
    if flattened in _COMPLIANCE_WAIVED:
        return ('waived', None, raw), None
    if flattened in _COMPLIANCE_NA:
        return ('na', None, raw), None
    if flattened in _COMPLIANCE_MISSING:
        return ('missing', None, raw), None
    if any(word in flattened for word in _COMPLIANCE_PENDING_WORDS):
        return ('pending', None, raw), None
    if item == 'medication':
        return ('done', None, raw), 'medication_named'
    return ('other', None, raw), 'compliance_unreadable'


# ─────────────────────────────────────────────────────────────────────────────
# Parsed shapes
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ParsedContact:
    priority: int
    name: str | None = None
    phone: str | None = None
    email: str | None = None

    def as_dict(self) -> dict:
        return {'priority': self.priority, 'name': self.name,
                'phone': self.phone, 'email': self.email}


@dataclass
class ParsedCompliance:
    item: str
    status: str
    recorded_on: date | None = None
    raw_value: str | None = None

    def as_dict(self) -> dict:
        return {
            'item': self.item,
            'status': self.status,
            'recorded_on': self.recorded_on.isoformat() if self.recorded_on else None,
            'raw_value': self.raw_value,
        }


@dataclass
class ParsedRegistration:
    day_of_week: str
    dismissal_time: int | None = None
    starred: bool = False
    classes: list = field(default_factory=list)

    def as_dict(self) -> dict:
        return {'day_of_week': self.day_of_week, 'dismissal_time': self.dismissal_time,
                'starred': self.starred,
                'classes': [c.as_dict() for c in self.classes]}


@dataclass
class ParsedChild:
    """One child, possibly assembled from several rows (rule R4)."""

    source_sheet: str
    source_rows: list = field(default_factory=list)
    status: str = 'active'

    last_name: str | None = None
    first_name: str | None = None
    name: str | None = None

    # Three forms of the same cell, each with a job: raw is what Heather typed
    # and is what the review screen quotes back at her; token is that with the
    # arrival suffix taken off, which is the school's actual name; norm is the
    # case- and punctuation-blind form that school_aliases is keyed on.
    school_raw: str | None = None
    school_token: str | None = None
    school_norm: str | None = None
    arrival_mode: str | None = None

    grade_label: str | None = None
    grade_num: int | None = None
    dob: date | None = None
    sex: str | None = None
    allergies: str | None = None
    bus_rider: bool | None = None
    notes: str | None = None
    roster_flag: str | None = None

    registrations: list = field(default_factory=list)
    contacts: list = field(default_factory=list)
    compliance: list = field(default_factory=list)
    notices: list = field(default_factory=list)

    @property
    def key(self):
        """Dedupe key: name + school + grade, per rule R4.

        Checked against every sheet in both workbooks — the count of distinct
        triples equals the count of distinct names on all of them, so no two
        children collide on it.
        """
        return (
            (self.last_name or '').casefold(),
            (self.first_name or '').casefold(),
            self.school_norm or '',
            self.grade_num,
        )

    @property
    def enrolled_days(self):
        return [r.day_of_week for r in self.registrations if r.dismissal_time is not None]

    @property
    def blocked(self) -> bool:
        return any(n.level == ERROR for n in self.notices)

    def as_dict(self) -> dict:
        return {
            'source_sheet': self.source_sheet,
            'source_rows': list(self.source_rows),
            'status': self.status,
            'last_name': self.last_name,
            'first_name': self.first_name,
            'name': self.name,
            'school_raw': self.school_raw,
            'school_token': self.school_token,
            'school_norm': self.school_norm,
            'arrival_mode': self.arrival_mode,
            'grade_label': self.grade_label,
            'grade_num': self.grade_num,
            'dob': self.dob.isoformat() if self.dob else None,
            'sex': self.sex,
            'allergies': self.allergies,
            'bus_rider': self.bus_rider,
            'notes': self.notes,
            'roster_flag': self.roster_flag,
            'registrations': [r.as_dict() for r in self.registrations],
            'contacts': [c.as_dict() for c in self.contacts],
            'compliance': [c.as_dict() for c in self.compliance],
            'notices': [n.as_dict() for n in self.notices],
        }


@dataclass
class SkippedRow:
    source_sheet: str
    source_row: int
    reason: str
    label: str | None = None

    def as_dict(self) -> dict:
        return {'source_sheet': self.source_sheet, 'source_row': self.source_row,
                'reason': self.reason, 'label': self.label}


@dataclass
class SheetParse:
    sheet: str
    status: str = 'active'
    columns: ColumnMap | None = None
    children: list = field(default_factory=list)
    skipped: list = field(default_factory=list)
    notices: list = field(default_factory=list)
    totals: dict = field(default_factory=dict)

    def as_dict(self) -> dict:
        return {
            'sheet': self.sheet,
            'status': self.status,
            'children': [c.as_dict() for c in self.children],
            'skipped': [s.as_dict() for s in self.skipped],
            'notices': [n.as_dict() for n in self.notices],
            'totals': dict(self.totals),
        }


@dataclass
class WorkbookParse:
    sheets: list = field(default_factory=list)
    notices: list = field(default_factory=list)
    totals: dict = field(default_factory=dict)

    @property
    def children(self):
        return [child for sheet in self.sheets for child in sheet.children]

    def as_dict(self) -> dict:
        return {
            'sheets': [s.as_dict() for s in self.sheets],
            'notices': [n.as_dict() for n in self.notices],
            'totals': dict(self.totals),
        }


# ─────────────────────────────────────────────────────────────────────────────
# Row classification
# ─────────────────────────────────────────────────────────────────────────────

ROW_CHILD = 'child'
ROW_BLANK = 'blank'
ROW_DRAFT = 'draft'
ROW_LEGEND = 'legend'


def classify_row(cells, columns: ColumnMap) -> str:
    """Decide what a row is. See WHAT COUNTS AS A CHILD in the module docstring."""
    if not any(cell_text(cell) is not None for cell in cells):
        return ROW_BLANK

    school = cell_text(columns.get(cells, 'school'))
    if school is None:
        # The scratch block at the foot of the sheet: names typed the other way
        # round, no grade, no contact. Everything the file repeats lives here.
        return ROW_DRAFT

    name = cell_text(columns.get(cells, 'name'))
    has_comma = bool(name and ',' in name)
    has_dob = columns.at(cells, columns.fields.get('dob')) is not None
    has_contact = any(
        cell_text(columns.at(cells, block.name)) is not None
        or cell_text(columns.at(cells, block.phone)) is not None
        or cell_text(columns.at(cells, block.email)) is not None
        for block in columns.contacts
    )
    if not has_comma and not has_dob and not has_contact:
        # A legend line that happens to have something in School: `REG` /
        # `BUS ONLY - EHS`, `NUMBERS` / `Bball 3:15p-4p`.
        return ROW_LEGEND
    return ROW_CHILD


# ─────────────────────────────────────────────────────────────────────────────
# Row → child
# ─────────────────────────────────────────────────────────────────────────────

_MESSAGES = {
    'child_name_missing': "No child name.",
    'name_incomplete': "Name has a comma but only one half.",
    'name_not_last_first': "Name is not in \"Last, First\" form; read as a last name.",
    'school_missing': "No school.",
    'school_placeholder': "School is a placeholder and has to be resolved by hand.",
    'grade_missing': "No grade.",
    'grade_unrecognized': "Grade not recognised; stored as written, with no number to match care rules against.",
    'dismissal_unreadable': "Not one of the four dismissal hours; read as not enrolled that day.",
    'dismissal_out_of_range': "Dismissal hour outside 3–6; read as not enrolled that day.",
    'bus_unreadable': "Bus column not recognised.",
    'sex_unreadable': "Sex column not recognised.",
    'dob_unreadable': "Date of birth not readable.",
    'compliance_unreadable': "Checklist value not recognised; kept verbatim as 'other'.",
    'medication_named': "The medication column names a medication rather than "
                        "marking the form; read as on file, with the name kept.",
    'contact_missing_name': "Contact has no name; the email or phone was used instead.",
    'class_not_a_class': "Not a class, so no class was created; the child is "
                         "still enrolled that day.",
    'class_annotation_stripped': "A billing note was taken off the class name; "
                                 "the child is in the class itself.",
    'class_time_missing': "The class name carries no hours; its start and end "
                          "have to be filled in on the Classes screen.",
    'class_time_unreadable': "The hours in the class name could not be read; "
                             "the whole cell was kept as the class name.",
    'dismissal_annotation_stripped': "A billing note was taken off the dismissal "
                                     "hour; the hour itself was read.",
}

_LEVELS = {
    'child_name_missing': ERROR,
    'school_missing': ERROR,
    'school_placeholder': WARNING,
    'compliance_unreadable': INFO,
    'medication_named': INFO,
    'name_not_last_first': WARNING,
    # `NO CLASS` is Heather stating a fact, not a problem to look at.
    'class_not_a_class': INFO,
    'class_annotation_stripped': INFO,
    # A class named without its hours is the shape sql/35 was written for and
    # is still perfectly legal — the admin fills them in on the Classes screen.
    # Worth recording, not worth a warning on every class of an org whose
    # roster spells them that way.
    'class_time_missing': INFO,
    'class_time_unreadable': WARNING,
    'dismissal_annotation_stripped': INFO,
}


def _notice(code, sheet, row, column=None, value=None, message=None) -> Notice:
    return Notice(
        code,
        message or _MESSAGES.get(code, code),
        _LEVELS.get(code, WARNING),
        sheet=sheet, row=row, column=column,
        value=None if value is None else str(value),
    )


def parse_child_row(cells, columns: ColumnMap, sheet: str, row_number: int,
                    status: str = 'active') -> ParsedChild:
    """Turn one spreadsheet row into a ParsedChild. Pure; no lookups."""
    child = ParsedChild(source_sheet=sheet, source_rows=[row_number], status=status)

    def col(key):
        index = columns.fields.get(key)
        return column_letter(index) if index is not None else None

    (last, first), code = parse_person_name(columns.get(cells, 'name'))
    child.last_name, child.first_name = last, first
    child.name = display_name(last, first)
    if code:
        child.notices.append(_notice(code, sheet, row_number, col('name'),
                                     cell_text(columns.get(cells, 'name'))))

    school_raw = cell_text(columns.get(cells, 'school'))
    (token, arrival), code = parse_school(columns.get(cells, 'school'))
    child.school_raw = school_raw
    child.school_token = token
    child.school_norm = normalize_school_token(token)
    child.arrival_mode = arrival
    if code:
        child.notices.append(_notice(code, sheet, row_number, col('school'), school_raw))

    (grade_label, grade_num), code = parse_grade(columns.get(cells, 'grade'))
    child.grade_label, child.grade_num = grade_label, grade_num
    if code:
        child.notices.append(_notice(code, sheet, row_number, col('grade'), grade_label))

    bus, code = parse_bus(columns.get(cells, 'bus'))
    child.bus_rider = bus
    if code:
        child.notices.append(_notice(code, sheet, row_number, col('bus'),
                                     cell_text(columns.get(cells, 'bus'))))
    # Arrival mode is only claimed where the file states it: the "- Drop Off"
    # suffix, or an `x` in the Bus column. Bus = NO says the child is not on the
    # bus, which is not the same as saying a parent drives them.
    if child.arrival_mode is None and bus is True:
        child.arrival_mode = 'bus'

    sex, code = parse_sex(columns.get(cells, 'sex'))
    child.sex = sex
    if code:
        child.notices.append(_notice(code, sheet, row_number, col('sex'),
                                     cell_text(columns.get(cells, 'sex'))))

    dob, code = parse_dob(columns.get(cells, 'dob'))
    child.dob = dob
    if code:
        child.notices.append(_notice(code, sheet, row_number, col('dob'),
                                     cell_text(columns.get(cells, 'dob'))))

    child.allergies = cell_text(columns.get(cells, 'allergies'))
    child.notes = cell_text(columns.get(cells, 'notes'))
    child.roster_flag = cell_text(columns.get(cells, 'roster_flag'))

    for day in WEEKDAYS:
        day_index = columns.days.get(day)
        class_index = columns.classes.get(day)
        if day_index is None and class_index is None:
            continue
        (hour, starred), code = parse_dismissal(columns.at(cells, day_index))
        if code:
            child.notices.append(_notice(
                code, sheet, row_number, column_letter(day_index),
                cell_text(columns.at(cells, day_index)),
            ))
        raw_class = cell_text(columns.at(cells, class_index))
        parsed_class, class_code = parse_class_cell(columns.at(cells, class_index))
        if class_code:
            child.notices.append(_notice(
                class_code, sheet, row_number,
                column_letter(class_index) if class_index is not None else None,
                raw_class,
            ))
        if hour is None and not raw_class and not starred:
            continue
        child.registrations.append(ParsedRegistration(
            day_of_week=day, dismissal_time=hour, starred=starred,
            classes=[parsed_class] if parsed_class else [],
        ))

    for block in columns.contacts:
        name = cell_text(columns.at(cells, block.name))
        phone = cell_text(columns.at(cells, block.phone))
        email = cell_text(columns.at(cells, block.email))
        if not any((name, phone, email)):
            continue
        if not name:
            # child_contacts.name is NOT NULL, and a phone number with no label
            # is still the number to call. Fall back rather than drop it.
            name = email or phone
            child.notices.append(_notice(
                'contact_missing_name', sheet, row_number,
                column_letter(block.name) if block.name is not None else None,
            ))
        child.contacts.append(ParsedContact(block.priority, name, phone, email))

    for item, index in columns.compliance.items():
        parsed, code = parse_compliance(columns.at(cells, index), item=item)
        if parsed is None:
            continue
        status_value, recorded_on, raw = parsed
        child.compliance.append(ParsedCompliance(item, status_value, recorded_on, raw))
        if code:
            child.notices.append(_notice(code, sheet, row_number,
                                         column_letter(index), raw))

    return child


# ─────────────────────────────────────────────────────────────────────────────
# Merging duplicated rows (rule R4)
# ─────────────────────────────────────────────────────────────────────────────

_SCALARS = ('dob', 'sex', 'allergies', 'notes', 'roster_flag', 'bus_rider',
            'grade_label', 'grade_num', 'school_raw', 'school_token',
            'arrival_mode')


def merge_children(existing: ParsedChild, incoming: ParsedChild) -> ParsedChild:
    """Fold a repeated row into the child already built from an earlier one.

    Heather duplicates a child's row once per class, so the second row is the
    same child with a different class in it. The first row wins every scalar —
    it is the one she filled in — and later rows only contribute what the first
    left blank, plus their class enrollments. A disagreement is reported rather
    than resolved, because the two rows saying different dismissal times for
    the same day is a mistake in the file, not something to pick a winner for.
    """
    existing.source_rows.extend(incoming.source_rows)
    row_number = incoming.source_rows[0] if incoming.source_rows else None

    for attribute in _SCALARS:
        if getattr(existing, attribute) is None:
            setattr(existing, attribute, getattr(incoming, attribute))

    by_day = {r.day_of_week: r for r in existing.registrations}
    for registration in incoming.registrations:
        current = by_day.get(registration.day_of_week)
        if current is None:
            existing.registrations.append(registration)
            by_day[registration.day_of_week] = registration
            continue
        if current.dismissal_time is None:
            current.dismissal_time = registration.dismissal_time
        elif (registration.dismissal_time is not None
              and registration.dismissal_time != current.dismissal_time):
            existing.notices.append(Notice(
                'dismissal_conflict',
                f"{registration.day_of_week} is {current.dismissal_time} on row "
                f"{existing.source_rows[0]} and {registration.dismissal_time} on "
                f"row {row_number}; {current.dismissal_time} was used.",
                WARNING, sheet=existing.source_sheet, row=row_number,
            ))
        current.starred = current.starred or registration.starred
        # Deduped on (name, start time): the same class at two different hours
        # on one weekday is two classes — `HW Club 3p-4p` and `HW Club 4p-5p`
        # both run on Tuesday — while the same class repeated across two merged
        # rows is one enrolment.
        known_classes = {c.key for c in current.classes}
        for parsed_class in registration.classes:
            if parsed_class.key not in known_classes:
                current.classes.append(parsed_class)
                known_classes.add(parsed_class.key)

    known = {c.priority for c in existing.contacts}
    for contact in incoming.contacts:
        if contact.priority not in known:
            existing.contacts.append(contact)
            known.add(contact.priority)

    items = {c.item for c in existing.compliance}
    for entry in incoming.compliance:
        if entry.item not in items:
            existing.compliance.append(entry)
            items.add(entry.item)

    existing.notices.extend(incoming.notices)
    return existing


# ─────────────────────────────────────────────────────────────────────────────
# Sheets and workbooks
# ─────────────────────────────────────────────────────────────────────────────

def parse_sheet(rows, sheet: str, status: str = 'active') -> SheetParse:
    """Parse one sheet given as a list of rows of raw cell values."""
    result = SheetParse(sheet=sheet, status=status)
    rows = list(rows)

    header_index, header_cells = find_header_row(rows)
    if header_cells is None:
        result.notices.append(Notice(
            'header_not_found',
            "No header row with a School and a name column in the first 10 rows.",
            ERROR, sheet=sheet,
        ))
        result.totals = _sheet_totals(result, rows_read=0)
        return result

    columns, header_notices = resolve_columns(header_cells, sheet=sheet)
    result.columns = columns
    result.notices.extend(header_notices)
    if any(n.level == ERROR for n in header_notices):
        result.totals = _sheet_totals(result, rows_read=0)
        return result

    by_key = {}
    rows_read = 0
    merged = 0

    for offset, cells in enumerate(rows[header_index + 1:]):
        row_number = header_index + offset + 2  # 1-based, and past the header
        kind = classify_row(cells, columns)
        if kind == ROW_BLANK:
            continue
        rows_read += 1
        if kind in (ROW_DRAFT, ROW_LEGEND):
            result.skipped.append(SkippedRow(
                sheet, row_number, kind,
                cell_text(columns.get(cells, 'name')),
            ))
            continue

        child = parse_child_row(cells, columns, sheet, row_number, status=status)
        existing = by_key.get(child.key)
        if existing is None:
            by_key[child.key] = child
            result.children.append(child)
        else:
            merge_children(existing, child)
            merged += 1

    result.totals = _sheet_totals(result, rows_read=rows_read, merged=merged)
    return result


def _class_totals(children) -> tuple:
    """(distinct classes, total enrolments) over a list of children.

    A class is counted once per (weekday, name, start time) — the same key
    class_sessions is stored under — so `HW Club` at 3 and at 4 on a Tuesday
    count as the two classes they are.
    """
    sessions = set()
    enrolments = 0
    for child in children:
        for registration in child.registrations:
            for parsed_class in registration.classes:
                sessions.add((registration.day_of_week, *parsed_class.key))
                enrolments += 1
    return len(sessions), enrolments


def _sheet_totals(result: SheetParse, rows_read: int, merged: int = 0) -> dict:
    children = result.children
    schools = sorted({c.school_norm for c in children if c.school_norm})
    classes, class_enrolments = _class_totals(children)
    return {
        'rows_read': rows_read,
        'children': len(children),
        'rows_merged': merged,
        'draft_rows': sum(1 for s in result.skipped if s.reason == ROW_DRAFT),
        'legend_rows': sum(1 for s in result.skipped if s.reason == ROW_LEGEND),
        'blocked': sum(1 for c in children if c.blocked),
        # The 60 children with no day enrolled are not a bug and not an error:
        # they are the ones the billing column flags as REMOVE or as waiting on
        # paperwork. They import, and no derived daily view shows them. The
        # number has to be on the review screen or phase 2 looks broken.
        'children_without_days': sum(1 for c in children if not c.enrolled_days),
        'children_without_contacts': sum(1 for c in children if not c.contacts),
        'children_without_dob': sum(1 for c in children if c.dob is None),
        'school_tokens': schools,
        'warnings': sum(1 for c in children for n in c.notices if n.level == WARNING),
        # The activities. Counted here so the review screen can say the file's
        # classes were read before anything is committed — the half of the
        # spreadsheet that used to be parsed and thrown away.
        'classes': classes,
        'class_enrolments': class_enrolments,
    }


# Which sheet is which. Resolved by pattern because the workbook's own titles
# are hand-typed and inconsistent.
_WITHDRAWN_TITLE = re.compile(r'no\s*longer|withdraw|inactive|dropped', re.I)


def sheet_status(title: str) -> str:
    """`No Longer` is children who left — archived, never deleted.

    Only the *status* is read off the title, never whether the sheet is read at
    all — see is_roster_sheet.
    """
    return 'withdrawn' if _WITHDRAWN_TITLE.search(title or '') else 'active'


def is_roster_sheet(rows) -> bool:
    """True if this sheet is a roster, decided by what is in it.

    This used to be decided by the sheet's *title*, against a regex that wanted
    the word `roster` in it, with a fallback to "read everything" when no title
    matched. Heather's real workbook names its sheets for the season — `Spring
    2025-26` and `No Longer` — and that combination is the one the fallback
    cannot save: `No Longer` matched, so the filtered list came back non-empty,
    so the fallback never fired and the entire roster was dropped in silence.
    The import screen then showed ten withdrawn children and nothing else,
    which reads as "the file is unreadable" rather than "one sheet was skipped".

    A sheet that carries a School column and a name column is a roster whatever
    it is called, and next season's `Fall 2026-27` needs no code change. Scratch
    tabs stay out because they have no such header row.
    """
    return find_header_row(rows)[1] is not None


def parse_rows_by_sheet(sheets) -> WorkbookParse:
    """Parse `{title: rows}` (or a list of pairs) into a WorkbookParse.

    This is the whole importer minus openpyxl, which is what the tests drive.
    """
    if isinstance(sheets, dict):
        sheets = list(sheets.items())

    result = WorkbookParse()
    for title, rows in sheets:
        result.sheets.append(parse_sheet(rows, title, status=sheet_status(title)))

    for sheet in result.sheets:
        result.notices.extend(sheet.notices)

    active = [c for c in result.children if c.status == 'active']
    withdrawn = [c for c in result.children if c.status == 'withdrawn']
    # Counted over the active children only: a class nobody still enrolled
    # attends is not a class this season, and the withdrawn sheet keeps last
    # term's for the record.
    classes, class_enrolments = _class_totals(active)
    result.totals = {
        'sheets': len(result.sheets),
        'children': len(active),
        'withdrawn': len(withdrawn),
        'blocked': sum(1 for c in result.children if c.blocked),
        'children_without_days': sum(1 for c in active if not c.enrolled_days),
        'children_without_contacts': sum(1 for c in active if not c.contacts),
        'school_tokens': sorted({c.school_norm for c in result.children if c.school_norm}),
        'rows_read': sum(s.totals.get('rows_read', 0) for s in result.sheets),
        'rows_merged': sum(s.totals.get('rows_merged', 0) for s in result.sheets),
        'draft_rows': sum(s.totals.get('draft_rows', 0) for s in result.sheets),
        'legend_rows': sum(s.totals.get('legend_rows', 0) for s in result.sheets),
        'classes': classes,
        'class_enrolments': class_enrolments,
    }
    return result


def parse_workbook(source, sheet_titles=None) -> WorkbookParse:
    """Parse an .xlsx roster. `source` is a path or a file-like object.

    The only function here that touches openpyxl; everything above it works on
    lists of values, so the parsing rules are testable without a binary fixture
    and without a real child's name on disk.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        titles = sheet_titles or workbook.sheetnames
        sheets = []
        skipped = []
        for title in titles:
            worksheet = workbook[title]
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            # A workbook of scratch tabs would otherwise import as children.
            if is_roster_sheet(rows):
                sheets.append((title, rows))
            else:
                skipped.append(title)
    finally:
        workbook.close()

    result = parse_rows_by_sheet(sheets)
    # Never let a sheet disappear without saying so: silently dropping one is
    # the bug this whole function was rewritten for.
    for title in skipped:
        result.notices.append(Notice(
            'sheet_skipped',
            f"Sheet {title!r} has no School and name columns, so it was not "
            f"read as a roster.",
            INFO, sheet=title,
        ))
    return result
